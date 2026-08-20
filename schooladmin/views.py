from decimal import Decimal, InvalidOperation
import csv
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from school.decorators import feature_required, perm_required, FeatureRequiredMixin, PermRequiredMixin
from school.features import has_feature, has_perm
from school.nepali_utils import to_bs_display
import nepali_datetime
from handle.models import AttendingClassification, member
from django.views.generic.list import ListView
from handle.models import AttendanceRecord
import datetime
from datetime import datetime as dt
from handle.models import Classification, PaySlip
from handle.models import Device
from management.models import CustomUser, LeaveReport, LeaveType, WifiBased
from handle.forms import PayrollAdjustmentForm, PayrollPolicyForm, PaySlipForm, ProbationReviewForm
from django.urls import reverse
from django.contrib import messages
from .forms import OrgFormSchool
from handle.models import (
    PayrollAdjustment,
    PayrollPolicy,
    ProvidentFundRecord,
    ProbationReview,
    SocialSecurityFundRecord,
    Staff,
    member,
    Classification,
    Branch,
    Section,
    Course,
    CourseAttendance,
    AttendanceGap,
    FinancialTransaction,
    TransactionCategory,
    StockItem,
    StockMovement,
    StockCategory,
    Subject,
    ExamTerm,
    ResultRecord,
    Event,
    EventStockUsage,
    Complaint,
    ResignationRecord,
    StaffDocument,
    CustomerBill,
    CustomerBillPayment,
    CustomerContract,
    CustomerProposal,
    CustomerDocument,
    Book,
    BookIssue,
    LibraryCategory,
    LibraryAuthor,
    LibraryPublisher,
    LibraryRack,
    LibraryShelf,
    LibrarySettings,
    Account,
    JournalEntry,
    JournalEntryLine,
    AcademicYear,
    Faculty,
    Semester,
    CourseTeacherAssignment,
    SubjectTeacherAssignment,
    StudentCourseEnrollment,
    Assignment,
    AssignmentAttachment,
    AssignmentSubmission,
    AssignmentSubmissionAttachment,
    AssignmentSubmissionHistory,
    Homework,
    HomeworkAttachment,
    HomeworkStatus,
    CourseMaterial,
    CourseMaterialAccess,
    TeachingLog,
    TeachingLogAttachment,
    SubjectAttendanceRecord,
    RoutinePeriod,
    InAppNotification,
)
from handle.forms import (
    BookForm, BookIssueForm, LibraryCategoryForm, LibraryAuthorForm,
    LibraryPublisherForm, LibraryRackForm, LibraryShelfForm, LibrarySettingsForm,
    AccountForm, JournalEntryForm, JournalEntryLineFormSet,
    AcademicYearForm, FacultyForm, SemesterForm, AssignmentForm,
    AssignmentSubmissionForm, AssignmentGradeForm, HomeworkForm,
    CourseMaterialForm, TeachingLogForm, RoutinePeriodForm,
    SubjectTeacherAssignmentForm, StudentCourseEnrollmentForm,
)
from management.models import Holiday, Occasion, OrganizationShiftOverride
from django.conf import settings
from collections import defaultdict
from django.db.models.functions import TruncDate
from datetime import timedelta
from django.utils import timezone
from django.db.models import Min, Max, Sum, Count, Q, F
from django.db import transaction
import nepali_datetime
from django.core.mail import send_mail
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from management.models import LocationBased, QRCode, AutoCheckin

from .forms import LocationForm, QRCodeForm, AutoCheckinForm


def _require_admin(request):
    """Return an error response if the user is not a schooladmin (user_type=2).
    Returns None if access is allowed.
    Usage: err = _require_admin(request); if err: return err
    """
    if not request.user.is_authenticated:
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())
    if request.user.user_type not in ('1', '2'):
        messages.error(request, "You do not have permission to access this page.")
        return redirect('management:homepage')
    return None


class AdminRequiredMixin:
    """Mixin that blocks non-admin users from any admin view."""
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('management:homepage')
        if request.user.user_type not in ('1', '2'):
            messages.error(request, "Access denied. Admins only.")
            return redirect('management:homepage')
        return super().dispatch(request, *args, **kwargs)


class PayrollSettingsView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'payroll'
    required_perm = 'can_manage_payroll_cfg'
    template_name = 'admin/payroll_settings.html'

    def get_org(self, request):
        if request.user.user_type == "2":
            return request.user.schooladmin.org
        if request.user.user_type == "3":
            return request.user.staff.org
        return None

    def get(self, request, *args, **kwargs):
        org = self.get_org(request)
        policy, _ = PayrollPolicy.objects.get_or_create(org=org)
        today = timezone.localdate()
        reminders_until = today + timedelta(days=policy.probation_reminder_days)
        probation_members = member.objects.filter(
            org=org,
            probation_end_date__isnull=False,
            probation_end_date__gte=today,
            probation_end_date__lte=reminders_until,
        ).exclude(status='dumped').order_by('probation_end_date')

        context = {
            'org': org,
            'policy': policy,
            'policy_form': PayrollPolicyForm(instance=policy),
            'adjustment_form': PayrollAdjustmentForm(org=org),
            'review_form': ProbationReviewForm(org=org),
            'adjustments': PayrollAdjustment.objects.filter(org=org).select_related('member', 'created_by')[:30],
            'probation_reviews': ProbationReview.objects.filter(org=org).select_related('member', 'reviewer')[:30],
            'probation_members': probation_members,
            'pf_records': ProvidentFundRecord.objects.filter(org=org).select_related('member', 'payslip')[:20],
            'ssf_records': SocialSecurityFundRecord.objects.filter(org=org).select_related('member', 'payslip')[:20],
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        org = self.get_org(request)
        action = request.POST.get('action')
        policy, _ = PayrollPolicy.objects.get_or_create(org=org)

        if action == 'policy':
            form = PayrollPolicyForm(request.POST, instance=policy)
            if form.is_valid():
                form.save()
                messages.success(request, "Payroll policy updated successfully.")
            else:
                messages.error(request, "Could not update payroll policy: " + form.errors.as_text())

        elif action == 'adjustment':
            form = PayrollAdjustmentForm(request.POST, org=org)
            if form.is_valid():
                adjustment = form.save(commit=False)
                adjustment.org = org
                adjustment.created_by = request.user
                adjustment.save()
                messages.success(request, "Payroll adjustment saved successfully.")
            else:
                messages.error(request, "Could not save adjustment: " + form.errors.as_text())

        elif action == 'review':
            form = ProbationReviewForm(request.POST, org=org)
            if form.is_valid():
                review = form.save(commit=False)
                review.org = org
                review.reviewer = request.user
                review.save()
                review.member.probation_review_status = review.status
                if review.status == 'passed':
                    review.member.status = 'active'
                    review.member.staff_type = 'permanent'
                elif review.status == 'extended':
                    review.member.status = 'probation'
                    review.member.staff_type = 'probation'
                review.member.save(update_fields=['probation_review_status', 'status', 'staff_type'])
                messages.success(request, "Probation review saved successfully.")
            else:
                messages.error(request, "Could not save review: " + form.errors.as_text())

        else:
            messages.error(request, "Invalid payroll settings action.")

        return HttpResponseRedirect(reverse('schooladmin:payroll_settings'))


class ManageLeaveTypesView(FeatureRequiredMixin, View):
    required_feature = 'leave'
    template_name = 'admin/manage_leave_types.html'

    def get(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        # Fetch all leave policies for this organization
        leave_types = LeaveType.objects.filter(org=org).order_by('-id')
        
        dist = {
            'leave_types': leave_types
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        action = request.POST.get('action')

        try:
            # ACTION 1: Create New Leave Policy
            if action == 'add':
                name = request.POST.get('name')
                allocation = request.POST.get('annual_allocation')
                
                LeaveType.objects.create(
                    org=org,
                    name=name,
                    annual_allocation=allocation,
                    is_paid=request.POST.get('is_paid') == 'on',
                )
                messages.success(request, f"Policy '{name}' created successfully.")

            # ACTION 2: Edit Existing Policy
            elif action == 'edit':
                leave_id = request.POST.get('leave_id')
                name = request.POST.get('name')
                allocation = request.POST.get('annual_allocation')

                l_type = LeaveType.objects.get(id=leave_id, org=org)
                l_type.name = name
                l_type.annual_allocation = allocation
                l_type.is_paid = request.POST.get('is_paid') == 'on'
                l_type.save()
                messages.success(request, f"Policy '{name}' updated successfully.")

            # ACTION 3: Delete Policy
            elif action == 'delete':
                leave_id = request.POST.get('leave_id')
                l_type = LeaveType.objects.get(id=leave_id, org=org)
                policy_name = l_type.name
                l_type.delete()
                messages.warning(request, f"Policy '{policy_name}' has been permanently deleted.")

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

        return redirect('schooladmin:manage_leave_types')
    
class MasterLeaveReportView(FeatureRequiredMixin, View):
    required_feature = 'leave'
    template_name = 'admin/master_leave_report.html'

    def get(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        members = member.objects.filter(org=org).exclude(status='dumped')
        leave_types = LeaveType.objects.filter(org=org)

        master_data = []
        for mem in members:
            # Gather balance for every policy for this specific member
            balances = []
            for l_type in leave_types:
                balances.append({
                    'type_name': l_type.name,
                    'data': mem.get_leave_balance(l_type.id)
                })
            
            master_data.append({
                'member': mem,
                'balances': balances
            })

        dist = {
            'leave_types': leave_types,
            'master_data': master_data
        }
        return render(request, self.template_name, dist)



def create_admin_leave(org, mem, l_type, start_date, end_date, reason):
    """Shared by AdminLogLeaveView (full-page form) and DayQuickActionView
    (calendar day-click AJAX) — same validation/balance-check/auto-approve
    behaviour either way. Returns (ok, message, leave_or_None)."""
    requested_days = (end_date - start_date).days + 1
    if requested_days <= 0:
        return False, "End date must be after or equal to Start date.", None

    balance = mem.get_leave_balance(l_type.id)
    if requested_days > balance['remaining']:
        return False, (
            f"Action Denied: {mem.name} only has {balance['remaining']} days of {l_type.name} left. "
            f"(Requested: {requested_days} days)"
        ), None

    leave = LeaveReport.objects.create(
        member=mem,
        org=org,
        leave_type=l_type,
        gap_start=start_date,
        gap_end=end_date,
        reason=reason,
        approved=True,  # Auto-approved because Admin logged it
        seen=True
    )
    return True, f"Successfully logged {requested_days} days of {l_type.name} for {mem.name}.", leave


class AdminLogLeaveView(FeatureRequiredMixin, View):
    required_feature = 'leave'
    template_name = 'admin/log_leave.html'

    def get(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        dist = {
            'members': member.objects.filter(org=org).exclude(status='dumped'),
            'leave_types': LeaveType.objects.filter(org=org),
            'today': datetime.date.today().strftime('%Y-%m-%d')
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        member_id = request.POST.get('member_id')
        leave_type_id = request.POST.get('leave_type_id')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        reason = request.POST.get('reason', 'Admin assigned leave.')

        if not all([member_id, leave_type_id, start_date_str, end_date_str]):
            messages.error(request, "All fields are required.")
            return redirect('schooladmin:log_leave')

        mem = member.objects.get(id=member_id, org=org)
        l_type = LeaveType.objects.get(id=leave_type_id, org=org)

        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()

        ok, msg, _leave = create_admin_leave(org, mem, l_type, start_date, end_date, reason)
        if ok:
            messages.success(request, msg)
        else:
            messages.error(request, msg)
        return redirect('schooladmin:log_leave')

def location_list(request):
    org = request.user.schooladmin.org
    locations = LocationBased.objects.filter(org = org)
    return render(request, 'admin/location_list.html', {'locations': locations})

def location_add(request):
    org = request.user.schooladmin.org
    print(org.id)

    if request.method == 'POST':
        form = LocationForm(request.POST)
        form.initial['org'] = org.id
        if form.is_valid():
            location = form.save(commit=False) 
            location.org = org  
            location.save()
            messages.success(request, "Succesfully added location data")
            return redirect('schooladmin:location_list')
    else:
        form = LocationForm()
    return render(request, 'admin/add_location.html', {'form': form})

def location_edit(request, id):
    org = request.user.schooladmin.org
    location = get_object_or_404(LocationBased, id=id, org=org)
    form = LocationForm(request.POST or None, instance=location)
    if form.is_valid():
        form.save()
        messages.success(request, "Succesfully edited location details")
        return redirect('schooladmin:location_list')
    return render(request, 'admin/add_location.html', {'form': form})

def location_delete(request, id):
    org = request.user.schooladmin.org
    if request.method != 'POST':
        return redirect('schooladmin:location_list')
    location = get_object_or_404(LocationBased, id=id, org=org)
    location.delete()
    messages.success(request, "Succesfully deleted location details")
    return redirect('schooladmin:location_list')


def qrcode_list(request):
    org = request.user.schooladmin.org
    qrcodes = QRCode.objects.filter(org = org)
    return render(request, 'admin/qrcode_list.html', {'qrcodes': qrcodes})

def qrcode_add(request):

    org = request.user.schooladmin.org
   
    if request.method == 'POST':
        form = QRCodeForm(request.POST, request.FILES)
        form.initial['org'] = org.id
        if form.is_valid():
            location = form.save(commit=False) 
            location.org = org  
            location.save()
            messages.success(request, "Succesfully added QR code data")
            return redirect('schooladmin:qrcode_list')
    else:
        form = QRCodeForm()
    return render(request, 'admin/add_qrcode.html', {'form': form})

def qrcode_edit(request, id):
    org = request.user.schooladmin.org
    qrcode = get_object_or_404(QRCode, id=id, org=org)
    form = QRCodeForm(request.POST or None, request.FILES or None, instance=qrcode)
    if form.is_valid():
        form.save()
        messages.success(request, "Succesfully edited qr code")
        return redirect('schooladmin:qrcode_list')

    return render(request, 'admin/add_qrcode.html', {'form': form})

def qrcode_delete(request, id):
    org = request.user.schooladmin.org
    if request.method != 'POST':
        return redirect('schooladmin:qrcode_list')
    qrcode = get_object_or_404(QRCode, id=id, org=org)
    qrcode.delete()
    messages.success(request, "Succesfully deleted QR code")
    return redirect('schooladmin:qrcode_list')

def auto_checkin_list(request):
    org = request.user.schooladmin.org
    records = AutoCheckin.objects.filter(org = org)
    return render(request, 'admin/autocheckin_list.html', {'records': records})


# Helper to sync AutoCheckin with AttendanceRecord
def sync_attendance_records(auto_obj, action='create'):
    # Remove old attendance records linked to this specific AutoCheckin record
    AttendanceRecord.objects.filter(mem=auto_obj.member, scanned_time=auto_obj.checkin_time).delete()
    AttendanceRecord.objects.filter(mem=auto_obj.member, scanned_time=auto_obj.checkout_time).delete()
    
    if action == 'create' or action == 'edit':
        # Create new ones directly. Do NOT use timezone.make_aware() 
        # because auto_obj.checkin_time is already timezone-aware.
        AttendanceRecord.objects.create(
            mem=auto_obj.member, 
            org=auto_obj.org, 
            scanned_time=auto_obj.checkin_time
        )
        AttendanceRecord.objects.create(
            mem=auto_obj.member, 
            org=auto_obj.org, 
            scanned_time=auto_obj.checkout_time
        )

def auto_checkin_add(request):
    org = request.user.schooladmin.org
    if request.method == 'POST':
        form = AutoCheckinForm(request.POST, org=org)
        if form.is_valid():
            auto = form.save(commit=False)
            auto.org = org
            auto.save()
            sync_attendance_records(auto, action='create')
            messages.success(request, "Auto-check-in log created successfully.")
            return redirect('schooladmin:auto_checkin_list')
    else:
        form = AutoCheckinForm(org=org)
    return render(request, 'admin/add_autocheckin.html', {'form': form, 'org': org})

def auto_checkin_edit(request, id):
    org = request.user.schooladmin.org
    record = get_object_or_404(AutoCheckin, id=id, org=org)
    if request.method == 'POST':
        form = AutoCheckinForm(request.POST, instance=record, org=org)
        if form.is_valid():
            # Update the AutoCheckin instance
            updated_record = form.save()
            # Re-sync with AttendanceRecords
            sync_attendance_records(updated_record, action='edit')
            messages.success(request, "Auto-check-in log updated.")
            return redirect('schooladmin:auto_checkin_list')
    else:
        form = AutoCheckinForm(instance=record, org=org)
    return render(request, 'admin/add_autocheckin.html', {'form': form, 'org': org})

def auto_checkin_delete(request, id):
    org = request.user.schooladmin.org
    if request.method != 'POST':
        return redirect('schooladmin:auto_checkin_list')
    record = get_object_or_404(AutoCheckin, id=id, org=org)
    # Cleanup AttendanceRecords
    sync_attendance_records(record, action='delete')
    record.delete()
    messages.success(request, "Auto-check-in log removed.")
    return redirect('schooladmin:auto_checkin_list')


def attendance_analytics(request):
    org = request.user.schooladmin.org if hasattr(request.user, 'schooladmin') else None
    nepali_enabled = getattr(org, 'nepali_date', False)

    # --- 1. DATE RANGE LOGIC ---
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    date_range = request.GET.get('date_range', '30days')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')
    classification_id = request.GET.get('classification')

    custom_start_np = request.GET.get('custom_start_np')
    custom_end_np = request.GET.get('custom_end_np')
    
    if date_range == 'custom':
        if nepali_enabled and custom_start_np and custom_end_np:
            try:
                y1, m1, d1 = map(int, custom_start_np.replace('/', '-').strip().split('-'))
                start_date_obj = nepali_datetime.date(y1, m1, d1).to_datetime_date()
                start_date = timezone.make_aware(datetime.datetime.combine(start_date_obj, datetime.datetime.min.time()))
                
                y2, m2, d2 = map(int, custom_end_np.replace('/', '-').strip().split('-'))
                end_date_obj = nepali_datetime.date(y2, m2, d2).to_datetime_date()
                end_date = timezone.make_aware(datetime.datetime.combine(end_date_obj, datetime.datetime.max.time()))
            except Exception:
                pass 
        elif custom_start and custom_end:
            try:
                start_date = timezone.make_aware(datetime.datetime.strptime(custom_start, '%Y-%m-%d'))
                end_date = timezone.make_aware(datetime.datetime.strptime(custom_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
            except Exception:
                pass
                
    elif date_range == '7days':
        start_date = end_date - timedelta(days=7)
    elif date_range == 'month':
        start_date = end_date.replace(day=1)
    elif date_range == 'year':
        start_date = end_date.replace(month=1, day=1)
    
    total_days = (end_date - start_date).days + 1
    if total_days <= 0: total_days = 1

    # --- 2. SUPER-FAST AGGREGATED QUERY ---
    records = AttendanceRecord.objects.filter(org=org, scanned_time__range=[start_date, end_date])
    if classification_id and classification_id != 'all':
        records = records.filter(mem__classification_id=classification_id)

    daily_stats = records.annotate(date=TruncDate('scanned_time')).values('mem', 'date').annotate(
        first_in=Min('scanned_time'),
        last_out=Max('scanned_time')
    )

    # --- 3. DATA PROCESSING & ANOMALY DETECTION ---
    raw_stats = defaultdict(lambda: {'days_present': 0, 'total_seconds': 0, 'incomplete_logs': 0})
    
    for stat in daily_stats:
        mem_id = stat['mem']
        raw_stats[mem_id]['days_present'] += 1
        
        # Check for Check-out Gaps
        if stat['last_out'] > stat['first_in']:
            delta = stat['last_out'] - stat['first_in']
            raw_stats[mem_id]['total_seconds'] += delta.total_seconds()
        else:
            # They punched in, but never punched out
            raw_stats[mem_id]['incomplete_logs'] += 1

    member_qs = member.objects.filter(org=org).exclude(status='dumped').select_related('classification')
    if classification_id and classification_id != 'all':
        member_qs = member_qs.filter(classification_id=classification_id)
    
    member_dict = {m.id: m for m in member_qs}
    member_stats = []
    attendance_summary = defaultdict(lambda: defaultdict(int))
    total_org_missed_punches = 0

    def format_time(seconds):
        hours, remainder = divmod(int(seconds), 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{hours}h {minutes}m"

    for mem_id, m in member_dict.items():
        data = raw_stats.get(mem_id, {'days_present': 0, 'total_seconds': 0, 'incomplete_logs': 0})
        
        days_present = data['days_present']
        days_absent = total_days - days_present
        if days_absent < 0: days_absent = 0
        
        total_seconds = data['total_seconds']
        incomplete_logs = data['incomplete_logs']
        total_org_missed_punches += incomplete_logs
        
        avg_seconds = total_seconds / days_present if days_present > 0 else 0
        
        attendance_percentage = round((days_present / total_days) * 100, 1)
        if attendance_percentage > 100: attendance_percentage = 100
        
        absence_percentage = round((days_absent / total_days) * 100, 1)
        if absence_percentage > 100: absence_percentage = 100

        dept_name = m.classification.name if m.classification else "Unclassified"
        if days_present > 0:
            attendance_summary[dept_name][m.name] = days_present

        member_stats.append({
            'member': m,
            'days_present': days_present,
            'days_absent': days_absent,
            'percentage': attendance_percentage,
            'absence_percentage': absence_percentage,
            'total_seconds': total_seconds,
            'formatted_total': format_time(total_seconds),
            'avg_seconds': avg_seconds,
            'formatted_avg': format_time(avg_seconds),
            'incomplete_logs': incomplete_logs
        })

    attendance_summary = {
        cls: dict(sorted(members_dict.items(), key=lambda item: item[1], reverse=True))
        for cls, members_dict in attendance_summary.items()
    }

    # --- 4. RANKING & SORTING ---
    most_present = sorted(member_stats, key=lambda x: x['days_present'], reverse=True)[:5]
    most_absent = sorted(member_stats, key=lambda x: x['days_absent'], reverse=True)[:5] # GAP ANALYSIS
    most_time_spent = sorted(member_stats, key=lambda x: x['total_seconds'], reverse=True)[:5]
    most_incomplete = sorted([m for m in member_stats if m['incomplete_logs'] > 0], key=lambda x: x['incomplete_logs'], reverse=True)[:5] # MISSED PUNCHES
    
    start_date_np_display = ""
    end_date_np_display = ""
    if nepali_enabled:
        try:
            start_date_np_display = str(nepali_datetime.date.from_datetime_date(start_date.date()))
            end_date_np_display = str(nepali_datetime.date.from_datetime_date(end_date.date()))
        except:
            pass
    
    # --- 5. ALL MEMBERS SORTED (for full absent leaderboard) ---
    all_absent_sorted = sorted(member_stats, key=lambda x: x['days_absent'], reverse=True)

    # --- 6. STUDY GAPS (AttendanceGap model) ---
    from django.db.models.functions import TruncMonth
    gap_qs = AttendanceGap.objects.filter(org=org, date__range=[start_date.date(), end_date.date()])
    if classification_id and classification_id != 'all':
        gap_qs = gap_qs.filter(classification_id=classification_id)
    top_gap_members = (
        gap_qs.values('member__id', 'member__name', 'member__classification__name')
        .annotate(gap_count=Count('id'))
        .order_by('-gap_count')[:10]
    )

    # --- 7. FINANCIAL SUMMARY ---
    sd_date = start_date.date() if hasattr(start_date, 'date') else start_date
    ed_date = end_date.date() if hasattr(end_date, 'date') else end_date
    fin_qs = FinancialTransaction.objects.filter(org=org, transaction_date__range=[sd_date, ed_date])
    total_income  = fin_qs.filter(transaction_type='income').aggregate(t=Sum('amount'))['t'] or 0
    total_expense = fin_qs.filter(transaction_type='expense').aggregate(t=Sum('amount'))['t'] or 0

    # Monthly income/expense for chart (last 6 months)
    import calendar as _cal
    today = datetime.date.today()
    monthly_finance = []
    for i in range(5, -1, -1):
        month_dt = today.replace(day=1) - timedelta(days=i * 30)
        month_dt = month_dt.replace(day=1)
        last_day = _cal.monthrange(month_dt.year, month_dt.month)[1]
        month_end = month_dt.replace(day=last_day)
        inc = FinancialTransaction.objects.filter(
            org=org, transaction_type='income',
            transaction_date__range=[month_dt, month_end]
        ).aggregate(t=Sum('amount'))['t'] or 0
        exp = FinancialTransaction.objects.filter(
            org=org, transaction_type='expense',
            transaction_date__range=[month_dt, month_end]
        ).aggregate(t=Sum('amount'))['t'] or 0
        monthly_finance.append({
            'month': month_dt.strftime('%b %Y'),
            'income': float(inc),
            'expense': float(exp),
        })

    # --- 8. PAYSLIP SUMMARY ---
    payslip_qs = PaySlip.objects.filter(org=org, from_date__gte=sd_date, to_date__lte=ed_date)
    total_salary_paid = payslip_qs.filter(status='paid').aggregate(t=Sum('net_payable'))['t'] or 0
    total_payslips    = payslip_qs.count()

    # --- 9. LEAVE SUMMARY ---
    leave_qs = LeaveReport.objects.filter(org=org, gap_start__gte=sd_date, gap_start__lte=ed_date)
    total_leaves   = leave_qs.count()
    approved_leaves= leave_qs.filter(approved=True).count()
    pending_leaves = leave_qs.filter(approved=False, rejected=False).count()

    # Top leave takers
    top_leave_members = (
        leave_qs.filter(approved=True)
        .values('member__id', 'member__name', 'member__classification__name')
        .annotate(leave_count=Count('id'))
        .order_by('-leave_count')[:8]
    )

    # Dept-wise absent for chart
    dept_absent_chart = {}
    for stat in member_stats:
        dept = stat['member'].classification.name if stat['member'].classification else 'Unclassified'
        dept_absent_chart[dept] = dept_absent_chart.get(dept, 0) + stat['days_absent']

    context = {
        'org': org,
        'classifications': Classification.objects.filter(org=org),
        'most_present': most_present,
        'most_absent': most_absent,
        'most_time_spent': most_time_spent,
        'most_incomplete': most_incomplete,
        'all_absent_sorted': all_absent_sorted[:20],
        'total_org_missed_punches': total_org_missed_punches,
        'total_days': total_days,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_np': start_date_np_display,
        'end_date_np': end_date_np_display,
        'custom_start_np': custom_start_np,
        'custom_end_np': custom_end_np,
        'nepali_enabled': nepali_enabled,
        'selected_range': date_range,
        'selected_classification': classification_id,
        'custom_start': custom_start,
        'custom_end': custom_end,
        'classifi': dict(attendance_summary),
        # gaps
        'top_gap_members': list(top_gap_members),
        'total_gaps': gap_qs.count(),
        # finance
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': float(total_income) - float(total_expense),
        'monthly_finance_json': json.dumps(monthly_finance),
        # payroll
        'total_salary_paid': total_salary_paid,
        'total_payslips': total_payslips,
        # leave
        'total_leaves': total_leaves,
        'approved_leaves': approved_leaves,
        'pending_leaves': pending_leaves,
        'top_leave_members': list(top_leave_members),
        # charts
        'dept_absent_json': json.dumps([
            {'dept': k, 'absent': v} for k, v in sorted(dept_absent_chart.items(), key=lambda x: x[1], reverse=True)[:8]
        ]),
        'absent_chart_json': json.dumps([
            {'name': s['member'].name, 'absent': s['days_absent'], 'pct': s['absence_percentage']}
            for s in all_absent_sorted[:10]
        ]),
    }

    return render(request, 'admin/highest.html', context)
class AllRecord(ListView):
    model = AttendanceRecord
    paginate_by = 100
    template_name = 'admin/allReport.html'

    def get_context_data(self, request, **kwargs):
        today_date = datetime.date.today()
        context = super().get_context_data(**kwargs)
        context["daily"] = self.model.objects.filter(scanned_time__date = today_date)
        context['member'] = member.objects.filter(org=org).exclude(status='dumped')
        auser = request.user
        org =  auser.schooladmin.org
        context['org'] = org
        return context



def getMember(request):
    qid = request.GET.get('questionid', None)




def orgDetail(request):
    user = request.user
    org = user.schooladmin.org
    form = OrgFormSchool(instance=org)
    holiday = Holiday.objects.filter(org = org)
    existing_holidays = [holiday.holiday for holiday in holiday]
    occasion = Occasion.objects.filter(org =org)
    shift_overrides = OrganizationShiftOverride.objects.filter(org=org).order_by('-date')[:20]

    from school.org_stats import get_org_stats

    dist = {
        'form':form,
        'org':org,
        'holiday':existing_holidays,
        'occasions':occasion,
        'shift_overrides':shift_overrides,
        'org_stats': get_org_stats(org),
    }
    if request.method == 'POST':
        org.name = request.POST.get('name', org.name)
        try:
            org.image = request.FILES['image'] or None
        except Exception:
            pass
        org.serial_key = request.POST.get('serial_key', org.serial_key)
        org.address = request.POST.get('address', org.address)
        org.email = request.POST.get('email', org.email or '').strip()
        org.nepali_date = request.POST.get('nepali_date') == 'on'
        default_start = request.POST.get('default_shift_start_time')
        default_end = request.POST.get('default_shift_end_time')
        if default_start:
            org.default_shift_start_time = default_start
        if default_end:
            org.default_shift_end_time = default_end
        org.save()
        messages.success(request, "Organization profile updated successfully.")
        return HttpResponseRedirect(reverse('schooladmin:orgDetail'))


    return render(request, "admin/orgDetail.html", dist)



class Dashboard(View):
    template_name = 'admin/Adashboard.html'

    def get(self, request, *args, **kwargs):
        user = request.user
        org = user.schooladmin.org
        if not org.activate:
            return render(request, 'admin/activate.html')

        today_date = datetime.date.today()
        member.date = today_date
        
        tm = member.objects.filter(org=org).exclude(status='dumped')
        total_member = tm.count()
        leave = LeaveReport.objects.filter(org=org).count()
        unseen_leave = LeaveReport.objects.filter(seen=False, org=org).count()

        # Bulk equivalent of "for each member, did first_daily_time() find a
        # scan today?" — avoids one AttendanceRecord query per member.
        present_member_ids = set(
            AttendanceRecord.objects.filter(
                org=org, scanned_time__date=today_date, mem_id__in=tm.values_list('id', flat=True),
            ).values_list('mem_id', flat=True).distinct()
        )
        present = len(present_member_ids)
        absent = total_member - present

        # 🔥 NEW: Critical Attendance Gap Logic (Last 7 Days)
        last_7_days = today_date - timedelta(days=7)
        
        # Members with at least one scan in last 7 days
        recent_active_ids = AttendanceRecord.objects.filter(
            org=org, scanned_time__gte=last_7_days
        ).values_list('mem_id', flat=True).distinct()

        # Members NOT seen in 7 days are "Critical Absentees"
        critical_absentees = tm.exclude(id__in=recent_active_ids).filter(status='active')[:5]

        today = datetime.date.today()
        month_start = today.replace(day=1)

        # ── Feature-sensitive dashboard queries ───────────────────────────────
        # Only run queries for enabled modules to avoid wasted DB load and to
        # ensure disabled modules return zero rather than stale data.

        total_income = total_expense = Decimal("0.00")
        if org.feature_finance:
            total_income  = FinancialTransaction.objects.filter(org=org, transaction_type='income',  transaction_date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0
            total_expense = FinancialTransaction.objects.filter(org=org, transaction_type='expense', transaction_date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0

        low_stock_count = 0
        if org.feature_stock:
            low_stock_count = StockItem.objects.filter(org=org, status='active', quantity__lte=F('low_stock_threshold')).count()

        library_overdue_count = 0
        if has_feature(org, 'library'):
            library_overdue_count = BookIssue.objects.filter(
                org=org, status__in=('issued', 'overdue'), due_date__lt=today
            ).count()

        pending_journal_count = 0
        if has_feature(org, 'accounting'):
            pending_journal_count = JournalEntry.objects.filter(org=org, status='pending').count()

        pending_complaints = 0
        if org.feature_complaints:
            pending_complaints = Complaint.objects.filter(org=org, status='pending').count()

        upcoming_events = 0
        if org.feature_events:
            upcoming_events = Event.objects.filter(org=org, start_date__gte=today, status='upcoming').count()

        pending_resignations = 0
        if org.feature_hrms:
            pending_resignations = ResignationRecord.objects.filter(org=org, status='pending').count()

        # Task stats
        from handle.models import TaskInstance
        task_total = task_pending = task_overdue = task_today = task_pending_approval = 0
        if org.feature_tasks:
            task_inst_qs = TaskInstance.objects.filter(task__org=org)
            task_total            = task_inst_qs.count()
            task_pending          = task_inst_qs.filter(status='pending').count()
            task_overdue          = task_inst_qs.filter(status__in=['overdue', 'missed_absence']).count()
            task_today            = task_inst_qs.filter(due_date=today).exclude(status='cancelled').count()
            task_pending_approval = task_inst_qs.filter(approval_status='pending_approval').count()

        # Payroll stats
        from handle.models import AdvanceSalary
        advance_total_outstanding = advance_count = draft_payslips = 0
        month_payroll_total = Decimal("0.00")
        if org.feature_payroll:
            active_advances = AdvanceSalary.objects.filter(org=org, status='active')
            advance_total_outstanding = active_advances.aggregate(t=Sum('remaining_balance'))['t'] or 0
            advance_count = active_advances.count()
            draft_payslips = PaySlip.objects.filter(org=org, status='draft', from_date__gte=month_start).count()
            month_payroll_total = PaySlip.objects.filter(org=org, from_date__gte=month_start).aggregate(t=Sum('net_payable'))['t'] or 0

        # Billing / student stats
        student_qs = member.objects.filter(org=org, member_type='student').exclude(status='dumped')
        expected_monthly_billing = paid_this_month = Decimal("0.00")
        bills_generated_this_month = bills_not_sent = 0
        pending_due = Decimal("0.00")
        recent_payments = recent_sent_bills = class_income = []
        results_published = results_pending_publish = 0
        if org.feature_billing:
            active_student_qs = student_qs.filter(status='active')
            expected_monthly_billing = sum((_compute_final_fee(m) for m in active_student_qs), Decimal("0.00"))
            this_month_bills = Bill.objects.filter(org=org, billing_month=today.month, billing_year=today.year)
            paid_this_month = this_month_bills.aggregate(t=Sum('amount_paid'))['t'] or 0
            bills_generated_this_month = this_month_bills.count()
            bills_not_sent = this_month_bills.filter(is_sent=False).count()
            bill_totals = Bill.objects.filter(org=org).exclude(status='Cancelled').aggregate(total=Sum('total_amount'), paid=Sum('amount_paid'))
            pending_due = max(Decimal("0.00"), _money(bill_totals.get('total')) - _money(bill_totals.get('paid')))
            recent_payments = Bill.objects.filter(org=org, amount_paid__gt=0).select_related('member').order_by('-issue_date')[:5]
            recent_sent_bills = BillSendLog.objects.filter(bill__org=org).select_related('bill', 'bill__member')[:5]
            class_income = (Bill.objects.filter(org=org).values('classification__name').annotate(total=Sum('total_amount'), paid=Sum('amount_paid')).order_by('-total')[:6])

        if org.feature_results:
            results_published       = ExamTerm.objects.filter(org=org, is_published=True).count()
            results_pending_publish = ExamTerm.objects.filter(org=org, is_published=False).exclude(status='archived').count()

        # Leave
        pending_leave_requests = 0
        if org.feature_leave:
            pending_leave_requests = LeaveReport.objects.filter(org=org, approved=False, rejected=False).count()

        # Branches
        total_branches = 0
        if org.feature_branches:
            total_branches = Branch.objects.filter(org=org, status='active').count()

        academic_dashboard = {
            'academic_current_year': None,
            'academic_course_count': 0,
            'academic_enrollment_count': 0,
            'academic_active_assignment_count': 0,
            'academic_unlinked_subject_count': 0,
            'academic_unmapped_routine_count': 0,
            'academic_today_period_count': 0,
            'academic_missing_session_count': 0,
            'academic_submitted_session_count': 0,
            'academic_draft_session_count': 0,
            'academic_rejected_session_count': 0,
            'academic_recent_sessions': [],
        }
        if has_feature(org, 'academic_management'):
            current_year = AcademicYear.objects.filter(
                org=org, is_current=True, status='active',
            ).order_by('-start_date', '-pk').first()
            active_assignments = SubjectTeacherAssignment.objects.filter(
                org=org, status='active', start_date__lte=today,
            ).filter(Q(end_date__isnull=True) | Q(end_date__gte=today))
            if current_year:
                active_assignments = active_assignments.filter(
                    Q(academic_year=current_year) | Q(academic_year__isnull=True)
                )
            today_weekday = (today.weekday() + 1) % 7
            today_periods = RoutinePeriod.objects.filter(
                org=org, day_of_week=today_weekday, is_active=True,
            )
            if current_year:
                today_periods = today_periods.filter(
                    Q(academic_year=current_year) | Q(academic_year__isnull=True)
                )
            completed_periods = TeachingLog.objects.filter(
                org=org,
                date=today,
                routine_period__in=today_periods,
                status__in=('submitted', 'approved'),
            ).values_list('routine_period_id', flat=True)
            recent_sessions = TeachingLog.objects.filter(org=org).select_related(
                'course', 'subject', 'classification', 'section', 'teacher',
            ).order_by('-date', '-pk')[:8]
            academic_dashboard.update({
                'academic_current_year': current_year,
                'academic_course_count': Course.objects.filter(org=org, status='active').count(),
                'academic_enrollment_count': StudentCourseEnrollment.objects.filter(
                    org=org, status='active', start_date__lte=today,
                ).filter(Q(end_date__isnull=True) | Q(end_date__gte=today)).count(),
                'academic_active_assignment_count': active_assignments.count(),
                'academic_unlinked_subject_count': Subject.objects.filter(
                    org=org, status='active', course__isnull=True,
                ).count(),
                'academic_unmapped_routine_count': RoutinePeriod.objects.filter(
                    org=org, is_active=True, teacher_assignment__isnull=True,
                ).count(),
                'academic_today_period_count': today_periods.count(),
                'academic_missing_session_count': today_periods.exclude(
                    pk__in=completed_periods,
                ).count(),
                'academic_submitted_session_count': TeachingLog.objects.filter(
                    org=org, status='submitted',
                ).count(),
                'academic_draft_session_count': TeachingLog.objects.filter(
                    org=org, status='draft',
                ).count(),
                'academic_rejected_session_count': TeachingLog.objects.filter(
                    org=org, status='rejected',
                ).count(),
                'academic_recent_sessions': recent_sessions,
            })

        dist = {
            'org': org,
            'tm': total_member,
            'absent': absent,
            'present': present,
            'leave': leave,
            'unseen_leave': unseen_leave,
            'pending_leave_requests': pending_leave_requests,
            'devices': Device.objects.filter(org=org),
            'critical_absentees': critical_absentees,
            'month_income': total_income,
            'month_expense': total_expense,
            'net_balance': total_income - total_expense,
            'low_stock_count': low_stock_count,
            'library_overdue_count': library_overdue_count,
            'pending_journal_count': pending_journal_count,
            'pending_complaints': pending_complaints,
            'upcoming_events': upcoming_events,
            'pending_resignations': pending_resignations,
            'total_branches': total_branches,
            'task_total': task_total,
            'task_pending': task_pending,
            'task_overdue': task_overdue,
            'task_today': task_today,
            'task_pending_approval': task_pending_approval,
            'advance_total_outstanding': advance_total_outstanding,
            'advance_count': advance_count,
            'draft_payslips': draft_payslips,
            'month_payroll_total': month_payroll_total,
            'total_students': student_qs.count() if org.feature_student_mgmt else 0,
            'expected_monthly_billing': expected_monthly_billing,
            'paid_this_month': paid_this_month,
            'pending_due': pending_due,
            'bills_generated_this_month': bills_generated_this_month,
            'bills_not_sent': bills_not_sent,
            'results_pending_publish': results_pending_publish,
            'results_published': results_published,
            'recent_payments': recent_payments,
            'recent_sent_bills': recent_sent_bills,
            'class_income': class_income,
            'dash_notices': self._recent_notices(org),
        }
        dist.update(academic_dashboard)
        return render(request, self.template_name, dist)

    @staticmethod
    def _recent_notices(org):
        """Latest live notices for the admin dashboard widget."""
        from handle.models import Notice
        from django.db.models import Q
        if not getattr(org, 'feature_notices', False):
            return []
        now = timezone.now()
        return list(
            Notice.objects.filter(org=org, publish_at__lte=now)
            .filter(Q(expires_at__isnull=True) | Q(expires_at__gt=now))[:5]
        )


class ManualAttendance(View):
    template_name = 'admin/manual_attendance.html'

    def get(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        today_date = datetime.date.today()
        
        # Set the class variable so first_daily_time works for today
        member.date = today_date 
        members = member.objects.filter(org=org).exclude(status='dumped')
        
        dist = {
            'members': members,
            'today_display': today_date.strftime('%A, %d %B %Y'),
            'today_val': today_date.strftime('%Y-%m-%d'),
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        org = request.user.schooladmin.org
        member_id = request.POST.get('member_id')
        action = request.POST.get('action') 
        
        if not member_id:
            messages.error(request, "Error: No member selected.")
            return redirect('schooladmin:manual_attendance')
            
        mem = member.objects.get(id=member_id, org=org)
        today_date = datetime.date.today()

        # Get all records for today, ordered from earliest to latest
        # (Replace 'member_record' if your related name is different, e.g., 'attendancerecord_set')
        records_today = mem.member_record.filter(scanned_time__date=today_date).order_by('scanned_time')

        # ACTION 1: 1-Click Mark Present
        if action == 'mark_present':
            # Use timezone.now() to respect Asia/Kathmandu
            aware_now = timezone.now() 
            mem.member_record.create(scanned_time=aware_now, org=org)
            messages.success(request, f"Marked {mem.name} as Present!")

        # ACTION 2: Update Exact Times
        elif action == 'update_times':
            in_time = request.POST.get('in_time')
            out_time = request.POST.get('out_time')
            
            if in_time:
                # 1. Create native time, 2. Make it timezone aware for Nepal
                naive_in = datetime.datetime.strptime(f"{today_date} {in_time}", "%Y-%m-%d %H:%M")
                aware_in = timezone.make_aware(naive_in)
                
                if records_today.exists():
                    # UPDATE the first punch of the day
                    first_rec = records_today.first()
                    first_rec.scanned_time = aware_in
                    first_rec.save()
                else:
                    # CREATE if they have no punches yet
                    mem.member_record.create(scanned_time=aware_in, org=org)
            
            # Refresh the records list in case we just created the first one above
            records_today = mem.member_record.filter(scanned_time__date=today_date).order_by('scanned_time')
            
            if out_time:
                naive_out = datetime.datetime.strptime(f"{today_date} {out_time}", "%Y-%m-%d %H:%M")
                aware_out = timezone.make_aware(naive_out)
                
                if records_today.count() > 1:
                    # UPDATE the last punch of the day
                    last_rec = records_today.last()
                    last_rec.scanned_time = aware_out
                    last_rec.save()
                elif records_today.count() == 1:
                    # If they only had a Check-In, CREATE this as the Check-Out
                    mem.member_record.create(scanned_time=aware_out, org=org)
                else:
                    # Fallback
                    mem.member_record.create(scanned_time=aware_out, org=org)
                    
            messages.success(request, f"Successfully updated time logs for {mem.name}.")

        return redirect('schooladmin:manual_attendance')




class DailyReport(ListView):
    template_name = 'admin/dailyReport.html'

    def get_organization(self):
        return self.request.user.schooladmin.org

    def get_target_date(self, request, org):
        """
        Convert the submitted AD or BS date into a Python AD date.
        """
        today = timezone.localdate()
        nepali_enabled = getattr(org, 'nepali_date', False)

        date_ad_string = request.POST.get('date', '').strip()
        date_np_string = request.POST.get('date_np', '').strip()

        target_date = today
        normalized_nepali_date = None

        if nepali_enabled and date_np_string:
            try:
                clean_date = date_np_string.replace('/', '-').strip()
                year, month, day = map(int, clean_date.split('-'))

                nepali_date = nepali_datetime.date(
                    year,
                    month,
                    day
                )

                target_date = nepali_date.to_datetime_date()
                normalized_nepali_date = clean_date

            except (ValueError, TypeError):
                if date_ad_string:
                    try:
                        target_date = datetime.datetime.strptime(
                            date_ad_string,
                            '%Y-%m-%d'
                        ).date()
                    except ValueError:
                        target_date = today

        elif date_ad_string:
            try:
                target_date = datetime.datetime.strptime(
                    date_ad_string,
                    '%Y-%m-%d'
                ).date()
            except ValueError:
                target_date = today

        if nepali_enabled and not normalized_nepali_date:
            normalized_nepali_date = str(
                nepali_datetime.date.from_datetime_date(target_date)
            )

        return target_date, normalized_nepali_date

    def get_members(self, org, selected_filter, branch_filter='All'):
        """
        Return organization members according to the selected classification and/or branch.
        """
        from school.hierarchy import get_accessible_members
        members = (
            get_accessible_members(self.request.user, org)
            .select_related('classification', 'branch')
            .order_by('name')
        )

        selected_classification_name = 'All'

        if selected_filter and selected_filter != 'All':
            members = members.filter(
                classification_id=selected_filter
            )

            selected_classification = (
                Classification.objects
                .filter(
                    id=selected_filter,
                    org=org
                )
                .first()
            )

            if selected_classification:
                selected_classification_name = (
                    selected_classification.name
                )
            else:
                selected_classification_name = 'Unknown'

        if branch_filter and branch_filter != 'All':
            members = members.filter(branch_id=branch_filter)

        return members, selected_classification_name

    def get_scan_queryset(self, report_member, target_date):
        """
        Get only the selected date's scans.
        Returns a list to avoid negative indexing issues.
        """
        member.date = target_date
        
        daily_source = report_member.alldataofdaily
        
        # Handle different types of daily_source
        if callable(daily_source):
            daily_source = daily_source()
        
        if daily_source is None:
            return []
        
        # If it's a related manager, get all
        if hasattr(daily_source, 'all') and not hasattr(daily_source, 'filter'):
            daily_source = daily_source.all()
        
        # QuerySet case - convert to list to avoid negative indexing issues
        if hasattr(daily_source, 'filter'):
            scans_qs = daily_source.filter(
                scanned_time__date=target_date
            ).order_by('scanned_time')
            return list(scans_qs)  # Return as list
        
        # Python list or iterable case
        filtered_scans = []
        
        for scan in daily_source:
            scanned_datetime = scan.scanned_time
        
            if timezone.is_aware(scanned_datetime):
                scanned_datetime = timezone.localtime(scanned_datetime)
        
            if scanned_datetime.date() == target_date:
                filtered_scans.append(scan)
        
        return sorted(
            filtered_scans,
            key=lambda scan: scan.scanned_time
        )

    @staticmethod
    def get_local_time(scan):
        """
        Return a scan's time in the current Django timezone.
        """
        if not scan:
            return None

        scanned_datetime = scan.scanned_time

        if timezone.is_aware(scanned_datetime):
            scanned_datetime = timezone.localtime(scanned_datetime)

        return scanned_datetime.time()

    @staticmethod
    def calculate_total_hours(first_time, last_time):
        """
        Calculate total hours between first and last scan times.
        """
        if not first_time or not last_time:
            return None
        
        try:
            # Convert times to datetime objects for calculation
            base_date = datetime.now().date()
            first_datetime = datetime.combine(base_date, first_time)
            last_datetime = datetime.combine(base_date, last_time)
            
            # Handle cases where last_time is on the next day (e.g., overnight shifts)
            if last_datetime < first_datetime:
                last_datetime += timedelta(days=1)
            
            duration = last_datetime - first_datetime
            hours = duration.seconds // 3600
            minutes = (duration.seconds % 3600) // 60
            
            return f"{hours:02d}:{minutes:02d}:00"
        except Exception:
            return None

    def prepare_report(self, members_queryset, target_date):
        """
        Attach selected-date attendance values to every member and
        calculate all dashboard counts. Shows BOTH present AND absent members.
        """
        member.date = target_date

        report_members = []
        issue_members = []
        absent_members = []
        present_members = []

        total_members = 0
        present_count = 0
        absent_count = 0
        late_in_count = 0
        early_out_count = 0

        for report_member in members_queryset:
            total_members += 1

            selected_date_scans = self.get_scan_queryset(
                report_member,
                target_date
            )  # This now returns a list

            is_present = len(selected_date_scans) > 0
            
            if is_present:
                present_count += 1
                
                first_scan = selected_date_scans[0] if selected_date_scans else None
                last_scan = selected_date_scans[-1] if selected_date_scans else None

                first_time = self.get_local_time(first_scan)
                last_time = self.get_local_time(last_scan)

                # Calculate total hours manually if not available from model
                total_hours = self.calculate_total_hours(first_time, last_time)
                
                # If model has hour_inside, use it; otherwise use calculated
                if hasattr(report_member, 'hour_inside') and report_member.hour_inside:
                    total_hours = report_member.hour_inside

                # Existing model properties
                late_in = report_member.late_in()
                early_in = report_member.early_in()
                early_out = report_member.early_out()
                late_out = report_member.late_out()

                is_late_in = bool(late_in)
                is_early_out = bool(early_out)

                # Temporary attributes for template
                report_member.report_scans = selected_date_scans
                report_member.report_first_time = first_time
                report_member.report_last_time = last_time
                report_member.report_late_in = late_in
                report_member.report_early_in = early_in
                report_member.report_early_out = early_out
                report_member.report_late_out = late_out
                report_member.report_total_hours = total_hours

                report_member.report_is_present = True
                report_member.report_is_absent = False
                report_member.report_is_late_in = is_late_in
                report_member.report_is_early_out = is_early_out

                report_member.report_has_issue = (
                    is_late_in or is_early_out
                )

                if is_late_in:
                    late_in_count += 1

                if is_early_out:
                    early_out_count += 1

                if report_member.report_has_issue:
                    issue_members.append(report_member)

                present_members.append(report_member)
                
            else:
                absent_count += 1
                # Store absent member info for display
                report_member.report_is_present = False
                report_member.report_is_absent = True
                report_member.report_first_time = None
                report_member.report_last_time = None
                report_member.report_total_hours = None
                report_member.report_late_in = None
                report_member.report_early_in = None
                report_member.report_early_out = None
                report_member.report_late_out = None
                report_member.report_scans = []
                report_member.report_has_issue = True
                report_member.report_is_late_in = False
                report_member.report_is_early_out = False
                absent_members.append(report_member)
                issue_members.append(report_member)

            # Add all members to report_members (both present and absent)
            report_members.append(report_member)

        attendance_percentage = 0

        if total_members:
            attendance_percentage = round(
                (present_count / total_members) * 100,
                1
            )

        return {
            'tm': report_members,  # ALL members (present + absent)
            'present_members': present_members,  # Only present members
            'absent_members': absent_members,  # Only absent members
            'issue_members': issue_members,
            'total_members': total_members,
            'present_count': present_count,
            'absent_count': absent_count,
            'late_in_count': late_in_count,
            'early_out_count': early_out_count,
            'issue_count': len(issue_members),
            'attendance_percentage': attendance_percentage,
        }

    def build_context(
        self,
        org,
        target_date,
        selected_filter='All',
        branch_filter='All',
        date_np=None
    ):
        nepali_enabled = getattr(org, 'nepali_date', False)

        members_queryset, selected_classification_name = (
            self.get_members(
                org,
                selected_filter,
                branch_filter,
            )
        )

        report_context = self.prepare_report(
            members_queryset,
            target_date
        )

        if nepali_enabled and not date_np:
            date_np = str(
                nepali_datetime.date.from_datetime_date(
                    target_date
                )
            )

        from school.hierarchy import get_accessible_branches
        from school.print_settings import get_print_preference
        context = {
            'date': target_date,
            'date_np': date_np if nepali_enabled else None,
            'org': org,
            'thisone': selected_classification_name,
            'selected_filter': selected_filter,
            'selected_branch': branch_filter,
            'clas': Classification.objects.filter(org=org).order_by('name'),
            'branches': get_accessible_branches(self.request.user, org).order_by('name'),
            'nepali_enabled': nepali_enabled,
            'print_preference': get_print_preference(self.request.user, 'daily_report', org=org),
        }

        context.update(report_context)

        return context

    def get(self, request, *args, **kwargs):
        org = self.get_organization()
        target_date = timezone.localdate()

        context = self.build_context(
            org=org,
            target_date=target_date,
            selected_filter='All',
            branch_filter='All',
        )

        return render(
            request,
            self.template_name,
            context
        )

    def post(self, request, *args, **kwargs):
        org = self.get_organization()

        selected_filter = request.POST.get('filter', 'All')
        branch_filter = request.POST.get('branch', 'All')

        target_date, date_np = self.get_target_date(
            request,
            org
        )

        context = self.build_context(
            org=org,
            target_date=target_date,
            selected_filter=selected_filter,
            branch_filter=branch_filter,
            date_np=date_np
        )

        return render(
            request,
            self.template_name,
            context
        )


# Export functionality with proper total hours calculation
def export_present_attendance(request):
    """
    Export attendance data to Excel or CSV format with proper total hours calculation.
    """
    org = request.user.schooladmin.org
    date_str = request.GET.get('date')
    fmt = request.GET.get('fmt', 'excel')
    
    # If no date, use today
    if not date_str:
        target_date = timezone.localdate()
    else:
        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = timezone.localdate()
    
    # Get all members
    members = member.objects.filter(org=org).exclude(status='dumped').select_related('classification')
    
    data = []
    
    for member_obj in members:
        member_obj.date = target_date
        
        # Get scans for the date
        daily_source = member_obj.alldataofdaily
        
        # Handle different types of daily_source
        if callable(daily_source):
            daily_source = daily_source()
        
        if daily_source is None:
            continue
        
        # If it's a related manager, get all
        if hasattr(daily_source, 'all') and not hasattr(daily_source, 'filter'):
            daily_source = daily_source.all()
        
        # Filter scans for the target date
        scans = []
        
        if hasattr(daily_source, 'filter'):
            # QuerySet case - convert to list to avoid negative indexing issues
            scans_qs = daily_source.filter(
                scanned_time__date=target_date
            ).order_by('scanned_time')
            scans = list(scans_qs)  # Convert to list
        else:
            # Python list or iterable case
            for scan in daily_source:
                scanned_datetime = scan.scanned_time
                if timezone.is_aware(scanned_datetime):
                    scanned_datetime = timezone.localtime(scanned_datetime)
                if scanned_datetime.date() == target_date:
                    scans.append(scan)
            scans = sorted(scans, key=lambda x: x.scanned_time)
        
        # Check if present
        is_present = len(scans) > 0
        
        if is_present:
            first_scan = scans[0] if scans else None
            last_scan = scans[-1] if scans else None
            
            first_time = first_scan.scanned_time if first_scan else None
            last_time = last_scan.scanned_time if last_scan else None
            
            # Calculate total hours
            total_hours = '00:00:00'
            if first_time and last_time:
                try:
                    # Convert to local time if needed
                    if timezone.is_aware(first_time):
                        first_time = timezone.localtime(first_time)
                    if timezone.is_aware(last_time):
                        last_time = timezone.localtime(last_time)
                    
                    # Calculate duration
                    duration = last_time - first_time
                    hours = duration.seconds // 3600
                    minutes = (duration.seconds % 3600) // 60
                    total_hours = f"{hours:02d}:{minutes:02d}:00"
                except Exception:
                    total_hours = '00:00:00'
            
            # Get attendance metrics
            late_in = member_obj.late_in() or ''
            early_out = member_obj.early_out() or ''
            
            # Use model's hour_inside if available and valid
            if hasattr(member_obj, 'hour_inside') and member_obj.hour_inside:
                total_hours = member_obj.hour_inside
        else:
            first_scan = None
            last_scan = None
            late_in = ''
            early_out = ''
            total_hours = '00:00:00'
        
        data.append({
            'Member Name': member_obj.name,
            'Department': member_obj.classification.name if member_obj.classification else 'N/A',
            'Date': target_date.strftime('%Y-%m-%d'),
            'Status': 'Present' if is_present else 'Absent',
            'Punch In': first_scan.scanned_time.strftime('%H:%M:%S') if first_scan else '',
            'Punch Out': last_scan.scanned_time.strftime('%H:%M:%S') if last_scan else '',
            'Late In': late_in if is_present else '',
            'Early Out': early_out if is_present else '',
            'Total Hours': total_hours,
            'Shift Start': member_obj.shift_start_time.strftime('%H:%M') if member_obj.shift_start_time else '',
            'Shift End': member_obj.shift_end_time.strftime('%H:%M') if member_obj.shift_end_time else '',
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{target_date}.csv"'
        
        writer = csv.writer(response)
        # Write headers
        writer.writerow(df.columns.tolist())
        # Write data
        for _, row in df.iterrows():
            writer.writerow(row.tolist())
        
        return response
    else:
        # Excel format
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{target_date}.xlsx"'
        
        df.to_excel(response, index=False, sheet_name='Attendance Report')
        return response

class PresentToday(ListView):
    template_name = 'admin/presentToday.html'

    def get(self, request, *args, **kwargs):
        user = request.user
        org = user.schooladmin.org
        today_date = datetime.date.today()
        
        # Set class variable so model properties format correctly
        member.date = today_date

        tm = member.objects.filter(org=org).exclude(status='dumped')
        today_attendance_data = AttendanceRecord.objects.filter(org=org, scanned_time__date=today_date)
        classifi = Classification.objects.filter(org=org)
        
        member_data = []
        for j in tm:
            for i in today_attendance_data:
                if i.mem == j:
                    member_data.append(j)
                    break
        
        nepali_enabled = getattr(org, 'nepali_date', False)
        
        dist = {
            'date': today_date,
            'date_np': '',  # Prevents template crash on initial load
            'tm': member_data,
            'org': org,
            'thisone': 'All',
            'clas': classifi,
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        today_date = datetime.date.today()
        user = request.user
        org = user.schooladmin.org
        nepali_enabled = getattr(org, 'nepali_date', False)

        # Safely extract POST data
        name = request.POST.get('filter', 'All')
        date_str = request.POST.get('date', '').strip()
        date_np = request.POST.get('date_np', '').strip()

        # `date` must become a real datetime.date - member.date (below) feeds
        # straight into member.late_in()/early_out() (via effective_shift_start,
        # which calls weekday_number(target_date).weekday()), which raises
        # AttributeError on a plain string rather than failing gracefully.
        date = None
        if nepali_enabled and date_np:
            try:
                y, m, d = map(int, date_np.replace('/', '-').strip().split('-'))
                date = nepali_datetime.date(y, m, d).to_datetime_date()
            except (ValueError, TypeError):
                date = None
        if date is None and date_str:
            try:
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                date = None
        if date is None:
            date = today_date

        # Set class variable for the requested date
        member.date = date

        classifi = Classification.objects.filter(org=org)
        today_attendance_data = AttendanceRecord.objects.filter(org=org, scanned_time__date=date)

        member_data = []

        # Filter by classification or show All
        if name != 'All':
            tm = member.objects.filter(org=org, classification=name).exclude(status='dumped')
            sn = Classification.objects.filter(id=name, org=org).values_list('name', flat=True).first() or 'Unknown'
        else:
            tm = member.objects.filter(org=org).exclude(status='dumped')
            sn = 'All'

        for j in tm:
            for i in today_attendance_data:
                if i.mem == j:
                    member_data.append(j)
                    break

        dist = {
            'date': date,
            'date_np': date_np,  # Passes the Nepali date string back to the form
            'tm': member_data,
            'thisone': sn,
            'org': org,
            'clas': classifi,
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)


class AbsentToday(ListView):
    template_name = 'admin/absentToday.html'

    def get(self, request, *args, **kwargs):
        user = request.user
        org = user.schooladmin.org
        today_date = datetime.date.today()
        
        # Set class variable
        member.date = today_date

        tm = member.objects.filter(org=org).exclude(status='dumped')
        today_attendance_data = AttendanceRecord.objects.filter(org=org, scanned_time__date=today_date)
        classifi = Classification.objects.filter(org=org)
        
        member_data = []
        nepali_enabled = getattr(org, 'nepali_date', False)
        
        for j in tm:
            for i in today_attendance_data:
                if i.mem == j:
                    member_data.append(j)
                    break
                    
        # Calculate who is absent
        new = list(set(member_data).symmetric_difference(set(tm)))
        
        dist = {
            'date': today_date,
            'date_np': '',  # Prevents template crash on initial load
            'tm': new,
            'org': org,
            'thisone': 'All',
            'clas': classifi,
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        today_date = datetime.date.today()
        user = request.user
        org = user.schooladmin.org
        nepali_enabled = getattr(org, 'nepali_date', False)

        # Safely extract POST data
        name = request.POST.get('filter', 'All')
        date_str = request.POST.get('date', '').strip()
        date_np = request.POST.get('date_np', '').strip()

        # `date` must become a real datetime.date - member.date (below) feeds
        # straight into member.late_in()/early_out() (via effective_shift_start,
        # which calls weekday_number(target_date).weekday()), which raises
        # AttributeError on a plain string rather than failing gracefully.
        date = None
        if nepali_enabled and date_np:
            try:
                y, m, d = map(int, date_np.replace('/', '-').strip().split('-'))
                date = nepali_datetime.date(y, m, d).to_datetime_date()
            except (ValueError, TypeError):
                date = None
        if date is None and date_str:
            try:
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                date = None
        if date is None:
            date = today_date

        # Set class variable for the requested date
        member.date = date

        classifi = Classification.objects.filter(org=org)
        today_attendance_data = AttendanceRecord.objects.filter(org=org, scanned_time__date=date)

        member_data = []

        # Filter by classification or show All
        if name != 'All':
            tm = member.objects.filter(org=org, classification=name).exclude(status='dumped')
            sn = Classification.objects.filter(id=name, org=org).values_list('name', flat=True).first() or 'Unknown'
        else:
            tm = member.objects.filter(org=org).exclude(status='dumped')
            sn = 'All'

        for j in tm:
            for i in today_attendance_data:
                if i.mem == j:
                    member_data.append(j)
                    break

        # Calculate who is absent
        new = list(set(member_data).symmetric_difference(set(tm)))

        dist = {
            'date': date,
            'date_np': date_np,  # Passes the Nepali date string back to the form
            'tm': new,
            'thisone': sn,
            'org': org,
            'clas': classifi,
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)






def parse_time(time_string):
    try:
        # Try parsing with microseconds
        return dt.strptime(time_string, '%H:%M:%S.%f')
    except ValueError:
        # Fallback to parsing without microseconds
        return dt.strptime(time_string.split('.')[0], '%H:%M:%S')
    


class GapReport(View):
    template_name = "admin/gapReport.html"

    def get(self, request, *args, **kwargs):
        user = request.user
        org = user.schooladmin.org
        tm = None
        today_date = datetime.date.today()
        clas = Classification.objects.filter(org=org)
        branches = Branch.objects.filter(org=org, status='active').order_by('name')

        nepali_enabled = org.nepali_date

        today_np = str(nepali_datetime.date.from_datetime_date(today_date)) if nepali_enabled else ""

        dist = {
            'first_date': today_date,
            'last_date': today_date,
            'first_date_np': today_np,
            'last_date_np': today_np,
            'tm': tm,
            'org': org,
            'thisone': 'All',
            'clas': clas,
            'branches': branches,
            'selected_branch': 'All',
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        user = request.user
        org = user.schooladmin.org
        clas = Classification.objects.filter(org=org)
        branches = Branch.objects.filter(org=org, status='active').order_by('name')
        name = request.POST.get('classification', 'All')
        branch_id = request.POST.get('branch', 'All')
        
        # 🔥 नयाँ मोडल फिल्ड
        nepali_enabled = org.nepali_date

        first_date_str = request.POST.get('first_date', '')
        last_date_str = request.POST.get('last_date', '')
        first_date_np_str = request.POST.get('first_date_np', '')
        last_date_np_str = request.POST.get('last_date_np', '')

        # 1. First Date कन्भर्सन
        if nepali_enabled and first_date_np_str:
            try:
                y, m, d = map(int, first_date_np_str.replace('/', '-').strip().split('-'))
                start_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except Exception:
                start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()
        else:
            start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()

        # 2. Last Date कन्भर्सन
        if nepali_enabled and last_date_np_str:
            try:
                y, m, d = map(int, last_date_np_str.replace('/', '-').strip().split('-'))
                end_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except Exception:
                end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()
        else:
            end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()

        delta = datetime.timedelta(days=1)
        later_date = start_date

        if name == 'All':
            mem = member.objects.filter(org=org).exclude(status='dumped')
            th = 'All'
        else:
            mem = member.objects.filter(org=org).exclude(status='dumped').filter(classification=name)
            th = Classification.objects.filter(id=name, org=org).values_list('name', flat=True).first() or 'Unknown'

        if branch_id and branch_id != 'All':
            mem = mem.filter(branch_id=branch_id)
            
        member_data = []
        holiday = Holiday.objects.filter(org=org)
        occasion = Occasion.objects.filter(org=org)
        
        for i in mem:
            current_loop_date = later_date 
            while current_loop_date <= end_date:
                i.date = current_loop_date
                aa = i.first_daily_time()
                try:
                    bb = i.last_daily_time()
                except:
                    bb = None
                    
                time_interval = None
                
                if aa is not None:
                    time_1 = i.parse_time(str(aa)) # member model function
                if bb:
                    time_2 = i.parse_time(str(bb))
                    time_interval = time_2 - time_1
                    
                holi = False
                oca = None
              
                for p in holiday:
                    if p.holiday == current_loop_date.strftime("%A"):
                        holi = True
                        break
                
                for n in occasion:
                    if not n.end_date:
                        if n.date == current_loop_date:
                            oca = n.name
                    else:
                        temp_date = n.date
                        while temp_date <= n.end_date:
                            if temp_date == current_loop_date:
                                oca = n.name
                                break
                            temp_date += datetime.timedelta(days=1)

                # 🔥 हरेक दिनको मितिलाई नेपालीमा बदल्ने
                loop_date_np = ""
                if nepali_enabled:
                    loop_date_np = str(nepali_datetime.date.from_datetime_date(current_loop_date))

                # 🔥 i.7 मा नेपाली मिति पठाउने (loop_date_np)
                member_data.append([current_loop_date, i.name, aa, bb, time_interval, holi, oca, loop_date_np])
                
                i.first_date = None
                i.last_date = None
                i.ft = None
                i.tt = None
                i.date = None
              
                current_loop_date += delta
      
        dist = {
            'first_date': start_date,
            'last_date': end_date,
            'first_date_np': first_date_np_str if nepali_enabled else None,
            'last_date_np': last_date_np_str if nepali_enabled else None,
            'tm': member_data,
            'thisone': th,
            'org': org,
            'clas': clas,
            'branches': branches,
            'selected_branch': branch_id,
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)
class MemberGapReport(View):
    template_name = "admin/memberRecord.html"
    
    def get(self, request, *args, **kwargs):
        id = self.kwargs['id']
        user = request.user
        org = user.schooladmin.org
        memb = get_object_or_404(member, id=id, org=org)
        today_date = datetime.date.today()
        memb.date = today_date
        member_data = []

        nepali_enabled = getattr(org, 'nepali_date', False)
        today_np = str(nepali_datetime.date.from_datetime_date(today_date)) if nepali_enabled else ""

        time_interval = None
        aa = memb.first_daily_time()
        try:
            bb = memb.last_daily_time()
        except:
            bb = None
     
        if aa is not None:
            time_1 = memb.parse_time(str(aa))
        if bb:
            time_2 = memb.parse_time(str(bb))
            time_interval = time_2 - time_1
      
        # 🔥 NEW: Calculate today's late in / early out for the GET request
        try:
            late_in = memb.late_in()
            early_out = memb.early_out()
        except:
            late_in = None
            early_out = None

        # Added late_in as index 9 and early_out as index 10
        member_data.append([today_date, memb.name, aa, bb, time_interval, False, None, None, today_np, late_in, early_out])

        dist = {
            'date': today_date,
            'first_date': today_date,
            'last_date': today_date,
            'tm': member_data,
            'org': org,
            'thisone': memb.name,
            'nepali_enabled': nepali_enabled,
            'first_date_np': today_np,
            'last_date_np': today_np,
            'total_late_in': late_in if late_in else "00:00",
            'total_early_out': early_out if early_out else "00:00",
            'current_member': memb,
            'all_members': member.objects.filter(org=org).exclude(status='dumped').order_by('name'),
            'leave_types': list(LeaveType.objects.filter(org=org).values('id', 'name', 'is_paid')),
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        id = self.kwargs['id']
        user = request.user
        org = user.schooladmin.org
        mem = get_object_or_404(member, id=id, org=org)
        nepali_enabled = getattr(org, 'nepali_date', False)

        first_date_str = request.POST.get('first_date', '')
        last_date_str = request.POST.get('last_date', '')
        first_date_np_str = request.POST.get('first_date_np', '')
        last_date_np_str = request.POST.get('last_date_np', '')

        # 1. First Date Conversion
        if nepali_enabled and first_date_np_str:
            try:
                y, m, d = map(int, first_date_np_str.replace('/', '-').strip().split('-'))
                start_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except Exception:
                start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()
        else:
            start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()

        # 2. Last Date Conversion
        if nepali_enabled and last_date_np_str:
            try:
                y, m, d = map(int, last_date_np_str.replace('/', '-').strip().split('-'))
                end_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except Exception:
                end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()
        else:
            end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()

        delta = datetime.timedelta(days=1)
        member_data = []
        holiday = Holiday.objects.filter(org=org)
        occasion = Occasion.objects.filter(org=org)
        
        total_days, total_absent_days, total_present_days = 0, 0, 0
        total_holidays, total_occasion_holidays, total_leave_days = 0, 0, 0
        
        # 🔥 NEW: Variables to sum up the total late and early hours
        total_late_seconds = 0
        total_early_out_seconds = 0
        
        total_leave = LeaveReport.objects.filter(member=mem).filter(approved=True)
        leave_date = []

        for leave_report in total_leave:
            total_leave_days += leave_report.total_leave_days()
                
        for i in total_leave:
            if not i.gap_end:
                leave_date.append(i.gap_start)
            else:
                current_date = i.gap_start
                while current_date <= i.gap_end:
                    leave_date.append(current_date)
                    current_date += datetime.timedelta(days=1)

        # To avoid altering the loop variable
        loop_date = start_date

        while loop_date <= end_date:
            total_days += 1
            i = mem
            i.date = loop_date
            aa = i.first_daily_time()
            bb = i.last_daily_time()
            time_interval = None
            if aa:
                time_1 = i.parse_time(str(aa))
            if bb:
                time_2 = i.parse_time(str(bb))
                time_interval = time_2 - time_1

            # 🔥 NEW: Calculate Late In and Early Out for this specific day
            try:
                late_in_time = i.late_in()
                early_out_time = i.early_out()
            except Exception:
                late_in_time = None
                early_out_time = None

            # Add to total seconds for the summary cards
            if late_in_time:
                h, m = map(int, late_in_time.split(':'))
                total_late_seconds += (h * 3600 + m * 60)
            
            if early_out_time:
                h, m = map(int, early_out_time.split(':'))
                total_early_out_seconds += (h * 3600 + m * 60)

            holi = False
            for p in holiday:
                if p.holiday == loop_date.strftime("%A"):
                    total_holidays += 1
                    holi = True
                    break
            
            oca = None
            for n in occasion:
                if not n.end_date:
                    if n.date == loop_date:
                        oca = n.name
                        total_occasion_holidays += 1
                        break
                else:
                    current_date = n.date
                    while current_date <= n.end_date:
                        if current_date == loop_date:
                            oca = n.name
                            total_occasion_holidays += 1
                            break
                        current_date += datetime.timedelta(days=1)

            if aa:
                total_present_days += 1
            else:
                if not holi and not oca:
                    total_absent_days += 1

            leave_status = 'On Leave' if loop_date in leave_date else None

            # Nepali date logic
            loop_date_np = ""
            if nepali_enabled:
                loop_date_np = str(nepali_datetime.date.from_datetime_date(loop_date))

            # 🔥 NEW: Appended late_in_time as i.9 and early_out_time as i.10
            member_data.append([loop_date, i.name, aa, bb, time_interval, holi, oca, leave_status, loop_date_np, late_in_time, early_out_time])
            
            i.first_date = None
            i.last_date = None
            i.ft = None
            i.tt = None
            i.date = None
              
            loop_date += delta

        # Convert total accumulated seconds back to HH:MM string format
        total_late_str = f"{total_late_seconds // 3600:02d}:{(total_late_seconds % 3600) // 60:02d}"
        total_early_out_str = f"{total_early_out_seconds // 3600:02d}:{(total_early_out_seconds % 3600) // 60:02d}"

        dist = {
            'first_date': start_date,
            'last_date': end_date,
            'first_date_np': first_date_np_str if nepali_enabled else None,
            'last_date_np': last_date_np_str if nepali_enabled else None,
            'tm': member_data,
            'org': org,
            'thisone': mem.name,
            'total_days': total_days,
            'total_absent_days': total_absent_days,
            'total_present_days': total_present_days,
            'total_holidays': total_holidays,
            'total_occasion_holidays': total_occasion_holidays,
            'total_leave_days': total_leave_days,
            'total_late_in': total_late_str,        # 🔥 Passed to template
            'total_early_out': total_early_out_str, # 🔥 Passed to template
            'nepali_enabled': nepali_enabled,
            'current_member': mem,
            'all_members': member.objects.filter(org=org).exclude(status='dumped').order_by('name'),
            'leave_types': list(LeaveType.objects.filter(org=org).values('id', 'name', 'is_paid')),
        }
        return render(request, self.template_name, dist)
    

class salaryReport(FeatureRequiredMixin, View):
    required_feature = 'payroll'
    template_name = "admin/salaryReport.html"
    
    def get(self, request, *args, **kwargs):
        id = self.kwargs['id']
        user = request.user
        org = user.schooladmin.org
        memb = get_object_or_404(member, id=id, org=org)

        nepali_enabled = getattr(org, 'nepali_date', False)
        today_date = datetime.date.today()
        today_np = str(nepali_datetime.date.from_datetime_date(today_date)) if nepali_enabled else ""
        
        memb.date = today_date
        aa = memb.first_daily_time()
        try:
            bb = memb.last_daily_time()
        except:
            bb = None
            
        time_interval = None
        time_interval_cost = 0
        total_hour = 0
        total_cost = 0
        
        if aa and bb:
            time_1 = memb.parse_time(str(aa))
            time_2 = memb.parse_time(str(bb))
            time_interval = time_2 - time_1
            tim = (time_interval.total_seconds() / 60) / 60
            times = float("{:.2f}".format(tim))
            time_interval_cost = times * memb.salary_per_hour
            total_cost += time_interval_cost

        # लिस्टको 6th Index मा नेपाली मिति पठाउँदै
        member_data = [[today_date, memb.name, aa, bb, time_interval, time_interval_cost, today_np]]
        
        dist = {
            'first_date': today_date.strftime("%Y-%m-%d"),
            'last_date': today_date.strftime("%Y-%m-%d"),
            'first_date_np': today_np,
            'last_date_np': today_np,
            'tm': member_data,
            'org': org,
            'thisone': memb.name,
            'mem_id': memb.id, # Payslip बटनको लागि
            'total_hour': total_hour,
            'total_cost': round(total_cost, 2),
            'allMember': member.objects.filter(org=org).exclude(status='dumped'),
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        id = self.kwargs['id']
        user = request.user
        org = user.schooladmin.org
        
        # यदि ड्रपडाउनबाट अर्को मेम्बर छानेको छ भने
        selected_member_id = request.POST.get('member', id)
        mem = get_object_or_404(member, id=selected_member_id, org=org)
        
        nepali_enabled = getattr(org, 'nepali_date', False)

        first_date_str = request.POST.get('first_date', '')
        last_date_str = request.POST.get('last_date', '')
        first_date_np_str = request.POST.get('first_date_np', '')
        last_date_np_str = request.POST.get('last_date_np', '')

        # 1. First Date कन्भर्सन
        if nepali_enabled and first_date_np_str:
            try:
                y, m, d = map(int, first_date_np_str.replace('/', '-').strip().split('-'))
                start_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except Exception:
                start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()
        else:
            start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()

        # 2. Last Date कन्भर्सन
        if nepali_enabled and last_date_np_str:
            try:
                y, m, d = map(int, last_date_np_str.replace('/', '-').strip().split('-'))
                end_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except Exception:
                end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()
        else:
            end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()

        delta = datetime.timedelta(days=1)
        member_data = []
        total_hour_in_sec = 0.0
        total_cost = 0
        
        loop_date = start_date
        while loop_date <= end_date:
            time_interval = None
            time_interval_cost = 0
            
            mem.date = loop_date
            aa = mem.first_daily_time()
            try:
                bb = mem.last_daily_time()
            except:
                bb = None
           
            if aa and bb:
                time_1 = mem.parse_time(str(aa))
                time_2 = mem.parse_time(str(bb))
                time_interval = time_2 - time_1
                
                tim = (time_interval.total_seconds() / 60) / 60
                total_hour_in_sec += tim
                time_interval_cost = tim * mem.salary_per_hour
                total_cost += round(time_interval_cost)

            # हरेक दिनको नेपाली मिति
            loop_date_np = str(nepali_datetime.date.from_datetime_date(loop_date)) if nepali_enabled else ""
            
            member_data.append([loop_date, mem.name, aa, bb, time_interval, round(time_interval_cost), loop_date_np])
            
            mem.first_date = None
            mem.last_date = None
            mem.ft = None
            mem.tt = None
            mem.date = None
            loop_date += delta
            
        dist = {
            'first_date': start_date.strftime("%Y-%m-%d"),
            'last_date': end_date.strftime("%Y-%m-%d"),
            'first_date_np': first_date_np_str if nepali_enabled else None,
            'last_date_np': last_date_np_str if nepali_enabled else None,
            'tm': member_data,
            'org': org,
            'allMember': member.objects.filter(org=org).exclude(status='dumped'),
            'thisone': mem.name,
            'mem_id': mem.id, # Payslip बटनको लागि
            'total_hour': int(total_hour_in_sec),
            'total_cost': total_cost,
            'nepali_enabled': nepali_enabled
        }
        return render(request, self.template_name, dist)

class salaryReportAll(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'payroll'
    required_perm = 'can_view_payroll'
    template_name = 'admin/salaryReportAll.html'

    def get(self, request, *args, **kwargs):
        auser = request.user
        if auser.user_type == "2":
            org =  auser.schooladmin.org
        elif auser.user_type == "3":
            org =  auser.staff.org
        else:
            org = None
  
        memb = member.objects.filter(org=org).exclude(status='dumped').order_by('-id')

        dist = {
            'mem':memb,
            'org':org,
            'clas':Classification.objects.filter(org=org),
            'thisone':'All'
        }
        return render(request, self.template_name, dist)

    def post(self, request, *agrs, **kwargs):
        clas = request.POST['classification']
        print('class', clas)
        auser = request.user
        if auser.user_type == "2":
            org =  auser.schooladmin.org
        elif auser.user_type == "3":
            org =  auser.staff.org
      
        cl=Classification.objects.filter(org=org)
        if clas == 'All':
            mem = member.objects.filter(org = org).exclude(status='dumped')
            th = 'All' 
        else:
            mem = member.objects.filter(org = org).exclude(status='dumped').filter(classification=clas)
            th = Classification.objects.filter(id=clas, org=org).values_list('name', flat=True).first() or 'Unknown'
        dist = {
            'mem':mem,
            'clas':cl,
            'org':org,
            'thisone': th
        }
        return render(request, self.template_name, dist)



class leaveReportView(FeatureRequiredMixin, View):
    required_feature = 'leave'
    template_name = "admin/leaveReportView.html"

    def get(self, request, *args, **kwargs):
        from django.core.paginator import Paginator
        user = request.user
        org = user.schooladmin.org
        nepali_enabled = getattr(org, 'nepali_date', False)
        today_date = datetime.date.today()

        # Filters
        status_filter = request.GET.get('status', 'all')
        branch_id = request.GET.get('branch', '')
        classification_id = request.GET.get('classification', '')
        leave_type_id = request.GET.get('leave_type', '')

        leave_reports = LeaveReport.objects.filter(org=org).select_related(
            'member', 'leave_type', 'member__branch', 'member__classification'
        ).order_by('-gap_start')

        if status_filter == 'pending':
            leave_reports = leave_reports.filter(approved=False, rejected=False, seen=False)
        elif status_filter == 'approved':
            leave_reports = leave_reports.filter(approved=True)
        elif status_filter == 'rejected':
            leave_reports = leave_reports.filter(rejected=True)

        if branch_id:
            leave_reports = leave_reports.filter(member__branch_id=branch_id)
        if classification_id:
            leave_reports = leave_reports.filter(member__classification_id=classification_id)
        if leave_type_id:
            leave_reports = leave_reports.filter(leave_type_id=leave_type_id)

        for leave in leave_reports:
            if nepali_enabled:
                leave.start_display = str(nepali_datetime.date.from_datetime_date(leave.gap_start)) if leave.gap_start else ""
                leave.end_display   = str(nepali_datetime.date.from_datetime_date(leave.gap_end))   if leave.gap_end   else ""
            else:
                leave.start_display = leave.gap_start.strftime("%Y-%m-%d") if leave.gap_start else ""
                leave.end_display   = leave.gap_end.strftime("%Y-%m-%d")   if leave.gap_end   else ""
            leave.is_paid_leave = getattr(leave.leave_type, 'is_paid', True) if leave.leave_type else True

        paginator = Paginator(leave_reports, 30)
        page_obj  = paginator.get_page(request.GET.get('page'))

        today_np = str(nepali_datetime.date.from_datetime_date(today_date)) if nepali_enabled else today_date.strftime("%Y-%m-%d")
        dist = {
            'org': org,
            'leave': page_obj,
            'page_obj': page_obj,
            'status_filter': status_filter,
            'nepali_enabled': nepali_enabled,
            'date': today_np if nepali_enabled else today_date.strftime("%Y-%m-%d"),
            'total_pending': LeaveReport.objects.filter(org=org, approved=False, rejected=False).count(),
            'branches': Branch.objects.filter(org=org, status='active').order_by('name'),
            'classifications': Classification.objects.filter(org=org).order_by('name'),
            'leave_types': LeaveType.objects.filter(org=org),
            'selected_branch': branch_id,
            'selected_classification': classification_id,
            'selected_leave_type': leave_type_id,
        }
        return render(request, self.template_name, dist)
    


def _handle_leave_status_change(request, report_id, status, org):
    """Shared logic for approving/rejecting a leave report and notifying the
    member by email. `status` accepts either 'accept'/'reject' (legacy) or
    'approved'/'rejected'. Returns (report, normalized_status)."""
    report = get_object_or_404(LeaveReport, id=report_id, org=org)
    norm_status = 'approved' if status in ('accept', 'approved') else 'rejected'
    report.approved = norm_status == 'approved'
    report.rejected = norm_status == 'rejected'
    report.seen = True
    report.save()

    admin_remarks = request.POST.get('remarks', '') if request.method == 'POST' else ''

    if report.member and report.member.email:
        lt_name = report.leave_type.name if report.leave_type else 'Leave'
        send_leave_status_email(
            email=report.member.email,
            name=report.member.name,
            status=norm_status,
            leave_type=lt_name,
            start=str(report.gap_start),
            end=str(report.gap_end),
            remarks=admin_remarks,
            org_name=org.name,
            org=org,
            related_object_id=report.id,
        )
    return report, norm_status


@perm_required('can_approve_leave')
def leaveStatus(request, id, status):
    org = _get_org(request)
    report, norm_status = _handle_leave_status_change(request, id, status, org)
    messages.success(request, f"Leave {norm_status}.")
    return HttpResponseRedirect(reverse('schooladmin:leaveReportView'))

   
@feature_required('payroll')
@perm_required('can_generate_payroll')
def playSlipView(request):
    auser = request.user
    if auser.user_type == "2":
        org = auser.schooladmin.org
    elif auser.user_type == "3":
        org = auser.staff.org
    else:
        org = None

    memb = member.objects.filter(org=org).exclude(status='dumped').order_by('-id')

    if request.method == 'POST':
        clas_id = request.POST.get('classification', 'All')
        branch_id = request.POST.get('branch', 'All')
        thisone = 'All'
        if clas_id and clas_id != 'All':
            memb = memb.filter(classification_id=clas_id)
            thisone = Classification.objects.filter(id=clas_id, org=org).values_list('name', flat=True).first() or thisone
        if branch_id and branch_id != 'All':
            memb = memb.filter(branch_id=branch_id)
    else:
        clas_id = 'All'
        branch_id = 'All'
        thisone = 'All'

    dist = {
        'mem': memb,
        'org': org,
        'clas': Classification.objects.filter(org=org),
        'branches': Branch.objects.filter(org=org, status='active').order_by('name'),
        'thisone': thisone,
        'selected_classification': clas_id,
        'selected_branch': branch_id,
    }
    return render(request, "admin/payslip.html", dist)


# Helper to convert "HH:MM" string to decimal hours for precise penalty math
from schooladmin.payroll_service import (
    calculate_attendance_stats,
    calculate_payroll_components,
    get_or_create_policy,
    time_to_decimal,
    decimal_money,
)

def sum_adjustments(queryset, adjustment_type):
    total = queryset.filter(adjustment_type=adjustment_type).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    return decimal_money(total)

@feature_required('payroll')
@perm_required('can_generate_payroll')
def paySlipDetailView(request, id):
    auser = request.user
    org = auser.schooladmin.org if auser.user_type == "2" else auser.staff.org
        
    memb = get_object_or_404(member, id=id, org=org)
    nepali_enabled = getattr(org, 'nepali_date', False)
    payroll_policy, _ = PayrollPolicy.objects.get_or_create(org=org)

    dist = {
        'mem': memb,
        'org': org,
        'clas': Classification.objects.filter(org=org),
        'paySlip': PaySlip.objects.filter(member__id=id).order_by('-id'),
        'nepali_enabled': nepali_enabled,
        'payroll_policy': payroll_policy,
        'is_generated': False
    }
    
    if request.method == 'POST':
        first_date_str = request.POST.get('first_date', '')
        last_date_str = request.POST.get('last_date', '')
        first_date_np_str = request.POST.get('first_date_np', '')
        last_date_np_str = request.POST.get('last_date_np', '')

        # Date Parsing
        if nepali_enabled and first_date_np_str:
            try:
                y, m, d = map(int, first_date_np_str.replace('/', '-').strip().split('-'))
                start_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except:
                start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()
        else:
            start_date = datetime.datetime.strptime(first_date_str, "%Y-%m-%d").date()

        if nepali_enabled and last_date_np_str:
            try:
                y, m, d = map(int, last_date_np_str.replace('/', '-').strip().split('-'))
                end_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except:
                end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()
        else:
            end_date = datetime.datetime.strptime(last_date_str, "%Y-%m-%d").date()

        # ── Attendance stats via payroll service ──────────────────────────
        stats, daily_logs = calculate_attendance_stats(
            memb, start_date, end_date, org, nepali_enabled=nepali_enabled
        )

        # ── Payroll components via payroll service ────────────────────────
        comps = calculate_payroll_components(memb, stats, org, payroll_policy, end_date, daily_logs=daily_logs)

        month_name = f"{start_date.strftime('%B')} - {end_date.strftime('%B %Y')}"

        dist.update({
            'start_date': start_date.strftime("%Y-%m-%d"),
            'end_date': end_date.strftime("%Y-%m-%d"),
            'first_date_np': first_date_np_str,
            'last_date_np': last_date_np_str,
            'month_name': month_name,
            'daily_logs': daily_logs,
            'stats': stats,
            **comps,
            'active_adjustments': comps['active_adjustments'],
            'is_generated': True,
        })

    from school.print_settings import get_print_preference
    dist['print_preference'] = get_print_preference(request.user, 'payroll', org=org)
    return render(request, "admin/generate_payslip.html", dist)

def generate(request, id):
    if request.method == 'POST':
        # 1. Identify the Organization of the logged-in user
        auser = request.user
        if auser.user_type == "2":
            org = auser.schooladmin.org
        elif auser.user_type == "3":
            org = auser.staff.org
        else:
            org = None

        # 2. The member is taken from the URL and scoped to this org — never
        #    trusted from POST — so a tampered form can't target another
        #    org's staff.
        memb = get_object_or_404(member, id=id, org=org)

        first_date_str = request.POST.get('first_date', '')
        last_date_str = request.POST.get('last_date', '')
        first_date_np_str = request.POST.get('first_date_np', '')
        last_date_np_str = request.POST.get('last_date_np', '')
        nepali_enabled = getattr(org, 'nepali_date', False)

        def _parse_date(greg_str, nepali_str):
            """Try the BS (Nepali) date first when the org uses it, then fall
            back to the Gregorian field, then give up (None) instead of
            crashing on an empty/malformed string."""
            if nepali_enabled and nepali_str:
                try:
                    y, m, d = map(int, nepali_str.replace('/', '-').strip().split('-'))
                    return nepali_datetime.date(y, m, d).to_datetime_date()
                except Exception:
                    pass
            if greg_str:
                try:
                    return datetime.datetime.strptime(greg_str, "%Y-%m-%d").date()
                except Exception:
                    pass
            return None

        start_date = _parse_date(first_date_str, first_date_np_str)
        end_date = _parse_date(last_date_str, last_date_np_str)

        if not start_date or not end_date:
            messages.error(request, "Please select both a start and end date before generating the payslip.")
            return HttpResponseRedirect(reverse('schooladmin:play-slip-detail', args=(id,)))

        # 3. Duplicate-generation guard — same as the bulk-payslip flow.
        if PaySlip.objects.filter(member=memb, org=org, from_date=start_date, to_date=end_date).exists():
            messages.error(request, "A payslip already exists for this member and period.")
            return HttpResponseRedirect(reverse('schooladmin:play-slip-detail', args=(id,)))

        # 4. Compute every financial figure server-side. Preview/hidden POST
        #    values are display-only and must never become payroll authority.
        payroll_policy, _ = PayrollPolicy.objects.get_or_create(org=org)
        stats, _daily_logs = calculate_attendance_stats(
            memb, start_date, end_date, org, nepali_enabled=nepali_enabled
        )
        comps = calculate_payroll_components(memb, stats, org, payroll_policy, end_date)
        month_name = f"{start_date.strftime('%B')} - {end_date.strftime('%B %Y')}"

        final_values = {
            'gross_salary': comps['gross_salary'],
            'overtime_hours': comps['overtime_hours'],
            'overtime_amount': comps['overtime_amount'],
            'overtime_rate_multiplier': comps['effective_ot_multiplier'],
            'allowance_total': comps['allowance_total'],
            'bonus_total': comps['bonus_total'],
            'other_deduction': comps['other_deduction'],
            'tax_deduction': comps['tax_amount'],
            'pf_employee': comps['pf_employee'],
            'pf_employer': comps['pf_employer'],
            'ssf_employee': comps['ssf_employee'],
            'ssf_employer': comps['ssf_employer'],
            'advance_deduction': comps['advance_deduction'],
            'loan_deduction': comps['loan_deduction'],
            'net_payable': comps['net_payable'],
        }
        adjustment_notes = []

        with transaction.atomic():
            payslip = PaySlip.objects.create(
                member=memb,
                org=org,
                from_date=start_date,
                to_date=end_date,
                month_name=month_name,
                total_days=stats['total_days'],
                present_days=stats['days_present'],
                paid_leaves=stats['days_paid_leave'],
                holidays=stats['days_holiday'],
                unpaid_absences=stats['days_unpaid_absent'],
                salary_type=(memb.salary_type or 'monthly'),
                gross_salary=final_values['gross_salary'],
                allowance_total=final_values['allowance_total'],
                bonus_total=final_values['bonus_total'],
                advance_deduction=final_values['advance_deduction'],
                loan_deduction=final_values['loan_deduction'],
                other_deduction=final_values['other_deduction'],
                tax_deduction=final_values['tax_deduction'],
                pf_employee=final_values['pf_employee'],
                pf_employer=final_values['pf_employer'],
                ssf_employee=final_values['ssf_employee'],
                ssf_employer=final_values['ssf_employer'],
                probation_adjustment=comps['probation_adjustment'],
                overtime_hours=final_values['overtime_hours'],
                overtime_amount=final_values['overtime_amount'],
                overtime_rate_multiplier=final_values['overtime_rate_multiplier'],
                net_payable=final_values['net_payable'],
                notes='; '.join(adjustment_notes) or None,
            )

            if payslip.pf_employee or payslip.pf_employer:
                ProvidentFundRecord.objects.create(
                    org=org,
                    member=payslip.member,
                    payslip=payslip,
                    month_name=payslip.month_name or '',
                    employee_contribution=payslip.pf_employee,
                    employer_contribution=payslip.pf_employer,
                )

            if payslip.ssf_employee or payslip.ssf_employer:
                SocialSecurityFundRecord.objects.create(
                    org=org,
                    member=payslip.member,
                    payslip=payslip,
                    month_name=payslip.month_name or '',
                    employee_contribution=payslip.ssf_employee,
                    employer_contribution=payslip.ssf_employer,
                )

            # Auto-log net salary as Finance expense
            if org and payslip.net_payable:
                sal_cat, _ = TransactionCategory.objects.get_or_create(
                    org=org, name='Salary Payment', transaction_type='expense'
                )
                FinancialTransaction.objects.create(
                    org=org,
                    transaction_type='expense',
                    title=f"Salary — {payslip.member.name} ({payslip.month_name or 'Payslip'})",
                    amount=payslip.net_payable,
                    category=sal_cat,
                    reference_number=str(payslip.id),
                    note=f"Auto-linked from payslip #{payslip.id}",
                    created_by=request.user,
                )

            # Mark advance/loan adjustments that fed this payslip as 'applied'
            if org and (payslip.advance_deduction or payslip.loan_deduction):
                PayrollAdjustment.objects.filter(
                    org=org,
                    member=payslip.member,
                    status='active',
                    adjustment_type__in=['advance', 'loan'],
                    effective_date__lte=payslip.to_date,
                ).update(status='applied')

        messages.success(request, "Successfully saved and generated Payslip!")
        return HttpResponseRedirect(reverse('schooladmin:play-slip-detail', args=(id,)))

    return HttpResponseRedirect(reverse('schooladmin:play-slip-detail', args=(id,)))


@feature_required('payroll')
@perm_required('can_generate_payroll')
def finalize_payslip(request, pk):
    """Toggle a payslip between draft → finalized → paid, or reset."""
    org = _get_org(request)
    slip = get_object_or_404(PaySlip, pk=pk, org=org)
    if request.method == 'POST':
        new_status = request.POST.get('status', 'finalized')
        slip.status = new_status
        slip.finalized_by = request.user
        update_fields = ['status', 'finalized_by', 'updated_at']
        if new_status == 'paid' and not slip.payment_date:
            slip.payment_date = timezone.localdate()
            update_fields.append('payment_date')
        slip.save(update_fields=update_fields)

        # When finalising, apply AdvanceSalary installments
        if new_status == 'finalized' and slip.advance_deduction:
            from handle.models import AdvanceSalary as _AdvSal
            for adv in _AdvSal.objects.filter(org=org, member=slip.member, status='active'):
                adv.apply_installment(payslip=slip)

        # Auto-accounting: accrual entry on finalize, payment entry when
        # actually marked paid — mirrors the two real transitions this view
        # already has, no invented workflow.
        if has_feature(org, 'accounting'):
            from handle.accounting import post_payroll_accrual_journal_entry, post_payroll_payment_journal_entry
            try:
                if new_status == 'finalized':
                    post_payroll_accrual_journal_entry(slip)
                elif new_status == 'paid':
                    post_payroll_payment_journal_entry(slip)
            except ValueError as e:
                messages.warning(request, f"Payslip updated, but the journal entry could not be posted: {e}")

        messages.success(request, f"Payslip marked as {new_status}.")
        return redirect('schooladmin:play-slip-detail', id=slip.member.id)
    return redirect('schooladmin:play-slip-detail', id=slip.member.id)


@feature_required('payroll')
@perm_required('can_view_payroll')
def paySlipViewDetail(request, pk):
    """Read-only detail page for one already-saved payslip."""
    org = _get_org(request)
    slip = get_object_or_404(PaySlip, pk=pk, org=org)

    if request.method == 'POST' and request.POST.get('action') == 'resend_email':
        if slip.member.email:
            from school.email_utils import send_payslip_email
            send_payslip_email(
                email=slip.member.email, name=slip.member.name,
                month_name=slip.month_name, net_payable=slip.net_payable,
                org_name=org.name,
                details={
                    'Gross Salary': f"Rs. {slip.gross_salary}",
                    'Allowances': f"Rs. {slip.allowance_total}",
                    'Deductions': f"Rs. {slip.advance_deduction + slip.loan_deduction}",
                    'PF': f"Rs. {slip.pf_employee}",
                    'Tax': f"Rs. {slip.tax_deduction}",
                    'Present Days': slip.present_days,
                    'Total Days': slip.total_days,
                },
                org=org, related_object_id=slip.id, force=True,
            )
            messages.success(request, f"Payslip emailed to {slip.member.email}.")
        else:
            messages.error(request, "Member has no email address on file.")
        return redirect('schooladmin:payslip_detail', pk=pk)

    memb = slip.member
    daily_logs = []
    if slip.from_date and slip.to_date:
        stats, daily_logs = calculate_attendance_stats(memb, slip.from_date, slip.to_date, org)
        # Enrich each day with an 'amount' key for display only — the saved
        # slip.* totals remain authoritative; this just shows how the saved
        # figures break down day by day.
        policy = PayrollPolicy.objects.filter(org=org).first()
        if policy:
            calculate_payroll_components(memb, stats, org, policy, slip.to_date, daily_logs=daily_logs)
    from school.print_settings import get_print_preference
    return render(request, 'admin/payslip_detail.html', {
        'org': org,
        'mem': memb,
        'slip': slip,
        'daily_logs': daily_logs,
        'nepali_enabled': getattr(org, 'nepali_date', False),
        'print_preference': get_print_preference(request.user, 'payslip', org=org),
    })


class PayslipListView(FeatureRequiredMixin, PermRequiredMixin, View):
    """All generated payslips for the org, filterable by branch, classification, and date range."""
    required_feature = 'payroll'
    required_perm = 'can_view_payroll'
    template_name = 'admin/payroll/payslip_list.html'

    def get(self, request, *args, **kwargs):
        from django.core.paginator import Paginator
        from django.db.models import Sum
        org = _get_org(request)

        branch_id = request.GET.get('branch', '')
        classification_id = request.GET.get('classification', '')
        status_filter = request.GET.get('status', '')
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')

        slips = PaySlip.objects.filter(org=org).select_related(
            'member', 'member__branch', 'member__classification'
        ).order_by('-from_date', '-id')

        if branch_id:
            slips = slips.filter(member__branch_id=branch_id)
        if classification_id:
            slips = slips.filter(member__classification_id=classification_id)
        if status_filter:
            slips = slips.filter(status=status_filter)
        if from_date:
            slips = slips.filter(from_date__gte=from_date)
        if to_date:
            slips = slips.filter(to_date__lte=to_date)

        total_net_payable = slips.aggregate(t=Sum('net_payable'))['t'] or 0
        paginator = Paginator(slips, 30)
        page_obj = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'org': org,
            'slips': page_obj,
            'page_obj': page_obj,
            'total_count': slips.count(),
            'total_net_payable': total_net_payable,
            'branches': Branch.objects.filter(org=org, status='active').order_by('name'),
            'classifications': Classification.objects.filter(org=org).order_by('name'),
            'status_choices': PaySlip.STATUS_CHOICES,
            'selected_branch': branch_id,
            'selected_classification': classification_id,
            'selected_status': status_filter,
            'from_date': from_date,
            'to_date': to_date,
        })

    def post(self, request, *args, **kwargs):
        org = _get_org(request)
        if request.POST.get('action') == 'bulk_resend_email':
            from school.email_utils import send_payslip_email
            ids = request.POST.getlist('slip_ids')
            slips = PaySlip.objects.filter(org=org, id__in=ids).select_related('member')
            sent, skipped = 0, 0
            for slip in slips:
                if not slip.member or not slip.member.email:
                    skipped += 1
                    continue
                send_payslip_email(
                    email=slip.member.email, name=slip.member.name,
                    month_name=slip.month_name, net_payable=slip.net_payable,
                    org_name=org.name,
                    details={
                        'Gross Salary': f"Rs. {slip.gross_salary}",
                        'Allowances': f"Rs. {slip.allowance_total}",
                        'Deductions': f"Rs. {slip.advance_deduction + slip.loan_deduction}",
                        'PF': f"Rs. {slip.pf_employee}",
                        'Tax': f"Rs. {slip.tax_deduction}",
                        'Present Days': slip.present_days,
                        'Total Days': slip.total_days,
                    },
                    org=org, related_object_id=slip.id, force=True,
                )
                sent += 1
            msg = f"Resent {sent} payslip email(s)."
            if skipped:
                msg += f" Skipped {skipped} (no member email)."
            messages.success(request, msg)
        return redirect('schooladmin:payslip_list')


def addHoliday(request):
    org =  request.user.schooladmin.org
    if request.method == 'POST':
        days = request.POST.getlist('day')
        for i in days:
            Holiday.objects.create(holiday = i, org = org)
        print(days)
        messages.success(request,"Successfully added holiday")
    return HttpResponseRedirect(reverse('schooladmin:orgDetail'))


def updateHoliday(request):
    org = request.user.schooladmin.org
    if request.method == 'POST':
        current_holidays = Holiday.objects.filter(org=org)
        selected_days = request.POST.getlist('day')

        # Convert the selected days and current holidays to sets for easier comparison
        selected_days_set = set(selected_days)
        current_holidays_set = set(holiday.holiday for holiday in current_holidays)

        # Holidays to add
        holidays_to_add = selected_days_set - current_holidays_set
        for day in holidays_to_add:
            # Check if the holiday already exists to avoid duplicates
            if not Holiday.objects.filter(holiday=day, org=org).exists():
                Holiday.objects.create(holiday=day, org=org)
                print(f"Adding holiday: {day}")

        # Holidays to remove
        holidays_to_remove = current_holidays_set - selected_days_set
        for day in holidays_to_remove:
            holiday_to_delete = Holiday.objects.filter(holiday=day, org=org).first()
            if holiday_to_delete:
                holiday_to_delete.delete()
                print(f"Deleting holiday: {day}")

        messages.success(request, "Successfully updated holidays")
        return HttpResponseRedirect(reverse('schooladmin:orgDetail'))


def addOccasion(request):
    org = request.user.schooladmin.org

    if request.method == 'POST':
        name = request.POST['occasion']
        start_date = request.POST['ocDate']
        holiday_type = request.POST['holidayType']
        end_date = request.POST.get('ocEndDate', None)

        if holiday_type == 'gap' and end_date:
            Occasion.objects.create(org=org, date=start_date, end_date=end_date, name=name)
        else:
            Occasion.objects.create(org=org, date=start_date, name=name)

        messages.success(request, "Successfully added occasion")
        return HttpResponseRedirect(reverse('schooladmin:orgDetail'))


def addOrgShiftOverride(request):
    """Company-wide shift-time change for one specific date (e.g. a half-day
    Friday) — applies to members on the plain default-shift path, not those
    with a Shift Management assignment. See handle.models.member.shift_windows_detailed."""
    org = request.user.schooladmin.org

    if request.method == 'POST':
        date = request.POST.get('override_date')
        start_time = request.POST.get('override_start_time')
        end_time = request.POST.get('override_end_time')
        note = request.POST.get('override_note', '').strip()

        if not (date and start_time and end_time):
            messages.error(request, "Date, start time and end time are all required.")
            return HttpResponseRedirect(reverse('schooladmin:orgDetail'))

        _, created = OrganizationShiftOverride.objects.update_or_create(
            org=org, date=date,
            defaults={'start_time': start_time, 'end_time': end_time, 'note': note},
        )
        messages.success(
            request,
            f"Company-wide shift for {date} {'set' if created else 'updated'} to {start_time}–{end_time}.",
        )
        return HttpResponseRedirect(reverse('schooladmin:orgDetail'))


def deleteOrgShiftOverride(request, pk):
    org = request.user.schooladmin.org
    if request.method == 'POST':
        OrganizationShiftOverride.objects.filter(pk=pk, org=org).delete()
        messages.success(request, "Company-wide shift change removed.")
    return HttpResponseRedirect(reverse('schooladmin:orgDetail'))


def staffMake(request):
    auser = request.user
    if auser.user_type == "2":
        org = auser.schooladmin.org
    elif auser.user_type == "3":
        org = auser.staff.org
    else:
        org = None
    memb = member.objects.filter(org=org).exclude(status='dumped')

    dist = {
        'member': memb,
        'classifications': Classification.objects.filter(org=org),
        'courses': Course.objects.filter(org=org, status='active').prefetch_related('classifications'),
        'org': org,
    }
    if request.method == 'POST':
        classification_ids = request.POST.getlist('classifications')
        course_ids = request.POST.getlist('courses')
        memId = request.POST['member']
        memb_obj = get_object_or_404(member, id=memId, org=org)
        classifications = Classification.objects.filter(id__in=classification_ids, org=org)

        if not CustomUser.objects.filter(email=memb_obj.email).exists():
            TypeOne = CustomUser.objects.create_user(
                first_name=memb_obj.name,
                last_name=memb_obj.name,
                email=memb_obj.email,
                username=memb_obj.email,
                password=str(memb_obj.phone)
            )
            TypeOne.user_type = "3"
            TypeOne.save()
            Staff.objects.create(member=memb_obj, admin=TypeOne, org=org, number=memb_obj.phone)
            for classification in classifications:
                AttendingClassification.objects.create(staff=TypeOne, classification=classification)
            # Assign teacher to selected courses
            if course_ids:
                Course.objects.filter(id__in=course_ids, org=org).update(teacher=TypeOne)
        else:
            existing_user = CustomUser.objects.get(email=memb_obj.email)
            for classification in classifications:
                AttendingClassification.objects.get_or_create(staff=existing_user, classification=classification)
            if course_ids:
                Course.objects.filter(id__in=course_ids, org=org).update(teacher=existing_user)

        messages.success(request, f"Portal access granted for {memb_obj.name}.")

    return render(request, "admin/addStaff.html", dist)
    


def updateClass(request, id):
    org = _get_org(request)
    mem = get_object_or_404(member, id=id, org=org)
    staf = get_object_or_404(Staff, admin__email=mem.email, org=org)
    current_classifications = AttendingClassification.objects.filter(staff=staf.admin)

    if request.method == 'POST':
        selected_classifications = request.POST.getlist('classifications')
        selected_courses = request.POST.getlist('courses')

        current_classifications = AttendingClassification.objects.filter(staff=staf.admin)
        current_classification_ids = [str(c.classification.id) for c in current_classifications]

        to_add = set(selected_classifications) - set(current_classification_ids)
        to_remove = set(current_classification_ids) - set(selected_classifications)

        for classification_id in to_add:
            classification = get_object_or_404(Classification, id=classification_id, org=org)
            AttendingClassification.objects.create(staff=staf.admin, classification=classification)

        for classification_id in to_remove:
            AttendingClassification.objects.filter(
                staff=staf.admin, classification__id=classification_id
            ).delete()

        # Unassign this teacher from all courses in the org, then reassign selected
        Course.objects.filter(org=org, teacher=staf.admin).update(teacher=None)
        if selected_courses:
            Course.objects.filter(id__in=selected_courses, org=org).update(teacher=staf.admin)

        messages.success(request, "Classifications and courses updated successfully.")
        return HttpResponseRedirect(reverse('schooladmin:updateClass', args=(mem.id,)))

    clas = Classification.objects.filter(org=org)
    current_class_ids = [c.classification.id for c in current_classifications]
    all_courses = Course.objects.filter(org=org, status='active').prefetch_related('classifications')
    assigned_course_ids = list(Course.objects.filter(org=org, teacher=staf.admin).values_list('id', flat=True))

    dist = {
        'clas': clas,
        'staf': staf,
        'cas': current_classifications,
        'mem': mem,
        'holiday': current_class_ids,
        'courses': all_courses,
        'assigned_course_ids': assigned_course_ids,
        'org': org,
    }

    return render(request, "admin/updateClass.html", dist)




class AdminWifiManageView(LoginRequiredMixin, View):
    template_name = 'admin/wifi_manage.html' # Adjust path to match your folder structure

    def get(self, request, *args, **kwargs):
        # Security: Ensure only Admins (type "2") can access this
        if request.user.user_type != "2":
            messages.error(request, "Access Denied. Only Administrators can manage WiFi networks.")
            return redirect('dashboard') # Redirect to their respective dashboard

        org = request.user.schooladmin.org
        networks = WifiBased.objects.filter(org=org)
        
        return render(request, self.template_name, {
            'networks': networks, 
            'org': org
        })

    def post(self, request, *args, **kwargs):
        if request.user.user_type != "2":
            return redirect('dashboard')

        org = request.user.schooladmin.org
        
        # Handle Delete Action
        if 'delete_id' in request.POST:
            WifiBased.objects.filter(id=request.POST.get('delete_id'), org=org).delete()
            messages.success(request, "Authorized WiFi network removed successfully.")
            return redirect('schooladmin:admin_wifi_manage') # Adjust namespace to match your urls.py

        # Handle Add Action
        name = request.POST.get('name')
        ssid = request.POST.get('ssid')
        bssid = request.POST.get('bssid', '').strip().upper() # Clean up BSSID format

        if name and ssid and bssid:
            # Prevent duplicates
            if WifiBased.objects.filter(org=org, bssid=bssid).exists():
                messages.warning(request, "This BSSID is already registered to your organization.")
            else:
                WifiBased.objects.create(org=org, name=name, ssid=ssid, bssid=bssid)
                messages.success(request, f"Network '{name}' is now authorized for mobile check-ins.")
        else:
            messages.error(request, "All fields (Location Name, SSID, BSSID) are required.")

        return redirect('schooladmin:admin_wifi_manage')


# =============================================================
# FINANCE MODULE
# =============================================================

def _get_org(request):
    if request.user.user_type == "2":
        return request.user.schooladmin.org
    if request.user.user_type == "3":
        return request.user.staff.org
    return None


class FinanceDashboardView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    template_name = 'admin/finance/dashboard.html'

    def get(self, request):
        org = _get_org(request)
        today = datetime.date.today()
        month_start = today.replace(day=1)

        branch_id = request.GET.get('branch', '')

        income_qs = FinancialTransaction.objects.filter(org=org, transaction_type='income')
        expense_qs = FinancialTransaction.objects.filter(org=org, transaction_type='expense')

        if branch_id:
            income_qs = income_qs.filter(branch_id=branch_id)
            expense_qs = expense_qs.filter(branch_id=branch_id)

        total_income = income_qs.aggregate(t=Sum('amount'))['t'] or 0
        total_expense = expense_qs.aggregate(t=Sum('amount'))['t'] or 0
        today_income = income_qs.filter(transaction_date=today).aggregate(t=Sum('amount'))['t'] or 0
        today_expense = expense_qs.filter(transaction_date=today).aggregate(t=Sum('amount'))['t'] or 0
        month_income = income_qs.filter(transaction_date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0
        month_expense = expense_qs.filter(transaction_date__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0

        recent_qs = FinancialTransaction.objects.filter(org=org).select_related('branch', 'category').order_by('-transaction_date')
        if branch_id:
            recent_qs = recent_qs.filter(branch_id=branch_id)
        recent_transactions = recent_qs[:10]
        categories = TransactionCategory.objects.filter(org=org)
        branches = Branch.objects.filter(org=org, status='active')

        context = {
            'org': org,
            'total_income': total_income,
            'total_expense': total_expense,
            'net_balance': total_income - total_expense,
            'today_income': today_income,
            'today_expense': today_expense,
            'month_income': month_income,
            'month_expense': month_expense,
            'recent_transactions': recent_transactions,
            'categories': categories,
            'branches': branches,
            'selected_branch': branch_id,
        }

        # Accounting-linked panel — read-only view into the real double-entry
        # ledger for Purchase/Sale-driven entries. No FinancialTransaction
        # mirroring: this is intentionally a separate, clearly-labeled view
        # into the ledger so nothing gets counted twice.
        if has_feature(org, 'accounting'):
            from handle.models import JournalEntryLine
            linked_lines = JournalEntryLine.objects.filter(
                entry__org=org, entry__status='approved', entry__source__in=['purchase', 'sale', 'stock_adjustment'],
            ).select_related('entry', 'account')
            purchase_total = linked_lines.filter(entry__source='purchase').aggregate(t=Sum('debit'))['t'] or 0
            sale_total = linked_lines.filter(entry__source='sale', account__account_type='income').aggregate(t=Sum('credit'))['t'] or 0
            context.update({
                'accounting_linked': True,
                'linked_purchase_total': purchase_total,
                'linked_sale_total': sale_total,
                'linked_recent_entries': linked_lines.order_by('-entry__entry_date', '-entry__id')[:10],
            })
        return render(request, self.template_name, context)


class IncomeListView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    template_name = 'admin/finance/income_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = FinancialTransaction.objects.filter(org=org, transaction_type='income').select_related('branch', 'category', 'created_by').order_by('-transaction_date')

        branch_id = request.GET.get('branch')
        category_id = request.GET.get('category')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        payment_method = request.GET.get('payment_method')

        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        if category_id:
            qs = qs.filter(category_id=category_id)
        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)
        if payment_method:
            qs = qs.filter(payment_method=payment_method)

        total = qs.aggregate(t=Sum('amount'))['t'] or 0
        nepali_enabled = getattr(org, 'nepali_date', False)
        transactions = list(qs)
        if nepali_enabled:
            for t in transactions:
                t.transaction_date_np = to_bs_display(t.transaction_date)
        context = {
            'org': org,
            'transactions': transactions,
            'total': total,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': TransactionCategory.objects.filter(org=org, transaction_type='income'),
            'selected_branch': branch_id,
            'selected_category': category_id,
            'date_from': date_from,
            'date_to': date_to,
            'payment_method': payment_method,
            'nepali_enabled': nepali_enabled,
        }
        return render(request, self.template_name, context)


class AddIncomeView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    template_name = 'admin/finance/add_income.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': TransactionCategory.objects.filter(org=org, transaction_type='income'),
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        try:
            title = request.POST.get('title', '').strip()
            amount = request.POST.get('amount')
            category_id = request.POST.get('category')
            branch_id = request.POST.get('branch')
            transaction_date = request.POST.get('transaction_date')
            payment_method = request.POST.get('payment_method', 'cash')
            note = request.POST.get('note', '')
            reference = request.POST.get('reference_number', '')

            if not all([title, amount, transaction_date]):
                messages.error(request, "Title, amount, and date are required.")
                return redirect('schooladmin:add_income')

            amount = Decimal(amount)
            tx = FinancialTransaction.objects.create(
                org=org,
                transaction_type='income',
                title=title,
                amount=amount,
                category_id=category_id if category_id else None,
                branch_id=branch_id if branch_id else None,
                transaction_date=transaction_date,
                payment_method=payment_method,
                note=note,
                reference_number=reference,
                created_by=request.user,
            )
            if has_feature(org, 'accounting'):
                from handle.accounting import post_income_journal_entry
                try:
                    post_income_journal_entry(tx)
                except ValueError as e:
                    messages.warning(request, f"Income saved, but the journal entry could not be posted: {e}")
            messages.success(request, f"Income '{title}' added successfully.")
            return redirect('schooladmin:income_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:add_income')


class EditIncomeView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    def get(self, request, pk):
        org = _get_org(request)
        tx = get_object_or_404(FinancialTransaction, pk=pk, org=org, transaction_type='income')
        context = {
            'org': org,
            'tx': tx,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': TransactionCategory.objects.filter(org=org, transaction_type='income'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, 'admin/finance/edit_income.html', context)

    def post(self, request, pk):
        org = _get_org(request)
        tx = get_object_or_404(FinancialTransaction, pk=pk, org=org, transaction_type='income')
        tx.title = request.POST.get('title', tx.title)
        tx.amount = request.POST.get('amount', tx.amount)
        tx.category_id = request.POST.get('category') or None
        tx.branch_id = request.POST.get('branch') or None
        tx.transaction_date = request.POST.get('transaction_date', tx.transaction_date)
        tx.payment_method = request.POST.get('payment_method', tx.payment_method)
        tx.note = request.POST.get('note', '')
        tx.reference_number = request.POST.get('reference_number', '')
        tx.save()
        messages.success(request, "Income updated successfully.")
        return redirect('schooladmin:income_list')


@feature_required('finance')
def delete_income(request, pk):
    org = _get_org(request)
    tx = get_object_or_404(FinancialTransaction, pk=pk, org=org, transaction_type='income')
    tx.delete()
    messages.success(request, "Income deleted.")
    return redirect('schooladmin:income_list')


class ExpenseListView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    template_name = 'admin/finance/expense_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = FinancialTransaction.objects.filter(org=org, transaction_type='expense').select_related('branch', 'category', 'created_by').order_by('-transaction_date')

        branch_id = request.GET.get('branch')
        category_id = request.GET.get('category')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        payment_method = request.GET.get('payment_method')

        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        if category_id:
            qs = qs.filter(category_id=category_id)
        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)
        if payment_method:
            qs = qs.filter(payment_method=payment_method)

        total = qs.aggregate(t=Sum('amount'))['t'] or 0
        nepali_enabled = getattr(org, 'nepali_date', False)
        transactions = list(qs)
        if nepali_enabled:
            for t in transactions:
                t.transaction_date_np = to_bs_display(t.transaction_date)
        context = {
            'org': org,
            'transactions': transactions,
            'total': total,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': TransactionCategory.objects.filter(org=org, transaction_type='expense'),
            'selected_branch': branch_id,
            'selected_category': category_id,
            'date_from': date_from,
            'date_to': date_to,
            'payment_method': payment_method,
            'nepali_enabled': nepali_enabled,
        }
        return render(request, self.template_name, context)


class AddExpenseView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': TransactionCategory.objects.filter(org=org, transaction_type='expense'),
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, 'admin/finance/add_expense.html', context)

    def post(self, request):
        org = _get_org(request)
        try:
            title = request.POST.get('title', '').strip()
            amount = request.POST.get('amount')
            category_id = request.POST.get('category')
            branch_id = request.POST.get('branch')
            transaction_date = request.POST.get('transaction_date')
            payment_method = request.POST.get('payment_method', 'cash')
            note = request.POST.get('note', '')
            reference = request.POST.get('reference_number', '')

            if not all([title, amount, transaction_date]):
                messages.error(request, "Title, amount, and date are required.")
                return redirect('schooladmin:add_expense')

            amount = Decimal(amount)
            tx = FinancialTransaction.objects.create(
                org=org,
                transaction_type='expense',
                title=title,
                amount=amount,
                category_id=category_id if category_id else None,
                branch_id=branch_id if branch_id else None,
                transaction_date=transaction_date,
                payment_method=payment_method,
                note=note,
                reference_number=reference,
                created_by=request.user,
            )
            if has_feature(org, 'accounting'):
                from handle.accounting import post_expense_journal_entry
                try:
                    post_expense_journal_entry(tx)
                except ValueError as e:
                    messages.warning(request, f"Expense saved, but the journal entry could not be posted: {e}")
            messages.success(request, f"Expense '{title}' added successfully.")
            return redirect('schooladmin:expense_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:add_expense')


class EditExpenseView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    def get(self, request, pk):
        org = _get_org(request)
        tx = get_object_or_404(FinancialTransaction, pk=pk, org=org, transaction_type='expense')
        context = {
            'org': org,
            'tx': tx,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': TransactionCategory.objects.filter(org=org, transaction_type='expense'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, 'admin/finance/edit_expense.html', context)

    def post(self, request, pk):
        org = _get_org(request)
        tx = get_object_or_404(FinancialTransaction, pk=pk, org=org, transaction_type='expense')
        tx.title = request.POST.get('title', tx.title)
        tx.amount = request.POST.get('amount', tx.amount)
        tx.category_id = request.POST.get('category') or None
        tx.branch_id = request.POST.get('branch') or None
        tx.transaction_date = request.POST.get('transaction_date', tx.transaction_date)
        tx.payment_method = request.POST.get('payment_method', tx.payment_method)
        tx.note = request.POST.get('note', '')
        tx.reference_number = request.POST.get('reference_number', '')
        tx.save()
        messages.success(request, "Expense updated successfully.")
        return redirect('schooladmin:expense_list')


@feature_required('finance')
def delete_expense(request, pk):
    org = _get_org(request)
    tx = get_object_or_404(FinancialTransaction, pk=pk, org=org, transaction_type='expense')
    tx.delete()
    messages.success(request, "Expense deleted.")
    return redirect('schooladmin:expense_list')


class FinanceCategoryView(FeatureRequiredMixin, View):
    required_feature = 'finance'
    template_name = 'admin/finance/categories.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'categories': TransactionCategory.objects.filter(org=org).order_by('transaction_type', 'name'),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            tx_type = request.POST.get('transaction_type', 'income')
            if name:
                TransactionCategory.objects.get_or_create(org=org, name=name, transaction_type=tx_type)
                messages.success(request, f"Category '{name}' added.")
            else:
                messages.error(request, "Category name is required.")
        elif action == 'delete':
            cat_id = request.POST.get('cat_id')
            TransactionCategory.objects.filter(pk=cat_id, org=org).delete()
            messages.success(request, "Category deleted.")
        return redirect('schooladmin:finance_categories')


# =============================================================
# STOCK MODULE
# =============================================================

class StockDashboardView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/dashboard.html'

    def get(self, request):
        org = _get_org(request)
        branch_id = request.GET.get('branch', '')
        category_id = request.GET.get('category', '')

        items = StockItem.objects.filter(org=org, status='active').select_related('category', 'branch')
        movements_qs = StockMovement.objects.filter(org=org).select_related('item', 'branch').order_by('-movement_date')

        if branch_id:
            items = items.filter(branch_id=branch_id)
            movements_qs = movements_qs.filter(branch_id=branch_id)
        if category_id:
            items = items.filter(category_id=category_id)

        low_stock = [i for i in items if i.is_low_stock]
        total_items = items.count()
        total_value = sum((i.quantity * (i.purchase_cost or 0)) for i in items)
        recent_movements = movements_qs[:10]
        categories = StockCategory.objects.filter(org=org)
        branches = Branch.objects.filter(org=org, status='active')
        context = {
            'org': org,
            'items': items,
            'low_stock': low_stock,
            'low_stock_count': len(low_stock),
            'total_items': total_items,
            'total_value': total_value,
            'recent_movements': recent_movements,
            'categories': categories,
            'branches': branches,
            'selected_branch': branch_id,
            'selected_category': category_id,
        }
        return render(request, self.template_name, context)


class StockItemListView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/item_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = StockItem.objects.filter(org=org).select_related('category', 'branch').order_by('name')
        branch_id = request.GET.get('branch')
        category_id = request.GET.get('category')
        status = request.GET.get('status')
        search = request.GET.get('search', '')
        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        if category_id:
            qs = qs.filter(category_id=category_id)
        if status:
            qs = qs.filter(status=status)
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(sku__icontains=search))
        context = {
            'org': org,
            'items': qs,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': StockCategory.objects.filter(org=org),
            'selected_branch': branch_id,
            'selected_category': category_id,
            'selected_status': status,
            'search': search,
        }
        return render(request, self.template_name, context)


class AddStockItemView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/add_item.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': StockCategory.objects.filter(org=org),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        try:
            name = request.POST.get('name', '').strip()
            if not name:
                messages.error(request, "Item name is required.")
                return redirect('schooladmin:add_stock_item')
            init_qty = Decimal(str(request.POST.get('quantity', 0) or 0))
            init_cost = Decimal(str(request.POST.get('purchase_cost', 0) or 0))
            item = StockItem.objects.create(
                org=org,
                name=name,
                sku=request.POST.get('sku', ''),
                unit=request.POST.get('unit', 'pcs'),
                category_id=request.POST.get('category') or None,
                branch_id=request.POST.get('branch') or None,
                quantity=Decimal('0'),
                low_stock_threshold=int(request.POST.get('low_stock_threshold', 5) or 5),
                supplier=request.POST.get('supplier', ''),
                purchase_cost=init_cost,
                status=request.POST.get('status', 'active'),
            )
            if init_qty > 0:
                StockMovement.objects.create(
                    org=org,
                    branch=item.branch,
                    item=item,
                    created_by=request.user,
                    movement_type='in',
                    quantity=init_qty,
                    unit_cost=init_cost,
                    movement_date=datetime.date.today(),
                    note='Initial stock',
                )
            messages.success(request, f"Stock item '{name}' added.")
            return redirect('schooladmin:stock_items')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:add_stock_item')


class EditStockItemView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    def get(self, request, pk):
        org = _get_org(request)
        item = get_object_or_404(StockItem, pk=pk, org=org)
        context = {
            'org': org,
            'item': item,
            'branches': Branch.objects.filter(org=org, status='active'),
            'categories': StockCategory.objects.filter(org=org),
        }
        return render(request, 'admin/stock/edit_item.html', context)

    def post(self, request, pk):
        org = _get_org(request)
        item = get_object_or_404(StockItem, pk=pk, org=org)
        item.name = request.POST.get('name', item.name)
        item.sku = request.POST.get('sku', item.sku)
        item.unit = request.POST.get('unit', item.unit)
        item.category_id = request.POST.get('category') or None
        item.branch_id = request.POST.get('branch') or None
        item.low_stock_threshold = int(request.POST.get('low_stock_threshold', item.low_stock_threshold) or 0)
        item.supplier = request.POST.get('supplier', item.supplier)
        item.purchase_cost = Decimal(str(request.POST.get('purchase_cost', item.purchase_cost) or 0))
        item.status = request.POST.get('status', item.status)
        item.save()
        messages.success(request, "Stock item updated.")
        return redirect('schooladmin:stock_items')


@feature_required('stock')
def delete_stock_item(request, pk):
    org = _get_org(request)
    item = get_object_or_404(StockItem, pk=pk, org=org)
    item.delete()
    messages.success(request, "Stock item deleted.")
    return redirect('schooladmin:stock_items')


class StockInView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/stock_in.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'items': StockItem.objects.filter(org=org, status='active').order_by('name'),
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        item_id = request.POST.get('item')
        quantity = request.POST.get('quantity')
        unit_cost = request.POST.get('unit_cost', 0)
        note = request.POST.get('note', '')
        movement_date = request.POST.get('movement_date', datetime.date.today())
        if not item_id or not quantity:
            messages.error(request, "Item and quantity are required.")
            return redirect('schooladmin:stock_in')
        item = get_object_or_404(StockItem, pk=item_id, org=org)
        movement = StockMovement.objects.create(
            org=org, branch=item.branch, item=item,
            created_by=request.user, movement_type='in',
            quantity=Decimal(str(quantity)), unit_cost=Decimal(str(unit_cost or 0)),
            movement_date=movement_date, note=note,
        )
        # Auto-log purchase cost: real double-entry journal entry when
        # Accounting is enabled for this org, otherwise the original flat
        # FinancialTransaction (unchanged, for backward compatibility).
        add_as_expense = request.POST.get('add_as_expense') == 'on'
        unit_cost_val = float(unit_cost or 0)
        if add_as_expense and unit_cost_val > 0:
            if has_feature(org, 'accounting'):
                from handle.accounting import post_stock_in_journal_entry
                try:
                    post_stock_in_journal_entry(
                        item, Decimal(str(quantity)), Decimal(str(unit_cost_val)),
                        request.POST.get('payment_method', 'cash'), org,
                        branch=item.branch, created_by=request.user,
                    )
                except ValueError as e:
                    messages.warning(request, f"Stock recorded, but the journal entry could not be posted: {e}")
            else:
                total_cost = float(quantity) * unit_cost_val
                stock_cat, _ = TransactionCategory.objects.get_or_create(
                    org=org, name='Stock Purchase', transaction_type='expense'
                )
                FinancialTransaction.objects.create(
                    org=org,
                    transaction_type='expense',
                    title=f"Stock Purchase — {item.name}",
                    amount=Decimal(str(total_cost)),
                    category=stock_cat,
                    transaction_date=movement_date if movement_date else datetime.date.today(),
                    note=f"Auto-linked: {quantity} × {item.name} @ Rs.{unit_cost_val}",
                    created_by=request.user,
                )
        messages.success(request, f"Stock in recorded for '{item.name}'.")
        return redirect('schooladmin:stock_items')


class StockOutView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/stock_out.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'items': StockItem.objects.filter(org=org, status='active').order_by('name'),
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        item_id = request.POST.get('item')
        quantity = Decimal(str(request.POST.get('quantity', 0) or 0))
        note = request.POST.get('note', '')
        movement_date = request.POST.get('movement_date', datetime.date.today())
        if not item_id or quantity <= 0:
            messages.error(request, "Item and valid quantity are required.")
            return redirect('schooladmin:stock_out')
        item = get_object_or_404(StockItem, pk=item_id, org=org)
        if item.quantity < quantity:
            messages.error(request, f"Insufficient stock. Available: {item.quantity} {item.unit}")
            return redirect('schooladmin:stock_out')
        StockMovement.objects.create(
            org=org, branch=item.branch, item=item,
            created_by=request.user, movement_type='out',
            quantity=quantity, unit_cost=item.purchase_cost or Decimal('0'),
            movement_date=movement_date, note=note,
        )
        messages.success(request, f"Stock out recorded for '{item.name}'.")
        return redirect('schooladmin:stock_items')


class StockMovementHistoryView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/movement_history.html'

    def get(self, request):
        org = _get_org(request)
        qs = StockMovement.objects.filter(org=org).select_related('item', 'branch', 'created_by').order_by('-movement_date', '-id')
        item_id = request.GET.get('item')
        movement_type = request.GET.get('movement_type')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if item_id:
            qs = qs.filter(item_id=item_id)
        if movement_type:
            qs = qs.filter(movement_type=movement_type)
        if date_from:
            qs = qs.filter(movement_date__gte=date_from)
        if date_to:
            qs = qs.filter(movement_date__lte=date_to)
        nepali_enabled = getattr(org, 'nepali_date', False)
        movements = list(qs)
        if nepali_enabled:
            for m in movements:
                m.movement_date_np = to_bs_display(m.movement_date)
        context = {
            'org': org,
            'movements': movements,
            'items': StockItem.objects.filter(org=org).order_by('name'),
            'selected_item': item_id,
            'selected_type': movement_type,
            'date_from': date_from,
            'date_to': date_to,
            'nepali_enabled': nepali_enabled,
        }
        return render(request, self.template_name, context)


class StockCategoryView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/categories.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {
            'org': org,
            'categories': StockCategory.objects.filter(org=org),
        })

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            desc = request.POST.get('description', '')
            if name:
                StockCategory.objects.get_or_create(org=org, name=name, defaults={'description': desc})
                messages.success(request, f"Category '{name}' added.")
            else:
                messages.error(request, "Name is required.")
        elif action == 'delete':
            StockCategory.objects.filter(pk=request.POST.get('cat_id'), org=org).delete()
            messages.success(request, "Category deleted.")
        return redirect('schooladmin:stock_categories')


# =============================================================
# SUPPLIER / PURCHASE / SALE (extends Stock — required_feature = 'stock').
# Auto-posting to the double-entry ledger is additionally gated behind
# has_feature(org, 'accounting') at each call site; if an org doesn't have
# Accounting, these flows still work fully for Stock purposes and the
# journal-posting step is silently skipped.
# =============================================================

from handle.models import (
    Supplier, SupplierDocument, SupplierPayment, Purchase, PurchaseItem, PurchaseReturn, PurchaseReturnItem,
    Sale, SaleItem, SalePayment, SalesReturn, SalesReturnItem, AssetPurchase, Client,
)
from handle.forms import (
    SupplierForm, SupplierDocumentForm, SupplierPaymentForm,
    PurchaseForm, PurchaseItemFormSet, PurchaseReturnForm, PurchaseReturnItemFormSet,
    SaleForm, SaleItemFormSet, SalePaymentForm, SalesReturnForm, SalesReturnItemFormSet,
    AssetPurchaseForm, clean_sale_item_formset_stock,
)


class SupplierListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_view_purchases'
    template_name = 'admin/stock/supplier_list.html'

    def get(self, request):
        org = _get_org(request)
        search = request.GET.get('search', '')
        qs = Supplier.objects.filter(org=org).order_by('name')
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(contact_person__icontains=search))
        return render(request, self.template_name, {'org': org, 'suppliers': qs, 'search': search})


class AddSupplierView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_manage_purchases'
    template_name = 'admin/stock/supplier_form.html'

    def get(self, request):
        org = _get_org(request)
        form = SupplierForm(org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = SupplierForm(request.POST, org=org)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.org = org
            supplier.created_by = request.user
            supplier.save()
            messages.success(request, f"Supplier '{supplier.name}' added.")
            return redirect('schooladmin:supplier_detail', pk=supplier.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})


class EditSupplierView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_manage_purchases'
    template_name = 'admin/stock/supplier_form.html'

    def get(self, request, pk):
        org = _get_org(request)
        supplier = get_object_or_404(Supplier, pk=pk, org=org)
        form = SupplierForm(instance=supplier, org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'supplier': supplier, 'mode': 'edit'})

    def post(self, request, pk):
        org = _get_org(request)
        supplier = get_object_or_404(Supplier, pk=pk, org=org)
        form = SupplierForm(request.POST, instance=supplier, org=org)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated.")
            return redirect('schooladmin:supplier_detail', pk=supplier.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'supplier': supplier, 'mode': 'edit'})


class SupplierDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = ('can_view_purchases', 'can_manage_purchases')
    template_name = 'admin/stock/supplier_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        supplier = get_object_or_404(Supplier, pk=pk, org=org)
        purchases = supplier.purchases.select_related('branch').prefetch_related('items__stock_item').order_by('-purchase_date')
        purchase_returns = supplier.returns.order_by('-return_date')
        payments = supplier.payments.order_by('-payment_date')
        documents = supplier.documents.all()

        # Ledger: chronological merge of Purchase/Return/Payment with a
        # running balance (sub-ledger pattern — the real GL keeps one shared
        # Accounts Payable control account, not a row per supplier).
        events = []
        for p in supplier.purchases.exclude(status='cancelled'):
            events.append((p.purchase_date, p.created_at, 'Purchase', p, p.total_amount))
        for r in purchase_returns.filter(status='completed'):
            events.append((r.return_date, r.created_at, 'Return', r, -r.total_amount))
        for pay in payments:
            events.append((pay.payment_date, pay.created_at, 'Payment', pay, -pay.amount))
        events.sort(key=lambda e: (e[0], e[1]))
        running = supplier.opening_balance
        ledger_rows = []
        for date, _, kind, ref, amount in events:
            running += amount
            ledger_rows.append({'date': date, 'kind': kind, 'ref': ref, 'amount': amount, 'balance': running})

        active_purchases = purchases.exclude(status='cancelled')
        supplied_quantity = PurchaseItem.objects.filter(
            purchase__supplier=supplier,
            purchase__org=org,
            purchase__status__in=('received', 'paid'),
            stock_item__isnull=False,
        ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
        distinct_items_supplied = PurchaseItem.objects.filter(
            purchase__supplier=supplier,
            purchase__org=org,
            purchase__status__in=('received', 'paid'),
            stock_item__isnull=False,
        ).values('stock_item_id').distinct().count()
        outstanding_balance = supplier.outstanding_balance()
        available_credit = max(supplier.credit_limit - max(outstanding_balance, Decimal('0')), Decimal('0'))

        return render(request, self.template_name, {
            'org': org, 'supplier': supplier, 'purchases': purchases,
            'purchase_returns': purchase_returns, 'payments': payments, 'documents': documents,
            'ledger_rows': ledger_rows, 'opening_balance': supplier.opening_balance,
            'total_purchases': supplier.total_purchases(),
            'total_returns': supplier.total_returns(),
            'total_paid': supplier.total_paid(),
            'outstanding_balance': outstanding_balance,
            'available_credit': available_credit,
            'purchase_count': active_purchases.count(),
            'draft_purchase_count': active_purchases.filter(status='draft').count(),
            'received_purchase_count': active_purchases.filter(status__in=('received', 'paid')).count(),
            'last_purchase': active_purchases.first(),
            'supplied_quantity': supplied_quantity,
            'distinct_items_supplied': distinct_items_supplied,
            'document_form': SupplierDocumentForm(),
            'payment_form': SupplierPaymentForm(org=org, supplier=supplier, initial={'payment_date': datetime.date.today()}),
        })


@feature_required('stock')
@perm_required('can_manage_purchases')
def delete_supplier(request, pk):
    org = _get_org(request)
    supplier = get_object_or_404(Supplier, pk=pk, org=org)
    if supplier.purchases.exists():
        messages.error(request, "Cannot delete a supplier with existing purchases. Mark as inactive instead.")
    else:
        supplier.delete()
        messages.success(request, "Supplier deleted.")
    return redirect('schooladmin:supplier_list')


class ConvertClientToSupplierView(FeatureRequiredMixin, PermRequiredMixin, View):
    """Creates an independent Supplier pre-filled from a Client's contact
    details. The Client row is never modified — source_client only records
    where the data came from."""
    required_feature = 'stock'
    required_perm = 'can_manage_purchases'
    template_name = 'admin/stock/convert_client_to_supplier.html'

    def get(self, request, pk):
        org = _get_org(request)
        client = get_object_or_404(Client, pk=pk, org=org)
        form = SupplierForm(org=org, initial={
            'name': client.client_org_name,
            'contact_person': client.contact_person,
            'phone': client.phone,
            'email': client.email,
            'address': client.address,
        })
        return render(request, self.template_name, {'org': org, 'client': client, 'form': form})

    def post(self, request, pk):
        org = _get_org(request)
        client = get_object_or_404(Client, pk=pk, org=org)
        form = SupplierForm(request.POST, org=org)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.org = org
            supplier.source_client = client
            supplier.created_by = request.user
            supplier.save()
            messages.success(request, f"'{client.client_org_name}' converted to a supplier.")
            return redirect('schooladmin:supplier_detail', pk=supplier.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'client': client, 'form': form})


class PurchaseListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_view_purchases'
    template_name = 'admin/stock/purchase_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Purchase.objects.filter(org=org).select_related('supplier').order_by('-purchase_date')
        status = request.GET.get('status', '')
        if status:
            qs = qs.filter(status=status)
        return render(request, self.template_name, {'org': org, 'purchases': qs, 'selected_status': status})


class AddPurchaseView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_manage_purchases'
    template_name = 'admin/stock/purchase_form.html'

    def get(self, request):
        org = _get_org(request)
        source_supplier = Supplier.objects.filter(
            org=org, status='active', pk=request.GET.get('supplier'),
        ).first()
        initial = {'purchase_date': datetime.date.today()}
        if source_supplier:
            initial['supplier'] = source_supplier
            if source_supplier.branch_id:
                initial['branch'] = source_supplier.branch
        form = PurchaseForm(org=org, initial=initial)
        formset = PurchaseItemFormSet()
        self._scope_item_formset(formset, org)
        return render(request, self.template_name, {
            'org': org, 'form': form, 'formset': formset, 'mode': 'create',
            'source_supplier': source_supplier,
        })

    def post(self, request):
        org = _get_org(request)
        form = PurchaseForm(request.POST, request.FILES, org=org)
        formset = PurchaseItemFormSet(request.POST)
        self._scope_item_formset(formset, org)
        source_supplier = Supplier.objects.filter(
            org=org, pk=request.POST.get('source_supplier'),
        ).first()

        if not (form.is_valid() and formset.is_valid()):
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {
                'org': org, 'form': form, 'formset': formset, 'mode': 'create',
                'source_supplier': source_supplier,
            })

        with transaction.atomic():
            purchase = form.save(commit=False)
            purchase.org = org
            purchase.created_by = request.user
            purchase.save()
            for line_form in formset.forms:
                if not line_form.cleaned_data or line_form.cleaned_data.get('DELETE'):
                    continue
                if not (line_form.cleaned_data.get('stock_item') or line_form.cleaned_data.get('description')):
                    continue
                item = line_form.save(commit=False)
                item.purchase = purchase
                item.save()
            purchase.recalc_totals()
        messages.success(request, f"Purchase #{purchase.pk} created.")
        return redirect('schooladmin:purchase_detail', pk=purchase.pk)

    @staticmethod
    def _scope_item_formset(formset, org):
        for f in formset.forms:
            f.fields['stock_item'].queryset = StockItem.objects.filter(org=org, status='active')


class PurchaseDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = (
        'can_view_purchases', 'can_manage_purchases', 'can_manage_purchase_returns',
    )
    template_name = 'admin/stock/purchase_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        purchase = get_object_or_404(Purchase, pk=pk, org=org)
        return render(request, self.template_name, {
            'org': org, 'purchase': purchase,
            'items': purchase.items.select_related('stock_item'),
            'payments': purchase.allocated_payments.order_by('-payment_date'),
            'payment_form': SupplierPaymentForm(
                org=org, supplier=purchase.supplier,
                initial={'payment_date': datetime.date.today(), 'purchase': purchase.id},
            ),
        })


@feature_required('stock')
@perm_required('can_manage_purchases')
def receive_purchase(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:purchase_detail', pk=pk)
    org = _get_org(request)
    purchase = get_object_or_404(Purchase, pk=pk, org=org)
    if purchase.status != 'draft':
        messages.error(request, "Only a draft purchase can be received.")
        return redirect('schooladmin:purchase_detail', pk=pk)

    with transaction.atomic():
        for item in purchase.items.select_related('stock_item'):
            if not item.stock_item:
                continue
            StockMovement.objects.create(
                org=org, branch=purchase.branch or item.stock_item.branch, item=item.stock_item,
                created_by=request.user, movement_type='in',
                quantity=item.quantity, unit_cost=item.unit_cost,
                movement_date=purchase.purchase_date, note=f"Purchase #{purchase.pk} — {purchase.supplier.name}",
            )
        purchase.status = 'received'
        purchase.save(update_fields=['status', 'updated_at'])
        if has_feature(org, 'accounting'):
            from handle.accounting import post_purchase_journal_entry
            try:
                post_purchase_journal_entry(purchase)
            except ValueError as e:
                messages.warning(request, f"Stock received, but the journal entry could not be posted: {e}")
    messages.success(request, f"Purchase #{purchase.pk} received into stock.")
    return redirect('schooladmin:purchase_detail', pk=pk)


class SaleListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_view_sales'
    template_name = 'admin/stock/sale_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Sale.objects.filter(org=org).select_related('client').order_by('-sale_date')
        status = request.GET.get('status', '')
        if status:
            qs = qs.filter(status=status)
        return render(request, self.template_name, {'org': org, 'sales': qs, 'selected_status': status})


class AddSaleView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_manage_sales'
    template_name = 'admin/stock/sale_form.html'

    def get(self, request):
        org = _get_org(request)
        form = SaleForm(org=org, initial={'sale_date': datetime.date.today()})
        formset = SaleItemFormSet()
        self._scope_item_formset(formset, org)
        return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = SaleForm(request.POST, org=org)
        formset = SaleItemFormSet(request.POST)
        self._scope_item_formset(formset, org)

        if not (form.is_valid() and formset.is_valid()):
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'create'})

        stock_errors = clean_sale_item_formset_stock(formset)
        if stock_errors:
            for err in stock_errors:
                messages.error(request, err)
            return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'create'})

        with transaction.atomic():
            sale = form.save(commit=False)
            sale.org = org
            sale.created_by = request.user
            sale.save()
            for line_form in formset.forms:
                if not line_form.cleaned_data or line_form.cleaned_data.get('DELETE'):
                    continue
                if not line_form.cleaned_data.get('stock_item'):
                    continue
                item = line_form.save(commit=False)
                item.sale = sale
                item.save()
            sale.recalc_totals()
        messages.success(request, f"Sale #{sale.pk} created.")
        return redirect('schooladmin:sale_detail', pk=sale.pk)

    @staticmethod
    def _scope_item_formset(formset, org):
        for f in formset.forms:
            f.fields['stock_item'].queryset = StockItem.objects.filter(org=org, status='active')


class SaleDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = ('can_view_sales', 'can_manage_sales', 'can_manage_sales_returns')
    template_name = 'admin/stock/sale_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        sale = get_object_or_404(Sale, pk=pk, org=org)
        return render(request, self.template_name, {
            'org': org, 'sale': sale, 'items': sale.items.select_related('stock_item'),
            'payments': sale.payments.order_by('-payment_date'),
            'sale_returns': sale.returns.order_by('-return_date'),
            'payment_form': SalePaymentForm(org=org, initial={'payment_date': datetime.date.today()}),
        })


@feature_required('stock')
@perm_required('can_manage_sales')
def complete_sale(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:sale_detail', pk=pk)
    org = _get_org(request)
    sale = get_object_or_404(Sale, pk=pk, org=org)
    if sale.status != 'draft':
        messages.error(request, "Only a draft sale can be completed.")
        return redirect('schooladmin:sale_detail', pk=pk)

    items = list(sale.items.select_related('stock_item'))
    for item in items:
        if item.quantity > item.stock_item.quantity:
            messages.error(request, f"Cannot complete sale — only {item.stock_item.quantity} {item.stock_item.unit} of '{item.stock_item.name}' available.")
            return redirect('schooladmin:sale_detail', pk=pk)

    with transaction.atomic():
        for item in items:
            StockMovement.objects.create(
                org=org, branch=sale.branch or item.stock_item.branch, item=item.stock_item,
                created_by=request.user, movement_type='out',
                quantity=item.quantity, unit_cost=item.stock_item.purchase_cost,
                movement_date=sale.sale_date, note=f"Sale #{sale.pk}",
            )
        sale.status = 'completed'
        sale.save(update_fields=['status', 'updated_at'])
        if has_feature(org, 'accounting'):
            from handle.accounting import post_sale_journal_entry
            try:
                post_sale_journal_entry(sale)
            except ValueError as e:
                messages.warning(request, f"Sale completed, but the journal entry could not be posted: {e}")
    messages.success(request, f"Sale #{sale.pk} completed.")
    return redirect('schooladmin:sale_detail', pk=pk)


@feature_required('stock')
@perm_required('can_manage_purchases')
def add_supplier_document(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:supplier_detail', pk=pk)
    org = _get_org(request)
    supplier = get_object_or_404(Supplier, pk=pk, org=org)
    form = SupplierDocumentForm(request.POST, request.FILES)
    if form.is_valid():
        doc = form.save(commit=False)
        doc.org = org
        doc.supplier = supplier
        doc.uploaded_by = request.user
        doc.save()
        messages.success(request, "Document uploaded.")
    else:
        messages.error(request, "Please choose a file and title.")
    return redirect(f"{reverse('schooladmin:supplier_detail', kwargs={'pk': pk})}?tab=documents")


@feature_required('stock')
@perm_required('can_manage_purchases')
def add_supplier_payment(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:supplier_detail', pk=pk)
    org = _get_org(request)
    supplier = get_object_or_404(Supplier, pk=pk, org=org)
    form = SupplierPaymentForm(request.POST, org=org, supplier=supplier)
    if form.is_valid():
        payment = form.save(commit=False)
        payment.org = org
        payment.supplier = supplier
        payment.created_by = request.user
        payment.save()
        if has_feature(org, 'accounting'):
            from handle.accounting import post_supplier_payment_journal_entry
            try:
                post_supplier_payment_journal_entry(payment)
            except ValueError as e:
                messages.warning(request, f"Payment recorded, but the journal entry could not be posted: {e}")
        messages.success(request, f"Payment of Rs. {payment.amount} recorded.")
        if payment.purchase_id:
            return redirect('schooladmin:purchase_detail', pk=payment.purchase_id)
    else:
        messages.error(request, "Please correct the errors below.")
        purchase_id = request.POST.get('purchase')
        if purchase_id:
            return redirect('schooladmin:purchase_detail', pk=purchase_id)
    return redirect(f"{reverse('schooladmin:supplier_detail', kwargs={'pk': pk})}?tab=payments")


class PurchaseReturnListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = ('can_view_purchases', 'can_manage_purchase_returns')
    template_name = 'admin/stock/purchase_return_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = PurchaseReturn.objects.filter(org=org).select_related('supplier', 'purchase').order_by('-return_date')
        return render(request, self.template_name, {'org': org, 'returns': qs})


class AddPurchaseReturnView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_manage_purchase_returns'
    template_name = 'admin/stock/purchase_return_form.html'

    def get(self, request, purchase_pk):
        org = _get_org(request)
        purchase = get_object_or_404(Purchase, pk=purchase_pk, org=org)
        form = PurchaseReturnForm(org=org, initial={'return_date': datetime.date.today()})
        formset = PurchaseReturnItemFormSet()
        self._scope_item_formset(formset, purchase)
        return render(request, self.template_name, {'org': org, 'purchase': purchase, 'form': form, 'formset': formset})

    def post(self, request, purchase_pk):
        org = _get_org(request)
        purchase = get_object_or_404(Purchase, pk=purchase_pk, org=org)
        form = PurchaseReturnForm(request.POST, org=org)
        formset = PurchaseReturnItemFormSet(request.POST)
        self._scope_item_formset(formset, purchase)

        if not (form.is_valid() and formset.is_valid()):
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'org': org, 'purchase': purchase, 'form': form, 'formset': formset})

        with transaction.atomic():
            return_doc = form.save(commit=False)
            return_doc.org = org
            return_doc.purchase = purchase
            return_doc.supplier = purchase.supplier
            return_doc.created_by = request.user
            return_doc.save()
            for line_form in formset.forms:
                if not line_form.cleaned_data or line_form.cleaned_data.get('DELETE'):
                    continue
                if not line_form.cleaned_data.get('stock_item'):
                    continue
                item = line_form.save(commit=False)
                item.return_doc = return_doc
                item.save()
            return_doc.recalc_totals()
        messages.success(request, f"Purchase Return #{return_doc.pk} created.")
        return redirect('schooladmin:purchase_return_detail', pk=return_doc.pk)

    @staticmethod
    def _scope_item_formset(formset, purchase):
        for f in formset.forms:
            f.fields['stock_item'].queryset = StockItem.objects.filter(org=purchase.org, status='active')


class PurchaseReturnDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = ('can_view_purchases', 'can_manage_purchase_returns')
    template_name = 'admin/stock/purchase_return_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        return_doc = get_object_or_404(PurchaseReturn, pk=pk, org=org)
        return render(request, self.template_name, {
            'org': org, 'return_doc': return_doc,
            'items': return_doc.items.select_related('stock_item'),
        })


@feature_required('stock')
@perm_required('can_manage_purchase_returns')
def complete_purchase_return(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:purchase_return_detail', pk=pk)
    org = _get_org(request)
    return_doc = get_object_or_404(PurchaseReturn, pk=pk, org=org)
    if return_doc.status != 'draft':
        messages.error(request, "Only a draft return can be completed.")
        return redirect('schooladmin:purchase_return_detail', pk=pk)

    with transaction.atomic():
        for item in return_doc.items.select_related('stock_item'):
            if not item.stock_item:
                continue
            StockMovement.objects.create(
                org=org, branch=return_doc.branch or item.stock_item.branch, item=item.stock_item,
                created_by=request.user, movement_type='out',
                quantity=item.quantity, unit_cost=item.unit_cost,
                movement_date=return_doc.return_date, note=f"Purchase Return #{return_doc.pk}",
            )
        return_doc.status = 'completed'
        return_doc.save(update_fields=['status', 'updated_at'])
        if has_feature(org, 'accounting'):
            from handle.accounting import post_purchase_return_journal_entry
            try:
                post_purchase_return_journal_entry(return_doc)
            except ValueError as e:
                messages.warning(request, f"Return completed, but the journal entry could not be posted: {e}")
    messages.success(request, f"Purchase Return #{return_doc.pk} completed.")
    return redirect('schooladmin:purchase_return_detail', pk=pk)


@feature_required('stock')
@perm_required('can_manage_sales')
def add_sale_payment(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:sale_detail', pk=pk)
    org = _get_org(request)
    sale = get_object_or_404(Sale, pk=pk, org=org)
    form = SalePaymentForm(request.POST, org=org)
    if form.is_valid():
        payment = form.save(commit=False)
        payment.org = org
        payment.sale = sale
        payment.client = sale.client
        payment.created_by = request.user
        payment.save()
        if has_feature(org, 'accounting'):
            from handle.accounting import post_sale_payment_journal_entry
            try:
                post_sale_payment_journal_entry(payment)
            except ValueError as e:
                messages.warning(request, f"Payment recorded, but the journal entry could not be posted: {e}")
        messages.success(request, f"Payment of Rs. {payment.amount} recorded.")
    else:
        messages.error(request, "Please correct the errors below.")
    return redirect('schooladmin:sale_detail', pk=pk)


class SalesReturnListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = ('can_view_sales', 'can_manage_sales_returns')
    template_name = 'admin/stock/sales_return_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = SalesReturn.objects.filter(org=org).select_related('sale', 'client').order_by('-return_date')
        return render(request, self.template_name, {'org': org, 'returns': qs})


class AddSalesReturnView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = 'can_manage_sales_returns'
    template_name = 'admin/stock/sales_return_form.html'

    def get(self, request, sale_pk):
        org = _get_org(request)
        sale = get_object_or_404(Sale, pk=sale_pk, org=org)
        form = SalesReturnForm(org=org, initial={'return_date': datetime.date.today()})
        formset = SalesReturnItemFormSet()
        self._scope_item_formset(formset, sale)
        return render(request, self.template_name, {'org': org, 'sale': sale, 'form': form, 'formset': formset})

    def post(self, request, sale_pk):
        org = _get_org(request)
        sale = get_object_or_404(Sale, pk=sale_pk, org=org)
        form = SalesReturnForm(request.POST, org=org)
        formset = SalesReturnItemFormSet(request.POST)
        self._scope_item_formset(formset, sale)

        if not (form.is_valid() and formset.is_valid()):
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'org': org, 'sale': sale, 'form': form, 'formset': formset})

        with transaction.atomic():
            return_doc = form.save(commit=False)
            return_doc.org = org
            return_doc.sale = sale
            return_doc.client = sale.client
            return_doc.created_by = request.user
            return_doc.save()
            for line_form in formset.forms:
                if not line_form.cleaned_data or line_form.cleaned_data.get('DELETE'):
                    continue
                if not line_form.cleaned_data.get('stock_item'):
                    continue
                item = line_form.save(commit=False)
                item.return_doc = return_doc
                item.save()
            return_doc.recalc_totals()
        messages.success(request, f"Sales Return #{return_doc.pk} created.")
        return redirect('schooladmin:sales_return_detail', pk=return_doc.pk)

    @staticmethod
    def _scope_item_formset(formset, sale):
        for f in formset.forms:
            f.fields['stock_item'].queryset = StockItem.objects.filter(org=sale.org, status='active')


class SalesReturnDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'stock'
    required_perm = ('can_view_sales', 'can_manage_sales_returns')
    template_name = 'admin/stock/sales_return_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        return_doc = get_object_or_404(SalesReturn, pk=pk, org=org)
        return render(request, self.template_name, {
            'org': org, 'return_doc': return_doc,
            'items': return_doc.items.select_related('stock_item'),
        })


@feature_required('stock')
@perm_required('can_manage_sales_returns')
def complete_sales_return(request, pk):
    if request.method != 'POST':
        return redirect('schooladmin:sales_return_detail', pk=pk)
    org = _get_org(request)
    return_doc = get_object_or_404(SalesReturn, pk=pk, org=org)
    if return_doc.status != 'draft':
        messages.error(request, "Only a draft return can be completed.")
        return redirect('schooladmin:sales_return_detail', pk=pk)

    with transaction.atomic():
        for item in return_doc.items.select_related('stock_item'):
            StockMovement.objects.create(
                org=org, branch=return_doc.branch or item.stock_item.branch, item=item.stock_item,
                created_by=request.user, movement_type='in',
                quantity=item.quantity, unit_cost=item.stock_item.purchase_cost,
                movement_date=return_doc.return_date, note=f"Sales Return #{return_doc.pk}",
            )
        return_doc.status = 'completed'
        return_doc.save(update_fields=['status', 'updated_at'])
        if has_feature(org, 'accounting'):
            from handle.accounting import post_sales_return_journal_entry
            try:
                post_sales_return_journal_entry(return_doc)
            except ValueError as e:
                messages.warning(request, f"Return completed, but the journal entry could not be posted: {e}")
    messages.success(request, f"Sales Return #{return_doc.pk} completed.")
    return redirect('schooladmin:sales_return_detail', pk=pk)


class AssetPurchaseListView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/asset_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = AssetPurchase.objects.filter(org=org).select_related('vendor').order_by('-purchase_date')
        return render(request, self.template_name, {'org': org, 'assets': qs})


class AddAssetPurchaseView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/asset_form.html'

    def get(self, request):
        org = _get_org(request)
        form = AssetPurchaseForm(org=org, initial={'purchase_date': datetime.date.today()})
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = AssetPurchaseForm(request.POST, org=org)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.org = org
            asset.created_by = request.user
            asset.save()
            if has_feature(org, 'accounting'):
                from handle.accounting import post_asset_purchase_journal_entry
                try:
                    post_asset_purchase_journal_entry(asset)
                except ValueError as e:
                    messages.warning(request, f"Asset saved, but the journal entry could not be posted: {e}")
            messages.success(request, f"Asset '{asset.name}' recorded.")
            return redirect('schooladmin:asset_detail', pk=asset.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})


class AssetPurchaseDetailView(FeatureRequiredMixin, View):
    required_feature = 'stock'
    template_name = 'admin/stock/asset_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        asset = get_object_or_404(AssetPurchase, pk=pk, org=org)
        return render(request, self.template_name, {'org': org, 'asset': asset})


class StockAdjustmentView(FeatureRequiredMixin, View):
    """Manual stock adjustment/damage entry — genuinely new UI, modeled on
    StockOutView. `direction='increase'` uses movement_type='adjustment'
    (StockMovement.signed_quantity() already treats this as positive);
    `direction='decrease'` uses movement_type='damage' (already treated as
    negative) — no change to that existing method was needed."""
    required_feature = 'stock'
    template_name = 'admin/stock/stock_adjustment.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'items': StockItem.objects.filter(org=org, status='active').order_by('name'),
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        item_id = request.POST.get('item')
        direction = request.POST.get('direction', 'increase')
        quantity = Decimal(str(request.POST.get('quantity', 0) or 0))
        note = request.POST.get('note', '')
        movement_date = request.POST.get('movement_date', datetime.date.today())
        if not item_id or quantity <= 0:
            messages.error(request, "Item and valid quantity are required.")
            return redirect('schooladmin:stock_adjustment')
        item = get_object_or_404(StockItem, pk=item_id, org=org)
        if direction == 'decrease' and item.quantity < quantity:
            messages.error(request, f"Cannot decrease below zero. Available: {item.quantity} {item.unit}")
            return redirect('schooladmin:stock_adjustment')

        movement_type = 'adjustment' if direction == 'increase' else 'damage'
        with transaction.atomic():
            movement = StockMovement.objects.create(
                org=org, branch=item.branch, item=item,
                created_by=request.user, movement_type=movement_type,
                quantity=quantity, unit_cost=item.purchase_cost or Decimal('0'),
                movement_date=movement_date, note=note,
            )
            if has_feature(org, 'accounting'):
                from handle.accounting import post_stock_adjustment_journal_entry
                try:
                    post_stock_adjustment_journal_entry(movement, direction)
                except ValueError as e:
                    messages.warning(request, f"Stock adjusted, but the journal entry could not be posted: {e}")
        messages.success(request, f"Stock adjustment recorded for '{item.name}'.")
        return redirect('schooladmin:stock_items')


# =============================================================
# LIBRARY MANAGEMENT (premium — required_feature = 'library')
# =============================================================

class LibraryDashboardView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = (
        'can_view_library', 'can_manage_library', 'can_issue_books', 'can_return_books',
    )
    template_name = 'admin/library/dashboard.html'

    def get(self, request):
        org = _get_org(request)
        today = datetime.date.today()

        books = Book.objects.filter(org=org)
        issues = BookIssue.objects.filter(org=org)

        total_books = books.aggregate(total=Sum('quantity'))['total'] or 0
        available = books.aggregate(total=Sum('available_quantity'))['total'] or 0
        lost = books.aggregate(total=Sum('lost_quantity'))['total'] or 0
        damaged = books.aggregate(total=Sum('damaged_quantity'))['total'] or 0
        issued_count = issues.filter(status__in=('issued', 'overdue')).count()
        returned_count = issues.filter(status='returned').count()
        overdue_count = issues.filter(status__in=('issued', 'overdue'), due_date__lt=today).count()
        today_issued = issues.filter(issue_date=today).count()
        today_returned = issues.filter(return_date=today).count()

        recent_issues = issues.select_related('book', 'member').order_by('-issue_date', '-id')[:10]

        context = {
            'org': org,
            'total_books': total_books,
            'available': available,
            'issued_count': issued_count,
            'returned_count': returned_count,
            'overdue_count': overdue_count,
            'lost': lost,
            'damaged': damaged,
            'today_issued': today_issued,
            'today_returned': today_returned,
            'recent_issues': recent_issues,
            'title_count': books.count(),
        }
        return render(request, self.template_name, context)


class BookListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = ('can_view_library', 'can_manage_library', 'can_issue_books')
    template_name = 'admin/library/book_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Book.objects.filter(org=org).select_related('category', 'author', 'publisher', 'branch').order_by('title')
        search = request.GET.get('search', '').strip()
        category_id = request.GET.get('category')
        status = request.GET.get('status')
        if search:
            qs = qs.filter(Q(title__icontains=search) | Q(book_code__icontains=search) | Q(isbn__icontains=search))
        if category_id:
            qs = qs.filter(category_id=category_id)
        if status:
            qs = qs.filter(status=status)
        context = {
            'org': org,
            'books': qs,
            'categories': LibraryCategory.objects.filter(org=org),
            'search': search,
            'selected_category': category_id,
            'selected_status': status,
        }
        return render(request, self.template_name, context)


class AddBookView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/add_book.html'

    def get(self, request):
        org = _get_org(request)
        form = BookForm(org=org)
        return render(request, self.template_name, {'org': org, 'form': form})

    def post(self, request):
        org = _get_org(request)
        form = BookForm(request.POST, request.FILES, org=org)
        if form.is_valid():
            book = form.save(commit=False)
            book.org = org
            book.available_quantity = book.quantity
            book.save()
            messages.success(request, f"Book '{book.title}' added.")
            return redirect('schooladmin:library_books')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form})


class EditBookView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/edit_book.html'

    def get(self, request, pk):
        org = _get_org(request)
        book = get_object_or_404(Book, pk=pk, org=org)
        form = BookForm(instance=book, org=org)
        return render(request, self.template_name, {'org': org, 'book': book, 'form': form})

    def post(self, request, pk):
        org = _get_org(request)
        book = get_object_or_404(Book, pk=pk, org=org)
        prior_quantity = book.quantity
        form = BookForm(request.POST, request.FILES, instance=book, org=org)
        if form.is_valid():
            book = form.save(commit=False)
            # Extra copies added/removed via the quantity field flow straight
            # through to what's on the shelf, so available count stays correct.
            delta = book.quantity - prior_quantity
            book.available_quantity = max(book.available_quantity + delta, 0)
            book.save()
            messages.success(request, "Book updated.")
            return redirect('schooladmin:library_books')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'book': book, 'form': form})


@feature_required('library')
@perm_required('can_manage_library')
def delete_book(request, pk):
    org = _get_org(request)
    book = get_object_or_404(Book, pk=pk, org=org)
    if book.issues.filter(status__in=('issued', 'overdue')).exists():
        messages.error(request, "Cannot delete a book with copies currently issued.")
    else:
        book.delete()
        messages.success(request, "Book deleted.")
    return redirect('schooladmin:library_books')


class IssueBookView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_issue_books'
    template_name = 'admin/library/issue_book.html'

    def get(self, request):
        org = _get_org(request)
        form = BookIssueForm(org=org)
        settings_obj = LibrarySettings.for_org(org)
        context = {
            'org': org,
            'form': form,
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'default_due': (datetime.date.today() + datetime.timedelta(days=settings_obj.loan_period_days)).strftime('%Y-%m-%d'),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        form = BookIssueForm(request.POST, org=org)
        if form.is_valid():
            member_obj = form.cleaned_data['member']
            settings_obj = LibrarySettings.for_org(org)
            active_count = BookIssue.objects.filter(
                org=org, member=member_obj, status__in=('issued', 'overdue')
            ).count()
            if active_count >= settings_obj.max_books_per_member:
                messages.error(
                    request,
                    f"{member_obj.name} already holds the maximum of {settings_obj.max_books_per_member} book(s)."
                )
                return render(request, self.template_name, {
                    'org': org, 'form': form,
                    'today': datetime.date.today().strftime('%Y-%m-%d'),
                    'default_due': (datetime.date.today() + datetime.timedelta(days=settings_obj.loan_period_days)).strftime('%Y-%m-%d'),
                })
            issue = form.save(commit=False)
            issue.org = org
            issue.issued_by = request.user
            issue.status = 'issued'
            issue.save()
            book = issue.book
            book.available_quantity = max(book.available_quantity - 1, 0)
            book.save(update_fields=['available_quantity'])
            messages.success(request, f"'{book.title}' issued to {member_obj.name}.")
            return redirect('schooladmin:library_issues')
        messages.error(request, "Please correct the errors below.")
        settings_obj = LibrarySettings.for_org(org)
        return render(request, self.template_name, {
            'org': org, 'form': form,
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'default_due': (datetime.date.today() + datetime.timedelta(days=settings_obj.loan_period_days)).strftime('%Y-%m-%d'),
        })


class BookIssueHistoryView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = ('can_view_library', 'can_issue_books', 'can_return_books')
    template_name = 'admin/library/issue_history.html'

    def get(self, request):
        org = _get_org(request)
        qs = BookIssue.objects.filter(org=org).select_related('book', 'member', 'branch', 'issued_by', 'returned_by').order_by('-issue_date', '-id')
        status = request.GET.get('status')
        book_id = request.GET.get('book')
        member_id = request.GET.get('member')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if status:
            qs = qs.filter(status=status)
        if book_id:
            qs = qs.filter(book_id=book_id)
        if member_id:
            qs = qs.filter(member_id=member_id)
        if date_from:
            qs = qs.filter(issue_date__gte=date_from)
        if date_to:
            qs = qs.filter(issue_date__lte=date_to)
        # Flag anything overdue-but-still-open as 'overdue' for display without a batch job.
        today = datetime.date.today()
        context = {
            'org': org,
            'issues': qs,
            'today': today,
            'books': Book.objects.filter(org=org).order_by('title'),
            'selected_status': status,
            'selected_book': book_id,
            'date_from': date_from,
            'date_to': date_to,
        }
        return render(request, self.template_name, context)


class ReturnBookView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_return_books'

    def get(self, request, pk):
        org = _get_org(request)
        issue = get_object_or_404(BookIssue, pk=pk, org=org)
        late_days, projected_fine = issue.compute_fine()
        context = {
            'org': org,
            'issue': issue,
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'late_days': late_days,
            'projected_fine': projected_fine,
        }
        return render(request, 'admin/library/return_book.html', context)

    def post(self, request, pk):
        org = _get_org(request)
        issue = get_object_or_404(BookIssue, pk=pk, org=org)
        if issue.status not in ('issued', 'overdue'):
            messages.error(request, "This copy has already been returned.")
            return redirect('schooladmin:library_issues')
        condition = request.POST.get('condition', 'good')
        fine_paid = Decimal(str(request.POST.get('fine_paid', 0) or 0))
        remarks = request.POST.get('remarks', '')
        issue.mark_returned(returned_by=request.user, condition=condition)
        if remarks:
            issue.remarks = remarks
        if fine_paid > 0:
            issue.fine_paid = min(fine_paid, issue.fine)
        issue.save()
        messages.success(request, f"'{issue.book.title}' returned by {issue.member.name}.")
        return redirect('schooladmin:library_issues')


class LibraryCategoryView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/categories.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {'org': org, 'categories': LibraryCategory.objects.filter(org=org)})

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            form = LibraryCategoryForm(request.POST)
            if form.is_valid():
                cat = form.save(commit=False)
                cat.org = org
                cat.save()
                messages.success(request, f"Category '{cat.name}' added.")
            else:
                messages.error(request, "Name is required and must be unique.")
        elif action == 'delete':
            LibraryCategory.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Category deleted.")
        return redirect('schooladmin:library_categories')


class LibraryAuthorView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/authors.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {'org': org, 'authors': LibraryAuthor.objects.filter(org=org)})

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            form = LibraryAuthorForm(request.POST)
            if form.is_valid():
                a = form.save(commit=False)
                a.org = org
                a.save()
                messages.success(request, f"Author '{a.name}' added.")
            else:
                messages.error(request, "Name is required and must be unique.")
        elif action == 'delete':
            LibraryAuthor.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Author deleted.")
        return redirect('schooladmin:library_authors')


class LibraryPublisherView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/publishers.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {'org': org, 'publishers': LibraryPublisher.objects.filter(org=org)})

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            form = LibraryPublisherForm(request.POST)
            if form.is_valid():
                p = form.save(commit=False)
                p.org = org
                p.save()
                messages.success(request, f"Publisher '{p.name}' added.")
            else:
                messages.error(request, "Name is required and must be unique.")
        elif action == 'delete':
            LibraryPublisher.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Publisher deleted.")
        return redirect('schooladmin:library_publishers')


class LibraryRackShelfView(FeatureRequiredMixin, PermRequiredMixin, View):
    """One page managing both racks and shelves — shelves nest under a rack."""
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/racks_shelves.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'racks': LibraryRack.objects.filter(org=org).prefetch_related('shelves'),
            'branches': Branch.objects.filter(org=org, status='active'),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add_rack':
            form = LibraryRackForm(request.POST, org=org)
            if form.is_valid():
                r = form.save(commit=False)
                r.org = org
                r.save()
                messages.success(request, f"Rack '{r.code}' added.")
            else:
                messages.error(request, "Rack code is required and must be unique.")
        elif action == 'add_shelf':
            form = LibraryShelfForm(request.POST, org=org)
            if form.is_valid():
                s = form.save(commit=False)
                s.org = org
                s.save()
                messages.success(request, f"Shelf '{s.code}' added.")
            else:
                messages.error(request, "Shelf code is required and must be unique within its rack.")
        elif action == 'delete_rack':
            LibraryRack.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Rack deleted.")
        elif action == 'delete_shelf':
            LibraryShelf.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Shelf deleted.")
        return redirect('schooladmin:library_racks_shelves')


class LibrarySettingsView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'library'
    required_perm = 'can_manage_library'
    template_name = 'admin/library/settings.html'

    def get(self, request):
        org = _get_org(request)
        settings_obj = LibrarySettings.for_org(org)
        form = LibrarySettingsForm(instance=settings_obj)
        return render(request, self.template_name, {'org': org, 'form': form})

    def post(self, request):
        org = _get_org(request)
        settings_obj = LibrarySettings.for_org(org)
        form = LibrarySettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Library settings updated.")
        else:
            messages.error(request, "Please correct the errors below.")
        return redirect('schooladmin:library_settings')


@login_required
@feature_required('library')
@perm_required('can_view_library')
def export_library(request):
    """CSV/Excel export of the book catalog or the issue history."""
    org = _get_org(request)
    fmt = request.GET.get('fmt', 'excel')
    export_type = request.GET.get('type', 'books')  # 'books' or 'issues'

    if export_type == 'issues':
        qs = BookIssue.objects.filter(org=org).select_related('book', 'member').order_by('-issue_date')
        headers = ['Book', 'Member', 'Issue Date', 'Due Date', 'Return Date', 'Status', 'Late Days', 'Fine', 'Fine Paid']
        rows = [[
            i.book.title, i.member.name, str(i.issue_date), str(i.due_date),
            str(i.return_date) if i.return_date else '', i.get_status_display(),
            i.late_days, float(i.fine), float(i.fine_paid),
        ] for i in qs]
        filename_base = 'library_issues'
    else:
        qs = Book.objects.filter(org=org).select_related('category', 'author', 'publisher').order_by('title')
        headers = ['Book Code', 'ISBN', 'Title', 'Category', 'Author', 'Publisher', 'Quantity', 'Available', 'Lost', 'Damaged', 'Status']
        rows = [[
            b.book_code, b.isbn or '', b.title,
            b.category.name if b.category else '', b.author.name if b.author else '',
            b.publisher.name if b.publisher else '', b.quantity, b.available_quantity,
            b.lost_quantity, b.damaged_quantity, b.get_status_display(),
        ] for b in qs]
        filename_base = 'library_books'

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Library'
    _style_header(ws, headers, fill_color='7C3AED')
    for ridx, row in enumerate(rows, 2):
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)
    resp = _excel_response(f'{filename_base}.xlsx')
    wb.save(resp)
    return resp


# =============================================================
# ACCOUNTING CORE (premium — required_feature = 'accounting')
# Chart of Accounts is schooladmin/superadmin-only (AdminRequiredMixin),
# not a grantable staff permission, per the confirmed Phase 1 scope.
# =============================================================

class AccountingDashboardView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_accounting'
    template_name = 'admin/accounting/dashboard.html'

    def get(self, request):
        org = _get_org(request)
        from handle.accounting import ensure_default_accounts, profit_and_loss
        if not Account.objects.filter(org=org).exists():
            ensure_default_accounts(org)
        today = datetime.date.today()
        month_start = today.replace(day=1)
        pending_count = JournalEntry.objects.filter(org=org, status='pending').count()
        pl = profit_and_loss(org, month_start, today)
        context = {
            'org': org,
            'pending_count': pending_count,
            'net_income': pl['net_income'],
            'income': pl['income'],
            'expense': pl['expense'],
            'account_count': Account.objects.filter(org=org, is_group=False).count(),
            'recent_entries': JournalEntry.objects.filter(org=org).select_related('created_by').order_by('-entry_date', '-id')[:10],
        }
        return render(request, self.template_name, context)


class ChartOfAccountsView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'accounting'
    template_name = 'admin/accounting/chart_of_accounts.html'

    def get(self, request):
        org = _get_org(request)
        roots = Account.objects.filter(org=org, parent__isnull=True).prefetch_related('children__children__children')
        return render(request, self.template_name, {'org': org, 'roots': roots})


class AddAccountView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'accounting'
    template_name = 'admin/accounting/add_account.html'

    def get(self, request):
        org = _get_org(request)
        form = AccountForm(org=org)
        return render(request, self.template_name, {'org': org, 'form': form})

    def post(self, request):
        org = _get_org(request)
        form = AccountForm(request.POST, org=org)
        if form.is_valid():
            account = form.save(commit=False)
            account.org = org
            account.save()
            messages.success(request, f"Account '{account.name}' added.")
            return redirect('schooladmin:accounting_accounts')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form})


class EditAccountView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'accounting'
    template_name = 'admin/accounting/edit_account.html'

    def get(self, request, pk):
        org = _get_org(request)
        account = get_object_or_404(Account, pk=pk, org=org)
        form = AccountForm(instance=account, org=org)
        return render(request, self.template_name, {'org': org, 'account': account, 'form': form})

    def post(self, request, pk):
        org = _get_org(request)
        account = get_object_or_404(Account, pk=pk, org=org)
        form = AccountForm(request.POST, instance=account, org=org)
        if form.is_valid():
            form.save()
            messages.success(request, "Account updated.")
            return redirect('schooladmin:accounting_accounts')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'account': account, 'form': form})


@feature_required('accounting')
def delete_account(request, pk):
    org = _get_org(request)
    account = get_object_or_404(Account, pk=pk, org=org)
    if request.user.user_type not in ('1', '2'):
        messages.error(request, "You don't have permission to manage the Chart of Accounts.")
        return redirect('schooladmin:accounting_accounts')
    if account.is_system:
        messages.error(request, f"'{account.name}' is a default account and can't be deleted — deactivate it instead.")
    elif account.children.exists():
        messages.error(request, f"'{account.name}' has sub-accounts — delete or reassign them first.")
    elif account.journal_lines.exists():
        messages.error(request, f"'{account.name}' has posted journal entries and can't be deleted — deactivate it instead.")
    else:
        account.delete()
        messages.success(request, "Account deleted.")
    return redirect('schooladmin:accounting_accounts')


class JournalEntryListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_accounting'
    template_name = 'admin/accounting/journal_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = JournalEntry.objects.filter(org=org).select_related('created_by', 'approved_by').order_by('-entry_date', '-id')
        status = request.GET.get('status')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if status:
            qs = qs.filter(status=status)
        if date_from:
            qs = qs.filter(entry_date__gte=date_from)
        if date_to:
            qs = qs.filter(entry_date__lte=date_to)
        context = {
            'org': org, 'entries': qs,
            'selected_status': status, 'date_from': date_from, 'date_to': date_to,
        }
        return render(request, self.template_name, context)


def _scope_line_formset_accounts(formset, org):
    qs = Account.objects.filter(org=org, is_group=False, is_active=True).order_by('code', 'name')
    for f in formset.forms:
        f.fields['account'].queryset = qs


class CreateJournalEntryView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_create_journals'
    template_name = 'admin/accounting/journal_form.html'

    def get(self, request):
        org = _get_org(request)
        form = JournalEntryForm(org=org, initial={'entry_date': datetime.date.today()})
        formset = JournalEntryLineFormSet()
        _scope_line_formset_accounts(formset, org)
        return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = JournalEntryForm(request.POST, request.FILES, org=org)
        formset = JournalEntryLineFormSet(request.POST)
        _scope_line_formset_accounts(formset, org)

        if not (form.is_valid() and formset.is_valid()):
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'create'})

        lines = []
        for line_form in formset.forms:
            if not line_form.cleaned_data or line_form.cleaned_data.get('DELETE'):
                continue
            account = line_form.cleaned_data.get('account')
            if not account:
                continue
            lines.append({
                'account': account,
                'debit': line_form.cleaned_data.get('debit') or Decimal('0'),
                'credit': line_form.cleaned_data.get('credit') or Decimal('0'),
                'remarks': line_form.cleaned_data.get('remarks', ''),
            })

        from handle.accounting import create_journal_entry
        try:
            entry = create_journal_entry(
                org, entry_date=form.cleaned_data['entry_date'],
                lines=lines, description=form.cleaned_data.get('description', ''),
                reference=form.cleaned_data.get('reference', ''),
                branch=form.cleaned_data.get('branch'),
                created_by=request.user, attachment=form.cleaned_data.get('attachment'),
                status='pending' if 'submit_for_approval' in request.POST else 'draft',
            )
        except ValueError as e:
            messages.error(request, str(e))
            return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'create'})

        messages.success(request, f"Journal entry {entry.voucher_number} saved.")
        return redirect('schooladmin:journal_detail', pk=entry.pk)


class EditJournalEntryView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_create_journals'
    template_name = 'admin/accounting/journal_form.html'

    def get(self, request, pk):
        org = _get_org(request)
        entry = get_object_or_404(JournalEntry, pk=pk, org=org)
        if entry.status not in ('draft', 'rejected'):
            messages.error(request, "Only draft or rejected entries can be edited. Approved entries are permanent — post a reversing entry instead.")
            return redirect('schooladmin:journal_detail', pk=entry.pk)
        form = JournalEntryForm(instance=entry, org=org)
        formset = JournalEntryLineFormSet(instance=entry)
        _scope_line_formset_accounts(formset, org)
        return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'edit', 'entry': entry})

    def post(self, request, pk):
        org = _get_org(request)
        entry = get_object_or_404(JournalEntry, pk=pk, org=org)
        if entry.status not in ('draft', 'rejected'):
            messages.error(request, "Only draft or rejected entries can be edited.")
            return redirect('schooladmin:journal_detail', pk=entry.pk)

        form = JournalEntryForm(request.POST, request.FILES, instance=entry, org=org)
        formset = JournalEntryLineFormSet(request.POST, instance=entry)
        _scope_line_formset_accounts(formset, org)

        if not (form.is_valid() and formset.is_valid()):
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'edit', 'entry': entry})

        total_debit = Decimal('0')
        total_credit = Decimal('0')
        usable = 0
        for line_form in formset.forms:
            if not line_form.cleaned_data or line_form.cleaned_data.get('DELETE'):
                continue
            if not line_form.cleaned_data.get('account'):
                continue
            usable += 1
            total_debit += line_form.cleaned_data.get('debit') or Decimal('0')
            total_credit += line_form.cleaned_data.get('credit') or Decimal('0')

        if usable < 2 or total_debit != total_credit or total_debit <= 0:
            messages.error(request, f"Journal entry does not balance: total debit {total_debit} != total credit {total_credit}.")
            return render(request, self.template_name, {'org': org, 'form': form, 'formset': formset, 'mode': 'edit', 'entry': entry})

        entry = form.save(commit=False)
        entry.org = org
        entry.status = 'pending' if 'submit_for_approval' in request.POST else 'draft'
        entry.rejection_reason = None
        entry.save()
        formset.save()
        messages.success(request, f"Journal entry {entry.voucher_number} updated.")
        return redirect('schooladmin:journal_detail', pk=entry.pk)


class JournalEntryDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_accounting'
    template_name = 'admin/accounting/journal_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        entry = get_object_or_404(JournalEntry.objects.select_related('created_by', 'approved_by', 'branch'), pk=pk, org=org)
        lines = entry.lines.select_related('account').order_by('line_order', 'id')
        # Computed explicitly via has_perm() (not the `staff_perms` context var) —
        # StaffPermission is a plain model with no can_create_journals/
        # can_approve_journals field, so accessing it directly would silently
        # evaluate false for staff granted these through the dynamic-permission
        # system even though has_perm() correctly resolves them.
        context = {
            'org': org, 'entry': entry, 'lines': lines,
            'can_create_journals': has_perm(request.user, 'can_create_journals'),
            'can_approve_journals': has_perm(request.user, 'can_approve_journals'),
        }
        return render(request, self.template_name, context)


@feature_required('accounting')
@perm_required('can_create_journals')
def submit_journal_entry(request, pk):
    org = _get_org(request)
    entry = get_object_or_404(JournalEntry, pk=pk, org=org)
    from handle.accounting import submit_for_approval
    try:
        submit_for_approval(entry, request.user)
        messages.success(request, f"Journal entry {entry.voucher_number} submitted for approval.")
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('schooladmin:journal_detail', pk=entry.pk)


@feature_required('accounting')
@perm_required('can_approve_journals')
def approve_journal_entry_view(request, pk):
    org = _get_org(request)
    entry = get_object_or_404(JournalEntry, pk=pk, org=org)
    from handle.accounting import approve_journal_entry
    try:
        approve_journal_entry(entry, request.user)
        messages.success(request, f"Journal entry {entry.voucher_number} approved.")
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('schooladmin:journal_detail', pk=entry.pk)


@feature_required('accounting')
@perm_required('can_approve_journals')
def reject_journal_entry_view(request, pk):
    org = _get_org(request)
    entry = get_object_or_404(JournalEntry, pk=pk, org=org)
    from handle.accounting import reject_journal_entry
    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, "A rejection reason is required.")
        return redirect('schooladmin:journal_detail', pk=entry.pk)
    try:
        reject_journal_entry(entry, request.user, reason)
        messages.success(request, f"Journal entry {entry.voucher_number} rejected.")
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('schooladmin:journal_detail', pk=entry.pk)


class GeneralLedgerView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_ledger'
    template_name = 'admin/accounting/general_ledger.html'

    def get(self, request):
        org = _get_org(request)
        accounts = Account.objects.filter(org=org, is_group=False, is_active=True).order_by('code', 'name')
        account_id = request.GET.get('account')
        today = datetime.date.today()
        date_from = request.GET.get('date_from') or today.replace(day=1).isoformat()
        date_to = request.GET.get('date_to') or today.isoformat()
        gl = None
        selected_account = None
        if account_id:
            selected_account = get_object_or_404(Account, pk=account_id, org=org)
            from handle.accounting import general_ledger
            gl = general_ledger(
                selected_account,
                datetime.date.fromisoformat(date_from),
                datetime.date.fromisoformat(date_to),
            )
        context = {
            'org': org, 'accounts': accounts, 'selected_account': selected_account,
            'gl': gl, 'date_from': date_from, 'date_to': date_to,
        }
        return render(request, self.template_name, context)


class TrialBalanceView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_accounting_reports'
    template_name = 'admin/accounting/trial_balance.html'

    def get(self, request):
        org = _get_org(request)
        as_of_date = request.GET.get('as_of_date') or datetime.date.today().isoformat()
        from handle.accounting import trial_balance
        tb = trial_balance(org, datetime.date.fromisoformat(as_of_date))
        return render(request, self.template_name, {'org': org, 'tb': tb, 'as_of_date': as_of_date})


class ProfitAndLossView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_accounting_reports'
    template_name = 'admin/accounting/profit_and_loss.html'

    def get(self, request):
        org = _get_org(request)
        today = datetime.date.today()
        date_from = request.GET.get('date_from') or today.replace(day=1).isoformat()
        date_to = request.GET.get('date_to') or today.isoformat()
        from handle.accounting import profit_and_loss
        pl = profit_and_loss(org, datetime.date.fromisoformat(date_from), datetime.date.fromisoformat(date_to))
        return render(request, self.template_name, {'org': org, 'pl': pl, 'date_from': date_from, 'date_to': date_to})


class BalanceSheetView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'accounting'
    required_perm = 'can_view_accounting_reports'
    template_name = 'admin/accounting/balance_sheet.html'

    def get(self, request):
        org = _get_org(request)
        as_of_date = request.GET.get('as_of_date') or datetime.date.today().isoformat()
        from handle.accounting import balance_sheet
        bs = balance_sheet(org, datetime.date.fromisoformat(as_of_date))
        return render(request, self.template_name, {'org': org, 'bs': bs, 'as_of_date': as_of_date})


@login_required
@feature_required('accounting')
def export_accounting(request):
    """CSV/Excel/PDF export for the ledger and the three core reports."""
    org = _get_org(request)
    export_type = request.GET.get('type', 'trial_balance')  # ledger|trial_balance|profit_loss|balance_sheet
    fmt = request.GET.get('fmt', 'excel')

    required_perm = 'can_view_ledger' if export_type == 'ledger' else 'can_view_accounting_reports'
    if request.user.user_type not in ('1', '2') and not has_perm(request.user, required_perm):
        return HttpResponse('Unauthorized', status=403)

    from handle import accounting as acct
    today = datetime.date.today()

    if export_type == 'ledger':
        account = get_object_or_404(Account, pk=request.GET.get('account'), org=org)
        date_from = datetime.date.fromisoformat(request.GET.get('date_from') or today.replace(day=1).isoformat())
        date_to = datetime.date.fromisoformat(request.GET.get('date_to') or today.isoformat())
        gl = acct.general_ledger(account, date_from, date_to)
        headers = ['Date', 'Voucher', 'Reference', 'Remarks', 'Debit', 'Credit', 'Running Balance']
        rows = [[r['date'], r['voucher_number'], r['reference'] or '', r['remarks'] or '',
                 float(r['debit']), float(r['credit']), float(r['running_balance'])] for r in gl['rows']]
        filename = f'general_ledger_{account.name}'
        pdf_context = {'org': org, 'account': account, 'gl': gl, 'date_from': date_from, 'date_to': date_to}
        pdf_template = 'admin/accounting/pdf/ledger.html'
    elif export_type == 'profit_loss':
        date_from = datetime.date.fromisoformat(request.GET.get('date_from') or today.replace(day=1).isoformat())
        date_to = datetime.date.fromisoformat(request.GET.get('date_to') or today.isoformat())
        pl = acct.profit_and_loss(org, date_from, date_to)
        headers = ['Line', 'Amount']
        rows = [['Income', float(pl['income'])], ['Cost of Goods Sold', float(pl['cogs'])],
                ['Gross Profit', float(pl['gross_profit'])], ['Expense', float(pl['expense'])],
                ['Net Income', float(pl['net_income'])]]
        filename = 'profit_and_loss'
        pdf_context = {'org': org, 'pl': pl, 'date_from': date_from, 'date_to': date_to}
        pdf_template = 'admin/accounting/pdf/profit_loss.html'
    elif export_type == 'balance_sheet':
        as_of_date = datetime.date.fromisoformat(request.GET.get('as_of_date') or today.isoformat())
        bs = acct.balance_sheet(org, as_of_date)
        headers = ['Account', 'Balance']
        rows = [[a['account'].full_path(), float(a['balance'])] for a in bs['assets']]
        rows += [[l['account'].full_path(), float(l['balance'])] for l in bs['liabilities']]
        rows += [[e['account'].full_path(), float(e['balance'])] for e in bs['equity']]
        rows.append(['Retained Earnings (current)', float(bs['retained_earnings_current'])])
        filename = 'balance_sheet'
        pdf_context = {'org': org, 'bs': bs, 'as_of_date': as_of_date}
        pdf_template = 'admin/accounting/pdf/balance_sheet.html'
    else:  # trial_balance
        as_of_date = datetime.date.fromisoformat(request.GET.get('as_of_date') or today.isoformat())
        tb = acct.trial_balance(org, as_of_date)
        headers = ['Account', 'Debit', 'Credit']
        rows = [[r['account'].full_path(), float(r['debit_total']), float(r['credit_total'])] for r in tb['rows']]
        filename = 'trial_balance'
        pdf_context = {'org': org, 'tb': tb, 'as_of_date': as_of_date}
        pdf_template = 'admin/accounting/pdf/trial_balance.html'

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return resp

    if fmt == 'pdf':
        from school.pdf_utils import render_to_pdf
        pdf_bytes = render_to_pdf(pdf_template, pdf_context)
        if pdf_bytes is None:
            return HttpResponse('PDF rendering failed', status=500)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:31]
    _style_header(ws, headers, fill_color='0F766E')
    for ridx, row in enumerate(rows, 2):
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)
    resp = _excel_response(f'{filename}.xlsx')
    wb.save(resp)
    return resp


# =============================================================
# ACADEMIC MANAGEMENT (premium — required_feature = 'academic_management')
# Course/Classification/Section/Subject stay exactly as they are —
# everything here is additive.
# =============================================================

# ── Academic Year / Faculty / Semester ──────────────────────────────────

class AcademicYearListView(FeatureRequiredMixin, View):
    required_feature = 'academic_management'
    template_name = 'admin/academic/academic_years.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {'org': org, 'years': AcademicYear.objects.filter(org=org)})

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            form = AcademicYearForm(request.POST)
            form.instance.org = org
            name = (request.POST.get('name') or '').strip()
            if name and AcademicYear.objects.filter(org=org, name=name).exists():
                messages.error(request, f"An academic year named '{name}' already exists for this organization.")
            elif form.is_valid():
                obj = form.save(commit=False)
                obj.org = org
                if obj.is_current:
                    AcademicYear.objects.filter(org=org, is_current=True).update(is_current=False)
                obj.save()
                messages.success(request, f"Academic year '{obj.name}' added.")
            else:
                print(form.errors)
                messages.error(request, "Please correct the errors below.")
        elif action == 'delete':
            AcademicYear.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Academic year deleted.")
        return redirect('schooladmin:academic_years')


class FacultyListView(FeatureRequiredMixin, View):
    required_feature = 'academic_management'
    template_name = 'admin/academic/faculties.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {'org': org, 'faculties': Faculty.objects.filter(org=org)})

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            form = FacultyForm(request.POST)
            form.instance.org = org
            name = (request.POST.get('name') or '').strip()
            if name and Faculty.objects.filter(org=org, name=name).exists():
                messages.error(request, f"A faculty named '{name}' already exists for this organization.")
            elif form.is_valid():
                obj = form.save(commit=False)
                obj.org = org
                obj.save()
                messages.success(request, f"Faculty '{obj.name}' added.")
            else:
                messages.error(request, "Please correct the errors below.")
        elif action == 'delete':
            Faculty.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Faculty deleted.")
        return redirect('schooladmin:faculty_list')


class SemesterListView(FeatureRequiredMixin, View):
    required_feature = 'academic_management'
    template_name = 'admin/academic/semesters.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'semesters': Semester.objects.filter(org=org).select_related('faculty', 'academic_year'),
            'faculties': Faculty.objects.filter(org=org, status='active'),
            'years': AcademicYear.objects.filter(org=org),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            form = SemesterForm(request.POST, org=org)
            form.instance.org = org
            if form.is_valid():
                obj = form.save(commit=False)
                obj.org = org
                obj.save()
                messages.success(request, f"Semester '{obj.name}' added.")
            else:
                messages.error(request, "Please correct the errors below.")
        elif action == 'delete':
            Semester.objects.filter(pk=request.POST.get('item_id'), org=org).delete()
            messages.success(request, "Semester deleted.")
        return redirect('schooladmin:semester_list')


# ── Course / Subject teacher assignment (additive — Course.teacher and
# Subject.teacher stay as the 'primary teacher' for every existing reader) ──

class CourseTeachersView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    template_name = 'admin/academic/course_teachers.html'

    def get(self, request, pk):
        org = _get_org(request)
        course = get_object_or_404(Course, pk=pk, org=org)
        return redirect(f"{reverse('schooladmin:subject_list')}?course={course.pk}")
        context = {
            'org': org, 'course': course,
            'assignments': course.teacher_assignments.select_related('teacher'),
            'staff_list': Staff.objects.filter(org=org).select_related('admin'),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        org = _get_org(request)
        course = get_object_or_404(Course, pk=pk, org=org)
        return redirect(f"{reverse('schooladmin:subject_list')}?course={course.pk}")
        action = request.POST.get('action')
        if action == 'add':
            teacher_id = request.POST.get('teacher')
            is_primary = request.POST.get('is_primary') == 'on'
            teacher = Staff.objects.filter(
                org=org, admin_id=teacher_id
            ).select_related('admin').first()
            if teacher:
                if is_primary:
                    CourseTeacherAssignment.objects.filter(course=course).update(is_primary=False)
                CourseTeacherAssignment.objects.update_or_create(
                    course=course, teacher=teacher.admin, defaults={'is_primary': is_primary}
                )
                if is_primary:
                    course.teacher = teacher.admin
                    course.save(update_fields=['teacher'])
                teacher_name = teacher.admin.get_full_name() or teacher.admin.email
                messages.success(
                    request,
                    f"{teacher_name} can now record attendance for this course."
                )
            else:
                messages.error(request, "Please select a staff member from your organization.")
        elif action == 'remove':
            assignment = CourseTeacherAssignment.objects.filter(
                pk=request.POST.get('item_id'), course=course
            ).first()
            if assignment:
                removed_teacher_id = assignment.teacher_id
                assignment.delete()
                if course.teacher_id == removed_teacher_id:
                    course.teacher = None
                    course.save(update_fields=['teacher'])
                messages.success(request, "Course attendance assignment removed.")
        return redirect('schooladmin:course_teachers', pk=course.pk)


class SubjectTeachersView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    template_name = 'admin/academic/subject_teachers.html'

    def _context(self, request, org, subject, form=None):
        from django.core.paginator import Paginator

        assignments = subject.teacher_assignments.select_related(
            'teacher', 'academic_year', 'course', 'classification', 'section',
            'assigned_by',
        )
        status = request.GET.get('status', '')
        academic_year_id = request.GET.get('academic_year', '')
        teacher_id = request.GET.get('teacher', '')
        if status:
            assignments = assignments.filter(status=status)
        if academic_year_id:
            assignments = assignments.filter(academic_year_id=academic_year_id)
        if teacher_id:
            assignments = assignments.filter(teacher_id=teacher_id)
        page_obj = Paginator(assignments, 25).get_page(request.GET.get('page'))
        return {
            'org': org,
            'subject': subject,
            'form': form or SubjectTeacherAssignmentForm(org=org, subject=subject),
            'assignments': page_obj,
            'page_obj': page_obj,
            'staff_list': Staff.objects.filter(org=org).select_related('admin'),
            'academic_years': AcademicYear.objects.filter(org=org, status='active'),
            'selected_status': status,
            'selected_academic_year': academic_year_id,
            'selected_teacher': teacher_id,
        }

    def get(self, request, pk):
        org = _get_org(request)
        subject = get_object_or_404(
            Subject.objects.select_related('course', 'classification', 'section'),
            pk=pk, org=org,
        )
        edit_id = request.GET.get('edit')
        instance = subject.teacher_assignments.filter(pk=edit_id).first() if edit_id else None
        form = SubjectTeacherAssignmentForm(
            org=org, subject=subject, instance=instance,
        ) if instance else None
        return render(
            request, self.template_name,
            self._context(request, org, subject, form=form),
        )

    def post(self, request, pk):
        org = _get_org(request)
        subject = get_object_or_404(
            Subject.objects.select_related('course', 'classification', 'section'),
            pk=pk, org=org,
        )
        action = request.POST.get('action')
        if action in ('add', 'edit'):
            instance = None
            if action == 'edit':
                instance = subject.teacher_assignments.filter(
                    pk=request.POST.get('item_id')
                ).first()
                if not instance:
                    messages.error(request, "Assignment not found.")
                    return redirect('schooladmin:subject_teachers', pk=subject.pk)
            form = SubjectTeacherAssignmentForm(
                request.POST, org=org, subject=subject, instance=instance,
            )
            if form.is_valid():
                assignment = form.save(commit=False)
                assignment.assigned_by = assignment.assigned_by or request.user
                if assignment.is_primary:
                    subject.teacher_assignments.exclude(pk=assignment.pk).update(is_primary=False)
                assignment.save()
                messages.success(
                    request,
                    f"Subject assignment for {assignment.teacher.get_full_name() or assignment.teacher.email} saved."
                )
            else:
                messages.error(request, "Please correct the assignment errors below.")
                return render(
                    request, self.template_name,
                    self._context(request, org, subject, form=form),
                )
        elif action == 'deactivate':
            assignment = SubjectTeacherAssignment.objects.filter(
                pk=request.POST.get('item_id'), subject=subject
            ).first()
            if assignment:
                assignment.status = 'inactive'
                if not assignment.end_date or assignment.end_date > timezone.localdate():
                    assignment.end_date = timezone.localdate()
                assignment.save(update_fields=['status', 'end_date', 'updated_at'])
                if subject.teacher_id == assignment.teacher_id:
                    subject.teacher = None
                    subject.save(update_fields=['teacher'])
                messages.success(request, "Future teaching and attendance access deactivated; history was preserved.")
        elif action == 'delete':
            assignment = SubjectTeacherAssignment.objects.filter(
                pk=request.POST.get('item_id'), subject=subject
            ).first()
            if assignment:
                if assignment.teaching_logs.exists():
                    messages.error(
                        request,
                        "This assignment has attendance history. Deactivate it instead of deleting it.",
                    )
                else:
                    assignment.delete()
                    messages.success(request, "Unused subject assignment deleted.")
        return redirect('schooladmin:subject_teachers', pk=subject.pk)


class CourseEnrollmentsView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    template_name = 'admin/academic/course_enrollments.html'

    def _context(self, request, org, course, form=None):
        from django.core.paginator import Paginator

        enrollments = course.student_enrollments.filter(org=org).select_related(
            'student', 'academic_year', 'classification', 'section', 'branch',
        )
        status = request.GET.get('status', '')
        academic_year_id = request.GET.get('academic_year', '')
        classification_id = request.GET.get('classification', '')
        section_id = request.GET.get('section', '')
        search = request.GET.get('search', '').strip()
        if status:
            enrollments = enrollments.filter(status=status)
        if academic_year_id:
            enrollments = enrollments.filter(academic_year_id=academic_year_id)
        if classification_id:
            enrollments = enrollments.filter(classification_id=classification_id)
        if section_id:
            enrollments = enrollments.filter(section_id=section_id)
        if search:
            enrollments = enrollments.filter(
                Q(student__name__icontains=search)
                | Q(student__card__icontains=search)
                | Q(student__roll_number__icontains=search)
            )
        page_obj = Paginator(enrollments, 30).get_page(request.GET.get('page'))
        return {
            'org': org,
            'course': course,
            'form': form or StudentCourseEnrollmentForm(org=org, course=course),
            'enrollments': page_obj,
            'page_obj': page_obj,
            'academic_years': AcademicYear.objects.filter(org=org, status='active'),
            'classifications': course.classifications.filter(org=org, status='active'),
            'sections': course.sections.filter(org=org, status='active').select_related('classification'),
            'status_choices': StudentCourseEnrollment.STATUS_CHOICES,
            'selected_status': status,
            'selected_academic_year': academic_year_id,
            'selected_classification': classification_id,
            'selected_section': section_id,
            'search': search,
        }

    def get(self, request, pk):
        org = _get_org(request)
        course = get_object_or_404(
            Course.objects.prefetch_related('classifications', 'sections'),
            pk=pk, org=org,
        )
        edit_id = request.GET.get('edit')
        instance = course.student_enrollments.filter(pk=edit_id, org=org).first() if edit_id else None
        form = StudentCourseEnrollmentForm(
            org=org, course=course, instance=instance,
        ) if instance else None
        return render(
            request, self.template_name,
            self._context(request, org, course, form=form),
        )

    def post(self, request, pk):
        org = _get_org(request)
        course = get_object_or_404(
            Course.objects.prefetch_related('classifications', 'sections'),
            pk=pk, org=org,
        )
        action = request.POST.get('action')
        if action in ('add', 'edit'):
            instance = None
            if action == 'edit':
                instance = course.student_enrollments.filter(
                    pk=request.POST.get('item_id'), org=org,
                ).first()
                if not instance:
                    messages.error(request, "Enrollment not found.")
                    return redirect('schooladmin:course_enrollments', pk=course.pk)
            form = StudentCourseEnrollmentForm(
                request.POST, org=org, course=course, instance=instance,
            )
            if form.is_valid():
                enrollment = form.save(commit=False)
                enrollment.created_by = enrollment.created_by or request.user
                enrollment.save()
                # Keep the legacy current-placement fields synchronized only
                # for active current enrollments.
                if enrollment.status == 'active':
                    student = enrollment.student
                    student.classification = enrollment.classification
                    student.section = enrollment.section
                    student.branch = enrollment.branch
                    student.save(update_fields=['classification', 'section', 'branch'])
                    student.courses.add(course)
                messages.success(request, f"Enrollment for {enrollment.student.name} saved.")
            else:
                messages.error(request, "Please correct the enrollment errors below.")
                return render(
                    request, self.template_name,
                    self._context(request, org, course, form=form),
                )
        elif action == 'close':
            enrollment = course.student_enrollments.filter(
                pk=request.POST.get('item_id'), org=org,
            ).first()
            if enrollment:
                enrollment.status = request.POST.get('status', 'completed')
                if enrollment.status not in ('completed', 'transferred', 'cancelled'):
                    enrollment.status = 'completed'
                enrollment.end_date = enrollment.end_date or timezone.localdate()
                enrollment.save(update_fields=['status', 'end_date', 'updated_at'])
                messages.success(request, "Enrollment closed; historical placement was preserved.")
        elif action == 'delete':
            enrollment = course.student_enrollments.filter(
                pk=request.POST.get('item_id'), org=org,
            ).first()
            if enrollment:
                has_history = SubjectAttendanceRecord.objects.filter(
                    member=enrollment.student,
                    teaching_log__course=course,
                    teaching_log__classification=enrollment.classification,
                    teaching_log__date__gte=enrollment.start_date,
                ).exists()
                if has_history:
                    messages.error(request, "This enrollment has attendance history. Close it instead.")
                else:
                    enrollment.delete()
                    messages.success(request, "Unused enrollment deleted.")
        return redirect('schooladmin:course_enrollments', pk=course.pk)


# ── Assignment Module ────────────────────────────────────────────────────

class AssignmentListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_assignments'
    template_name = 'admin/academic/assignment_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Assignment.objects.filter(org=org).select_related('classification', 'section', 'subject').order_by('-due_date')
        classification_id = request.GET.get('classification')
        status = request.GET.get('status')
        if classification_id:
            qs = qs.filter(classification_id=classification_id)
        if status:
            qs = qs.filter(status=status)
        context = {
            'org': org, 'assignments': qs,
            'classifications': Classification.objects.filter(org=org, status='active'),
            'selected_classification': classification_id, 'selected_status': status,
        }
        return render(request, self.template_name, context)


class AddAssignmentView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_assignments'
    template_name = 'admin/academic/assignment_form.html'

    def get(self, request):
        org = _get_org(request)
        form = AssignmentForm(org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = AssignmentForm(request.POST, org=org)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.org = org
            assignment.assigned_by = request.user
            assignment.save()
            for f in request.FILES.getlist('attachments'):
                AssignmentAttachment.objects.create(assignment=assignment, file=f)

            if assignment.visibility == 'published':
                from handle.models import member as member_model
                students = member_model.objects.filter(
                    org=org, classification=assignment.classification, status='active',
                    member_type__in=('student', 'trainee'),
                )
                if assignment.section_id:
                    students = students.filter(section_id=assignment.section_id)
                from handle.notifications import notify_many
                notify_many(
                    students, 'assignment_assigned', f"New Assignment: {assignment.title}",
                    body=f"Due {assignment.due_date}", link_url=f'/staff/assignments/{assignment.pk}/',
                )
                from school.email_utils import send_assignment_assigned_email
                for s in students:
                    if s.email:
                        send_assignment_assigned_email(
                            s.email, s.name, assignment.title, assignment.subject.name,
                            str(assignment.due_date), org.name, org=org, related_object_id=assignment.pk,
                        )

            messages.success(request, f"Assignment '{assignment.title}' created.")
            return redirect('schooladmin:assignment_detail', pk=assignment.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})


class EditAssignmentView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_assignments'
    template_name = 'admin/academic/assignment_form.html'

    def get(self, request, pk):
        org = _get_org(request)
        assignment = get_object_or_404(Assignment, pk=pk, org=org)
        form = AssignmentForm(instance=assignment, org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'assignment': assignment, 'mode': 'edit'})

    def post(self, request, pk):
        org = _get_org(request)
        assignment = get_object_or_404(Assignment, pk=pk, org=org)
        form = AssignmentForm(request.POST, instance=assignment, org=org)
        if form.is_valid():
            form.save()
            for f in request.FILES.getlist('attachments'):
                AssignmentAttachment.objects.create(assignment=assignment, file=f)
            messages.success(request, "Assignment updated.")
            return redirect('schooladmin:assignment_detail', pk=assignment.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'assignment': assignment, 'mode': 'edit'})


class AssignmentDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_assignments'
    template_name = 'admin/academic/assignment_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        assignment = get_object_or_404(Assignment, pk=pk, org=org)
        submissions = assignment.submissions.select_related('student').order_by('-submitted_at')
        context = {
            'org': org, 'assignment': assignment, 'submissions': submissions,
            'attachments': assignment.attachments.all(),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        """Grade a single submission (inline form on the detail page)."""
        org = _get_org(request)
        assignment = get_object_or_404(Assignment, pk=pk, org=org)
        submission = get_object_or_404(AssignmentSubmission, pk=request.POST.get('submission_id'), assignment=assignment)
        if not has_perm(request.user, 'can_grade_assignments'):
            messages.error(request, "You don't have permission to grade assignments.")
            return redirect('schooladmin:assignment_detail', pk=assignment.pk)

        form = AssignmentGradeForm(request.POST, instance=submission)
        if form.is_valid():
            submission = form.save(commit=False)
            submission.graded_by = request.user
            submission.graded_at = timezone.now()
            submission.save()
            AssignmentSubmissionHistory.objects.create(
                submission=submission, action='graded', status=submission.status,
                obtained_marks=submission.obtained_marks, remarks=submission.teacher_remarks,
                performed_by=request.user,
            )
            from handle.notifications import notify
            notify(
                submission.student, 'marks_published', f"Graded: {assignment.title}",
                body=f"Marks: {submission.obtained_marks}/{assignment.total_marks}",
                link_url=f'/staff/assignments/{assignment.pk}/',
            )
            if submission.student.email:
                from school.email_utils import send_assignment_graded_email
                send_assignment_graded_email(
                    submission.student.email, submission.student.name, assignment.title,
                    str(submission.obtained_marks or 0), str(assignment.total_marks), org.name,
                    org=org, related_object_id=submission.pk,
                )
            messages.success(request, f"Graded {submission.student.name}'s submission.")
        else:
            messages.error(request, "Please correct the errors below.")
        return redirect('schooladmin:assignment_detail', pk=assignment.pk)


@feature_required('academic_management')
@perm_required('can_manage_assignments')
def delete_assignment(request, pk):
    org = _get_org(request)
    assignment = get_object_or_404(Assignment, pk=pk, org=org)
    assignment.delete()
    messages.success(request, "Assignment deleted.")
    return redirect('schooladmin:assignment_list')


# ── Homework Module ──────────────────────────────────────────────────────

class HomeworkListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_homework'
    template_name = 'admin/academic/homework_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Homework.objects.filter(org=org).select_related('classification', 'section', 'subject').order_by('-due_date')
        classification_id = request.GET.get('classification')
        if classification_id:
            qs = qs.filter(classification_id=classification_id)
        context = {
            'org': org, 'homeworks': qs,
            'classifications': Classification.objects.filter(org=org, status='active'),
            'selected_classification': classification_id,
        }
        return render(request, self.template_name, context)


class AddHomeworkView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_homework'
    template_name = 'admin/academic/homework_form.html'

    def get(self, request):
        org = _get_org(request)
        form = HomeworkForm(org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = HomeworkForm(request.POST, org=org)
        if form.is_valid():
            hw = form.save(commit=False)
            hw.org = org
            hw.assigned_by = request.user
            hw.save()
            for f in request.FILES.getlist('attachments'):
                HomeworkAttachment.objects.create(homework=hw, file=f)

            from handle.models import member as member_model
            students = member_model.objects.filter(
                org=org, classification=hw.classification, status='active',
                member_type__in=('student', 'trainee'),
            )
            if hw.section_id:
                students = students.filter(section_id=hw.section_id)
            for s in students:
                AA = HomeworkStatus.objects.get_or_create(homework=hw, student=s)
                print(f"Created HomeworkStatus for {s.name}: {AA[0]} (created={AA[1]})")
            from handle.notifications import notify_many
            notify_many(
                students, 'homework_assigned', f"New Homework: {hw.subject.name}",
                body=f"Due {hw.due_date}", link_url='/staff/homework/',
            )
            messages.success(request, "Homework assigned.")
            return redirect('schooladmin:homework_detail', pk=hw.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})


class EditHomeworkView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_homework'
    template_name = 'admin/academic/homework_form.html'

    def get(self, request, pk):
        org = _get_org(request)
        hw = get_object_or_404(Homework, pk=pk, org=org)
        form = HomeworkForm(instance=hw, org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'homework': hw, 'mode': 'edit'})

    def post(self, request, pk):
        org = _get_org(request)
        hw = get_object_or_404(Homework, pk=pk, org=org)
        form = HomeworkForm(request.POST, instance=hw, org=org)
        if form.is_valid():
            form.save()
            messages.success(request, "Homework updated.")
            return redirect('schooladmin:homework_detail', pk=hw.pk)
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'homework': hw, 'mode': 'edit'})


class HomeworkDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_homework'
    template_name = 'admin/academic/homework_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        hw = get_object_or_404(Homework, pk=pk, org=org)
        statuses = hw.statuses.select_related('student').order_by('student__name')
        context = {'org': org, 'homework': hw, 'statuses': statuses, 'attachments': hw.attachments.all()}
        return render(request, self.template_name, context)

    def post(self, request, pk):
        """Teacher verifies a student's completion."""
        org = _get_org(request)
        hw = get_object_or_404(Homework, pk=pk, org=org)
        status_obj = get_object_or_404(HomeworkStatus, pk=request.POST.get('status_id'), homework=hw)
        status_obj.verified_by_teacher = True
        status_obj.verified_at = timezone.now()
        status_obj.save(update_fields=['verified_by_teacher', 'verified_at'])
        messages.success(request, f"Verified {status_obj.student.name}'s homework.")
        return redirect('schooladmin:homework_detail', pk=hw.pk)


@feature_required('academic_management')
@perm_required('can_manage_homework')
def delete_homework(request, pk):
    org = _get_org(request)
    hw = get_object_or_404(Homework, pk=pk, org=org)
    hw.delete()
    messages.success(request, "Homework deleted.")
    return redirect('schooladmin:homework_list')


# ── Course Material Module ───────────────────────────────────────────────

class CourseMaterialListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_course_materials'
    template_name = 'admin/academic/course_material_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = CourseMaterial.objects.filter(org=org).select_related('subject', 'faculty', 'course', 'semester').order_by('-created_at')
        subject_id = request.GET.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        context = {
            'org': org, 'materials': qs,
            'subjects': Subject.objects.filter(org=org, status='active'),
            'selected_subject': subject_id,
        }
        return render(request, self.template_name, context)


class AddCourseMaterialView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_course_materials'
    template_name = 'admin/academic/course_material_form.html'

    def get(self, request):
        org = _get_org(request)
        form = CourseMaterialForm(org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})

    def post(self, request):
        org = _get_org(request)
        form = CourseMaterialForm(request.POST, request.FILES, org=org)
        if form.is_valid():
            material = form.save(commit=False)
            material.org = org
            material.uploaded_by = request.user
            material.save()

            from handle.models import member as member_model
            students = member_model.objects.filter(
                org=org, classification=material.subject.classification, status='active',
                member_type__in=('student', 'trainee'),
            )
            from handle.notifications import notify_many
            notify_many(
                students, 'course_material_added', f"New Material: {material.title}",
                body=material.subject.name, link_url='/staff/course-materials/',
            )
            messages.success(request, f"Course material '{material.title}' added.")
            return redirect('schooladmin:course_material_list')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'mode': 'create'})


class EditCourseMaterialView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_course_materials'
    template_name = 'admin/academic/course_material_form.html'

    def get(self, request, pk):
        org = _get_org(request)
        material = get_object_or_404(CourseMaterial, pk=pk, org=org)
        form = CourseMaterialForm(instance=material, org=org)
        return render(request, self.template_name, {'org': org, 'form': form, 'material': material, 'mode': 'edit'})

    def post(self, request, pk):
        org = _get_org(request)
        material = get_object_or_404(CourseMaterial, pk=pk, org=org)
        form = CourseMaterialForm(request.POST, request.FILES, instance=material, org=org)
        if form.is_valid():
            form.save()
            messages.success(request, "Course material updated.")
            return redirect('schooladmin:course_material_list')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'org': org, 'form': form, 'material': material, 'mode': 'edit'})


@feature_required('academic_management')
@perm_required('can_manage_course_materials')
def delete_course_material(request, pk):
    org = _get_org(request)
    material = get_object_or_404(CourseMaterial, pk=pk, org=org)
    material.delete()
    messages.success(request, "Course material deleted.")
    return redirect('schooladmin:course_material_list')


# ── Daily Teaching Log ───────────────────────────────────────────────────

class TeachingLogListView(FeatureRequiredMixin, PermRequiredMixin, View):
    """Admin review queue — filterable/searchable per the spec."""
    required_feature = 'academic_management'
    required_perm = 'can_approve_teaching_logs'
    template_name = 'admin/academic/teaching_log_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = TeachingLog.objects.filter(org=org).select_related('teacher', 'subject', 'classification', 'section').order_by('-date')
        status = request.GET.get('status')
        teacher_id = request.GET.get('teacher')
        search = request.GET.get('search', '').strip()
        if status:
            qs = qs.filter(status=status)
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        if search:
            qs = qs.filter(Q(topic_covered__icontains=search) | Q(chapter__icontains=search))
        context = {
            'org': org, 'logs': qs, 'selected_status': status, 'selected_teacher': teacher_id, 'search': search,
            'teachers': Staff.objects.filter(org=org).select_related('admin'),
        }
        return render(request, self.template_name, context)


class TeachingLogDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_approve_teaching_logs'
    template_name = 'admin/academic/teaching_log_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        log = get_object_or_404(TeachingLog, pk=pk, org=org)
        return render(request, self.template_name, {'org': org, 'log': log, 'attachments': log.attachments.all()})


@feature_required('academic_management')
@perm_required('can_approve_teaching_logs')
@require_POST
def approve_teaching_log(request, pk):
    org = _get_org(request)
    log = get_object_or_404(TeachingLog, pk=pk, org=org)
    if log.status not in ('submitted', 'rejected'):
        messages.error(request, "Only submitted attendance can be approved.")
        return redirect('schooladmin:teaching_log_detail', pk=log.pk)
    now = timezone.now()
    log.status = 'approved'
    log.approved_by = request.user
    log.approved_at = now
    log.reviewed_by = request.user
    log.reviewed_at = now
    log.rejection_reason = None
    log.save(update_fields=[
        'status', 'approved_by', 'approved_at', 'reviewed_by',
        'reviewed_at', 'rejection_reason',
    ])
    messages.success(request, "Teaching log approved.")
    return redirect('schooladmin:teaching_log_detail', pk=log.pk)


@feature_required('academic_management')
@perm_required('can_approve_teaching_logs')
@require_POST
def reject_teaching_log(request, pk):
    org = _get_org(request)
    log = get_object_or_404(TeachingLog, pk=pk, org=org)
    reason = request.POST.get('reason', '').strip()
    if not reason:
        messages.error(request, "A rejection reason is required.")
        return redirect('schooladmin:teaching_log_detail', pk=log.pk)
    if log.status != 'submitted':
        messages.error(request, "Only submitted attendance can be rejected.")
        return redirect('schooladmin:teaching_log_detail', pk=log.pk)
    log.status = 'rejected'
    log.approved_by = None
    log.approved_at = None
    log.reviewed_by = request.user
    log.reviewed_at = timezone.now()
    log.rejection_reason = reason
    log.save(update_fields=[
        'status', 'approved_by', 'approved_at', 'reviewed_by',
        'reviewed_at', 'rejection_reason',
    ])
    messages.success(request, "Teaching log rejected.")
    return redirect('schooladmin:teaching_log_detail', pk=log.pk)


class TeacherTeachingLogListView(FeatureRequiredMixin, PermRequiredMixin, View):
    """A teacher's own submitted logs."""
    required_feature = 'academic_management'
    required_perm = 'can_manage_teaching_logs'
    template_name = 'admin/academic/my_teaching_logs.html'

    def get(self, request):
        org = _get_org(request)
        logs = TeachingLog.objects.filter(org=org, teacher=request.user).select_related('subject', 'classification').order_by('-date')
        return render(request, self.template_name, {'org': org, 'logs': logs})


class AddTeachingLogView(FeatureRequiredMixin, PermRequiredMixin, View):
    """Submitting today's Teaching Log also marks per-student subject
    attendance for that period — auto-filled from the Class Routine
    (RoutinePeriod) when available, with a manual fallback. Core logic
    lives in handle/academics.py since real teachers (user_type 3) submit
    this from the staff portal, not from here — see staff/views.py's
    TeachingLogSubjectAttendanceView for that counterpart. This view mainly
    exists for admins to backfill/oversee."""
    required_feature = 'academic_management'
    required_perm = 'can_manage_teaching_logs'
    template_name = 'admin/academic/teaching_log_form.html'

    def get(self, request):
        from handle.academics import todays_routine_period_options, roster_for_subject

        org = _get_org(request)
        is_admin = request.user.user_type == '2'
        today = timezone.localdate()

        period_options = todays_routine_period_options(org, request.user, is_admin)

        manual_classification_id = request.GET.get('classification')
        manual_section_id = request.GET.get('section')
        manual_subject_id = request.GET.get('subject')
        manual_roster = None
        manual_subject = None
        if manual_classification_id and manual_subject_id:
            classification = Classification.objects.filter(org=org, pk=manual_classification_id).first()
            if classification:
                section = Section.objects.filter(org=org, pk=manual_section_id).first() if manual_section_id else None
                subject = Subject.objects.filter(
                    org=org, pk=manual_subject_id,
                    classification=classification,
                ).first()
                if subject and (
                    not subject.section_id
                    or subject.section_id == getattr(section, 'pk', None)
                ):
                    manual_roster = roster_for_subject(
                        org, subject, classification, section,
                        attendance_date=today,
                    )
                    manual_subject = subject

        form = TeachingLogForm(org=org, initial={
            'classification': manual_classification_id or None,
            'section': manual_section_id or None,
            'subject': request.GET.get('subject') or None,
            'date': request.GET.get('date') or today,
        })

        context = {
            'org': org, 'form': form, 'today': today, 'is_admin': is_admin,
            'period_options': period_options,
            'manual_roster': manual_roster,
            'manual_classification_id': manual_classification_id,
            'manual_section_id': manual_section_id,
            'manual_subject_id': manual_subject_id,
            'manual_subject': manual_subject,
            'teachers': Staff.objects.filter(org=org).select_related('admin') if is_admin else None,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        from handle.academics import submit_teaching_log_and_attendance

        org = _get_org(request)
        is_admin = request.user.user_type == '2'
        log, error = submit_teaching_log_and_attendance(org, request.user, is_admin, request.POST)
        if error:
            messages.error(request, error)
            return redirect('schooladmin:add_teaching_log')

        for f in request.FILES.getlist('attachments'):
            TeachingLogAttachment.objects.create(log=log, file=f)

        messages.success(request, "Teaching log & attendance submitted.")
        return redirect('schooladmin:my_teaching_logs')


class SubjectAttendanceReportView(FeatureRequiredMixin, PermRequiredMixin, View):
    """Per-subject-period attendance — one row per student per period, so a
    student's 5-6 daily subject periods each show as their own row, unlike
    the single whole-day DailyReport."""
    required_feature = 'academic_management'
    required_perm = 'can_approve_teaching_logs'
    template_name = 'admin/academic/subject_attendance_report.html'

    def get(self, request):
        org = _get_org(request)
        today = timezone.localdate()

        def _parse_date(value, default):
            try:
                return datetime.datetime.strptime(value, '%Y-%m-%d').date()
            except (TypeError, ValueError):
                return default

        date_from = _parse_date(request.GET.get('date_from'), today)
        date_to = _parse_date(request.GET.get('date_to'), today)
        if date_from > date_to:
            date_from, date_to = date_to, date_from

        academic_year_id = request.GET.get('academic_year')
        course_id = request.GET.get('course')
        classification_id = request.GET.get('classification')
        section_id = request.GET.get('section')
        subject_id = request.GET.get('subject')
        teacher_id = request.GET.get('teacher')
        student_id = request.GET.get('student')
        attendance_status = request.GET.get('attendance_status')
        session_status = request.GET.get('session_status')

        qs = SubjectAttendanceRecord.objects.filter(
            org=org, teaching_log__date__gte=date_from, teaching_log__date__lte=date_to,
        ).select_related(
            'teaching_log', 'teaching_log__academic_year', 'teaching_log__course',
            'teaching_log__subject', 'teaching_log__classification',
            'teaching_log__section', 'teaching_log__teacher', 'member',
        )
        if academic_year_id:
            qs = qs.filter(teaching_log__academic_year_id=academic_year_id)
        if course_id:
            qs = qs.filter(teaching_log__course_id=course_id)
        if classification_id:
            qs = qs.filter(teaching_log__classification_id=classification_id)
        if section_id:
            qs = qs.filter(teaching_log__section_id=section_id)
        if subject_id:
            qs = qs.filter(teaching_log__subject_id=subject_id)
        if teacher_id:
            qs = qs.filter(teaching_log__teacher_id=teacher_id)
        if student_id:
            qs = qs.filter(member_id=student_id)
        if attendance_status in dict(SubjectAttendanceRecord.STATUS_CHOICES):
            qs = qs.filter(status=attendance_status)
        if session_status in dict(TeachingLog.STATUS_CHOICES):
            qs = qs.filter(teaching_log__status=session_status)
        qs = qs.order_by('-teaching_log__date', 'teaching_log__period', 'member__name')

        summary = {}
        for rec in qs:
            row = summary.setdefault(rec.member_id, {
                'member': rec.member, 'present': 0, 'absent': 0,
                'late': 0, 'excused': 0, 'leave': 0,
            })
            row[rec.status] += 1
        summary_rows = sorted(summary.values(), key=lambda r: r['member'].name)
        for row in summary_rows:
            row['total'] = sum(row[key] for key in ('present', 'absent', 'late', 'excused', 'leave'))
            row['pct'] = round(
                ((row['present'] + row['late']) / row['total']) * 100, 1
            ) if row['total'] else 0

        context = {
            'org': org, 'records': qs, 'summary_rows': summary_rows,
            'date_from': date_from, 'date_to': date_to,
            'academic_years': AcademicYear.objects.filter(org=org, status='active'),
            'courses': Course.objects.filter(org=org, status='active'),
            'classifications': Classification.objects.filter(org=org, status='active'),
            'sections': Section.objects.filter(org=org, status='active'),
            'subjects': Subject.objects.filter(org=org, status='active').select_related('course'),
            'teachers': Staff.objects.filter(org=org).select_related('admin'),
            'students': member.objects.filter(
                org=org, status='active', member_type__in=('student', 'trainee'),
            ).order_by('name'),
            'attendance_status_choices': SubjectAttendanceRecord.STATUS_CHOICES,
            'session_status_choices': TeachingLog.STATUS_CHOICES,
            'selected_academic_year': academic_year_id,
            'selected_course': course_id,
            'selected_classification': classification_id,
            'selected_section': section_id,
            'selected_subject': subject_id,
            'selected_teacher': teacher_id,
            'selected_student': student_id,
            'selected_attendance_status': attendance_status,
            'selected_session_status': session_status,
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, self.template_name, context)


# ── Class Routine ────────────────────────────────────────────────────────

class RoutineListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_routine'
    template_name = 'admin/academic/routine_grid.html'

    def get(self, request):
        org = _get_org(request)
        classification_id = request.GET.get('classification', '')
        section_id = request.GET.get('section', '')
        teacher_id = request.GET.get('teacher', '')
        day_id = request.GET.get('day', '')
        qs = RoutinePeriod.objects.filter(org=org, is_active=True).select_related(
            'academic_year',
            'teacher_assignment',
            'subject__course',
            'teacher',
            'classification',
            'section',
        )
        if classification_id:
            qs = qs.filter(classification_id=classification_id)
        if section_id:
            qs = qs.filter(section_id=section_id)
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        if day_id.isdigit():
            qs = qs.filter(day_of_week=int(day_id))

        days = list(RoutinePeriod.DAY_CHOICES)
        visible_days = [
            day for day in days
            if not day_id.isdigit() or day[0] == int(day_id)
        ]
        periods = list(qs.order_by('period_number', 'start_time', 'day_of_week'))
        period_numbers = sorted({period.period_number for period in periods})
        schedule_rows = []
        for period_number in period_numbers:
            schedule_rows.append({
                'period_number': period_number,
                'by_day': {
                    day_num: [
                        period for period in periods
                        if period.period_number == period_number
                        and period.day_of_week == day_num
                    ]
                    for day_num, _ in visible_days
                },
            })
        current_day = (timezone.localdate().weekday() + 1) % 7
        context = {
            'org': org,
            'days': visible_days,
            'day_choices': days,
            'schedule_rows': schedule_rows,
            'period_count': len(periods),
            'teacher_count': len({period.teacher_id for period in periods}),
            'class_count': len({period.classification_id for period in periods}),
            'unmapped_count': RoutinePeriod.objects.filter(
                org=org, is_active=True, teacher_assignment__isnull=True,
            ).count(),
            'current_day': current_day,
            'classifications': Classification.objects.filter(org=org, status='active'),
            'sections': Section.objects.filter(
                org=org, status='active',
            ).select_related('classification'),
            'teachers': Staff.objects.filter(
                org=org,
                admin__subject_assignments__status='active',
            ).select_related('admin', 'member').distinct().order_by('member__name'),
            'selected_classification': classification_id,
            'selected_section': section_id,
            'selected_teacher': teacher_id,
            'selected_day': day_id,
        }
        return render(request, self.template_name, context)


class AddRoutinePeriodView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_routine'
    template_name = 'admin/academic/routine_form.html'

    def get(self, request):
        org = _get_org(request)
        initial = {}
        period = request.GET.get('period')
        if period and period.isdigit():
            initial['period_number'] = period
        form = RoutinePeriodForm(org=org, initial=initial)
        preselected_days = []
        day = request.GET.get('day')
        if day and day.isdigit():
            preselected_days = [day]
        return render(request, self.template_name, {
            'org': org,
            'form': form,
            'mode': 'create',
            'assignment_count': form.fields['teacher_assignment'].queryset.count(),
            'day_choices': RoutinePeriod.DAY_CHOICES,
            'preselected_days': preselected_days,
        })

    def post(self, request):
        org = _get_org(request)
        days = request.POST.getlist('days')
        if not days and request.POST.get('day_of_week'):
            # Back-compat: a single-day submission (the old form shape, still
            # used by anything posting `day_of_week` directly) behaves like a
            # one-day selection instead of being rejected.
            days = [request.POST.get('day_of_week')]
        common_ctx = {
            'org': org,
            'mode': 'create',
            'day_choices': RoutinePeriod.DAY_CHOICES,
        }

        if not days:
            messages.error(request, "Select at least one day.")
            form = RoutinePeriodForm(request.POST, org=org)
            return render(request, self.template_name, {
                **common_ctx,
                'form': form,
                'assignment_count': form.fields['teacher_assignment'].queryset.count(),
                'preselected_days': [],
            })

        # One shared period (subject/teacher/time/room) applied across every
        # selected day - each day is validated and saved independently so a
        # conflict on one day (e.g. the teacher/room/class is already booked)
        # only skips that day's cell instead of blocking the whole batch.
        created, skipped = [], []
        last_invalid_form = None
        for day in days:
            day_data = request.POST.copy()
            day_data['day_of_week'] = day
            form = RoutinePeriodForm(day_data, org=org)
            if form.is_valid():
                period = form.save(commit=False)
                period.org = org
                period.save()
                created.append(period)
            else:
                last_invalid_form = form
                day_label = dict(RoutinePeriod.DAY_CHOICES).get(int(day), day)
                reason = ' '.join(str(e) for e in form.non_field_errors())
                if not reason:
                    reason = ' '.join(
                        f"{field}: {' '.join(errs)}" for field, errs in form.errors.items()
                    ) or 'could not be scheduled'
                skipped.append(f"{day_label} — {reason}")

        if created:
            day_names = ', '.join(p.get_day_of_week_display() for p in created)
            messages.success(
                request,
                f"Period {created[0].period_number} ({created[0].subject.name}) created for {day_names}.",
            )
        if skipped:
            messages.warning(
                request,
                "Left empty — a conflicting period already exists: " + "; ".join(skipped),
            )
        if created:
            return redirect('schooladmin:routine_grid')

        messages.error(request, "Please correct the errors below.")
        form = last_invalid_form or RoutinePeriodForm(request.POST, org=org)
        return render(request, self.template_name, {
            **common_ctx,
            'form': form,
            'assignment_count': form.fields['teacher_assignment'].queryset.count(),
            'preselected_days': days,
        })


class EditRoutinePeriodView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'academic_management'
    required_perm = 'can_manage_routine'
    template_name = 'admin/academic/routine_form.html'

    def get(self, request, pk):
        org = _get_org(request)
        period = get_object_or_404(RoutinePeriod, pk=pk, org=org)
        form = RoutinePeriodForm(instance=period, org=org)
        return render(request, self.template_name, {
            'org': org,
            'form': form,
            'period': period,
            'mode': 'edit',
            'assignment_count': form.fields['teacher_assignment'].queryset.count(),
        })

    def post(self, request, pk):
        org = _get_org(request)
        period = get_object_or_404(RoutinePeriod, pk=pk, org=org)
        form = RoutinePeriodForm(request.POST, instance=period, org=org)
        if form.is_valid():
            form.save()
            messages.success(request, "Routine period updated.")
            return redirect('schooladmin:routine_grid')
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {
            'org': org,
            'form': form,
            'period': period,
            'mode': 'edit',
            'assignment_count': form.fields['teacher_assignment'].queryset.count(),
        })


@feature_required('academic_management')
@perm_required('can_manage_routine')
def delete_routine_period(request, pk):
    org = _get_org(request)
    RoutinePeriod.objects.filter(pk=pk, org=org).delete()
    messages.success(request, "Routine period deleted.")
    return redirect('schooladmin:routine_grid')


class TeacherRoutineView(FeatureRequiredMixin, PermRequiredMixin, View):
    """A teacher's own weekly schedule, read-only."""
    required_feature = 'academic_management'
    required_perm = 'can_manage_routine'
    template_name = 'admin/academic/teacher_routine.html'

    def get(self, request):
        org = _get_org(request)
        qs = RoutinePeriod.objects.filter(org=org, teacher=request.user, is_active=True).select_related('subject', 'classification', 'section')
        days = list(RoutinePeriod.DAY_CHOICES)
        grid = {d[0]: list(qs.filter(day_of_week=d[0]).order_by('period_number')) for d in days}
        return render(request, self.template_name, {'org': org, 'days': days, 'grid': grid})


# ── Reports ───────────────────────────────────────────────────────────────

@login_required
@feature_required('academic_management')
def export_academic(request):
    """CSV/Excel/PDF export for the six Academic Management reports."""
    org = _get_org(request)
    export_type = request.GET.get('type', 'assignment_submission')
    fmt = request.GET.get('fmt', 'excel')

    if export_type == 'assignment_submission':
        qs = AssignmentSubmission.objects.filter(assignment__org=org).select_related('assignment', 'student')
        headers = ['Assignment', 'Student', 'Submitted', 'Late', 'Status', 'Marks']
        rows = [[s.assignment.title, s.student.name, str(s.submitted_at or ''), 'Yes' if s.is_late else 'No',
                 s.get_status_display(), str(s.obtained_marks or '')] for s in qs]
        filename = 'assignment_submission_report'
    elif export_type == 'homework_completion':
        qs = HomeworkStatus.objects.filter(homework__org=org).select_related('homework', 'student')
        headers = ['Homework', 'Student', 'Status', 'Completed At', 'Verified']
        rows = [[h.homework.subject.name, h.student.name, h.get_status_display(), str(h.completed_at or ''),
                 'Yes' if h.verified_by_teacher else 'No'] for h in qs]
        filename = 'homework_completion_report'
    elif export_type == 'teaching_logs':
        qs = TeachingLog.objects.filter(org=org).select_related('teacher', 'subject', 'classification')
        headers = ['Date', 'Teacher', 'Subject', 'Class', 'Topic', 'Status']
        rows = [[str(l.date), str(l.teacher), l.subject.name, str(l.classification), l.topic_covered, l.get_status_display()] for l in qs]
        filename = 'teaching_log_report'
    elif export_type == 'teacher_performance':
        qs = TeachingLog.objects.filter(org=org).values('teacher__email').annotate(
            total=Count('id'), approved=Count('id', filter=Q(status='approved')), rejected=Count('id', filter=Q(status='rejected'))
        )
        headers = ['Teacher', 'Total Logs', 'Approved', 'Rejected']
        rows = [[r['teacher__email'], r['total'], r['approved'], r['rejected']] for r in qs]
        filename = 'teacher_performance_report'
    elif export_type == 'student_performance':
        qs = AssignmentSubmission.objects.filter(assignment__org=org, status='graded').values('student__name').annotate(
            total=Count('id'), avg_marks=Sum('obtained_marks')
        )
        headers = ['Student', 'Graded Assignments', 'Total Marks']
        rows = [[r['student__name'], r['total'], str(r['avg_marks'] or 0)] for r in qs]
        filename = 'student_performance_report'
    elif export_type == 'subject_attendance':
        qs = SubjectAttendanceRecord.objects.filter(org=org).select_related(
            'teaching_log', 'teaching_log__academic_year', 'teaching_log__course',
            'teaching_log__subject', 'teaching_log__classification',
            'teaching_log__section', 'teaching_log__teacher', 'member',
        )
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from:
            qs = qs.filter(teaching_log__date__gte=date_from)
        if date_to:
            qs = qs.filter(teaching_log__date__lte=date_to)
        for field, param in (
            ('teaching_log__academic_year_id', 'academic_year'),
            ('teaching_log__course_id', 'course'),
            ('teaching_log__classification_id', 'classification'),
            ('teaching_log__section_id', 'section'),
            ('teaching_log__subject_id', 'subject'),
            ('teaching_log__teacher_id', 'teacher'),
            ('member_id', 'student'),
            ('status', 'attendance_status'),
            ('teaching_log__status', 'session_status'),
        ):
            value = request.GET.get(param)
            if value:
                qs = qs.filter(**{field: value})
        qs = qs.order_by('-teaching_log__date', 'teaching_log__period', 'member__name')
        headers = [
            'Date', 'Academic Year', 'Course', 'Period', 'Subject', 'Class',
            'Section', 'Student', 'Attendance Status', 'Session Status', 'Teacher',
        ]
        rows = [[
            str(r.teaching_log.date), str(r.teaching_log.academic_year or ''),
            str(r.teaching_log.course or ''), r.teaching_log.period or '',
            r.teaching_log.subject.name,
            str(r.teaching_log.classification), str(r.teaching_log.section or ''), r.member.name,
            r.get_status_display(), r.teaching_log.get_status_display(),
            str(r.teaching_log.teacher),
        ] for r in qs]
        filename = 'subject_attendance_report'
    else:  # course_material_usage
        qs = CourseMaterial.objects.filter(org=org)
        headers = ['Material', 'Subject', 'Views', 'Downloads']
        rows = [[m.title, m.subject.name, m.view_count(), m.download_count()] for m in qs]
        filename = 'course_material_usage_report'

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return resp

    if fmt == 'pdf':
        from school.pdf_utils import render_to_pdf
        pdf_bytes = render_to_pdf('admin/academic/pdf/generic_report.html', {
            'org': org, 'title': filename.replace('_', ' ').title(), 'headers': headers, 'rows': rows,
        })
        if pdf_bytes is None:
            return HttpResponse('PDF rendering failed', status=500)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:31]
    _style_header(ws, headers, fill_color='4338CA')
    for ridx, row in enumerate(rows, 2):
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)
    resp = _excel_response(f'{filename}.xlsx')
    wb.save(resp)
    return resp


# =============================================================
# EVENT MANAGEMENT
# =============================================================

class EventListView(FeatureRequiredMixin, View):
    required_feature = 'events'
    template_name = 'admin/events/event_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Event.objects.filter(org=org).select_related('branch', 'responsible_staff').order_by('-start_date')
        status = request.GET.get('status')
        event_type = request.GET.get('event_type')
        if status:
            qs = qs.filter(status=status)
        if event_type:
            qs = qs.filter(event_type=event_type)
        context = {
            'org': org,
            'events': qs,
            'branches': Branch.objects.filter(org=org, status='active'),
            'selected_status': status,
            'selected_type': event_type,
            'event_types': Event.EVENT_TYPE_CHOICES,
            'status_choices': Event.STATUS_CHOICES,
        }
        return render(request, self.template_name, context)


class AddEventView(FeatureRequiredMixin, View):
    required_feature = 'events'
    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'staff_list': Staff.objects.filter(org=org).select_related('admin'),
            'event_types': Event.EVENT_TYPE_CHOICES,
            'today': datetime.date.today().strftime('%Y-%m-%d'),
        }
        return render(request, 'admin/events/add_event.html', context)

    def post(self, request):
        org = _get_org(request)
        try:
            event = Event.objects.create(
                org=org,
                title=request.POST.get('title'),
                event_type=request.POST.get('event_type', 'other'),
                branch_id=request.POST.get('branch') or None,
                start_date=request.POST.get('start_date'),
                end_date=request.POST.get('end_date'),
                location=request.POST.get('location', ''),
                description=request.POST.get('description', ''),
                responsible_staff_id=request.POST.get('responsible_staff') or None,
                status=request.POST.get('status', 'upcoming'),
            )
            messages.success(request, f"Event '{event.title}' created.")
            return redirect('schooladmin:event_detail', pk=event.pk)
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:add_event')


class EventDetailView(FeatureRequiredMixin, View):
    required_feature = 'events'
    template_name = 'admin/events/event_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        event = get_object_or_404(Event, pk=pk, org=org)
        stock_usages = EventStockUsage.objects.filter(event=event).select_related('item')
        available_items = StockItem.objects.filter(org=org, status='active').order_by('name')
        context = {
            'org': org,
            'event': event,
            'stock_usages': stock_usages,
            'available_items': available_items,
            'branches': Branch.objects.filter(org=org, status='active'),
            'staff_list': Staff.objects.filter(org=org).select_related('admin'),
            'event_types': Event.EVENT_TYPE_CHOICES,
            'status_choices': Event.STATUS_CHOICES,
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        org = _get_org(request)
        event = get_object_or_404(Event, pk=pk, org=org)
        action = request.POST.get('action')

        if action == 'update_event':
            event.title = request.POST.get('title', event.title)
            event.event_type = request.POST.get('event_type', event.event_type)
            event.branch_id = request.POST.get('branch') or None
            event.start_date = request.POST.get('start_date', event.start_date)
            event.end_date = request.POST.get('end_date', event.end_date)
            event.location = request.POST.get('location', event.location)
            event.description = request.POST.get('description', event.description)
            event.responsible_staff_id = request.POST.get('responsible_staff') or None
            event.status = request.POST.get('status', event.status)
            event.save()
            messages.success(request, "Event updated.")

        elif action == 'add_stock':
            item_id = request.POST.get('item_id')
            quantity = Decimal(str(request.POST.get('quantity', 0) or 0))
            if item_id and quantity > 0:
                item = get_object_or_404(StockItem, pk=item_id, org=org)
                if item.quantity < quantity:
                    messages.error(request, f"Not enough stock for '{item.name}'. Available: {item.quantity}")
                else:
                    EventStockUsage.objects.create(event=event, item=item, quantity_used=quantity, note=request.POST.get('note', ''))
                    messages.success(request, f"Stock usage recorded for '{item.name}'.")
            else:
                messages.error(request, "Select an item and enter valid quantity.")

        elif action == 'remove_stock':
            usage_id = request.POST.get('usage_id')
            usage = get_object_or_404(EventStockUsage, pk=usage_id, event=event)
            usage.delete()
            messages.success(request, "Stock usage removed and quantity restored.")

        return redirect('schooladmin:event_detail', pk=pk)


@feature_required('events')
def delete_event(request, pk):
    org = _get_org(request)
    event = get_object_or_404(Event, pk=pk, org=org)
    event.delete()
    messages.success(request, "Event deleted.")
    return redirect('schooladmin:event_list')


# =============================================================
# COURSE MANAGEMENT
# =============================================================

class CourseListView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    template_name = 'admin/courses/course_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Course.objects.filter(org=org).select_related(
            'branch', 'teacher'
        ).prefetch_related(
            'classifications', 'sections', 'teacher_assignments__teacher'
        ).order_by('name')
        context = {
            'org': org,
            'courses': qs,
            'branches': Branch.objects.filter(org=org, status='active'),
        }
        return render(request, self.template_name, context)


class AddCourseView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'classifications': Classification.objects.filter(org=org),
            'sections': Section.objects.filter(org=org),
            'staff_list': Staff.objects.filter(org=org).select_related('admin'),
        }
        return render(request, 'admin/courses/add_course.html', context)

    def post(self, request):
        org = _get_org(request)
        try:
            teacher_id = request.POST.get('teacher') or None
            teacher = None
            if teacher_id:
                teacher = Staff.objects.filter(
                    org=org, admin_id=teacher_id
                ).select_related('admin').first()
                if not teacher:
                    messages.error(request, "Please select a staff member from your organization.")
                    return redirect('schooladmin:add_course')
            course = Course.objects.create(
                org=org,
                name=request.POST.get('name'),
                code=request.POST.get('code', ''),
                branch_id=request.POST.get('branch') or None,
                teacher=teacher.admin if teacher else None,
                description=request.POST.get('description', ''),
                credit_hour=request.POST.get('credit_hour', 0),
                status=request.POST.get('status', 'active'),
            )
            classification_ids = request.POST.getlist('classifications')
            section_ids = request.POST.getlist('sections')
            if classification_ids:
                course.classifications.set(classification_ids)
            if section_ids:
                course.sections.set(section_ids)
            messages.success(request, f"Course '{course.name}' created.")
            return redirect('schooladmin:course_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:add_course')


@feature_required('courses')
def delete_course(request, pk):
    org = _get_org(request)
    course = get_object_or_404(Course, pk=pk, org=org)
    course.delete()
    messages.success(request, "Course deleted.")
    return redirect('schooladmin:course_list')


class CourseAttendanceView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    template_name = 'admin/courses/course_attendance.html'

    def get(self, request):
        return redirect('schooladmin:subject_list')
        org = _get_org(request)
        courses = Course.objects.filter(org=org, status='active')
        selected_course_id = request.GET.get('course')
        selected_date = request.GET.get('date', datetime.date.today().strftime('%Y-%m-%d'))
        members_list = []
        course = None
        attendance_records = {}

        if selected_course_id:
            course = get_object_or_404(Course, pk=selected_course_id, org=org)
            members_list = member.objects.filter(org=org, courses__id=selected_course_id).exclude(status='dumped')
            records = CourseAttendance.objects.filter(course=course, attendance_date=selected_date)
            attendance_records = {r.pk: r for r in records}

        context = {
            'org': org,
            'courses': courses,
            'selected_course': course,
            'selected_date': selected_date,
            'members_list': members_list,
            'attendance_records': attendance_records,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        return redirect('schooladmin:subject_list')
        org = _get_org(request)
        course_id = request.POST.get('course_id')
        attendance_date = request.POST.get('attendance_date')
        topic_taught = request.POST.get('topic_taught', '')
        gap_note = request.POST.get('gap_note', '')
        present_ids = request.POST.getlist('present_members')

        course = get_object_or_404(Course, pk=course_id, org=org)
        all_members = member.objects.filter(org=org, courses__id=course_id).exclude(status='dumped')

        for mem in all_members:
            is_present = str(mem.id) in present_ids
            if not is_present:
                AttendanceGap.objects.get_or_create(
                    org=org, member=mem, course=course, date=attendance_date,
                    defaults={
                        'branch': course.branch,
                        'teacher': request.user,
                        'topic_missed': topic_taught,
                        'reason': 'Absent from course session',
                        'recovery_status': 'pending',
                    }
                )

        if topic_taught:
            CourseAttendance.objects.update_or_create(
                org=org, course=course, attendance_date=attendance_date,
                staff=request.user,
                defaults={'topic_taught': topic_taught, 'gap_note': gap_note,
                          'branch': course.branch, 'classification': None, 'section': None}
            )

        messages.success(request, f"Course attendance saved for {attendance_date}.")
        return redirect(f"{request.path}?course={course_id}&date={attendance_date}")


# =============================================================
# STUDY GAP / TEACHING LOG
# =============================================================

class StudyGapListView(FeatureRequiredMixin, View):
    required_feature = 'study_gap'
    template_name = 'admin/study_gap/list.html'

    def get(self, request):
        org = _get_org(request)
        qs = AttendanceGap.objects.filter(org=org).select_related('member', 'course', 'branch').order_by('-date')
        course_id = request.GET.get('course')
        status = request.GET.get('recovery_status')
        member_id = request.GET.get('member')
        if course_id:
            qs = qs.filter(course_id=course_id)
        if status:
            qs = qs.filter(recovery_status=status)
        if member_id:
            qs = qs.filter(member_id=member_id)
        context = {
            'org': org,
            'gaps': qs,
            'courses': Course.objects.filter(org=org),
            'members': member.objects.filter(org=org).exclude(status='dumped'),
            'selected_course': course_id,
            'selected_status': status,
            'selected_member': member_id,
            'status_choices': AttendanceGap._meta.get_field('recovery_status').choices,
        }
        return render(request, self.template_name, context)


@feature_required('study_gap')
def update_gap_status(request, pk):
    org = _get_org(request)
    gap = get_object_or_404(AttendanceGap, pk=pk, org=org)
    new_status = request.POST.get('status')
    if new_status:
        gap.recovery_status = new_status
        if new_status == 'covered':
            gap.covered_at = timezone.now()
        gap.save()
        messages.success(request, "Study gap status updated.")
    return redirect('schooladmin:study_gap_list')


# =============================================================
# RESULT MANAGEMENT
# =============================================================
from handle.models import compute_grade as _compute_grade
from handle.models import Section as _Section


class SubjectListView(FeatureRequiredMixin, View):
    required_feature = 'courses'
    template_name = 'admin/results/subjects.html'

    def _get_subjects(self, org, classification_id=None, section_id=None, course_id=None):
        qs = Subject.objects.filter(org=org).select_related(
            'course', 'classification', 'section', 'teacher'
        ).prefetch_related('teacher_assignments__teacher')
        if course_id:
            qs = qs.filter(course_id=course_id)
        if classification_id:
            qs = qs.filter(classification_id=classification_id)
        if section_id:
            qs = qs.filter(section_id=section_id)
        return qs

    def get(self, request):
        org = _get_org(request)
        classification_id = request.GET.get('classification')
        section_id = request.GET.get('section')
        course_id = request.GET.get('course')
        classifications = Classification.objects.filter(org=org)
        sections = _Section.objects.filter(org=org, classification_id=classification_id) if classification_id else _Section.objects.none()
        from management.models import CustomUser
        teachers = CustomUser.objects.filter(staff__org=org)
        context = {
            'org': org,
            'subjects': self._get_subjects(org, classification_id, section_id, course_id),
            'courses': Course.objects.filter(org=org, status='active').prefetch_related(
                'classifications', 'sections'
            ),
            'classifications': classifications,
            'sections': sections,
            'teachers': teachers,
            'selected_course': course_id,
            'selected_classification': classification_id,
            'selected_section': section_id,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            course_id = request.POST.get('course') or None
            cls_id = request.POST.get('classification') or None
            sec_id = request.POST.get('section') or None
            teacher_id = request.POST.get('teacher') or None
            name = request.POST.get('name', '').strip()
            try:
                full_m = float(request.POST.get('full_marks', 100))
                pass_m = float(request.POST.get('pass_marks', 40))
            except (TypeError, ValueError):
                full_m = pass_m = -1
            course = Course.objects.filter(org=org, pk=course_id, status='active').first()
            classification = Classification.objects.filter(org=org, pk=cls_id).first()
            section = _Section.objects.filter(org=org, pk=sec_id).first() if sec_id else None
            teacher = Staff.objects.filter(org=org, admin_id=teacher_id).select_related('admin').first() if teacher_id else None
            if not course:
                messages.error(request, "Course / program is required.")
            elif not classification:
                messages.error(request, "Classification is required.")
            elif not course.classifications.filter(pk=classification.pk).exists():
                messages.error(request, "The selected classification is not linked to this course.")
            elif section and section.classification_id != classification.pk:
                messages.error(request, "The selected section does not belong to this classification.")
            elif section and course.sections.exists() and not course.sections.filter(pk=section.pk).exists():
                messages.error(request, "The selected section is not linked to this course.")
            elif teacher_id and not teacher:
                messages.error(request, "Please select a teacher from your organization.")
            elif not name:
                messages.error(request, "Subject name is required.")
            elif pass_m < 0:
                messages.error(request, "Pass marks cannot be negative.")
            elif full_m <= pass_m:
                messages.error(request, "Full marks must be greater than pass marks.")
            elif Subject.objects.filter(
                org=org, course=course, classification_id=cls_id,
                section_id=sec_id, name=name,
            ).exists():
                messages.error(request, "A subject with this name already exists in this course/classification/section.")
            else:
                subject = Subject.objects.create(
                    org=org,
                    course=course,
                    name=name,
                    code=request.POST.get('code', '').strip() or None,
                    description=request.POST.get('description', '').strip() or None,
                    credit_hour=request.POST.get('credit_hour') or None,
                    classification=classification,
                    section=section,
                    teacher=teacher.admin if teacher else None,
                    full_marks=full_m,
                    pass_marks=pass_m,
                    monthly_fee=_money(request.POST.get('monthly_fee') or 0),
                    one_time_fee=_money(request.POST.get('one_time_fee') or 0),
                    status=request.POST.get('status', 'active'),
                )
                if teacher:
                    SubjectTeacherAssignment.objects.get_or_create(
                        subject=subject,
                        teacher=teacher.admin,
                        defaults={'is_primary': True},
                    )
                messages.success(request, f"Subject '{name}' added successfully.")
        elif action == 'edit':
            subj = get_object_or_404(Subject, pk=request.POST.get('subject_id'), org=org)
            course = Course.objects.filter(
                org=org, pk=request.POST.get('course'), status='active'
            ).first()
            teacher_id = request.POST.get('teacher') or None
            teacher = Staff.objects.filter(
                org=org, admin_id=teacher_id
            ).select_related('admin').first() if teacher_id else None
            if not course:
                messages.error(request, "Please select a valid course / program.")
                return redirect('schooladmin:subject_list')
            if not course.classifications.filter(pk=subj.classification_id).exists():
                messages.error(request, "This subject's classification is not linked to that course.")
                return redirect('schooladmin:subject_list')
            if subj.section_id and course.sections.exists() and not course.sections.filter(pk=subj.section_id).exists():
                messages.error(request, "This subject's section is not linked to that course.")
                return redirect('schooladmin:subject_list')
            if teacher_id and not teacher:
                messages.error(request, "Please select a teacher from your organization.")
                return redirect('schooladmin:subject_list')
            subj.course = course
            subj.name = request.POST.get('name', subj.name).strip()
            subj.code = request.POST.get('code', '').strip() or None
            subj.teacher = teacher.admin if teacher else None
            subj.full_marks = float(request.POST.get('full_marks', subj.full_marks))
            subj.pass_marks = float(request.POST.get('pass_marks', subj.pass_marks))
            subj.monthly_fee = _money(request.POST.get('monthly_fee', subj.monthly_fee))
            subj.one_time_fee = _money(request.POST.get('one_time_fee', subj.one_time_fee))
            subj.status = request.POST.get('status', subj.status)
            subj.credit_hour = request.POST.get('credit_hour') or None
            subj.save()
            if teacher:
                SubjectTeacherAssignment.objects.filter(
                    subject=subj
                ).exclude(teacher=teacher.admin).update(is_primary=False)
                assignment = SubjectTeacherAssignment.objects.filter(
                    subject=subj, teacher=teacher.admin, status='active',
                ).order_by('-start_date', '-pk').first()
                if assignment:
                    assignment.is_primary = True
                    assignment.save(update_fields=['is_primary', 'updated_at'])
                else:
                    SubjectTeacherAssignment.objects.create(
                        subject=subj,
                        teacher=teacher.admin,
                        is_primary=True,
                        assigned_by=request.user,
                    )
            messages.success(request, "Subject updated.")
        elif action == 'delete':
            Subject.objects.filter(pk=request.POST.get('subject_id'), org=org).delete()
            messages.success(request, "Subject deleted.")
        params = ''
        query_parts = []
        if request.POST.get('course'):
            query_parts.append(f"course={request.POST.get('course')}")
        if request.POST.get('classification'):
            query_parts.append(f"classification={request.POST.get('classification')}")
        if query_parts:
            params = f"?{'&'.join(query_parts)}"
        return redirect(f"{reverse('schooladmin:subject_list')}{params}")


class ExamTermListView(FeatureRequiredMixin, View):
    required_feature = 'results'
    template_name = 'admin/results/exam_terms.html'

    def get(self, request):
        org = _get_org(request)
        qs = ExamTerm.objects.filter(org=org).select_related('classification', 'section')
        classification_id = request.GET.get('classification')
        if classification_id:
            qs = qs.filter(classification_id=classification_id)
        total_students = member.objects.filter(org=org).exclude(status='dumped').count()
        # Annotate each exam with entry/publish stats
        exams_data = []
        for exam in qs:
            entry_count = ResultRecord.objects.filter(exam=exam).values('student').distinct().count()
            exams_data.append({'exam': exam, 'entry_count': entry_count, 'total_students': total_students})
        context = {
            'org': org,
            'exams_data': exams_data,
            'classifications': Classification.objects.filter(org=org),
            'selected_classification': classification_id,
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            ExamTerm.objects.create(
                org=org,
                name=request.POST.get('name'),
                academic_year=request.POST.get('academic_year', '').strip() or None,
                classification_id=request.POST.get('classification') or None,
                section_id=request.POST.get('section') or None,
                start_date=request.POST.get('start_date') or None,
                end_date=request.POST.get('end_date') or None,
                status=request.POST.get('status', 'draft'),
            )
            messages.success(request, "Exam created.")
        elif action == 'delete':
            ExamTerm.objects.filter(pk=request.POST.get('exam_id'), org=org).delete()
            messages.success(request, "Exam deleted.")
        elif action == 'set_status':
            exam = get_object_or_404(ExamTerm, pk=request.POST.get('exam_id'), org=org)
            new_status = request.POST.get('status')
            exam.status = new_status
            if new_status == 'published':
                exam.is_published = True
            elif new_status in ('draft', 'archived'):
                exam.is_published = False
            exam.save()
            messages.success(request, f"Exam status updated to {exam.get_status_display()}.")
        elif action == 'toggle_publish':
            exam = get_object_or_404(ExamTerm, pk=request.POST.get('exam_id'), org=org)
            exam.is_published = not exam.is_published
            exam.status = 'published' if exam.is_published else 'marks_entry'
            exam.save()
            messages.success(request, f"Result {'published' if exam.is_published else 'unpublished'}.")
        return redirect('schooladmin:exam_terms')


class ResultEntryView(FeatureRequiredMixin, View):
    required_feature = 'results'
    template_name = 'admin/results/result_entry.html'

    def _build_members_data(self, org, exam, classification, section_id=None):
        subjects_qs = Subject.objects.filter(org=org, classification=classification, status='active')
        if section_id:
            subjects_qs = subjects_qs.filter(Q(section_id=section_id) | Q(section__isnull=True))
        subjects = list(subjects_qs)

        members_filter = {'org': org, 'classification': classification}
        if section_id:
            members_filter['section_id'] = section_id
        members_qs = member.objects.filter(**members_filter).exclude(status='dumped').order_by('name')

        members_data = []
        for mem in members_qs:
            existing = {r.subject_id: r for r in ResultRecord.objects.filter(student=mem, exam=exam).select_related('subject')} if exam else {}
            members_data.append({'member': mem, 'records': existing})
        return subjects, members_data

    def get(self, request):
        org = _get_org(request)
        exam_id = request.GET.get('exam')
        classification_id = request.GET.get('classification')
        section_id = request.GET.get('section')
        exam = classification = None
        subjects = []
        members_data = []
        sections = []

        if exam_id:
            exam = get_object_or_404(ExamTerm, pk=exam_id, org=org)
        if classification_id:
            classification = get_object_or_404(Classification, pk=classification_id, org=org)
            sections = list(_Section.objects.filter(org=org, classification=classification))
            if exam:
                subjects, members_data = self._build_members_data(org, exam, classification, section_id)

        context = {
            'org': org,
            'exams': ExamTerm.objects.filter(org=org),
            'classifications': Classification.objects.filter(org=org),
            'sections': sections,
            'selected_exam': exam,
            'selected_classification': classification,
            'selected_section': section_id,
            'subjects': subjects,
            'members_data': members_data,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        exam_id = request.POST.get('exam_id')
        classification_id = request.POST.get('classification_id')
        section_id = request.POST.get('section_id') or None
        exam = get_object_or_404(ExamTerm, pk=exam_id, org=org)
        subjects_qs = Subject.objects.filter(org=org, classification_id=classification_id, status='active')
        if section_id:
            subjects_qs = subjects_qs.filter(Q(section_id=section_id) | Q(section__isnull=True))
        members_filter = {'org': org, 'classification_id': classification_id}
        if section_id:
            members_filter['section_id'] = section_id
        members_qs = member.objects.filter(**members_filter).exclude(status='dumped')

        saved = 0
        for mem in members_qs:
            for subj in subjects_qs:
                marks_key = f"marks_{mem.id}_{subj.id}"
                absent_key = f"absent_{mem.id}_{subj.id}"
                marks_val = request.POST.get(marks_key, '').strip()
                is_absent = request.POST.get(absent_key) == 'on'
                if marks_val != '' or is_absent:
                    try:
                        marks_float = 0.0 if is_absent else float(marks_val)
                        ResultRecord.objects.update_or_create(
                            student=mem, exam=exam, subject=subj,
                            defaults={
                                'obtained_marks': marks_float,
                                'is_absent': is_absent,
                                'remarks': request.POST.get(f"remarks_{mem.id}_{subj.id}", '').strip() or ('Absent' if is_absent else None),
                                'updated_by': request.user,
                            }
                        )
                        saved += 1
                    except (ValueError, Exception):
                        pass

        # Auto-set exam to marks_entry status
        if exam.status == 'draft' and saved:
            exam.status = 'marks_entry'
            exam.save(update_fields=['status'])

        messages.success(request, f"Saved {saved} result records.")
        qs = f"?exam={exam_id}&classification={classification_id}"
        if section_id:
            qs += f"&section={section_id}"
        return redirect(f"{reverse('schooladmin:result_entry')}{qs}")


class ResultReportView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'results'
    required_perm = 'can_view_result_report'
    template_name = 'admin/results/result_report.html'

    def get(self, request):
        org = _get_org(request)
        exam_id = request.GET.get('exam')
        classification_id = request.GET.get('classification')
        section_id = request.GET.get('section')
        exam = classification = None
        report_data = []
        subjects = []
        sections = []

        if exam_id and classification_id:
            exam = get_object_or_404(ExamTerm, pk=exam_id, org=org)
            classification = get_object_or_404(Classification, pk=classification_id, org=org)
            sections = list(_Section.objects.filter(org=org, classification=classification))
            subjects_qs = Subject.objects.filter(org=org, classification=classification, status='active')
            if section_id:
                subjects_qs = subjects_qs.filter(Q(section_id=section_id) | Q(section__isnull=True))
            subjects = list(subjects_qs)
            members_filter = {'org': org, 'classification': classification}
            if section_id:
                members_filter['section_id'] = section_id
            members_qs = member.objects.filter(**members_filter).exclude(status='dumped').order_by('name')

            pass_count = fail_count = absent_count = entry_count = 0
            for mem in members_qs:
                results = {r.subject_id: r for r in ResultRecord.objects.filter(student=mem, exam=exam).select_related('subject')}
                total_obt = float(sum(r.obtained_marks for r in results.values()))
                full_total = float(sum(s.full_marks for s in subjects))
                pct = round(total_obt / full_total * 100, 1) if full_total else 0
                passed = all(r.is_passed for r in results.values()) if results else False
                has_absent = any(r.is_absent for r in results.values())
                if results:
                    entry_count += 1
                    if has_absent:
                        absent_count += 1
                    elif passed:
                        pass_count += 1
                    else:
                        fail_count += 1
                report_data.append({
                    'member': mem,
                    'results': results,
                    'total': total_obt,
                    'full_total': full_total,
                    'percentage': pct,
                    'passed': passed,
                    'grade': _compute_grade(pct),
                    'has_absent': has_absent,
                    'rank': 0,
                })

            # Assign ranks (by total marks desc)
            ranked = sorted([r for r in report_data if r['results']], key=lambda x: x['total'], reverse=True)
            for i, r in enumerate(ranked, 1):
                r['rank'] = i

            stat_summary = {
                'total_students': len(report_data),
                'entry_count': entry_count,
                'pass_count': pass_count,
                'fail_count': fail_count,
                'absent_count': absent_count,
                'pending_count': len(report_data) - entry_count,
            }
        else:
            stat_summary = {}

        context = {
            'org': org,
            'exams': ExamTerm.objects.filter(org=org),
            'classifications': Classification.objects.filter(org=org),
            'sections': sections,
            'selected_exam': exam,
            'selected_classification': classification,
            # The dropdowns must remember a PARTIAL choice. `exam`/`classification`
            # above are only resolved once BOTH are supplied, so binding the
            # <option selected> to them made each dropdown reset the other:
            # pick an exam -> re-render with selected_exam=None -> the exam box
            # goes blank -> picking a class then submits exam='' and the report
            # could never load. Bind the options to the raw ids instead.
            'selected_exam_id': exam_id or '',
            'selected_classification_id': classification_id or '',
            'selected_section': section_id,
            'subjects': subjects,
            'report_data': report_data,
            'stat_summary': stat_summary,
        }
        return render(request, self.template_name, context)


class ResultPublishSummaryView(FeatureRequiredMixin, View):
    required_feature = 'results'
    """Pre-publish summary page for an exam."""

    def get(self, request, pk):
        org = _get_org(request)
        exam = get_object_or_404(ExamTerm, pk=pk, org=org)
        classifications = Classification.objects.filter(org=org)
        class_summaries = []
        pending_entry = 0
        for cls in classifications:
            students = list(member.objects.filter(org=org, classification=cls).exclude(status='dumped'))
            if not students:
                continue
            entry_count = ResultRecord.objects.filter(exam=exam, student__in=students).values('student').distinct().count()
            pass_count = fail_count = absent_count = 0
            for mem in students:
                recs = list(ResultRecord.objects.filter(exam=exam, student=mem))
                if not recs:
                    continue
                if any(r.is_absent for r in recs):
                    absent_count += 1
                elif all(r.is_passed for r in recs):
                    pass_count += 1
                else:
                    fail_count += 1
            pending_for_cls = len(students) - entry_count
            if pending_for_cls > 0:
                pending_entry += 1
            class_summaries.append({
                'classification': cls.name,
                'section': None,
                'total_students': len(students),
                'entry_count': entry_count,
                'pass_count': pass_count,
                'fail_count': fail_count,
                'absent_count': absent_count,
            })
        context = {
            'org': org,
            'exam': exam,
            'class_summaries': class_summaries,
            'pending_entry': pending_entry,
            'total_records': ResultRecord.objects.filter(exam=exam).count(),
        }
        return render(request, 'admin/results/publish_summary.html', context)

    def post(self, request, pk):
        org = _get_org(request)
        exam = get_object_or_404(ExamTerm, pk=pk, org=org)
        exam.is_published = not exam.is_published
        exam.status = 'published' if exam.is_published else 'marks_entry'
        exam.save(update_fields=['is_published', 'status'])
        action_word = 'published' if exam.is_published else 'unpublished'
        messages.success(request, f"'{exam.name}' has been {action_word}.")
        return redirect('schooladmin:result_publish_summary', pk=pk)


class MarksheetView(FeatureRequiredMixin, View):
    required_feature = 'results'
    """Printable marksheet for one student for one exam."""

    def get(self, request, exam_pk, member_pk):
        org = _get_org(request)
        exam = get_object_or_404(ExamTerm, pk=exam_pk, org=org)
        mem = get_object_or_404(member, pk=member_pk, org=org)
        subjects = _student_subjects(mem)
        results = {r.subject_id: r for r in ResultRecord.objects.filter(student=mem, exam=exam).select_related('subject')}
        rows = []
        total_obt = total_full = 0
        for subj in subjects:
            r = results.get(subj.id)
            rows.append({'subject': subj, 'record': r})
            total_full += float(subj.full_marks)
            if r:
                total_obt += float(r.obtained_marks)
        pct = round(total_obt / total_full * 100, 1) if total_full else 0
        passed = all(r['record'].is_passed for r in rows if r['record']) and bool(results)
        # Rank within classification
        all_members = member.objects.filter(org=org, classification=mem.classification).exclude(status='dumped')
        totals = []
        for m2 in all_members:
            t = float(sum(r.obtained_marks for r in ResultRecord.objects.filter(student=m2, exam=exam)))
            totals.append((m2.id, t))
        totals.sort(key=lambda x: x[1], reverse=True)
        rank = next((i+1 for i, (mid, _) in enumerate(totals) if mid == mem.id), '—')

        pass_count = sum(1 for r in rows if r['record'] and not r['record'].is_absent and r['record'].is_passed)
        fail_count = sum(1 for r in rows if r['record'] and not r['record'].is_absent and not r['record'].is_passed)
        absent_count = sum(1 for r in rows if r['record'] and r['record'].is_absent)
        context = {
            'org': org,
            'exam': exam,
            'member': mem,
            'result_rows': rows,
            'total_obtained': total_obt,
            'total_full': total_full,
            'percentage': pct,
            'overall_pass': passed,
            'overall_grade': _compute_grade(pct),
            'rank': rank,
            'pass_count': pass_count,
            'fail_count': fail_count,
            'absent_count': absent_count,
        }
        return render(request, 'admin/results/marksheet.html', context)


# =============================================================
# COMPLAINT SYSTEM
# =============================================================

class ComplaintListView(FeatureRequiredMixin, View):
    required_feature = 'complaints'
    template_name = 'admin/complaints/list.html'

    def get(self, request):
        org = _get_org(request)
        qs = Complaint.objects.filter(org=org).select_related('filed_by', 'branch', 'filed_by__classification').order_by('-created_at')
        status = request.GET.get('status')
        priority = request.GET.get('priority')
        branch_id = request.GET.get('branch')
        classification_id = request.GET.get('classification')
        if status:
            qs = qs.filter(status=status)
        if priority:
            qs = qs.filter(priority=priority)
        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        if classification_id:
            qs = qs.filter(filed_by__classification_id=classification_id)
        context = {
            'org': org,
            'complaints': qs,
            'status_choices': Complaint.STATUS_CHOICES,
            'priority_choices': Complaint.PRIORITY_CHOICES,
            'selected_status': status,
            'selected_priority': priority,
            'selected_branch': branch_id,
            'selected_classification': classification_id,
            'branches': Branch.objects.filter(org=org, status='active').order_by('name'),
            'classifications': Classification.objects.filter(org=org).order_by('name'),
            'pending_count': Complaint.objects.filter(org=org, status='pending').count(),
        }
        return render(request, self.template_name, context)


class ComplaintDetailView(FeatureRequiredMixin, View):
    required_feature = 'complaints'
    template_name = 'admin/complaints/detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        complaint = get_object_or_404(
            Complaint.objects.prefetch_related('messages'),
            pk=pk,
            org=org,
        )
        return render(request, self.template_name, {'org': org, 'complaint': complaint, 'status_choices': Complaint.STATUS_CHOICES})

    def post(self, request, pk):
        from handle.models import ComplaintMessage
        org = _get_org(request)
        complaint = get_object_or_404(Complaint, pk=pk, org=org)
        complaint.status = request.POST.get('status', complaint.status)
        complaint.admin_remarks = request.POST.get('admin_remarks', '')
        resolution_date = request.POST.get('resolution_date')
        if resolution_date:
            complaint.resolution_date = resolution_date
        complaint.save()
        reply_message = request.POST.get('reply_message', '').strip()
        if reply_message:
            ComplaintMessage.objects.create(
                complaint=complaint,
                author=request.user,
                message=reply_message,
                is_staff_reply=True,
            )
        messages.success(request, "Complaint updated.")
        return redirect('schooladmin:complaint_detail', pk=pk)


class FileComplaintView(FeatureRequiredMixin, View):
    required_feature = 'complaints'
    template_name = 'admin/complaints/file_complaint.html'

    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'members': member.objects.filter(org=org).exclude(status='dumped'),
            'branches': Branch.objects.filter(org=org, status='active'),
            'priority_choices': Complaint.PRIORITY_CHOICES,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        try:
            member_id = request.POST.get('filed_by')
            filed_by = get_object_or_404(member, pk=member_id, org=org)
            Complaint.objects.create(
                org=org,
                branch_id=request.POST.get('branch') or None,
                filed_by=filed_by,
                complaint_type=request.POST.get('complaint_type', ''),
                subject=request.POST.get('subject'),
                description=request.POST.get('description'),
                priority=request.POST.get('priority', 'medium'),
            )
            messages.success(request, "Complaint filed successfully.")
            return redirect('schooladmin:complaint_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:file_complaint')


# =============================================================
# HRMS EXTENDED - RESIGNATION
# =============================================================

class ResignationListView(FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/hrms/resignation_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = ResignationRecord.objects.filter(org=org).select_related('member').order_by('-created_at')
        status = request.GET.get('status')
        self_applied_filter = request.GET.get('self_applied')
        if status:
            qs = qs.filter(status=status)
        if self_applied_filter == '1':
            qs = qs.filter(self_applied=True)
        elif self_applied_filter == '0':
            qs = qs.filter(self_applied=False)
        all_qs = ResignationRecord.objects.filter(org=org)
        context = {
            'org': org,
            'resignations': qs,
            'status_choices': ResignationRecord.STATUS_CHOICES,
            'selected_status': status,
            'pending_count':       all_qs.filter(status='pending').count(),
            'approved_count':      all_qs.filter(status='approved').count(),
            'completed_count':     all_qs.filter(status='completed').count(),
            'self_applied_count':  all_qs.filter(self_applied=True).count(),
            'self_applied_pending': all_qs.filter(self_applied=True, status='pending').count(),
        }
        return render(request, self.template_name, context)


class AddResignationView(FeatureRequiredMixin, View):
    required_feature = 'hrms'
    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'members': member.objects.filter(org=org).exclude(status='dumped'),
            'today': datetime.date.today().strftime('%Y-%m-%d'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        }
        return render(request, 'admin/hrms/add_resignation.html', context)

    def post(self, request):
        org = _get_org(request)
        try:
            mem = get_object_or_404(member, pk=request.POST.get('member_id'), org=org)
            notice_days = int(request.POST.get('notice_period_days', 30))
            resignation_date = datetime.datetime.strptime(request.POST.get('resignation_date'), '%Y-%m-%d').date()
            last_working_day_str = request.POST.get('last_working_day')
            last_working_day = datetime.datetime.strptime(last_working_day_str, '%Y-%m-%d').date() if last_working_day_str else resignation_date + timedelta(days=notice_days)
            ResignationRecord.objects.create(
                org=org,
                member=mem,
                resignation_date=resignation_date,
                notice_period_days=notice_days,
                last_working_day=last_working_day,
                reason=request.POST.get('reason', ''),
            )
            messages.success(request, f"Resignation record created for {mem.name}.")
            return redirect('schooladmin:resignation_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:add_resignation')


@feature_required('hrms')
def update_resignation_status(request, pk):
    org = _get_org(request)
    res = get_object_or_404(ResignationRecord, pk=pk, org=org)
    new_status = request.POST.get('status')
    if new_status:
        res.status = new_status
        res.exit_interview_note = request.POST.get('exit_interview_note', res.exit_interview_note or '')
        res.clearance_status = request.POST.get('clearance_status') == 'on'
        res.final_settlement_status = request.POST.get('final_settlement_status') == 'on'
        res.save()
        messages.success(request, "Resignation status updated.")
    return redirect('schooladmin:resignation_list')


# =============================================================
# HRMS EXTENDED - STAFF DOCUMENTS
# =============================================================

class StaffDocumentListView(FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/hrms/document_list.html'

    def get(self, request):
        org = _get_org(request)
        qs = StaffDocument.objects.filter(org=org).select_related('member').order_by('-uploaded_at')
        member_id = request.GET.get('member')
        if member_id:
            qs = qs.filter(member_id=member_id)
        context = {
            'org': org,
            'documents': qs,
            'members': member.objects.filter(org=org).exclude(status='dumped'),
            'selected_member': member_id,
        }
        return render(request, self.template_name, context)


class UploadStaffDocumentView(FeatureRequiredMixin, View):
    required_feature = 'hrms'
    def get(self, request):
        org = _get_org(request)
        context = {
            'org': org,
            'members': member.objects.filter(org=org).exclude(status='dumped'),
            'doc_types': StaffDocument.DOC_TYPE_CHOICES,
        }
        return render(request, 'admin/hrms/upload_document.html', context)

    def post(self, request):
        org = _get_org(request)
        try:
            mem = get_object_or_404(member, pk=request.POST.get('member_id'), org=org)
            file = request.FILES.get('file')
            if not file:
                messages.error(request, "Please select a file.")
                return redirect('schooladmin:upload_document')
            expiry_str = request.POST.get('expiry_date')
            StaffDocument.objects.create(
                org=org,
                member=mem,
                document_type=request.POST.get('document_type', 'other'),
                title=request.POST.get('title'),
                file=file,
                expiry_date=expiry_str if expiry_str else None,
            )
            messages.success(request, "Document uploaded successfully.")
            return redirect('schooladmin:document_list')
        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect('schooladmin:upload_document')


@feature_required('hrms')
def delete_document(request, pk):
    org = _get_org(request)
    doc = get_object_or_404(StaffDocument, pk=pk, org=org)
    doc.file.delete(save=False)
    doc.delete()
    messages.success(request, "Document deleted.")
    return redirect('schooladmin:document_list')


# =============================================================
# BRANCH MANAGEMENT
# =============================================================

class BranchListView(FeatureRequiredMixin, View):
    required_feature = 'branches'
    template_name = 'admin/branches/branch_list.html'

    def get(self, request):
        org = _get_org(request)
        branches = Branch.objects.filter(org=org).annotate(
            member_count=Count('members', filter=Q(members__status='active'))
        ).order_by('name')
        context = {'org': org, 'branches': branches}
        return render(request, self.template_name, context)

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            code = request.POST.get('code', '').strip()
            if not name or not code:
                messages.error(request, "Name and code are required.")
            else:
                Branch.objects.get_or_create(org=org, code=code, defaults={
                    'name': name,
                    'address': request.POST.get('address', ''),
                    'phone': request.POST.get('phone', ''),
                    'email': request.POST.get('email', ''),
                })
                messages.success(request, f"Branch '{name}' added.")
        elif action == 'delete':
            Branch.objects.filter(pk=request.POST.get('branch_id'), org=org).delete()
            messages.success(request, "Branch deleted.")
        return redirect('schooladmin:branch_list')


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# NEW PHASE 2 VIEWS — Privilege, Billing, Absence Correction, Bulk Payslip
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from handle.models import AbsenceCorrection, Bill, BillItem, PRIVILEGE_LEVEL_CHOICES, EmailLog
from school.email_utils import (
    send_welcome_email, send_leave_status_email, send_bill_email,
    send_result_email, send_resignation_status_email,
    send_payslip_email, send_complaint_update_email,
)


# ── Absence Correction (Mark present-as-absent) ───────────────────────────────

class AbsenceCorrectionView(View):
    """Admin marks a member absent on a date where they have an incorrect present record."""
    template_name = 'admin/attendance/absence_correction.html'

    def get(self, request):
        org = _get_org(request)
        date_str = request.GET.get('date', timezone.localdate().strftime('%Y-%m-%d'))
        members_q = request.GET.get('q', '')
        members_qs = member.objects.filter(org=org).exclude(status='dumped')
        if members_q:
            members_qs = members_qs.filter(name__icontains=members_q)

        try:
            selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.localdate()

        present_ids = set(
            AttendanceRecord.objects.filter(org=org, scanned_time__date=selected_date)
            .values_list('mem_id', flat=True)
        )

        corrections = AbsenceCorrection.objects.filter(org=org).order_by('-corrected_at')[:30]

        return render(request, self.template_name, {
            'org': org,
            'selected_date': selected_date,
            'date_str': date_str,
            'members': members_qs,
            'present_ids': present_ids,
            'corrections': corrections,
            'q': members_q,
        })

    def post(self, request):
        org = _get_org(request)
        member_id = request.POST.get('member_id')
        date_str = request.POST.get('date')
        reason = request.POST.get('reason', '').strip()
        if not reason:
            messages.error(request, "Reason is required.")
            return redirect(f"{request.path}?date={date_str}")

        try:
            selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date.")
            return redirect('schooladmin:absence_correction')

        m = get_object_or_404(member, pk=member_id, org=org)
        records = AttendanceRecord.objects.filter(mem=m, org=org, scanned_time__date=selected_date)
        first_scan = records.first()
        original_time = first_scan.scanned_time if first_scan else None
        deleted = records.delete()[0]

        if deleted:
            AbsenceCorrection.objects.create(
                org=org, member=m, date=selected_date, reason=reason,
                corrected_by=request.user, original_scan_time=original_time,
            )
            messages.success(request, f"{m.name} marked absent on {selected_date}. Reason logged.")
        else:
            messages.warning(request, f"{m.name} had no attendance record on {selected_date}.")

        return redirect(f"{request.path}?date={date_str}")


# ── Bill Management ────────────────────────────────────────────────────────────

class BillListView(FeatureRequiredMixin, View):
    required_feature = 'billing'
    template_name = 'admin/billing/bill_list.html'

    def get(self, request):
        org = _get_org(request)
        bills = Bill.objects.filter(org=org).select_related('member').order_by('-issue_date')
        members_qs = member.objects.filter(org=org).exclude(status='dumped')
        status_filter = request.GET.get('status', '')
        member_filter = request.GET.get('member', '')
        if status_filter:
            bills = bills.filter(status=status_filter)
        if member_filter:
            bills = bills.filter(member_id=member_filter)
        total_billed = bills.aggregate(t=Sum('total_amount'))['t'] or 0
        total_paid = bills.aggregate(t=Sum('amount_paid'))['t'] or 0
        return render(request, self.template_name, {
            'org': org,
            'bills': bills,
            'members': members_qs,
            'status_choices': Bill.STATUS_CHOICES,
            'selected_status': status_filter,
            'selected_member': member_filter,
            'total_billed': total_billed,
            'total_paid': total_paid,
            'total_due': total_billed - total_paid,
        })


class CreateBillView(FeatureRequiredMixin, View):
    required_feature = 'billing'
    template_name = 'admin/billing/create_bill.html'

    def get(self, request):
        org = _get_org(request)
        return render(request, self.template_name, {
            'org': org,
            'members': member.objects.filter(org=org).exclude(status='dumped').order_by('name'),
            'nepali_enabled': getattr(org, 'nepali_date', False),
        })

    def post(self, request):
        org = _get_org(request)
        member_id = request.POST.get('member_id')
        due_date = request.POST.get('due_date')
        remarks = request.POST.get('remarks', '')
        send_email = request.POST.get('send_email') == 'on'
        descriptions = request.POST.getlist('description')
        amounts = request.POST.getlist('amount')

        if not member_id or not due_date or not descriptions:
            messages.error(request, "Member, due date and at least one item are required.")
            return redirect('schooladmin:create_bill')

        m = get_object_or_404(member, pk=member_id, org=org)
        import random, string
        invoice_no = 'INV-' + ''.join(random.choices(string.digits, k=8))

        items = []
        total = Decimal('0')
        for desc, amt in zip(descriptions, amounts):
            desc = desc.strip()
            if desc and amt:
                try:
                    amt_d = Decimal(str(amt))
                    items.append({'desc': desc, 'amount': amt_d})
                    total += amt_d
                except Exception:
                    pass

        if not items:
            messages.error(request, "Add at least one valid line item.")
            return redirect('schooladmin:create_bill')

        with transaction.atomic():
            bill = Bill.objects.create(
                org=org, member=m, invoice_number=invoice_no,
                due_date=due_date, total_amount=total, remarks=remarks,
            )
            for it in items:
                BillItem.objects.create(bill=bill, description=it['desc'], amount=it['amount'])

        if send_email and m.email:
            send_bill_email(
                email=m.email, name=m.name,
                invoice_number=invoice_no, total_amount=total,
                due_date=due_date, items=items, org_name=org.name, remarks=remarks,
            )

        messages.success(request, f"Bill {invoice_no} created for {m.name}.")
        return redirect('schooladmin:bill_list')


class BillDetailView(FeatureRequiredMixin, View):
    required_feature = 'billing'
    template_name = 'admin/billing/bill_detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        bill = get_object_or_404(Bill, pk=pk, org=org)
        return render(request, self.template_name, {'org': org, 'bill': bill})

    def post(self, request, pk):
        org = _get_org(request)
        bill = get_object_or_404(Bill, pk=pk, org=org)
        action = request.POST.get('action')
        if action == 'update_payment':
            amount_paid_raw = request.POST.get('amount_paid', '0')
            requested_status = request.POST.get('status', bill.status)

            try:
                new_amount = Decimal(amount_paid_raw)
            except Exception:
                messages.error(request, "Invalid amount.")
                return redirect('schooladmin:bill_detail', pk=pk)

            # Never trust a client-supplied absolute payment amount blindly:
            # it must be within [0, total_amount]. A negative value would
            # corrupt amount_paid; an amount above total_amount would let a
            # client fabricate an overpayment.
            if new_amount < 0 or new_amount > bill.total_amount:
                messages.error(
                    request,
                    f"Amount must be between 0 and the bill total (Rs. {bill.total_amount})."
                )
                return redirect('schooladmin:bill_detail', pk=pk)

            valid_statuses = dict(Bill.STATUS_CHOICES)
            if requested_status == 'Cancelled' and requested_status in valid_statuses:
                # Explicit admin cancellation is allowed regardless of amount.
                new_status = 'Cancelled'
            elif new_amount >= bill.total_amount:
                new_status = 'Paid'
            elif new_amount > 0:
                new_status = 'Partial'
            else:
                new_status = 'Unpaid'

            with transaction.atomic():
                delta = new_amount - bill.amount_paid
                bill.amount_paid = new_amount
                bill.status = new_status
                bill.save()
                # Auto-log payment delta as Finance income
                if delta > 0:
                    bill_cat, _ = TransactionCategory.objects.get_or_create(
                        org=org, name='Bill Collection', transaction_type='income'
                    )
                    FinancialTransaction.objects.create(
                        org=org,
                        transaction_type='income',
                        title=f"Bill Payment — {bill.invoice_number} ({bill.member.name})",
                        amount=delta,
                        category=bill_cat,
                        reference_number=bill.invoice_number,
                        note=f"Auto-linked from invoice #{bill.invoice_number}",
                        created_by=request.user,
                    )
            messages.success(request, "Payment updated.")
        elif action == 'resend_email':
            if bill.member.email:
                items = [{'desc': i.description, 'amount': i.amount} for i in bill.items.all()]
                send_bill_email(
                    email=bill.member.email, name=bill.member.name,
                    invoice_number=bill.invoice_number, total_amount=bill.total_amount,
                    due_date=bill.due_date, items=items, org_name=org.name, remarks=bill.remarks or '',
                    org=org, related_object_id=bill.id, force=True,
                )
                messages.success(request, f"Invoice emailed to {bill.member.email}.")
            else:
                messages.error(request, "Member has no email address on file.")
        return redirect('schooladmin:bill_detail', pk=pk)


@feature_required('billing')
def delete_bill(request, pk):
    org = _get_org(request)
    bill = get_object_or_404(Bill, pk=pk, org=org)
    bill.delete()
    messages.success(request, "Bill deleted.")
    return redirect('schooladmin:bill_list')


# ── Bulk Payslip ────────────────────────────────────────────────────────────────

class BulkPayslipView(FeatureRequiredMixin, PermRequiredMixin, View):
    """Three steps: (1) GET renders the member-selection form; (2) POST
    step=preview computes every selected member's payroll with the
    org/member defaults — no PaySlip rows written yet — and renders an
    editable table (item 30/31 of the upgrade spec); (3) POST step=generate
    re-runs the *same* trusted calculation (schooladmin.payroll_service.
    calculate_payroll_components) with whatever overrides were on the
    preview form and actually creates the PaySlip rows. The client never
    gets to hand the server a final gross/net figure — only the percentage/
    amount overrides that function already accepts as first-class
    parameters, so "editable bulk payroll" never means "trust the browser's
    arithmetic" for money."""

    required_feature = 'payroll'
    required_perm = 'can_generate_payroll'
    template_name = 'admin/payroll/bulk_payslip.html'
    preview_template_name = 'admin/payroll/bulk_payslip_preview.html'

    def get(self, request):
        org = _get_org(request)
        members_qs = member.objects.filter(org=org, salary_amount__gt=0).exclude(status='dumped').select_related('branch').order_by('name')
        classifications = Classification.objects.filter(org=org)
        branches = Branch.objects.filter(org=org, status='active').order_by('name')
        return render(request, self.template_name, {
            'org': org,
            'members': members_qs,
            'classifications': classifications,
            'branches': branches,
        })

    def post(self, request):
        if request.POST.get('step') == 'generate':
            return self._generate(request)
        return self._preview(request)

    # ── shared helpers ──────────────────────────────────────────────────

    def _parse_period(self, request):
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')
        month_name = (request.POST.get('month_name') or '').strip()
        if not from_date_str or not to_date_str or not month_name:
            return None, None, None
        try:
            from_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return None, None, None
        return from_date, to_date, month_name

    def _row_overrides(self, request, member_id):
        """Every override is optional — an untouched field reproduces
        exactly what plain `calculate_payroll_components(...)` (no kwargs)
        would already compute, so previewing without editing anything is a
        no-op relative to the pre-existing bulk-generate behavior."""
        def _amount(field):
            raw = (request.POST.get(f'{field}_{member_id}') or '').strip()
            if not raw:
                return Decimal('0')
            try:
                return Decimal(raw)
            except InvalidOperation:
                return Decimal('0')

        def _pct(field):
            raw = (request.POST.get(f'{field}_{member_id}') or '').strip()
            if raw == '':
                return None
            try:
                return Decimal(raw)
            except InvalidOperation:
                return None

        return {
            'extra_bonus': _amount('bonus'),
            'extra_allowance': _amount('allowance'),
            'extra_deduction': _amount('deduction'),
            'pf_employee_pct_override': _pct('pf'),
            'ssf_employee_pct_override': _pct('ssf'),
            'tds_pct_override': _pct('tds'),
        }

    def _totals(self, rows):
        keys = (
            'gross_salary', 'overtime_amount', 'allowance_total', 'bonus_total',
            'pf_employee', 'pf_employer', 'ssf_employee', 'ssf_employer',
            'tax_amount', 'other_deduction', 'net_payable',
        )
        totals = {k: Decimal('0.00') for k in keys}
        for row in rows:
            for k in keys:
                totals[k] += row['comps'][k]
        return totals

    # ── step 2: preview / recalculate ───────────────────────────────────

    def _preview(self, request):
        org = _get_org(request)
        member_ids = request.POST.getlist('member_ids')
        from_date, to_date, month_name = self._parse_period(request)
        send_emails = request.POST.get('send_email') == 'on'

        if not member_ids or not from_date or not to_date or not month_name:
            messages.error(request, "Please fill all required fields and select at least one member.")
            return redirect('schooladmin:bulk_payslip')

        policy = get_or_create_policy(org)
        rows = []
        for mid in member_ids:
            m = member.objects.filter(pk=mid, org=org).select_related('branch', 'classification').first()
            if not m:
                continue
            stats, _ = calculate_attendance_stats(m, from_date, to_date, org)
            overrides = self._row_overrides(request, mid)
            comps = calculate_payroll_components(m, stats, org, policy, to_date, **overrides)
            rows.append({
                'member': m,
                'stats': stats,
                'comps': comps,
                'overrides': overrides,
                'already_exists': PaySlip.objects.filter(member=m, org=org, from_date=from_date, to_date=to_date).exists(),
            })

        return render(request, self.preview_template_name, {
            'org': org, 'rows': rows, 'totals': self._totals(rows),
            'from_date': from_date, 'to_date': to_date, 'month_name': month_name,
            'send_emails': send_emails, 'member_ids': member_ids,
        })

    # ── step 3: generate ─────────────────────────────────────────────────

    def _generate(self, request):
        org = _get_org(request)
        member_ids = request.POST.getlist('member_ids')
        from_date, to_date, month_name = self._parse_period(request)
        send_emails = request.POST.get('send_email') == 'on'

        if not member_ids or not from_date or not to_date or not month_name:
            messages.error(request, "Please fill all required fields and select at least one member.")
            return redirect('schooladmin:bulk_payslip')

        policy = get_or_create_policy(org)
        generated = 0
        skipped = 0
        errors = 0

        for mid in member_ids:
            try:
                m = member.objects.get(pk=mid, org=org)
            except member.DoesNotExist:
                continue

            if PaySlip.objects.filter(member=m, org=org, from_date=from_date, to_date=to_date).exists():
                skipped += 1
                continue

            try:
                stats, _ = calculate_attendance_stats(m, from_date, to_date, org)
                overrides = self._row_overrides(request, mid)
                comps = calculate_payroll_components(m, stats, org, policy, to_date, **overrides)

                with transaction.atomic():
                    slip = PaySlip.objects.create(
                        member=m, org=org,
                        from_date=from_date, to_date=to_date, month_name=month_name,
                        total_days=stats['total_days'],
                        present_days=stats['days_present'],
                        paid_leaves=stats['days_paid_leave'],
                        holidays=stats['days_holiday'],
                        unpaid_absences=stats['days_unpaid_absent'],
                        salary_type=m.salary_type,
                        gross_salary=comps['gross_salary'],
                        allowance_total=comps['allowance_total'],
                        bonus_total=comps['bonus_total'],
                        advance_deduction=comps['advance_deduction'],
                        loan_deduction=comps['loan_deduction'],
                        other_deduction=comps['other_deduction'],
                        tax_deduction=comps['tax_amount'],
                        pf_employee=comps['pf_employee'],
                        pf_employer=comps['pf_employer'],
                        ssf_employee=comps['ssf_employee'],
                        ssf_employer=comps['ssf_employer'],
                        probation_adjustment=comps['probation_adjustment'],
                        overtime_hours=comps['overtime_hours'],
                        overtime_amount=comps['overtime_amount'],
                        net_payable=comps['net_payable'],
                        status='draft',
                    )

                    # PF / SSF records
                    if slip.pf_employee or slip.pf_employer:
                        ProvidentFundRecord.objects.create(
                            org=org, member=m, payslip=slip, month_name=month_name or '',
                            employee_contribution=slip.pf_employee,
                            employer_contribution=slip.pf_employer,
                        )
                    if slip.ssf_employee or slip.ssf_employer:
                        SocialSecurityFundRecord.objects.create(
                            org=org, member=m, payslip=slip, month_name=month_name or '',
                            employee_contribution=slip.ssf_employee,
                            employer_contribution=slip.ssf_employer,
                        )

                    # Auto-log salary as finance expense
                    sal_cat, _ = TransactionCategory.objects.get_or_create(
                        org=org, name='Salary Payment', transaction_type='expense'
                    )
                    FinancialTransaction.objects.create(
                        org=org, transaction_type='expense',
                        title=f"Salary — {m.name} ({month_name})",
                        amount=slip.net_payable,
                        category=sal_cat,
                        reference_number=str(slip.id),
                        note=f"Auto-linked from bulk payslip #{slip.id}",
                        created_by=request.user,
                    )

                generated += 1

                if send_emails and m.email:
                    send_payslip_email(
                        email=m.email, name=m.name,
                        month_name=month_name, net_payable=comps['net_payable'],
                        org_name=org.name,
                        details={
                            'Gross Salary': f"Rs. {comps['gross_salary']}",
                            'Allowances': f"Rs. {comps['allowance_total']}",
                            'Deductions': f"Rs. {comps['advance_deduction'] + comps['loan_deduction']}",
                            'PF': f"Rs. {comps['pf_employee']}",
                            'Tax': f"Rs. {comps['tax_amount']}",
                            'Present Days': stats['days_present'],
                            'Total Days': stats['total_days'],
                        },
                        org=org, related_object_id=slip.id,
                    )

            except Exception as exc:
                errors += 1
                import traceback; traceback.print_exc()

        msg = f"Generated {generated} payslip(s). Skipped {skipped} (already exist)."
        if errors:
            msg += f" {errors} member(s) had errors — check terminal log."
        messages.success(request, msg)
        return redirect('schooladmin:bulk_payslip')


# ── Enhanced Leave Status (with email) ────────────────────────────────────────

@perm_required('can_approve_leave')
def leave_status_with_email(request, id, status):
    org = _get_org(request)
    report, norm_status = _handle_leave_status_change(request, id, status, org)

    messages.success(request, f"Leave {norm_status}.")
    return redirect('schooladmin:leaveReportView')


# ── Complaint Detail (with email on status change) ────────────────────────────

class ComplaintDetailViewWithEmail(FeatureRequiredMixin, View):
    required_feature = 'complaints'
    """Override the existing ComplaintDetailView to send emails on status change."""
    template_name = 'admin/complaints/detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        complaint = get_object_or_404(Complaint, pk=pk, org=org)
        return render(request, self.template_name, {'org': org, 'complaint': complaint})

    def post(self, request, pk):
        org = _get_org(request)
        complaint = get_object_or_404(Complaint, pk=pk, org=org)
        old_status = complaint.status
        complaint.status = request.POST.get('status', complaint.status)
        complaint.admin_remarks = request.POST.get('admin_remarks', '')
        if complaint.status == 'resolved' and not complaint.resolution_date:
            from django.utils.timezone import now
            complaint.resolution_date = now().date()
        complaint.save()

        if complaint.status != old_status and complaint.filed_by.email:
            send_complaint_update_email(
                email=complaint.filed_by.email,
                name=complaint.filed_by.name,
                subject_text=complaint.subject,
                status=complaint.status,
                remarks=complaint.admin_remarks,
                org_name=org.name,
            )

        messages.success(request, "Complaint updated.")
        return redirect('schooladmin:complaint_detail', pk=pk)


# ── Resignation Update (with email) ───────────────────────────────────────────

@feature_required('hrms')
def update_resignation_with_email(request, pk):
    org = _get_org(request)
    rec = get_object_or_404(ResignationRecord, pk=pk, org=org)
    old_status = rec.status
    if request.method == 'POST':
        rec.status = request.POST.get('status', rec.status)
        rec.exit_interview_note = request.POST.get('exit_interview_note', '')
        rec.clearance_status = request.POST.get('clearance_status') == 'on'
        rec.final_settlement_status = request.POST.get('final_settlement_status') == 'on'
        rec.save()

        if rec.status != old_status and rec.member.email:
            send_resignation_status_email(
                email=rec.member.email,
                name=rec.member.name,
                status=rec.status,
                last_working_day=str(rec.last_working_day) if rec.last_working_day else '',
                org_name=org.name,
            )
        messages.success(request, "Resignation record updated.")
    return redirect('schooladmin:resignation_list')


# ── Result Publish (with email) ───────────────────────────────────────────────

@feature_required('results')
def publish_exam_with_email(request, pk):
    org = _get_org(request)
    exam = get_object_or_404(ExamTerm, pk=pk, org=org)
    exam.is_published = not exam.is_published
    exam.save()

    if exam.is_published:
        results = ResultRecord.objects.filter(exam=exam).select_related('student', 'subject')
        student_map = {}
        for r in results:
            sid = r.student_id
            if sid not in student_map:
                student_map[sid] = {'member': r.student, 'results': []}
            student_map[sid]['results'].append({
                'subject': r.subject.name,
                'marks': r.obtained_marks,
                'full': r.subject.full_marks,
                'passed': r.is_passed,
            })
        notified = 0
        for sid, data in student_map.items():
            m = data['member']
            if m.email:
                send_result_email(
                    email=m.email, name=m.name,
                    exam_name=exam.name, results=data['results'], org_name=org.name,
                    org=org, related_object_id=exam.id,
                )
                notified += 1
            guardian_email = getattr(m, 'guardian_email', '')
            if guardian_email and guardian_email != m.email:
                send_result_email(
                    email=guardian_email, name=getattr(m, 'guardian_name', '') or m.name,
                    exam_name=exam.name, results=data['results'], org_name=org.name,
                    org=org, related_object_id=exam.id,
                )
        messages.success(request, f"Exam published and emails sent to {notified} student(s) (+ guardians where available).")
    else:
        messages.success(request, "Exam unpublished.")

    return redirect('schooladmin:exam_terms')


# ── Email Log (delivery status + admin resend) ────────────────────────────────

class EmailLogListView(View):
    """Shows every outgoing email for the org (sent/failed/pending), filterable
    by status and type. Admin can resend a failed email — it's re-derived from
    the related object's current state (see school.email_utils.resend_email_log)
    rather than replayed from a stale snapshot."""
    template_name = 'admin/email_logs.html'

    def get(self, request, *args, **kwargs):
        org = _get_org(request)
        status_filter = request.GET.get('status', 'failed')
        type_filter = request.GET.get('type', '')

        logs = EmailLog.objects.filter(org=org).order_by('-created_at')
        if status_filter:
            logs = logs.filter(status=status_filter)
        if type_filter:
            logs = logs.filter(email_type=type_filter)

        from django.core.paginator import Paginator
        paginator = Paginator(logs, 50)
        page_obj = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'org': org,
            'logs': page_obj,
            'page_obj': page_obj,
            'failed_count': EmailLog.objects.filter(org=org, status='failed').count(),
            'sent_count': EmailLog.objects.filter(org=org, status='sent').count(),
            'status_choices': EmailLog.STATUS_CHOICES,
            'type_choices': EmailLog.EMAIL_TYPE_CHOICES,
            'selected_status': status_filter,
            'selected_type': type_filter,
        })

    def post(self, request, *args, **kwargs):
        org = _get_org(request)
        from school.email_utils import resend_email_log
        log = get_object_or_404(EmailLog, pk=request.POST.get('log_id'), org=org)
        ok, msg = resend_email_log(log)
        if ok:
            messages.success(request, f"{log.recipient_email}: {msg}")
        else:
            messages.error(request, f"{log.recipient_email}: {msg}")
        redirect_qs = request.POST.get('redirect_qs', '')
        base = reverse('schooladmin:email_logs')
        return redirect(f"{base}?{redirect_qs}" if redirect_qs else base)


# ── Super Admin: Cross-Org Attendance Report & Date-Range Delete ──────────────

class SuperAttendanceReportView(View):
    """Super admin can view and delete attendance records by org + date range."""
    template_name = 'super_admin/attendance_report.html'

    def get(self, request):
        from management.models import Organization
        orgs = Organization.objects.all().order_by('name')
        selected_org = request.GET.get('org')
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')
        records = AttendanceRecord.objects.none()
        org_obj = None

        if selected_org:
            try:
                org_obj = Organization.objects.get(pk=selected_org)
                records = AttendanceRecord.objects.filter(org=org_obj).select_related('mem').order_by('-scanned_time')
                if from_date:
                    records = records.filter(scanned_time__date__gte=from_date)
                if to_date:
                    records = records.filter(scanned_time__date__lte=to_date)
            except Organization.DoesNotExist:
                pass

        return render(request, self.template_name, {
            'orgs': orgs,
            'records': records[:500],
            'total_records': records.count() if org_obj else 0,
            'selected_org': selected_org,
            'org_obj': org_obj,
            'from_date': from_date,
            'to_date': to_date,
        })

    def post(self, request):
        selected_org = request.POST.get('org')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        confirm = request.POST.get('confirm') == 'yes'

        if not confirm:
            messages.error(request, "Please check the confirmation box before deleting.")
            return redirect(f"/superadmin/attendance-report/?org={selected_org}&from_date={from_date}&to_date={to_date}")

        try:
            org_obj = Organization.objects.get(pk=selected_org)
            qs = AttendanceRecord.objects.filter(org=org_obj)
            if from_date:
                qs = qs.filter(scanned_time__date__gte=from_date)
            if to_date:
                qs = qs.filter(scanned_time__date__lte=to_date)
            count = qs.count()
            qs.delete()
            messages.success(request, f"Deleted {count} attendance records from {org_obj.name}.")
        except Exception as e:
            messages.error(request, f"Error: {e}")

        return redirect('superadmin:attendance_report')



# ============================================================
# TASK MANAGEMENT — Admin Views
# ============================================================

from handle.models import Task, TaskInstance, TaskUpdateLog, TaskAttachment
from school.email_utils import (
    send_task_assigned_email, send_task_completed_email,
    send_task_overdue_email, send_task_approval_email,
)


def _task_org(request):
    if request.user.user_type == "2":
        return request.user.schooladmin.org
    if request.user.user_type == "3":
        return request.user.staff.org
    return None


class TaskDashboardView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'tasks'
    template_name = 'admin/tasks/dashboard.html'

    def get(self, request):
        org = _task_org(request)
        today = datetime.date.today()

        # Auto-refresh overdue statuses
        stale = TaskInstance.objects.filter(
            task__org=org, due_date__lt=today, status__in=['pending', 'in_progress']
        )
        for inst in stale:
            inst.refresh_overdue_status()

        all_inst = TaskInstance.objects.filter(task__org=org)
        total = all_inst.count()
        pending = all_inst.filter(status='pending').count()
        in_progress = all_inst.filter(status='in_progress').count()
        completed = all_inst.filter(status='completed').count()
        not_completed = all_inst.filter(status='not_completed').count()
        overdue = all_inst.filter(status='overdue').count()
        missed = all_inst.filter(status='missed_absence').count()
        rework = all_inst.filter(status='rework_required').count()
        pending_approval = all_inst.filter(approval_status='pending_approval').count()

        today_tasks = all_inst.filter(due_date=today).select_related('task', 'assigned_member')
        urgent_tasks = all_inst.filter(
            task__priority='urgent', status__in=['pending', 'in_progress', 'overdue']
        ).select_related('task', 'assigned_member')[:10]

        # Staff-wise progress
        from django.db.models import Count, Case, When, IntegerField
        staff_stats = (
            all_inst
            .values('assigned_member__name', 'assigned_member__id')
            .annotate(
                total=Count('id'),
                done=Count(Case(When(status='completed', then=1), output_field=IntegerField())),
                pending_c=Count(Case(When(status='pending', then=1), output_field=IntegerField())),
                overdue_c=Count(Case(When(status__in=['overdue', 'missed_absence'], then=1), output_field=IntegerField())),
            )
            .order_by('-total')[:10]
        )

        # Priority breakdown
        priority_stats = (
            TaskInstance.objects.filter(task__org=org)
            .values('task__priority')
            .annotate(count=Count('id'))
        )

        ctx = dict(
            org=org, today=today,
            total=total, pending=pending, in_progress=in_progress,
            completed=completed, not_completed=not_completed,
            overdue=overdue, missed=missed, rework=rework,
            pending_approval=pending_approval,
            today_tasks=today_tasks,
            urgent_tasks=urgent_tasks,
            staff_stats=staff_stats,
            priority_stats=priority_stats,
        )
        return render(request, self.template_name, ctx)


class TaskListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'tasks'
    template_name = 'admin/tasks/task_list.html'

    def get(self, request):
        org = _task_org(request)
        today = datetime.date.today()
        tasks = Task.objects.filter(org=org, is_active=True).prefetch_related('assigned_to', 'instances')

        # Filters
        status_f = request.GET.get('status', '')
        priority_f = request.GET.get('priority', '')
        member_id = request.GET.get('member', '')
        task_type_f = request.GET.get('task_type', '')
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')

        inst_qs = TaskInstance.objects.filter(task__org=org)
        if status_f:
            inst_qs = inst_qs.filter(status=status_f)
        if priority_f:
            inst_qs = inst_qs.filter(task__priority=priority_f)
        if member_id:
            inst_qs = inst_qs.filter(assigned_member_id=member_id)
        if task_type_f:
            inst_qs = inst_qs.filter(task__task_type=task_type_f)
        if from_date:
            inst_qs = inst_qs.filter(due_date__gte=from_date)
        if to_date:
            inst_qs = inst_qs.filter(due_date__lte=to_date)

        inst_qs = inst_qs.select_related('task', 'assigned_member').order_by('-due_date')[:200]

        members = member.objects.filter(org=org, status='active').order_by('name')
        ctx = dict(
            org=org, today=today,
            instances=inst_qs, members=members,
            status_choices=TaskInstance.STATUS_CHOICES,
            priority_choices=Task.PRIORITY_CHOICES,
            type_choices=Task.TASK_TYPE_CHOICES,
            filters=dict(status=status_f, priority=priority_f, member=member_id,
                         task_type=task_type_f, from_date=from_date, to_date=to_date),
            nepali_enabled=getattr(org, 'nepali_date', False),
        )
        return render(request, self.template_name, ctx)


class CreateTaskView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'tasks'
    template_name = 'admin/tasks/create_task.html'

    def get(self, request):
        org = _task_org(request)
        members = member.objects.filter(org=org, status='active').order_by('name')
        branches = Branch.objects.filter(org=org)
        ctx = dict(
            org=org, members=members, branches=branches,
            priority_choices=Task.PRIORITY_CHOICES,
            type_choices=Task.TASK_TYPE_CHOICES,
            today=datetime.date.today(),
            nepali_enabled=getattr(org, 'nepali_date', False),
        )
        return render(request, self.template_name, ctx)

    def post(self, request):
        org = _task_org(request)
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '')
        priority = request.POST.get('priority', 'medium')
        task_type = request.POST.get('task_type', 'one_time')
        start_date = request.POST.get('start_date')
        due_date = request.POST.get('due_date')
        due_time = request.POST.get('due_time') or None
        branch_id = request.POST.get('branch') or None
        notes = request.POST.get('notes', '')
        requires_approval = request.POST.get('requires_approval') == 'on'
        member_ids = request.POST.getlist('assigned_to')

        if not title or not start_date or not due_date or not member_ids:
            messages.error(request, "Title, dates, and at least one assignee are required.")
            return redirect('schooladmin:create_task')

        import datetime as _dt
        try:
            sd = _dt.date.fromisoformat(start_date)
            dd = _dt.date.fromisoformat(due_date)
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect('schooladmin:create_task')

        if dd < sd:
            messages.error(request, "Due date must be on or after start date.")
            return redirect('schooladmin:create_task')

        task = Task.objects.create(
            org=org,
            branch_id=branch_id,
            title=title,
            description=description,
            priority=priority,
            task_type=task_type,
            start_date=sd,
            due_date=dd,
            due_time=due_time,
            notes=notes,
            requires_approval=requires_approval,
            created_by=request.user,
        )

        if 'attachment' in request.FILES:
            task.attachment = request.FILES['attachment']
            task.save(update_fields=['attachment'])

        assigned_members = member.objects.filter(id__in=member_ids, org=org).exclude(status='dumped')
        task.assigned_to.set(assigned_members)
        task.generate_instances()

        # Email notifications
        assigned_by = request.user.get_full_name() or request.user.username
        for m in assigned_members:
            if m.email:
                send_task_assigned_email(
                    m.email, m.name, task.title,
                    str(dd), priority, org.name, assigned_by
                )
        from handle.notifications import notify_task_assigned
        notify_task_assigned(task, assigned_members, actor=request.user)

        messages.success(request, f"Task '{title}' created and assigned to {assigned_members.count()} member(s).")
        return redirect('schooladmin:task_list')


class TaskDetailAdminView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'tasks'
    template_name = 'admin/tasks/task_detail.html'

    def get(self, request, pk):
        org = _task_org(request)
        task = get_object_or_404(Task, pk=pk, org=org)
        instances = task.instances.select_related('assigned_member', 'approved_by').prefetch_related('update_logs', 'attachments').order_by('due_date', 'assigned_member__name')
        members = member.objects.filter(org=org, status='active').order_by('name')
        ctx = dict(org=org, task=task, instances=instances, members=members)
        return render(request, self.template_name, ctx)

    def post(self, request, pk):
        org = _task_org(request)
        task = get_object_or_404(Task, pk=pk, org=org)
        action = request.POST.get('action')

        if action == 'approve_instance':
            inst_id = request.POST.get('instance_id')
            inst = get_object_or_404(TaskInstance, pk=inst_id, task=task)
            inst.approval_status = 'approved'
            inst.status = 'completed'
            inst.approved_by = request.user
            inst.approved_at = timezone.now()
            inst.save()
            log = TaskUpdateLog.objects.create(instance=inst, changed_by=request.user, old_status='completed', new_status='completed', note='Admin approved completion.')
            if inst.assigned_member.email:
                send_task_approval_email(inst.assigned_member.email, inst.assigned_member.name, task.title, 'approved', '', org.name)
            from handle.notifications import notify_task_assignee
            notify_task_assignee(
                inst, 'task_approved', f'Task approved: {task.title}',
                'Your task completion was approved.',
                actor=request.user, log_id=log.pk, priority='normal',
            )
            messages.success(request, "Task completion approved.")

        elif action == 'reject_instance':
            inst_id = request.POST.get('instance_id')
            reason = request.POST.get('rejection_reason', '')
            inst = get_object_or_404(TaskInstance, pk=inst_id, task=task)
            old_status = inst.status
            inst.approval_status = 'rejected'
            inst.status = 'rework_required'
            inst.rejection_reason = reason
            inst.approved_by = request.user
            inst.approved_at = timezone.now()
            inst.save()
            log = TaskUpdateLog.objects.create(instance=inst, changed_by=request.user, old_status=old_status, new_status='rework_required', note=f'Rejected: {reason}')
            if inst.assigned_member.email:
                send_task_approval_email(inst.assigned_member.email, inst.assigned_member.name, task.title, 'rejected', reason, org.name)
            from handle.notifications import notify_task_assignee
            notify_task_assignee(
                inst, 'task_rejected', f'Rework required: {task.title}',
                reason or 'Your completion was returned for correction.',
                actor=request.user, log_id=log.pk, priority='urgent',
            )
            messages.success(request, "Task rejected and returned to staff.")

        elif action == 'reassign':
            inst_id = request.POST.get('instance_id')
            new_member_id = request.POST.get('new_member')
            inst = get_object_or_404(TaskInstance, pk=inst_id, task=task)
            new_m = get_object_or_404(member, pk=new_member_id, org=org)
            old_member = inst.assigned_member
            old_status = inst.status
            inst.assigned_member = new_m
            inst.status = 'pending'
            inst.completion_note = ''
            inst.save()
            task.assigned_to.add(new_m)
            log = TaskUpdateLog.objects.create(
                instance=inst, changed_by=request.user,
                old_status=old_status, new_status='pending',
                note=f'Reassigned from {old_member.name} to {new_m.name}.',
            )
            if new_m.email:
                send_task_assigned_email(new_m.email, new_m.name, task.title, str(inst.due_date), task.priority, org.name, 'Admin (Reassigned)')
            from handle.notifications import notify, notify_task_assignee
            if old_member.pk != new_m.pk:
                notify(
                    old_member, 'task_reassigned', f'Task reassigned: {task.title}',
                    f'This task was reassigned to {new_m.name}.',
                    reverse('staff:my_tasks'), priority='normal',
                    actor=request.user,
                    dedupe_key=f'task-reassigned-old:{inst.pk}:{log.pk}:{old_member.pk}',
                )
            notify_task_assignee(
                inst, 'task_reassigned', f'Task reassigned to you: {task.title}',
                f'Due {inst.due_date}.',
                actor=request.user, log_id=log.pk,
            )
            messages.success(request, f"Task reassigned to {new_m.name}.")

        elif action == 'cancel_instance':
            inst_id = request.POST.get('instance_id')
            inst = get_object_or_404(TaskInstance, pk=inst_id, task=task)
            old = inst.status
            inst.status = 'cancelled'
            inst.save()
            log = TaskUpdateLog.objects.create(instance=inst, changed_by=request.user, old_status=old, new_status='cancelled', note='Cancelled by admin.')
            from handle.notifications import notify_task_assignee
            notify_task_assignee(
                inst, 'task_cancelled', f'Task cancelled: {task.title}',
                f'The task due {inst.due_date} was cancelled.',
                actor=request.user, log_id=log.pk, priority='normal',
            )
            messages.success(request, "Task instance cancelled.")

        elif action == 'deactivate':
            task.is_active = False
            task.save(update_fields=['is_active'])
            from handle.notifications import notify
            for assigned_member in task.assigned_to.all():
                notify(
                    assigned_member, 'task_cancelled',
                    f'Task deactivated: {task.title}',
                    'This task is no longer active.',
                    reverse('staff:my_tasks'), priority='normal',
                    actor=request.user,
                    dedupe_key=f'task-deactivated:{task.pk}:member:{assigned_member.pk}',
                )
            messages.success(request, "Task deactivated.")
            return redirect('schooladmin:task_list')

        return redirect('schooladmin:task_detail', pk=pk)


class TaskReportView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'tasks'
    template_name = 'admin/tasks/task_report.html'

    def get(self, request):
        org = _task_org(request)
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')
        member_id = request.GET.get('member', '')
        status_f = request.GET.get('status', '')
        priority_f = request.GET.get('priority', '')
        task_type_f = request.GET.get('task_type', '')

        qs = TaskInstance.objects.filter(task__org=org).select_related('task', 'assigned_member', 'approved_by')

        if from_date:
            qs = qs.filter(due_date__gte=from_date)
        if to_date:
            qs = qs.filter(due_date__lte=to_date)
        if member_id:
            qs = qs.filter(assigned_member_id=member_id)
        if status_f:
            qs = qs.filter(status=status_f)
        if priority_f:
            qs = qs.filter(task__priority=priority_f)
        if task_type_f:
            qs = qs.filter(task__task_type=task_type_f)

        qs = qs.order_by('-due_date')[:500]

        members = member.objects.filter(org=org, status='active').order_by('name')
        ctx = dict(
            org=org, instances=qs, members=members,
            status_choices=TaskInstance.STATUS_CHOICES,
            priority_choices=Task.PRIORITY_CHOICES,
            type_choices=Task.TASK_TYPE_CHOICES,
            filters=dict(from_date=from_date, to_date=to_date, member=member_id,
                         status=status_f, priority=priority_f, task_type=task_type_f),
            nepali_enabled=getattr(org, 'nepali_date', False),
        )
        return render(request, self.template_name, ctx)

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _excel_response(filename):
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

def _style_header(ws, headers, fill_color='2563EB'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 20

def _np_str(d):
    """Convert a date/datetime to Nepali BS string if possible."""
    try:
        if hasattr(d, 'date'):
            d = d.date()
        return str(nepali_datetime.date.from_datetime_date(d))
    except Exception:
        return str(d) if d else ''

def _org_from_req(request):
    if hasattr(request.user, 'schooladmin'):
        return request.user.schooladmin.org
    if hasattr(request.user, 'superadmin'):
        # superadmin can pass org_id param
        org_id = request.GET.get('org_id')
        if org_id:
            from management.models import Organization
            return Organization.objects.filter(pk=org_id).first()
    return None


# ── 1. ATTENDANCE EXPORT ──────────────────────────────────────────────────────

@login_required
@perm_required('can_export_reports')
def export_attendance(request):
    org = _org_from_req(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    nepali_enabled = getattr(org, 'nepali_date', False)

    qs = AttendanceRecord.objects.filter(org=org).select_related('mem', 'mem__classification').order_by('scanned_time')
    if from_date:
        try:
            qs = qs.filter(scanned_time__date__gte=datetime.date.fromisoformat(from_date))
        except Exception:
            pass
    if to_date:
        try:
            qs = qs.filter(scanned_time__date__lte=datetime.date.fromisoformat(to_date))
        except Exception:
            pass

    headers = ['Member Name', 'Member Type', 'Class/Branch', 'Date (AD)', 'Date (BS)' if nepali_enabled else '', 'Check-in Time', 'Status']
    headers = [h for h in headers if h]

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for r in qs:
            row = [
                r.mem.name if r.mem else '',
                r.mem.get_member_type_display() if r.mem else '',
                r.mem.classification.name if r.mem and r.mem.classification else '',
                r.scanned_time.date(),
            ]
            if nepali_enabled:
                row.append(_np_str(r.scanned_time.date()))
            row += [r.scanned_time.strftime('%H:%M:%S'), 'Present']
            writer.writerow(row)
        return resp

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    _style_header(ws, headers)
    for ridx, r in enumerate(qs, 2):
        row = [
            r.mem.name if r.mem else '',
            r.mem.get_member_type_display() if r.mem else '',
            r.mem.classification.name if r.mem and r.mem.classification else '',
            str(r.scanned_time.date()),
        ]
        if nepali_enabled:
            row.append(_np_str(r.scanned_time.date()))
        row += [r.scanned_time.strftime('%H:%M:%S'), 'Present']
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    resp = _excel_response('attendance_report.xlsx')
    wb.save(resp)
    return resp


# ── 2. PAYSLIP EXPORT ────────────────────────────────────────────────────────

@login_required
@feature_required('bulk_export')
@perm_required('can_bulk_export')
def export_payslips(request):
    org = _org_from_req(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    nepali_enabled = getattr(org, 'nepali_date', False)

    qs = PaySlip.objects.filter(org=org).select_related('member').order_by('-from_date')
    if from_date:
        try:
            qs = qs.filter(from_date__gte=from_date)
        except Exception:
            pass
    if to_date:
        try:
            qs = qs.filter(to_date__lte=to_date)
        except Exception:
            pass

    headers = ['Member', 'Period', 'Month', 'From (AD)', 'To (AD)']
    if nepali_enabled:
        headers += ['From (BS)', 'To (BS)']
    headers += ['Present Days', 'Gross Salary', 'Allowances', 'Deductions', 'Net Payable', 'Generated On']

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="payslips_report.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for ps in qs:
            total_ded = float(ps.advance_deduction or 0) + float(ps.loan_deduction or 0) + float(ps.other_deduction or 0) + float(ps.tax_deduction or 0) + float(ps.pf_employee or 0) + float(ps.ssf_employee or 0)
            row = [ps.member.name, ps.month_name or '', ps.month_name or '', str(ps.from_date), str(ps.to_date)]
            if nepali_enabled:
                row += [_np_str(ps.from_date), _np_str(ps.to_date)]
            row += [ps.present_days, float(ps.gross_salary), float(ps.allowance_total), round(total_ded, 2), float(ps.net_payable), str(ps.generated_on.date() if ps.generated_on else '')]
            writer.writerow(row)
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payslips'
    _style_header(ws, headers, fill_color='16A34A')
    for ridx, ps in enumerate(qs, 2):
        total_ded = float(ps.advance_deduction or 0) + float(ps.loan_deduction or 0) + float(ps.other_deduction or 0) + float(ps.tax_deduction or 0) + float(ps.pf_employee or 0) + float(ps.ssf_employee or 0)
        row = [ps.member.name, ps.month_name or '', ps.month_name or '', str(ps.from_date), str(ps.to_date)]
        if nepali_enabled:
            row += [_np_str(ps.from_date), _np_str(ps.to_date)]
        row += [ps.present_days, float(ps.gross_salary), float(ps.allowance_total), round(total_ded, 2), float(ps.net_payable), str(ps.generated_on.date() if ps.generated_on else '')]
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)

    ws.column_dimensions['A'].width = 25
    resp = _excel_response('payslips_report.xlsx')
    wb.save(resp)
    return resp


# ── 3. STOCK EXPORT ──────────────────────────────────────────────────────────

@login_required
@feature_required('stock')
@perm_required('can_view_stock')
def export_stock(request):
    from handle.models import StockMovement, StockItem
    org = _org_from_req(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    export_type = request.GET.get('type', 'items')  # items or movements
    nepali_enabled = getattr(org, 'nepali_date', False)

    if export_type == 'movements':
        qs = StockMovement.objects.filter(org=org).select_related('item', 'created_by').order_by('-movement_date')
        headers = ['Item', 'Type', 'Quantity', 'Unit', 'Date (AD)']
        if nepali_enabled:
            headers.append('Date (BS)')
        headers += ['Note', 'Performed By']

        if fmt == 'csv':
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = 'attachment; filename="stock_movements.csv"'
            writer = csv.writer(resp)
            writer.writerow(headers)
            for m in qs:
                row = [m.item.name if m.item else '', m.movement_type, float(m.quantity), m.item.unit if m.item else '', str(m.movement_date)]
                if nepali_enabled:
                    row.append(_np_str(m.movement_date))
                row += [m.note or '', m.created_by.get_full_name() if m.created_by else '']
                writer.writerow(row)
            return resp

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Movements'
        _style_header(ws, headers, fill_color='D97706')
        for ridx, mv in enumerate(qs, 2):
            row = [mv.item.name if mv.item else '', mv.movement_type, float(mv.quantity), mv.item.unit if mv.item else '', str(mv.movement_date)]
            if nepali_enabled:
                row.append(_np_str(mv.movement_date))
            row += [mv.note or '', mv.created_by.get_full_name() if mv.created_by else '']
            for cidx, val in enumerate(row, 1):
                ws.cell(row=ridx, column=cidx, value=val)
    else:
        qs = StockItem.objects.filter(org=org).select_related('category').order_by('name')
        headers = ['Name', 'Category', 'Unit', 'Quantity', 'Low Stock Threshold', 'Unit Price', 'Status']

        if fmt == 'csv':
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = 'attachment; filename="stock_items.csv"'
            writer = csv.writer(resp)
            writer.writerow(headers)
            for item in qs:
                writer.writerow([item.name, item.category.name if item.category else '', item.unit, float(item.quantity), float(item.low_stock_threshold), float(item.purchase_cost or 0), item.status])
            return resp

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Items'
        _style_header(ws, headers, fill_color='D97706')
        for ridx, item in enumerate(qs, 2):
            row = [item.name, item.category.name if item.category else '', item.unit, float(item.quantity), float(item.low_stock_threshold), float(item.purchase_cost or 0), item.status]
            for cidx, val in enumerate(row, 1):
                ws.cell(row=ridx, column=cidx, value=val)

    ws.column_dimensions['A'].width = 25
    resp = _excel_response(f'stock_{export_type}.xlsx')
    wb.save(resp)
    return resp


# ── 4. FINANCE EXPORT ────────────────────────────────────────────────────────

@login_required
@feature_required('finance')
@perm_required('can_view_finance')
def export_finance(request):
    from handle.models import FinancialTransaction
    org = _org_from_req(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    txn_type = request.GET.get('type', 'all')  # income / expense / all
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    nepali_enabled = getattr(org, 'nepali_date', False)

    qs = FinancialTransaction.objects.filter(org=org).order_by('-transaction_date')
    if txn_type in ('income', 'expense'):
        qs = qs.filter(transaction_type=txn_type)
    if from_date:
        try:
            qs = qs.filter(transaction_date__gte=from_date)
        except Exception:
            pass
    if to_date:
        try:
            qs = qs.filter(transaction_date__lte=to_date)
        except Exception:
            pass

    headers = ['Type', 'Category', 'Amount', 'Date (AD)']
    if nepali_enabled:
        headers.append('Date (BS)')
    headers += ['Description', 'Recorded By']

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="finance_report.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for t in qs:
            row = [t.transaction_type, t.category.name if hasattr(t, 'category') and t.category else '', float(t.amount), str(t.transaction_date)]
            if nepali_enabled:
                row.append(_np_str(t.transaction_date))
            row += [t.description or '', t.created_by.get_full_name() if hasattr(t, 'created_by') and t.created_by else '']
            writer.writerow(row)
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Finance'
    _style_header(ws, headers, fill_color='15803D')
    for ridx, t in enumerate(qs, 2):
        row = [t.transaction_type, t.category.name if hasattr(t, 'category') and t.category else '', float(t.amount), str(t.transaction_date)]
        if nepali_enabled:
            row.append(_np_str(t.transaction_date))
        row += [t.description or '', t.created_by.get_full_name() if hasattr(t, 'created_by') and t.created_by else '']
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)

    ws.column_dimensions['C'].width = 15
    resp = _excel_response('finance_report.xlsx')
    wb.save(resp)
    return resp


# ── 5. LEAVE EXPORT ──────────────────────────────────────────────────────────

@login_required
@feature_required('leave')
@perm_required('can_view_leave_report')
def export_leave(request):
    org = _org_from_req(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    nepali_enabled = getattr(org, 'nepali_date', False)

    qs = LeaveReport.objects.filter(org=org).select_related('member', 'leave_type').order_by('-gap_start')

    headers = ['Member', 'Leave Type', 'From (AD)', 'To (AD)']
    if nepali_enabled:
        headers += ['From (BS)', 'To (BS)']
    headers += ['Days', 'Reason', 'Status']

    def leave_status(lr):
        if lr.approved:
            return 'Approved'
        if lr.rejected:
            return 'Rejected'
        return 'Pending'

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="leave_report.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for lr in qs:
            row = [lr.member.name, lr.leave_type.name if lr.leave_type else 'General', str(lr.gap_start or ''), str(lr.gap_end or '')]
            if nepali_enabled:
                row += [_np_str(lr.gap_start) if lr.gap_start else '', _np_str(lr.gap_end) if lr.gap_end else '']
            row += [lr.total_leave_days(), lr.reason, leave_status(lr)]
            writer.writerow(row)
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leave Report'
    _style_header(ws, headers, fill_color='7C3AED')
    for ridx, lr in enumerate(qs, 2):
        row = [lr.member.name, lr.leave_type.name if lr.leave_type else 'General', str(lr.gap_start or ''), str(lr.gap_end or '')]
        if nepali_enabled:
            row += [_np_str(lr.gap_start) if lr.gap_start else '', _np_str(lr.gap_end) if lr.gap_end else '']
        row += [lr.total_leave_days(), lr.reason, leave_status(lr)]
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)

    ws.column_dimensions['A'].width = 25
    resp = _excel_response('leave_report.xlsx')
    wb.save(resp)
    return resp


# ── 6. MEMBER LIST EXPORT ────────────────────────────────────────────────────

@login_required
@perm_required('can_export_reports')
def export_members(request):
    org = _org_from_req(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    nepali_enabled = getattr(org, 'nepali_date', False)

    qs = member.objects.filter(org=org).exclude(status='dumped').select_related('classification').order_by('name')

    headers = ['Name', 'Type', 'Class/Dept', 'Phone', 'Email', 'Status', 'Joined (AD)']
    if nepali_enabled:
        headers.append('Joined (BS)')
    headers.append('Salary')

    def _joined_date(m):
        # Prefer the explicit Joined Date field; fall back to the account's
        # creation timestamp for older records where it was never set.
        if m.admission_date:
            return m.admission_date
        return m.created_at.date() if hasattr(m, 'created_at') and m.created_at else None

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="members.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for m in qs:
            joined = _joined_date(m)
            row = [m.name, m.get_member_type_display(), m.classification.name if m.classification else '', m.phone or '', m.email or '', m.status, str(joined) if joined else '']
            if nepali_enabled:
                row.append(_np_str(joined) if joined else '')
            row.append(float(m.salary_amount or 0))
            writer.writerow(row)
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'
    _style_header(ws, headers, fill_color='0369A1')
    for ridx, m in enumerate(qs, 2):
        joined = _joined_date(m)
        row = [m.name, m.get_member_type_display(), m.classification.name if m.classification else '', m.phone or '', m.email or '', m.status, str(joined) if joined else '']
        if nepali_enabled:
            row.append(_np_str(joined) if joined else '')
        row.append(float(m.salary_amount or 0))
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)

    for col in ['A', 'D', 'E']:
        ws.column_dimensions[col].width = 25
    resp = _excel_response('members.xlsx')
    wb.save(resp)
    return resp


# ── 7. TASK EXPORT ───────────────────────────────────────────────────────────

@login_required
@perm_required('can_export_reports')
def export_tasks(request):
    org = _task_org(request)
    if not org:
        return HttpResponse('Unauthorized', status=403)

    fmt = request.GET.get('fmt', 'excel')
    nepali_enabled = getattr(org, 'nepali_date', False)

    qs = TaskInstance.objects.filter(task__org=org).select_related('task', 'assigned_member').order_by('-due_date')

    headers = ['Task Title', 'Member', 'Priority', 'Type', 'Status', 'Due Date (AD)']
    if nepali_enabled:
        headers.append('Due Date (BS)')
    headers += ['Completed At', 'Note']

    if fmt == 'csv':
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="tasks_report.csv"'
        writer = csv.writer(resp)
        writer.writerow(headers)
        for ti in qs:
            row = [ti.task.title, ti.assigned_member.name if ti.assigned_member else '', ti.task.get_priority_display(), ti.task.get_task_type_display(), ti.get_status_display(), str(ti.due_date)]
            if nepali_enabled:
                row.append(_np_str(ti.due_date) if ti.due_date else '')
            row += [str(ti.completed_at.date() if ti.completed_at else ''), ti.completion_note or '']
            writer.writerow(row)
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tasks'
    _style_header(ws, headers, fill_color='6D28D9')
    for ridx, ti in enumerate(qs, 2):
        row = [ti.task.title, ti.assigned_member.name if ti.assigned_member else '', ti.task.get_priority_display(), ti.task.get_task_type_display(), ti.get_status_display(), str(ti.due_date)]
        if nepali_enabled:
            row.append(_np_str(ti.due_date) if ti.due_date else '')
        row += [str(ti.completed_at.date() if ti.completed_at else ''), ti.completion_note or '']
        for cidx, val in enumerate(row, 1):
            ws.cell(row=ridx, column=cidx, value=val)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    resp = _excel_response('tasks_report.xlsx')
    wb.save(resp)
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT HUB PAGE
# ══════════════════════════════════════════════════════════════════════════════

class ExportHubView(AdminRequiredMixin, View):
    template_name = 'admin/export_hub.html'

    def get(self, request):
        if hasattr(request.user, 'schooladmin'):
            org = request.user.schooladmin.org
        elif request.user.user_type == '1':
            from management.models import Organization
            orgs = Organization.objects.all()
            return render(request, self.template_name, {'orgs': orgs, 'org': None, 'nepali_enabled': False})
        else:
            return redirect('management:homepage')

        nepali_enabled = getattr(org, 'nepali_date', False)
        return render(request, self.template_name, {'org': org, 'nepali_enabled': nepali_enabled})


# ══════════════════════════════════════════════════════════════════════════════
# CALENDAR PAGE + API
# ══════════════════════════════════════════════════════════════════════════════

class CalendarView(AdminRequiredMixin, View):
    template_name = 'admin/calendar.html'

    def get(self, request):
        if hasattr(request.user, 'schooladmin'):
            org = request.user.schooladmin.org
        else:
            return redirect('management:homepage')
        nepali_enabled = getattr(org, 'nepali_date', False)
        return render(request, self.template_name, {'org': org, 'nepali_enabled': nepali_enabled})


@login_required
def api_calendar_events(request):
    """Return all events + holidays + occasions as FullCalendar JSON."""
    if hasattr(request.user, 'schooladmin'):
        org = request.user.schooladmin.org
    else:
        return JsonResponse([], safe=False)

    nepali_enabled = getattr(org, 'nepali_date', False)
    events = []

    # Handle holidays (management.Holiday)
    for h in Holiday.objects.filter(org=org):
        events.append({
            'id': f'holiday_{h.pk}',
            'title': h.holiday,
            'color': '#ef4444',
            'allDay': True,
            'extendedProps': {'type': 'holiday', 'pk': h.pk},
        })

    # Occasions (management.Occasion) have dates
    for oc in Occasion.objects.filter(org=org):
        item = {
            'id': f'occasion_{oc.pk}',
            'title': oc.name,
            'start': str(oc.date),
            'color': '#f59e0b',
            'allDay': True,
            'extendedProps': {'type': 'occasion', 'pk': oc.pk},
        }
        if oc.end_date:
            import datetime as _dt
            end = oc.end_date + _dt.timedelta(days=1)  # FullCalendar end is exclusive
            item['end'] = str(end)
        events.append(item)

    # Events (handle.Event)
    for ev in Event.objects.filter(org=org):
        color_map = {
            'sports': '#3b82f6', 'seminar': '#8b5cf6', 'meeting': '#06b6d4',
            'exam': '#ef4444', 'program': '#22c55e', 'holiday': '#ef4444', 'other': '#6b7280',
        }
        import datetime as _dt
        events.append({
            'id': f'event_{ev.pk}',
            'title': ev.title,
            'start': str(ev.start_date),
            'end': str(ev.end_date + _dt.timedelta(days=1)),
            'color': color_map.get(ev.event_type, '#6b7280'),
            'allDay': True,
            'extendedProps': {'type': 'event', 'pk': ev.pk, 'event_type': ev.event_type, 'status': ev.status},
        })

    return JsonResponse(events, safe=False)


@login_required
def api_calendar_add(request):
    """Add a holiday, occasion, or event from calendar click."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    if hasattr(request.user, 'schooladmin'):
        org = request.user.schooladmin.org
    else:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    data = json.loads(request.body)
    item_type = data.get('type')
    title = data.get('title', '').strip()
    start = data.get('start')
    end = data.get('end') or start

    if not title or not start:
        return JsonResponse({'error': 'Title and date required'}, status=400)

    import datetime as _dt

    try:
        start_date = _dt.date.fromisoformat(start)
        end_date = _dt.date.fromisoformat(end) if end else start_date
    except Exception:
        return JsonResponse({'error': 'Invalid date'}, status=400)

    if item_type == 'holiday':
        # Holiday model has no date field; save as Event so it appears on the calendar
        ev = Event.objects.create(org=org, title=title, event_type='holiday', start_date=start_date, end_date=start_date, status='upcoming')
        return JsonResponse({'ok': True, 'id': f'event_{ev.pk}'})

    elif item_type == 'occasion':
        oc, created = Occasion.objects.get_or_create(
            org=org, name=title, date=start_date,
            defaults={'end_date': end_date if end_date != start_date else None},
        )
        if not created:
            return JsonResponse({'error': 'This occasion already exists on that date.'}, status=409)
        return JsonResponse({'ok': True, 'id': f'occasion_{oc.pk}'})

    elif item_type == 'event':
        event_type = data.get('event_type', 'other')
        ev = Event.objects.create(org=org, title=title, event_type=event_type, start_date=start_date, end_date=end_date, status='upcoming')
        return JsonResponse({'ok': True, 'id': f'event_{ev.pk}'})

    return JsonResponse({'error': 'Unknown type'}, status=400)


@login_required
def api_calendar_delete(request, item_type, pk):
    """Delete a holiday/occasion/event."""
    if hasattr(request.user, 'schooladmin'):
        org = request.user.schooladmin.org
    else:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if item_type == 'holiday':
        Holiday.objects.filter(pk=pk, org=org).delete()
    elif item_type == 'occasion':
        Occasion.objects.filter(pk=pk, org=org).delete()
    elif item_type == 'event':
        Event.objects.filter(pk=pk, org=org).delete()
    else:
        return JsonResponse({'error': 'Unknown type'}, status=400)

    return JsonResponse({'ok': True})


# ── Full Member Profile ─────────────────────────────────────────────────────────

class MemberProfileView(AdminRequiredMixin, View):
    def get(self, request, pk):
        org = request.user.schooladmin.org
        mem = get_object_or_404(member, pk=pk, org=org)

        # ── Attendance stats ───────────────────────────────────────────────
        from collections import defaultdict
        records_qs = AttendanceRecord.objects.filter(mem=mem).order_by('scanned_time')
        daily = defaultdict(list)
        for r in records_qs:
            d = timezone.localtime(r.scanned_time).date()
            daily[d].append(timezone.localtime(r.scanned_time).time())

        total_present = len(daily)
        total_late_minutes = 0
        late_days = 0

        for date, times in daily.items():
            times_sorted = sorted(times)
            first_punch = times_sorted[0]
            if first_punch > mem.shift_start_time:
                import datetime as _dt
                diff = _dt.datetime.combine(date, first_punch) - _dt.datetime.combine(date, mem.shift_start_time)
                total_late_minutes += int(diff.total_seconds() / 60)
                late_days += 1

        late_hours = total_late_minutes // 60
        late_mins = total_late_minutes % 60

        # Heatmap – last 90 days
        import datetime as _dt
        today = _dt.date.today()
        heatmap = {}
        for i in range(89, -1, -1):
            d = today - _dt.timedelta(days=i)
            heatmap[d.isoformat()] = 'present' if d in daily else 'absent'

        # ── Leave ─────────────────────────────────────────────────────────
        nepali_enabled = org.nepali_date
        leaves = list(LeaveReport.objects.filter(member=mem, org=org).order_by('-id')[:30])
        for leave in leaves:
            if nepali_enabled:
                leave.start_display = to_bs_display(leave.gap_start)
                leave.end_display = to_bs_display(leave.gap_end)
            else:
                leave.start_display = leave.gap_start.strftime("%Y-%m-%d") if leave.gap_start else ""
                leave.end_display = leave.gap_end.strftime("%Y-%m-%d") if leave.gap_end else ""
        leave_qs = LeaveReport.objects.filter(member=mem, org=org)
        leave_approved = leave_qs.filter(approved=True).count()
        leave_pending = leave_qs.filter(approved=False, rejected=False).count()
        leave_rejected = leave_qs.filter(rejected=True).count()

        # ── Payslips ──────────────────────────────────────────────────────
        payslips = PaySlip.objects.filter(member=mem, org=org).order_by('-generated_on')
        total_net_paid = sum(p.net_payable for p in payslips if p.status == 'paid')
        total_net_payable = sum(p.net_payable for p in payslips)

        # ── Tasks ─────────────────────────────────────────────────────────
        from handle.models import TaskInstance
        task_instances = TaskInstance.objects.filter(assigned_member=mem).select_related('task').order_by('-due_date')
        task_total    = task_instances.count()
        task_completed = task_instances.filter(status='completed').count()
        task_pending  = task_instances.filter(status__in=['pending', 'in_progress']).count()
        task_overdue  = task_instances.filter(status='overdue').count()
        task_missed   = task_instances.filter(status='missed_absence').count()
        task_on_time_pct = round(task_completed / task_total * 100) if task_total else 0

        # ── Complaints ────────────────────────────────────────────────────
        complaints = Complaint.objects.filter(filed_by=mem, org=org).order_by('-created_at')
        complaint_pending  = complaints.filter(status='pending').count()
        complaint_resolved = complaints.filter(status='resolved').count()

        # ── Bills ─────────────────────────────────────────────────────────
        bills = Bill.objects.filter(member=mem, org=org).order_by('-issue_date')
        total_billed = sum(b.total_amount for b in bills)
        total_paid   = sum(b.amount_paid for b in bills)
        total_due    = total_billed - total_paid

        # ── Academic Results ───────────────────────────────────────────────
        results = ResultRecord.objects.filter(student=mem).select_related('exam', 'subject').order_by('-exam__id')

        # ── HRMS ──────────────────────────────────────────────────────────
        resignation = mem.resignations.first()
        documents   = mem.documents.all().order_by('-id')

        # ── Complete member timeline and location/work evidence ──────────
        from handle.models import MemberHistory, LocationPing, LiveTrackingSession
        history = MemberHistory.objects.filter(org=org, member=mem).select_related('changed_by')[:30]
        enrollments = mem.course_enrollments.filter(org=org).select_related(
            'academic_year', 'course', 'classification', 'section', 'branch',
        ).order_by('-start_date')
        field_visits = mem.field_visits.filter(org=org).select_related(
            'client', 'reviewed_by',
        ).order_by('-visited_at')[:30]
        tracking_sessions = LiveTrackingSession.objects.filter(
            org=org, member=mem,
        ).order_by('-started_at')[:30]
        latest_pings = LocationPing.objects.filter(
            org=org, member=mem,
        ).order_by('-tracked_at')[:30]
        weekday_shifts = mem.weekday_shifts.select_related('shift').order_by('weekday')

        # ── Library — borrowing history (skip the query entirely if the
        # feature isn't enabled for this org) ──────────────────────────────
        book_issues = []
        if has_feature(org, 'library'):
            book_issues = list(
                mem.book_issues.select_related('book').order_by('-issue_date')[:30]
            )

        # ── Absence Corrections ───────────────────────────────────────────
        corrections_qs = AbsenceCorrection.objects.filter(member=mem, org=org).order_by('-date')
        corrections_count = corrections_qs.count()
        corrections = corrections_qs[:30]

        # ── Payroll Adjustments (allowances / advances / loans) ───────────
        adjustments_active = PayrollAdjustment.objects.filter(
            org=org, member=mem, status='active'
        ).order_by('-effective_date')
        adjustments_applied = PayrollAdjustment.objects.filter(
            org=org, member=mem, status='applied'
        ).order_by('-effective_date')[:10]
        adj_allowance_total = sum(a.amount for a in adjustments_active if a.adjustment_type == 'allowance')
        adj_advance_total   = sum(a.amount for a in adjustments_active if a.adjustment_type == 'advance')
        adj_loan_total      = sum(a.amount for a in adjustments_active if a.adjustment_type == 'loan')

        ctx = {
            'mem': mem,
            'org': org,
            'total_present': total_present,
            'late_days': late_days,
            'late_hours': late_hours,
            'late_mins': late_mins,
            'heatmap_json': json.dumps(heatmap),
            'leaves': leaves,
            'leave_approved': leave_approved,
            'leave_pending': leave_pending,
            'leave_rejected': leave_rejected,
            'payslips': payslips,
            'total_net_paid': total_net_paid,
            'total_net_payable': total_net_payable,
            'task_instances': task_instances[:30],
            'task_total': task_total,
            'task_completed': task_completed,
            'task_pending': task_pending,
            'task_overdue': task_overdue,
            'task_missed': task_missed,
            'task_on_time_pct': task_on_time_pct,
            'complaints': complaints,
            'complaint_pending': complaint_pending,
            'complaint_resolved': complaint_resolved,
            'bills': bills,
            'total_billed': total_billed,
            'total_paid': total_paid,
            'total_due': total_due,
            'results': results,
            'resignation': resignation,
            'documents': documents,
            'history': history,
            'enrollments': enrollments,
            'field_visits': field_visits,
            'tracking_sessions': tracking_sessions,
            'latest_pings': latest_pings,
            'weekday_shifts': weekday_shifts,
            'book_issues': book_issues,
            'corrections': corrections,
            'corrections_count': corrections_count,
            'adjustments_active': adjustments_active,
            'adjustments_applied': adjustments_applied,
            'adj_allowance_total': adj_allowance_total,
            'adj_advance_total': adj_advance_total,
            'adj_loan_total': adj_loan_total,
            'nepali_enabled': nepali_enabled,
            'today': _dt.date.today().strftime('%Y-%m-%d'),
        }
        return render(request, 'admin/member_profile.html', ctx)

    def post(self, request, pk):
        org = request.user.schooladmin.org
        mem = get_object_or_404(member, pk=pk, org=org)
        action = request.POST.get('action')

        if action == 'add_adjustment':
            from handle.models import MemberHistory
            adj_type  = request.POST.get('adjustment_type', '')
            title     = request.POST.get('title', '').strip()
            amount    = request.POST.get('amount', '0')
            eff_date  = request.POST.get('effective_date') or timezone.localdate()
            notes     = request.POST.get('notes', '')
            if title and adj_type:
                try:
                    adjustment = PayrollAdjustment.objects.create(
                        org=org, member=mem,
                        adjustment_type=adj_type,
                        title=title,
                        amount=Decimal(str(amount)),
                        effective_date=eff_date,
                        notes=notes,
                        created_by=request.user,
                        status='active',
                    )
                    MemberHistory.objects.create(
                        org=org, member=mem, action='payroll_adjustment_added',
                        field_name='payroll_adjustment', changed_by=request.user,
                        new_value=f'{adjustment.get_adjustment_type_display()}: Rs. {adjustment.amount}',
                        description=f'Payroll adjustment added: {title}',
                        metadata={'adjustment_id': adjustment.pk},
                    )
                    messages.success(request, f"{title} added to {mem.name}'s adjustments.")
                except Exception as e:
                    messages.error(request, f"Error: {e}")
            else:
                messages.error(request, "Title and type are required.")

        elif action == 'cancel_adjustment':
            from handle.models import MemberHistory
            adj_pk = request.POST.get('adj_pk')
            adjustment = PayrollAdjustment.objects.filter(pk=adj_pk, org=org, member=mem).first()
            if adjustment:
                adjustment.status = 'cancelled'
                adjustment.save(update_fields=['status'])
                MemberHistory.objects.create(
                    org=org, member=mem, action='payroll_adjustment_cancelled',
                    field_name='payroll_adjustment', changed_by=request.user,
                    old_value=f'{adjustment.get_adjustment_type_display()}: Rs. {adjustment.amount}',
                    description=f'Payroll adjustment cancelled: {adjustment.title}',
                    metadata={'adjustment_id': adjustment.pk},
                )
                messages.success(request, "Adjustment cancelled.")

        return redirect('schooladmin:member_profile', pk=pk)


class MemberHistoryView(AdminRequiredMixin, View):
    """Organization-scoped, paginated member lifecycle ledger."""
    template_name = 'admin/member_history.html'

    def get(self, request, pk):
        from django.core.paginator import Paginator
        from handle.models import MemberHistory
        org = _get_org(request)
        memb = get_object_or_404(member, pk=pk, org=org)
        history = MemberHistory.objects.filter(
            org=org, member=memb,
        ).select_related('changed_by')
        page_obj = Paginator(history, 50).get_page(request.GET.get('page'))
        return render(request, self.template_name, {
            'org': org, 'mem': memb, 'history': page_obj, 'page_obj': page_obj,
        })


# ── Advance Salary ─────────────────────────────────────────────────────────────

from handle.models import AdvanceSalary


class AdvanceSalaryView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'payroll'
    template_name = 'admin/payroll/advance_salary.html'

    def get(self, request):
        org = _get_org(request)
        advances = AdvanceSalary.objects.filter(org=org).select_related('member', 'approved_by').prefetch_related('payments').order_by('-created_at')
        members_qs = member.objects.filter(org=org, salary_amount__gt=0).exclude(status='dumped').order_by('name')
        return render(request, self.template_name, {
            'org': org,
            'advances': advances,
            'members': members_qs,
            'today': datetime.date.today().strftime('%Y-%m-%d'),
        })

    def post(self, request):
        org = _get_org(request)
        action = request.POST.get('action')

        if action == 'create':
            member_id = request.POST.get('member_id')
            total_amount = request.POST.get('total_amount')
            num_installments = int(request.POST.get('num_installments', 1))
            purpose = request.POST.get('purpose', '')
            effective_date = request.POST.get('effective_date') or datetime.date.today()
            notes = request.POST.get('notes', '')
            try:
                mem = get_object_or_404(member, pk=member_id, org=org)
                total = Decimal(str(total_amount))
                installment = (total / num_installments).quantize(Decimal('0.01'))
                adv = AdvanceSalary.objects.create(
                    org=org, member=mem,
                    total_amount=total,
                    num_installments=num_installments,
                    installment_amount=installment,
                    remaining_balance=total,
                    purpose=purpose,
                    approved_by=request.user,
                    effective_date=effective_date,
                    notes=notes,
                    status='active',
                )
                if has_feature(org, 'accounting'):
                    from handle.accounting import post_advance_salary_journal_entry
                    try:
                        post_advance_salary_journal_entry(adv)
                    except ValueError as e:
                        messages.warning(request, f"Advance recorded, but the journal entry could not be posted: {e}")
                messages.success(request, f"Advance Rs. {total} approved for {mem.name}. Recovering Rs. {installment}/month over {num_installments} month(s).")
            except Exception as e:
                messages.error(request, f"Error: {e}")

        elif action == 'apply_installment':
            adv_pk = request.POST.get('adv_pk')
            adv = get_object_or_404(AdvanceSalary, pk=adv_pk, org=org)
            adv.apply_installment()
            messages.success(request, f"Installment of Rs. {adv.installment_amount} applied. Remaining: Rs. {adv.remaining_balance}")

        elif action == 'cancel':
            adv_pk = request.POST.get('adv_pk')
            AdvanceSalary.objects.filter(pk=adv_pk, org=org).update(status='cancelled')
            messages.success(request, "Advance salary cancelled.")

        return redirect('schooladmin:advance_salary')


# ── Appointment Management ─────────────────────────────────────────────────────

from management.models import AppointmentType, Appointment, CustomForm, FormField, FormSubmission, FieldResponse


class AppointmentTypeView(AdminRequiredMixin, View):
    def get(self, request):
        org = request.user.schooladmin.org
        types = AppointmentType.objects.filter(org=org)
        appts = Appointment.objects.filter(org=org).order_by('-date', '-time')
        return render(request, 'admin/appointments.html', {'org': org, 'types': types, 'appts': appts})

    def post(self, request):
        org = request.user.schooladmin.org
        action = request.POST.get('action')
        if action == 'create_type':
            AppointmentType.objects.create(
                org=org,
                name=request.POST.get('name', ''),
                description=request.POST.get('description', ''),
                duration_minutes=int(request.POST.get('duration_minutes', 30)),
                color=request.POST.get('color', '#e11d48'),
            )
            messages.success(request, 'Appointment type created.')
        elif action == 'delete_type':
            AppointmentType.objects.filter(pk=request.POST.get('pk'), org=org).delete()
            messages.success(request, 'Appointment type deleted.')
        elif action == 'update_status':
            appt = get_object_or_404(Appointment, pk=request.POST.get('pk'), org=org)
            appt.status = request.POST.get('status', 'confirmed')
            appt.admin_note = request.POST.get('admin_note', '')
            appt.save()
            messages.success(request, f'Appointment {appt.status}.')
        return redirect(request.path)


class FormBuilderView(AdminRequiredMixin, View):
    def get(self, request):
        org = request.user.schooladmin.org
        forms = CustomForm.objects.filter(org=org)
        return render(request, 'admin/form_builder_list.html', {'org': org, 'forms': forms})


class FormBuilderEditView(AdminRequiredMixin, View):
    def get(self, request, pk=None):
        org = request.user.schooladmin.org
        form_obj = get_object_or_404(CustomForm, pk=pk, org=org) if pk else None
        fields = form_obj.fields.all() if form_obj else []
        return render(request, 'admin/form_builder_edit.html', {'org': org, 'form_obj': form_obj, 'fields': fields})

    def post(self, request, pk=None):
        org = request.user.schooladmin.org
        action = request.POST.get('action', 'save_form')

        if action == 'save_form':
            if pk:
                form_obj = get_object_or_404(CustomForm, pk=pk, org=org)
            else:
                form_obj = CustomForm(org=org)
            form_obj.title = request.POST.get('title', 'Untitled Form')
            form_obj.description = request.POST.get('description', '')
            form_obj.is_active = request.POST.get('is_active') == 'on'
            form_obj.success_message = request.POST.get('success_message', 'Thank you for your submission!')
            form_obj.save()
            messages.success(request, 'Form saved.')
            return redirect('schooladmin:form_builder_edit', pk=form_obj.pk)

        elif action == 'add_field':
            form_obj = get_object_or_404(CustomForm, pk=pk, org=org)
            order = form_obj.fields.count()
            FormField.objects.create(
                form=form_obj,
                label=request.POST.get('label', 'Field'),
                field_type=request.POST.get('field_type', 'text'),
                placeholder=request.POST.get('placeholder', ''),
                options=request.POST.get('options', ''),
                required=request.POST.get('required') == 'on',
                order=order,
            )
            messages.success(request, 'Field added.')
            return redirect('schooladmin:form_builder_edit', pk=pk)

        elif action == 'delete_field':
            FormField.objects.filter(pk=request.POST.get('field_pk'), form__org=org).delete()
            messages.success(request, 'Field removed.')
            return redirect('schooladmin:form_builder_edit', pk=pk)

        elif action == 'delete_form':
            get_object_or_404(CustomForm, pk=pk, org=org).delete()
            messages.success(request, 'Form deleted.')
            return redirect('schooladmin:form_builder')

        return redirect(request.path)


class FormSubmissionsView(AdminRequiredMixin, View):
    def get(self, request, pk):
        org = request.user.schooladmin.org
        form_obj = get_object_or_404(CustomForm, pk=pk, org=org)
        submissions = form_obj.submissions.prefetch_related('responses').all()
        return render(request, 'admin/form_submissions.html', {'org': org, 'form_obj': form_obj, 'submissions': submissions})


# =============================================================
# CLASSIFICATION DETAIL HUB
# =============================================================

import calendar as _cal
from handle.models import Bill, BillItem, BillSendLog, ResultSendLog

def _money(value):
    try:
        return Decimal(str(value or 0)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _student_subjects(mem):
    """Subjects/courses applicable to this student for result and fee use."""
    if not mem or not getattr(mem, 'org_id', None) or not getattr(mem, 'classification_id', None):
        return Subject.objects.none()
    qs = Subject.objects.filter(
        org_id=mem.org_id,
        classification_id=mem.classification_id,
        status='active',
    )
    if mem.section_id:
        return qs.filter(Q(section_id=mem.section_id) | Q(section__isnull=True)).order_by('name')
    return qs.filter(section__isnull=True).order_by('name')


def _fee_breakdown(mem):
    """Return a billing snapshot without mutating old bills."""
    billing_type = mem.billing_type or 'monthly_fee'
    base_amount = Decimal("0.00")
    course_fee = Decimal("0.00")
    subjects = []

    if billing_type == 'course_wise':
        subjects = list(_student_subjects(mem))
        course_fee = sum((_money(subject.monthly_fee) for subject in subjects), Decimal("0.00"))
    elif billing_type == 'scholarship':
        base_amount = Decimal("0.00")
    else:
        base_amount = _money(mem.monthly_fee)

    subtotal = base_amount + course_fee
    discount = Decimal("0.00")
    if mem.discount_type == 'fixed':
        discount = _money(mem.discount_amount)
    elif mem.discount_type == 'percentage':
        discount = (subtotal * _money(mem.discount_amount) / Decimal("100")).quantize(Decimal("0.01"))
    scholarship = _money(mem.scholarship_amount)
    final_amount = max(Decimal("0.00"), subtotal - discount - scholarship)
    if billing_type == 'scholarship' and not mem.monthly_fee:
        final_amount = Decimal("0.00")

    return {
        'billing_type': billing_type,
        'base_amount': base_amount,
        'course_fee_amount': course_fee,
        'discount_amount': discount,
        'scholarship_amount': scholarship,
        'final_amount': final_amount,
        'subjects': subjects,
    }


def _compute_final_fee(mem):
    """Return the final monthly payable for a member based on billing setup."""
    return _fee_breakdown(mem)['final_amount']


def _bill_due_queryset(org, mem, month=None, year=None):
    qs = Bill.objects.filter(org=org, member=mem).exclude(status='Cancelled')
    if month and year:
        qs = qs.exclude(billing_month=month, billing_year=year)
    return qs


def _bill_due_total(qs):
    totals = qs.aggregate(total=Sum('total_amount'), paid=Sum('amount_paid'))
    return max(Decimal("0.00"), _money(totals.get('total')) - _money(totals.get('paid')))


def _format_message_template(template, values):
    try:
        return template.format(**values)
    except Exception:
        return template


class ClassificationDetailView(AdminRequiredMixin, View):
    template_name = 'admin/classification/detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        cls = get_object_or_404(Classification, pk=pk, org=org)
        tab = request.GET.get('tab', 'overview')
        section_filter = request.GET.get('section', '')
        search_q = request.GET.get('q', '').strip()

        sections = list(Section.objects.filter(org=org, classification=cls).order_by('name'))

        # Student queryset
        students_qs = member.objects.filter(org=org, classification=cls).exclude(status='dumped').select_related('section')
        if section_filter:
            students_qs = students_qs.filter(section_id=section_filter)
        if search_q:
            students_qs = students_qs.filter(
                Q(name__icontains=search_q) | Q(card__icontains=search_q) |
                Q(phone__icontains=search_q) | Q(guardian_name__icontains=search_q)
            )
        students = list(students_qs.order_by('name'))

        # Stats
        total_students  = member.objects.filter(org=org, classification=cls).exclude(status='dumped').count()
        active_students = member.objects.filter(org=org, classification=cls, status='active').count()

        # Billing stats
        today = timezone.localdate()
        bills_qs = Bill.objects.filter(org=org, classification=cls)
        expected_monthly = sum(_compute_final_fee(m) for m in member.objects.filter(org=org, classification=cls, status='active'))
        total_billed = _money(bills_qs.aggregate(s=Sum('total_amount'))['s'])
        total_paid   = _money(bills_qs.aggregate(s=Sum('amount_paid'))['s'])
        total_due    = max(Decimal("0.00"), total_billed - total_paid)
        monthly_fee_count = member.objects.filter(org=org, classification=cls, billing_type='monthly_fee').exclude(status='dumped').count()
        course_wise_count = member.objects.filter(org=org, classification=cls, billing_type='course_wise').exclude(status='dumped').count()
        custom_fee_count = member.objects.filter(org=org, classification=cls, billing_type='custom').exclude(status='dumped').count()
        scholarship_count = member.objects.filter(org=org, classification=cls, billing_type='scholarship').exclude(status='dumped').count()

        # Subject/course stats
        subjects_count = Subject.objects.filter(org=org, classification=cls, status='active').count()

        # Exam stats
        exams = ExamTerm.objects.filter(org=org, classification=cls).order_by('-start_date')
        published_exams = exams.filter(is_published=True).count()

        # Attendance summary (today present %)
        today_records = AttendanceRecord.objects.filter(
            org=org, mem__classification=cls, scanned_time__date=today
        ).values('mem_id').distinct().count()
        attend_pct = round(today_records / max(active_students, 1) * 100)

        # Per-student billing summary for table
        student_rows = []
        for s in students:
            final = _compute_final_fee(s)
            s_due = _bill_due_total(_bill_due_queryset(org, s))
            # Last result
            last_result = ResultRecord.objects.filter(student=s).select_related('exam').order_by('-exam__start_date').first()
            present_days = AttendanceRecord.objects.filter(
                org=org, mem=s, scanned_time__date__gte=today - timedelta(days=30)
            ).values('scanned_time__date').distinct().count()
            student_rows.append({
                'member': s,
                'final_fee': final,
                'due': s_due,
                'last_result': last_result,
                'attendance_pct': round(present_days / 30 * 100),
            })
        from management.models import CustomUser as _CU

        context = {
            'org': org,
            'cls': cls,
            'tab': tab,
            'sections': sections,
            'section_filter': section_filter,
            'search_q': search_q,
            'students': student_rows,
            'total_students': total_students,
            'active_students': active_students,
            'sections_count': len(sections),
            'subjects_count': subjects_count,
            'expected_monthly': expected_monthly,
            'total_billed': total_billed,
            'total_paid': total_paid,
            'total_due': total_due,
            'monthly_fee_count': monthly_fee_count,
            'course_wise_count': course_wise_count,
            'custom_fee_count': custom_fee_count,
            'scholarship_count': scholarship_count,
            'published_exams': published_exams,
            'exams': exams,
            'attend_pct': attend_pct,
            'today_present': today_records,
            # Subject list
            'subjects': Subject.objects.filter(org=org, classification=cls).select_related('section', 'teacher').order_by('name'),
            'teachers': _CU.objects.filter(Q(staff__org=org) | Q(schooladmin__org=org)).distinct(),
            'bill_send_logs': BillSendLog.objects.filter(bill__org=org, bill__classification=cls).select_related('bill', 'bill__member', 'sent_by')[:10],
            'result_send_logs': ResultSendLog.objects.filter(exam__org=org, member__classification=cls).select_related('exam', 'member', 'sent_by')[:10],
            'student_status_choices': member.STATUS_CHOICES,
            'billing_type_choices': member.BILLING_TYPE_CHOICES,
            'discount_type_choices': member.DISCOUNT_TYPE_CHOICES,
            'months': [(str(i), _cal.month_name[i]) for i in range(1, 13)],
            'current_month': str(today.month),
            'current_year': str(today.year),
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        org = _get_org(request)
        cls = get_object_or_404(Classification, pk=pk, org=org)
        action = request.POST.get('action', '')

        if action == 'add_section':
            name = request.POST.get('name', '').strip()
            if name:
                Section.objects.get_or_create(
                    org=org,
                    classification=cls,
                    name=name,
                    defaults={
                        'status': request.POST.get('status', 'active'),
                        'code': request.POST.get('code', '').strip() or None,
                        'branch': cls.branch,
                    },
                )
                messages.success(request, f"Section '{name}' added.")
            else:
                messages.error(request, "Section name required.")

        elif action == 'edit_section':
            sid = request.POST.get('section_id')
            sec = get_object_or_404(Section, pk=sid, org=org, classification=cls)
            sec.name = request.POST.get('name', sec.name).strip() or sec.name
            sec.code = request.POST.get('code', '').strip() or None
            sec.status = request.POST.get('status', sec.status)
            sec.save(update_fields=['name', 'code', 'status', 'updated_at'])
            messages.success(request, "Section updated.")

        elif action == 'deactivate_section':
            sid = request.POST.get('section_id')
            Section.objects.filter(pk=sid, org=org, classification=cls).update(status='inactive')
            messages.success(request, "Section deactivated.")

        elif action == 'delete_section':
            sid = request.POST.get('section_id')
            Section.objects.filter(pk=sid, org=org, classification=cls).delete()
            messages.success(request, "Section deleted.")

        elif action == 'add_subject':
            from handle.models import Subject as _Subj
            name     = request.POST.get('name', '').strip()
            sec_id   = request.POST.get('section') or None
            full_m   = _money(request.POST.get('full_marks', 100))
            pass_m   = _money(request.POST.get('pass_marks', 40))
            monthly  = _money(request.POST.get('monthly_fee') or 0)
            one_time = _money(request.POST.get('one_time_fee') or 0)
            if not name:
                messages.error(request, "Subject name required.")
            elif full_m <= pass_m:
                messages.error(request, "Full marks must be greater than pass marks.")
            elif _Subj.objects.filter(org=org, classification=cls, section_id=sec_id, name=name).exists():
                messages.error(request, "Subject already exists.")
            else:
                _Subj.objects.create(
                    org=org, classification=cls, section_id=sec_id,
                    name=name, code=request.POST.get('code', '').strip() or None,
                    full_marks=full_m, pass_marks=pass_m,
                    monthly_fee=monthly, one_time_fee=one_time,
                    teacher_id=request.POST.get('teacher') or None,
                    status='active',
                )
                messages.success(request, f"Subject '{name}' added.")

        elif action == 'edit_subject':
            subj = get_object_or_404(Subject, pk=request.POST.get('subject_id'), org=org, classification=cls)
            subj.name = request.POST.get('name', subj.name).strip() or subj.name
            subj.code = request.POST.get('code', '').strip() or None
            subj.section_id = request.POST.get('section') or None
            subj.teacher_id = request.POST.get('teacher') or None
            subj.full_marks = _money(request.POST.get('full_marks', subj.full_marks))
            subj.pass_marks = _money(request.POST.get('pass_marks', subj.pass_marks))
            subj.monthly_fee = _money(request.POST.get('monthly_fee', subj.monthly_fee))
            subj.one_time_fee = _money(request.POST.get('one_time_fee', subj.one_time_fee))
            subj.status = request.POST.get('status', subj.status)
            subj.save()
            messages.success(request, "Course / Subject updated.")

        elif action == 'deactivate_subject':
            Subject.objects.filter(pk=request.POST.get('subject_id'), org=org, classification=cls).update(status='inactive')
            messages.success(request, "Course / Subject deactivated.")

        elif action == 'delete_subject':
            Subject.objects.filter(pk=request.POST.get('subject_id'), org=org, classification=cls).delete()
            messages.success(request, "Subject deleted.")

        return redirect(f"{reverse('schooladmin:classification_detail', args=[pk])}?tab={request.POST.get('return_tab','overview')}")


# =============================================================
# STUDENT / MEMBER MANAGEMENT (Add / Edit with billing)
# =============================================================

class StudentAddEditView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'student_mgmt'
    template_name = 'admin/students/add_edit.html'

    def _get_context(self, org, instance=None):
        from management.models import CustomUser as _CU
        final_fee = _compute_final_fee(instance) if instance else Decimal("0.00")
        total_due = _bill_due_total(_bill_due_queryset(org, instance)) if instance else Decimal("0.00")
        return {
            'org': org,
            'instance': instance,
            'classifications': Classification.objects.filter(org=org),
            'sections': Section.objects.filter(org=org).select_related('classification'),
            'teachers': _CU.objects.filter(Q(staff__org=org) | Q(schooladmin__org=org)).distinct(),
            'status_choices': member.STATUS_CHOICES,
            'billing_type_choices': member.BILLING_TYPE_CHOICES,
            'discount_type_choices': member.DISCOUNT_TYPE_CHOICES,
            'final_fee': final_fee,
            'total_due': total_due,
        }

    def get(self, request, pk=None):
        org = _get_org(request)
        instance = get_object_or_404(member, pk=pk, org=org) if pk else None
        return render(request, self.template_name, self._get_context(org, instance))

    def post(self, request, pk=None):
        org = _get_org(request)
        instance = get_object_or_404(member, pk=pk, org=org) if pk else None

        name     = request.POST.get('name', '').strip()
        gender   = request.POST.get('gender', 'Male')
        def _int_or_none(value):
            digits = ''.join(ch for ch in str(value or '') if ch.isdigit())
            return int(digits) if digits else None

        phone    = request.POST.get('phone', '').strip() or None
        email    = request.POST.get('email', '').strip() or None
        address  = request.POST.get('address', '').strip() or None
        card     = request.POST.get('card', '').strip() or None
        cls_id   = request.POST.get('classification') or None
        sec_id   = request.POST.get('section') or None
        mtype    = request.POST.get('member_type', 'student')
        status   = request.POST.get('status', 'active')
        dob      = request.POST.get('date_of_birth') or None
        adm_date = request.POST.get('admission_date') or None

        # Guardian
        g_name  = request.POST.get('guardian_name', '').strip() or None
        g_phone = request.POST.get('guardian_phone', '').strip() or None
        g_email = request.POST.get('guardian_email', '').strip() or None

        # Billing
        billing_type    = request.POST.get('billing_type', 'monthly_fee')
        monthly_fee_val = request.POST.get('monthly_fee', '0').strip() or 0
        disc_type       = request.POST.get('discount_type') or None
        disc_amount     = request.POST.get('discount_amount', '0').strip() or 0
        schol_amount    = request.POST.get('scholarship_amount', '0').strip() or 0
        billing_start   = request.POST.get('billing_start_date') or None
        try:
            due_day = min(max(int(request.POST.get('due_day', '15') or 15), 1), 31)
        except (TypeError, ValueError):
            due_day = 15

        if not name:
            messages.error(request, "Name is required.")
            return render(request, self.template_name, self._get_context(org, instance))
        if email and member.objects.filter(email=email).exclude(status='dumped').exclude(pk=instance.pk if instance else None).exists():
            messages.error(request, "A member with this email already exists.")
            return render(request, self.template_name, self._get_context(org, instance))

        # Build or update
        if instance:
            mem = instance
        else:
            mem = member(org=org)

        mem.name          = name
        mem.gender        = gender
        mem.phone         = _int_or_none(phone)
        mem.email         = email
        mem.address       = address
        mem.card          = card
        mem.classification_id = cls_id
        mem.section_id    = sec_id
        mem.member_type   = mtype
        mem.status        = status
        mem.date_of_birth = dob
        mem.admission_date = adm_date
        mem.guardian_name  = g_name
        mem.guardian_phone = _int_or_none(g_phone)
        mem.guardian_email = g_email
        mem.billing_type      = billing_type
        mem.monthly_fee       = _money(monthly_fee_val)
        mem.discount_type     = disc_type
        mem.discount_amount   = _money(disc_amount)
        mem.scholarship_amount = _money(schol_amount)
        mem.billing_start_date = billing_start
        mem.due_day            = due_day
        # Compute final fee
        mem.final_monthly_fee = _compute_final_fee(mem)
        mem.save()
        messages.success(request, f"{'Updated' if pk else 'Added'} member '{mem.name}' successfully.")
        if cls_id:
            return redirect(f"{reverse('schooladmin:classification_detail', args=[cls_id])}?tab=students")
        return redirect('schooladmin:student_list')


class StudentListView(AdminRequiredMixin, FeatureRequiredMixin, View):
    """Global student list (all classifications)."""
    required_feature = 'student_mgmt'
    template_name = 'admin/students/list.html'

    def get(self, request):
        org = _get_org(request)
        cls_filter = request.GET.get('classification', '')
        sec_filter = request.GET.get('section', '')
        type_filter = request.GET.get('member_type', '')
        search_q = request.GET.get('q', '').strip()

        qs = member.objects.filter(org=org).exclude(status='dumped').select_related('classification', 'section').order_by('classification__name', 'name')
        if cls_filter:
            qs = qs.filter(classification_id=cls_filter)
        if sec_filter:
            qs = qs.filter(section_id=sec_filter)
        if type_filter:
            qs = qs.filter(member_type=type_filter)
        if search_q:
            qs = qs.filter(
                Q(name__icontains=search_q) | Q(card__icontains=search_q) |
                Q(phone__icontains=search_q) | Q(guardian_name__icontains=search_q)
            )

        rows = []
        for mem_obj in qs:
            rows.append({
                'member': mem_obj,
                'final_fee': _compute_final_fee(mem_obj),
                'due': _bill_due_total(_bill_due_queryset(org, mem_obj)),
            })

        context = {
            'org': org,
            'members': qs,
            'member_rows': rows,
            'classifications': Classification.objects.filter(org=org),
            'sections': Section.objects.filter(org=org),
            'cls_filter': cls_filter,
            'sec_filter': sec_filter,
            'type_filter': type_filter,
            'search_q': search_q,
            'total': len(rows),
        }
        return render(request, self.template_name, context)


# =============================================================
# BULK BILL GENERATION
# =============================================================

class BulkBillGenerateView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'billing'
    template_name = 'admin/billing/bulk_generate.html'

    def get(self, request):
        org = _get_org(request)
        cls_id = request.GET.get('classification', '')
        sec_id = request.GET.get('section', '')
        month  = request.GET.get('month', '')
        year   = request.GET.get('year', str(timezone.localdate().year))
        billing_type = request.GET.get('billing_type', '')
        preview = request.GET.get('preview', '')

        classifications = Classification.objects.filter(org=org)
        sections = Section.objects.filter(org=org, classification_id=cls_id) if cls_id else Section.objects.none()

        preview_rows = []
        month_int = int(month) if month else None
        year_int  = int(year)  if year  else None

        if preview and cls_id and month_int and year_int:
            students = member.objects.filter(
                org=org, classification_id=cls_id, status='active'
            ).select_related('classification', 'section')
            if sec_id:
                students = students.filter(section_id=sec_id)
            if billing_type:
                students = students.filter(billing_type=billing_type)

            for s in students:
                breakdown = _fee_breakdown(s)
                # Check existing bill
                existing = Bill.objects.filter(
                    org=org, member=s, billing_month=month_int, billing_year=year_int
                ).first()
                prev_due = _bill_due_total(_bill_due_queryset(org, s, month_int, year_int))

                preview_rows.append({
                    'member': s,
                    'base_amount': breakdown['base_amount'],
                    'course_fee_amount': breakdown['course_fee_amount'],
                    'discount_amount': breakdown['discount_amount'],
                    'scholarship_amount': breakdown['scholarship_amount'],
                    'final_fee': breakdown['final_amount'],
                    'prev_due': prev_due,
                    'total_payable': breakdown['final_amount'] + prev_due,
                    'existing_bill': existing,
                    'existing_status': existing.status if existing else None,
                })

        context = {
            'org': org,
            'classifications': classifications,
            'sections': sections,
            'cls_id': cls_id,
            'sec_id': sec_id,
            'month': month,
            'year': year,
            'billing_type': billing_type,
            'billing_type_choices': member.BILLING_TYPE_CHOICES,
            'months': [(str(i), _cal.month_name[i]) for i in range(1, 13)],
            'years': [str(y) for y in range(timezone.localdate().year - 2, timezone.localdate().year + 2)],
            'preview_rows': preview_rows,
            'preview': bool(preview),
        }
        return render(request, self.template_name, context)

    def post(self, request):
        from django.db import transaction as _tx
        org = _get_org(request)
        cls_id    = request.POST.get('classification')
        sec_id    = request.POST.get('section') or None
        billing_type = request.POST.get('billing_type') or None
        month_int = int(request.POST.get('month', 0))
        year_int  = int(request.POST.get('year', 0))
        regenerate = request.POST.get('regenerate') == '1'

        if not cls_id or not month_int or not year_int:
            messages.error(request, "Classification, month and year are required.")
            return redirect('schooladmin:bulk_bill_generate')

        students = member.objects.filter(org=org, classification_id=cls_id, status='active')
        if sec_id:
            students = students.filter(section_id=sec_id)
        if billing_type:
            students = students.filter(billing_type=billing_type)

        generated = skipped = 0
        month_name = _cal.month_name[month_int]
        import uuid as _uuid

        with _tx.atomic():
            for s in students:
                existing = Bill.objects.filter(org=org, member=s, billing_month=month_int, billing_year=year_int).first()
                if existing and not regenerate:
                    skipped += 1
                    continue
                if existing and regenerate:
                    existing.delete()

                breakdown = _fee_breakdown(s)
                final_fee = breakdown['final_amount']
                prev_due = _bill_due_total(_bill_due_queryset(org, s))

                due_day_val = s.due_day or 15
                import calendar as cal2
                last_day = cal2.monthrange(year_int, month_int)[1]
                due_day_actual = min(due_day_val, last_day)
                import datetime as _dt
                due_date = _dt.date(year_int, month_int, due_day_actual)

                inv_num = f"BILL-{year_int}{month_int:02d}-{s.id}-{_uuid.uuid4().hex[:6].upper()}"
                bill = Bill.objects.create(
                    org=org,
                    member=s,
                    classification_id=cls_id,
                    section_id=s.section_id,
                    invoice_number=inv_num,
                    billing_month=month_int,
                    billing_year=year_int,
                    billing_type=breakdown['billing_type'],
                    base_amount=breakdown['base_amount'],
                    course_fee_amount=breakdown['course_fee_amount'],
                    previous_due=prev_due,
                    discount_amount=breakdown['discount_amount'],
                    scholarship_amount=breakdown['scholarship_amount'],
                    total_amount=final_fee + prev_due,
                    due_date=due_date,
                    generated_by=request.user,
                )
                # Bill items
                if breakdown['billing_type'] == 'course_wise':
                    for subject in breakdown['subjects']:
                        amount = _money(subject.monthly_fee)
                        if amount > 0:
                            BillItem.objects.create(
                                bill=bill,
                                subject=subject,
                                description=f"{subject.name} Course Fee - {month_name} {year_int}",
                                fee_type='course',
                                amount=amount,
                            )
                elif breakdown['base_amount'] > 0:
                    title = "Custom Fee" if breakdown['billing_type'] == 'custom' else "Monthly Fee"
                    BillItem.objects.create(
                        bill=bill,
                        description=f"{month_name} {year_int} {title}",
                        fee_type='monthly',
                        amount=breakdown['base_amount'],
                    )
                if breakdown['discount_amount'] > 0:
                    BillItem.objects.create(
                        bill=bill,
                        description="Discount",
                        fee_type='misc',
                        amount=0,
                        discount=breakdown['discount_amount'],
                    )
                if breakdown['scholarship_amount'] > 0:
                    BillItem.objects.create(
                        bill=bill,
                        description="Scholarship",
                        fee_type='misc',
                        amount=0,
                        discount=breakdown['scholarship_amount'],
                    )
                if prev_due > 0:
                    BillItem.objects.create(bill=bill, description="Previous Due", fee_type='misc', amount=prev_due)
                generated += 1

        messages.success(request, f"Generated {generated} bills. Skipped {skipped} (already exist).")
        send_url = f"{reverse('schooladmin:bulk_bill_send')}?classification={cls_id}&month={month_int}&year={year_int}"
        if sec_id:
            send_url += f"&section={sec_id}"
        return redirect(send_url)


# =============================================================
# BULK BILL SEND
# =============================================================

class BulkBillSendView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'billing'
    template_name = 'admin/billing/bulk_send.html'

    def get(self, request):
        org = _get_org(request)
        cls_id   = request.GET.get('classification', '')
        sec_id   = request.GET.get('section', '')
        month    = request.GET.get('month', '')
        year     = request.GET.get('year', str(timezone.localdate().year))

        classifications = Classification.objects.filter(org=org)
        sections = Section.objects.filter(org=org, classification_id=cls_id) if cls_id else Section.objects.none()

        bills = Bill.objects.none()
        if cls_id and month and year:
            bills = Bill.objects.filter(
                org=org, classification_id=cls_id,
                billing_month=int(month), billing_year=int(year)
            ).select_related('member', 'member__section')
            if sec_id:
                bills = bills.filter(section_id=sec_id)
            bills = bills.order_by('member__name')

        context = {
            'org': org,
            'classifications': classifications,
            'sections': sections,
            'bills': bills,
            'cls_id': cls_id, 'sec_id': sec_id,
            'month': month, 'year': year,
            'months': [(str(i), _cal.month_name[i]) for i in range(1, 13)],
            'years': [str(y) for y in range(timezone.localdate().year - 2, timezone.localdate().year + 2)],
        }
        return render(request, self.template_name, context)

    def post(self, request):
        from django.db import transaction as _tx
        org      = _get_org(request)
        bill_ids = request.POST.getlist('bill_ids')
        method   = request.POST.get('send_method', 'email')
        custom_msg = request.POST.get('custom_message', '').strip()

        sent = failed = 0
        with _tx.atomic():
            for bid in bill_ids:
                try:
                    bill = Bill.objects.get(pk=bid, org=org)
                    mem  = bill.member
                    month_name = _cal.month_name[bill.billing_month] if bill.billing_month else "—"

                    guardian = mem.guardian_name or mem.name
                    values = {
                        'guardian_name': guardian,
                        'student_name': mem.name,
                        'month': f"{month_name} {bill.billing_year or ''}".strip(),
                        'total_payable': bill.total_amount,
                        'paid_amount': bill.amount_paid,
                        'due_amount': bill.balance_due,
                        'due_date': bill.due_date,
                        'organization_name': org.name,
                    }
                    default_msg = (
                        f"Dear {guardian},\n"
                        f"Bill for {mem.name} for {month_name} {bill.billing_year} has been generated.\n"
                        f"Total Amount: Rs. {bill.total_amount}\n"
                        f"Paid: Rs. {bill.amount_paid}\n"
                        f"Due: Rs. {bill.balance_due}\n"
                        f"Due Date: {bill.due_date}\n"
                        f"Please clear the payment on time.\n\n"
                        f"— {org.name}"
                    )
                    msg = _format_message_template(custom_msg, values) if custom_msg else default_msg

                    log_status = 'pending'
                    err_msg = None

                    if method == 'email':
                        email_target = mem.guardian_email or mem.email
                        if email_target:
                            try:
                                send_mail(
                                    subject=f"Fee Bill – {month_name} {bill.billing_year} | {org.name}",
                                    message=msg,
                                    from_email=settings.DEFAULT_FROM_EMAIL,
                                    recipient_list=[email_target],
                                    fail_silently=False,
                                )
                                log_status = 'sent'
                                bill.is_sent = True
                                bill.sent_at = timezone.now()
                                bill.sent_method = 'email'
                                bill.save(update_fields=['is_sent', 'sent_at', 'sent_method'])
                                sent += 1
                            except Exception as e:
                                log_status = 'failed'
                                err_msg = str(e)
                                failed += 1
                        else:
                            log_status = 'failed'
                            err_msg = 'No email address'
                            failed += 1
                    else:
                        # PDF / other: just mark as ready
                        log_status = 'pending'
                        bill.sent_method = method
                        bill.save(update_fields=['sent_method'])
                        sent += 1

                    BillSendLog.objects.create(
                        bill=bill,
                        sent_to_email=mem.guardian_email or mem.email,
                        sent_to_phone=str(mem.guardian_phone or mem.phone or ''),
                        sent_method=method,
                        message_body=msg,
                        status=log_status,
                        sent_by=request.user,
                        error_message=err_msg,
                    )
                except Exception:
                    failed += 1

        if method == 'email':
            messages.success(request, f"Sent {sent} emails. Failed: {failed}.")
        else:
            messages.success(request, f"Marked {sent} bills as ready-to-send via {method}.")
        return redirect(request.path + f"?classification={request.POST.get('cls_id','')}&month={request.POST.get('month','')}&year={request.POST.get('year','')}")


# =============================================================
# BULK RESULT SEND
# =============================================================

class BulkResultSendView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'results'
    template_name = 'admin/results/bulk_send.html'

    def _build_msg(self, org, mem, exam, total_obtained, total_full, pct, grade, passed):
        guardian = mem.guardian_name or mem.name
        return (
            f"Dear {guardian},\n"
            f"Result of {mem.name} for {exam.name} has been published.\n"
            f"Total: {total_obtained}/{total_full}\n"
            f"Percentage: {pct}%\n"
            f"Grade: {grade}\n"
            f"Status: {'Pass' if passed else 'Fail'}\n"
            f"Please check the full marksheet from the school portal.\n\n"
            f"— {org.name}"
        )

    def get(self, request):
        org    = _get_org(request)
        exam_id = request.GET.get('exam', '')
        cls_id  = request.GET.get('classification', '')
        sec_id  = request.GET.get('section', '')

        exams           = ExamTerm.objects.filter(org=org, is_published=True).order_by('-start_date')
        classifications = Classification.objects.filter(org=org)
        sections        = Section.objects.filter(org=org, classification_id=cls_id) if cls_id else Section.objects.none()

        preview_rows = []
        exam = None
        if exam_id and cls_id:
            exam = get_object_or_404(ExamTerm, pk=exam_id, org=org)
            students = member.objects.filter(org=org, classification_id=cls_id).exclude(status='dumped')
            if sec_id:
                students = students.filter(section_id=sec_id)

            for s in students:
                subjects = list(_student_subjects(s))
                recs = {r.subject_id: r for r in ResultRecord.objects.filter(student=s, exam=exam, subject__in=subjects).select_related('subject')}
                total_obt = float(sum(r.obtained_marks for r in recs.values() if not r.is_absent))
                total_full_m = float(sum(sub.full_marks for sub in subjects))
                pct = round(total_obt / max(total_full_m, 1) * 100, 1)
                passed = bool(subjects) and len(recs) >= len(subjects) and all(r.is_passed for r in recs.values())
                grade = _compute_grade(pct) if recs else '—'
                already_sent = ResultSendLog.objects.filter(exam=exam, member=s, status='sent').exists()
                preview_rows.append({
                    'member': s,
                    'total_obt': total_obt,
                    'total_full': total_full_m,
                    'pct': pct,
                    'grade': grade,
                    'passed': passed,
                    'has_results': bool(recs),
                    'is_complete': bool(subjects) and len(recs) >= len(subjects),
                    'subjects_count': len(subjects),
                    'already_sent': already_sent,
                })

        context = {
            'org': org,
            'exams': exams,
            'classifications': classifications,
            'sections': sections,
            'exam_id': exam_id,
            'cls_id': cls_id,
            'sec_id': sec_id,
            'exam': exam,
            'preview_rows': preview_rows,
        }
        return render(request, self.template_name, context)

    def post(self, request):
        from django.db import transaction as _tx
        org        = _get_org(request)
        exam_id    = request.POST.get('exam_id')
        cls_id     = request.POST.get('cls_id') or None
        member_ids = request.POST.getlist('member_ids')
        method     = request.POST.get('send_method', 'email')
        custom_msg = request.POST.get('custom_message', '').strip()

        exam = get_object_or_404(ExamTerm, pk=exam_id, org=org)

        sent = failed = 0
        with _tx.atomic():
            for mid in member_ids:
                try:
                    mem = member.objects.get(pk=mid, org=org)
                    subjects = list(_student_subjects(mem))
                    recs = list(ResultRecord.objects.filter(student=mem, exam=exam, subject__in=subjects))
                    total_obt  = float(sum(r.obtained_marks for r in recs if not r.is_absent))
                    total_full = float(sum(sub.full_marks for sub in subjects))
                    pct        = round(total_obt / max(total_full, 1) * 100, 1)
                    passed     = bool(subjects) and len(recs) >= len(subjects) and all(r.is_passed for r in recs)
                    grade      = _compute_grade(pct) if recs else 'NG'

                    values = {
                        'guardian_name': mem.guardian_name or mem.name,
                        'student_name': mem.name,
                        'exam_title': exam.name,
                        'obtained_marks': total_obt,
                        'full_marks': total_full,
                        'percentage': pct,
                        'grade': grade,
                        'pass_fail_status': 'Pass' if passed else 'Fail',
                        'organization_name': org.name,
                    }
                    msg = _format_message_template(custom_msg, values) if custom_msg else self._build_msg(org, mem, exam, total_obt, total_full, pct, grade, passed)
                    log_status = 'pending'
                    err_msg = None

                    if method == 'email':
                        email_target = mem.guardian_email or mem.email
                        if email_target:
                            try:
                                send_mail(
                                    subject=f"Result Published – {exam.name} | {org.name}",
                                    message=msg,
                                    from_email=settings.DEFAULT_FROM_EMAIL,
                                    recipient_list=[email_target],
                                    fail_silently=False,
                                )
                                log_status = 'sent'
                                sent += 1
                            except Exception as e:
                                log_status = 'failed'
                                err_msg = str(e)
                                failed += 1
                        else:
                            log_status = 'failed'
                            err_msg = 'No email address'
                            failed += 1
                    else:
                        log_status = 'pending'
                        sent += 1

                    ResultSendLog.objects.create(
                        exam=exam, member=mem,
                        sent_to_email=mem.guardian_email or mem.email or '',
                        sent_to_phone=str(mem.guardian_phone or mem.phone or ''),
                        sent_method=method,
                        message_body=msg,
                        status=log_status,
                        sent_by=request.user,
                        error_message=err_msg,
                    )
                except Exception:
                    failed += 1

        if method == 'email':
            messages.success(request, f"Sent {sent} result emails. Failed: {failed}.")
        else:
            messages.success(request, f"Logged {sent} result send records.")
        return redirect(request.path + f"?exam={exam_id}&classification={cls_id or ''}")


# =============================================================
# STUDENT PROFILE (upgraded)
# =============================================================

class StudentProfileView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'student_mgmt'
    template_name = 'admin/students/profile.html'

    def get(self, request, pk):
        org = _get_org(request)
        mem = get_object_or_404(member, pk=pk, org=org)
        tab = request.GET.get('tab', 'overview')

        bills    = Bill.objects.filter(org=org, member=mem).prefetch_related('items').order_by('-issue_date')
        results  = ResultRecord.objects.filter(student=mem).select_related('exam', 'subject').order_by('-exam__start_date')
        exams    = ExamTerm.objects.filter(org=org, is_published=True)
        subjects = _student_subjects(mem)
        docs     = StaffDocument.objects.filter(org=org, member=mem)
        gaps     = AttendanceGap.objects.filter(org=org, member=mem).order_by('-date')[:10]

        # Finance summary
        total_billed = _money(bills.aggregate(s=Sum('total_amount'))['s'])
        total_paid   = _money(bills.aggregate(s=Sum('amount_paid'))['s'])
        total_due    = max(Decimal("0.00"), total_billed - total_paid)

        # Attendance summary (last 30 days)
        today = timezone.localdate()
        from_date = today - timedelta(days=30)
        att_count = AttendanceRecord.objects.filter(
            mem=mem, scanned_time__date__gte=from_date
        ).values('scanned_time__date').distinct().count()

        final_fee = _compute_final_fee(mem)

        context = {
            'org': org,
            'mem': mem,
            'tab': tab,
            'bills': bills,
            'results': results,
            'subjects': subjects,
            'docs': docs,
            'gaps': gaps,
            'final_fee': final_fee,
            'total_billed': total_billed,
            'total_paid': total_paid,
            'total_due': total_due,
            'att_count': att_count,
        }
        return render(request, self.template_name, context)


# ── Organization Feature Settings ────────────────────────────────────────────

# All boolean fields on Organization that can be toggled via the Feature Settings page
_ORG_FEATURE_FIELDS = [
    'rfid_based', 'qr_based', 'location_based', 'manual_attendance', 'auto_checkin',
    'course_based_attendance', 'nepali_date', 'mutifeature_enable',
    'feature_finance', 'feature_billing', 'feature_stock', 'feature_tasks',
    'feature_results', 'feature_hrms', 'feature_payroll', 'feature_complaints',
    'feature_events', 'feature_branches', 'feature_leave', 'feature_study_gap',
    'feature_bulk_export', 'feature_notifications', 'feature_courses',
    'feature_student_mgmt', 'feature_member_mgmt', 'enable_qr_attendance',
    'feature_timesheet', 'feature_id_cards',
    'feature_field_visits', 'feature_clients', 'feature_face_attendance',
]


class OrgFeaturesView(AdminRequiredMixin, View):
    """Let schooladmin toggle which modules are active for their organization."""
    template_name = 'admin/org/features.html'

    def _get_org(self, request):
        if request.user.user_type == '2':
            return request.user.schooladmin.org
        return None

    def get(self, request):
        org = self._get_org(request)
        if org is None:
            messages.error(request, "No organization found.")
            return redirect('schooladmin:dashboard')
        return render(request, self.template_name, {'org': org})

    def post(self, request):
        org = self._get_org(request)
        if org is None:
            messages.error(request, "No organization found.")
            return redirect('schooladmin:dashboard')

        from school.features import FIELD_TO_KEY_MAP, is_feature_allowed
        for field in _ORG_FEATURE_FIELDS:
            feature_key = FIELD_TO_KEY_MAP.get(field)
            if feature_key is not None and not is_feature_allowed(org, feature_key):
                continue  # not on this org's plan — ignore even if somehow POSTed
            value = request.POST.get(field) == '1'
            setattr(org, field, value)

        # Dependency enforcement: disable dependent features when parent is off
        if not org.feature_courses:
            org.feature_results = False
            org.feature_study_gap = False
        if not org.feature_student_mgmt:
            org.feature_billing = False

        org.save()
        messages.success(request, "Feature settings saved successfully.")
        return redirect('schooladmin:org_features')


# ── Staff Permission Management ───────────────────────────────────────────────


class RolesPermissionsListView(AdminRequiredMixin, View):
    """Unified list: every member with their privilege level + a link to edit permissions.

    Replaces the old separate PrivilegeManageView (member.privilege) and
    StaffPermissionsView (StaffPermission flags) pages.
    """
    template_name = 'admin/hrms/roles_permissions_list.html'

    def get(self, request):
        from django.core.paginator import Paginator
        org = _get_org(request)
        q = request.GET.get('q', '').strip()
        qs = member.objects.filter(org=org).exclude(status='dumped').order_by('name')
        if q:
            qs = qs.filter(name__icontains=q)
        paginator = Paginator(qs, 25)
        page_obj = paginator.get_page(request.GET.get('page'))
        return render(request, self.template_name, {
            'org': org,
            'members': page_obj,
            'page_obj': page_obj,
            'q': q,
            'privilege_choices': PRIVILEGE_LEVEL_CHOICES,
        })


class RolesPermissionsEditView(AdminRequiredMixin, View):
    """Edit a single staff member's role preset + granular permissions in one page.

    Replaces the old EditStaffPermissionsView (StaffPermission flags only) and
    folds in PrivilegeManageView's role dropdown as a pure UX preset-picker:
    applying a preset only bulk-checks the boxes below (still editable before
    save) and remembers the choice on member.privilege for next visit. Access
    control never reads member.privilege — only the StaffPermission flags.
    """
    template_name = 'admin/hrms/roles_permissions_edit.html'

    def _get_or_create_sp(self, mem):
        from handle.models import StaffPermission
        sp, _ = StaffPermission.objects.get_or_create(
            member=mem,
            defaults={'org': mem.org},
        )
        return sp

    def _dynamic_categories_for_org(self, org, mem):
        """Build {category_slug: {...}} for enabled dynamic features, mirroring PERMISSION_REGISTRY's shape."""
        from handle.models import DynamicFeature, OrganizationFeatureGrant, StaffPermissionGrant
        enabled_keys = set(
            OrganizationFeatureGrant.objects.filter(org=org, enabled=True, feature__is_active=True)
            .values_list('feature__key', flat=True)
        )
        if not enabled_keys:
            return {}
        features = DynamicFeature.objects.filter(key__in=enabled_keys, is_active=True).prefetch_related('permissions')
        granted_map = dict(
            StaffPermissionGrant.objects.filter(member=mem, permission__feature__in=features)
            .values_list('permission__flag', 'granted')
        )
        by_category = {}
        for feature in features:
            perms = [{
                'flag': perm.flag, 'label': perm.label, 'icon': perm.icon or 'fa-check-circle',
                'visible': True, 'dynamic': True, 'checked': granted_map.get(perm.flag, False),
            } for perm in feature.permissions.all()]
            if not perms:
                continue
            cat_slug = feature.category or 'custom'
            bucket = by_category.setdefault(cat_slug, {'slug': cat_slug, 'staff_perms': []})
            bucket['staff_perms'].extend(perms)
        return by_category

    def get(self, request, member_id):
        from school.permissions import PERMISSION_REGISTRY, PRESET_DEFINITIONS, is_perm_visible
        import copy, json
        org = _get_org(request)
        mem = get_object_or_404(member, pk=member_id, org=org)
        sp = self._get_or_create_sp(mem)

        registry = copy.deepcopy(PERMISSION_REGISTRY)
        dynamic_by_cat = self._dynamic_categories_for_org(org, mem)
        for cat in registry:
            for perm in cat['staff_perms']:
                perm['visible'] = is_perm_visible(org, perm)
            dyn_bucket = dynamic_by_cat.pop(cat['slug'], None)
            if dyn_bucket:
                cat['staff_perms'].extend(dyn_bucket['staff_perms'])
            visible_perms = [p for p in cat['staff_perms'] if p['visible']]
            # A category only earns a place on the page if at least one of its
            # permissions is actually usable for this org's enabled features.
            cat['visible'] = bool(visible_perms)
            cat['visible_count'] = len(visible_perms)

        # Any dynamic features whose category didn't match an existing registry
        # slug (or were left blank) get a generic "Custom Modules" category.
        leftover_perms = [p for bucket in dynamic_by_cat.values() for p in bucket['staff_perms']]
        if leftover_perms:
            registry.append({
                'slug': 'custom', 'label': 'Custom Modules', 'icon': 'fa-puzzle-piece',
                'staff_perms': leftover_perms, 'visible': True, 'visible_count': len(leftover_perms),
            })

        presets_json = json.dumps({str(k): sorted(v) for k, v in PRESET_DEFINITIONS.items()})
        return render(request, self.template_name, {
            'org': org,
            'member': mem,
            'sp': sp,
            'registry': registry,
            'privilege_choices': PRIVILEGE_LEVEL_CHOICES,
            'presets_json': presets_json,
        })

    def post(self, request, member_id):
        from school.permissions import PERMISSION_REGISTRY, is_perm_visible
        from handle.models import StaffPermissionGrant, DynamicPermission
        from school.features import invalidate_member_perm_cache
        org = _get_org(request)
        mem = get_object_or_404(member, pk=member_id, org=org)
        sp = self._get_or_create_sp(mem)

        # Only touch flags that were actually rendered as checkboxes for this
        # org's currently-enabled features. A flag hidden because its feature
        # is off right now must keep its stored value untouched - otherwise
        # saving any OTHER permission on this page would silently wipe out
        # permissions tied to features that are simply disabled at the moment
        # (e.g. re-enabling Payroll later would find every staff member's
        # payroll permissions reset to off).
        for cat in PERMISSION_REGISTRY:
            for perm in cat['staff_perms']:
                if is_perm_visible(org, perm):
                    field = perm['flag']
                    setattr(sp, field, request.POST.get(field) == '1')
        sp.save()

        # Same anti-wipe rule for dynamic permissions: only write flags whose
        # feature is currently granted+active for this org (i.e. was actually
        # rendered), leaving anything else's stored grant untouched.
        dynamic_by_cat = self._dynamic_categories_for_org(org, mem)
        rendered_flags = [p['flag'] for bucket in dynamic_by_cat.values() for p in bucket['staff_perms']]
        if rendered_flags:
            perms_by_flag = {p.flag: p for p in DynamicPermission.objects.filter(flag__in=rendered_flags)}
            for flag in rendered_flags:
                StaffPermissionGrant.objects.update_or_create(
                    member=mem, permission=perms_by_flag[flag],
                    defaults={'granted': request.POST.get(flag) == '1'},
                )
            invalidate_member_perm_cache(mem.id)

        privilege = request.POST.get('privilege')
        if privilege:
            mem.privilege = int(privilege)
            mem.save(update_fields=['privilege'])

        messages.success(request, f"Permissions updated for {mem.name}.")
        return redirect('schooladmin:roles_permissions_list')


# ─── Dynamic QR Attendance ────────────────────────────────────────────────────

import secrets
from django.utils import timezone as tz
from handle.models import QRAttendanceSession, QRAttendanceScanLog, Branch


def _qr_png_data_uri(value, box_size=5):
    """Build a self-contained QR image without relying on a browser CDN."""
    import base64
    import qrcode

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=box_size,
        border=3,
    )
    qr.add_data(value)
    qr.make(fit=True)
    image = qr.make_image(fill_color="#0f172a", back_color="white")
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


class QRAttendancePageView(AdminRequiredMixin, View):
    template_name = 'admin/qr_attendance.html'

    def get(self, request):
        org = _get_org(request)
        if not org or not org.enable_qr_attendance:
            messages.error(request, "QR Attendance is not enabled for your organization.")
            return redirect('schooladmin:dashboard')
        # Expire any stale active sessions
        stale = QRAttendanceSession.objects.filter(
            org=org,
            session_type='dynamic',
            status='active',
            expires_at__lt=tz.now(),
        )
        stale.update(status='expired')

        active_session = QRAttendanceSession.objects.filter(
            org=org, session_type='dynamic', status='active'
        ).order_by('-created_at').first()
        permanent_session = QRAttendanceSession.objects.filter(
            org=org, session_type='permanent', status='active'
        ).order_by('-created_at').first()

        from django.core.paginator import Paginator
        permanent_history = QRAttendanceSession.objects.filter(
            org=org,
            session_type='permanent',
        ).select_related('branch', 'generated_by').order_by('-created_at')
        permanent_sessions = Paginator(permanent_history, 8).get_page(
            request.GET.get('qr_page')
        )
        for saved_session in permanent_sessions.object_list:
            saved_session.qr_data_uri = _qr_png_data_uri(saved_session.token)

        branches = Branch.objects.filter(org=org, status='active') if org.feature_branches else []
        from management.models import LocationBased
        locations = LocationBased.objects.filter(org=org).order_by('name')
        return render(request, self.template_name, {
            'org': org,
            'active_session': active_session,
            'permanent_session': permanent_session,
            'permanent_sessions': permanent_sessions,
            'branches': branches,
            'locations': locations,
        })


class QRPermanentPrintView(AdminRequiredMixin, View):
    template_name = 'admin/qr_permanent_print.html'

    def get(self, request, session_id):
        org = _get_org(request)
        if not org or not org.enable_qr_attendance:
            messages.error(request, "QR Attendance is not enabled for your organization.")
            return redirect('schooladmin:dashboard')

        session = get_object_or_404(
            QRAttendanceSession.objects.select_related('branch', 'generated_by'),
            pk=session_id,
            org=org,
            session_type='permanent',
        )
        return render(request, self.template_name, {
            'org': org,
            'session': session,
            'qr_data_uri': _qr_png_data_uri(session.token, box_size=12),
        })


class QRGenerateView(AdminRequiredMixin, View):
    def post(self, request):
        org = _get_org(request)
        if not org or not org.enable_qr_attendance:
            return JsonResponse({'error': 'QR Attendance not enabled.'}, status=403)

        # Close any existing active sessions for this org
        QRAttendanceSession.objects.filter(
            org=org, session_type='dynamic', status='active'
        ).update(
            status='closed', closed_at=tz.now()
        )

        try:
            expiry_minutes = int(request.POST.get('expiry_minutes', 2))
        except (TypeError, ValueError):
            return JsonResponse(
                {'error': 'Expiry must be a whole number of minutes.'},
                status=400,
            )
        expiry_minutes = max(1, min(expiry_minutes, 30))  # clamp 1–30

        branch_id = request.POST.get('branch_id')
        branch = None
        if branch_id and org.feature_branches:
            branch = Branch.objects.filter(id=branch_id, org=org).first()

        now = tz.now()
        token = secrets.token_urlsafe(48)

        session = QRAttendanceSession.objects.create(
            org=org,
            branch=branch,
            generated_by=request.user,
            token=token,
            session_type='dynamic',
            status='active',
            valid_from=now,
            expires_at=now + datetime.timedelta(minutes=expiry_minutes),
            date=now.date(),
        )

        return JsonResponse({
            'session_id': session.id,
            'token': token,
            'expires_at': session.expires_at.isoformat(),
            'expires_at_ts': int(session.expires_at.timestamp() * 1000),
            'expiry_minutes': expiry_minutes,
        })


class QRPermanentGenerateView(AdminRequiredMixin, View):
    """Create one printable organisation QR protected by a required geofence."""

    def post(self, request):
        org = _get_org(request)
        if not org or not org.enable_qr_attendance:
            return JsonResponse({'error': 'QR Attendance not enabled.'}, status=403)

        from management.models import LocationBased

        branch_id = request.POST.get('branch_id')
        branch = None
        if branch_id and org.feature_branches:
            branch = Branch.objects.filter(pk=branch_id, org=org).first()
            if branch is None:
                return JsonResponse({'error': 'Invalid branch.'}, status=400)

        location = None
        location_id = request.POST.get('location_id')
        if location_id:
            location = LocationBased.objects.filter(
                pk=location_id, org=org,
            ).first()
        try:
            latitude = (
                location.latitude
                if location
                else float(request.POST.get('latitude', ''))
            )
            longitude = (
                location.longitude
                if location
                else float(request.POST.get('longitude', ''))
            )
            radius = int(
                location.radius
                if location
                else request.POST.get('radius_meters', 100)
            )
        except (TypeError, ValueError):
            return JsonResponse(
                {'error': 'Choose an attendance location or share valid GPS coordinates.'},
                status=400,
            )
        import math

        if (
            not math.isfinite(latitude)
            or not math.isfinite(longitude)
            or not -90 <= latitude <= 90
            or not -180 <= longitude <= 180
        ):
            return JsonResponse({'error': 'GPS coordinates are out of range.'}, status=400)
        radius = max(20, min(radius, 2000))
        location_name = (
            location.name
            if location
            else (request.POST.get('location_name') or 'Organisation premises').strip()
        )[:200]

        now = tz.now()
        with transaction.atomic():
            QRAttendanceSession.objects.filter(
                org=org,
                session_type='permanent',
                status='active',
            ).update(status='closed', closed_at=now)
            session = QRAttendanceSession.objects.create(
                org=org,
                branch=branch,
                generated_by=request.user,
                token=secrets.token_urlsafe(48),
                session_type='permanent',
                status='active',
                valid_from=now,
                expires_at=None,
                date=None,
                location_name=location_name,
                latitude=latitude,
                longitude=longitude,
                radius_meters=radius,
            )
        return JsonResponse({
            'session_id': session.id,
            'token': session.token,
            'location_name': session.location_name,
            'latitude': session.latitude,
            'longitude': session.longitude,
            'radius_meters': session.radius_meters,
        }, status=201)


class TransportManagementView(AdminRequiredMixin, View):
    template_name = 'admin/transport/dashboard.html'

    def get(self, request):
        from handle.models import (
            BusTrackingSession,
            SchoolBus,
            StudentBusAssignment,
            member,
        )

        org = _get_org(request)
        buses = SchoolBus.objects.filter(org=org).select_related(
            'branch', 'driver',
        ).prefetch_related('student_assignments')
        active_sessions = {
            item.bus_id: item
            for item in BusTrackingSession.objects.filter(
                org=org, status='active',
            ).select_related('bus', 'driver')
        }
        return render(request, self.template_name, {
            'org': org,
            'buses': buses,
            'drivers': member.objects.filter(
                org=org, member_type='driver', status='active',
            ).order_by('name'),
            'students': member.objects.filter(
                org=org,
                member_type__in=('student', 'trainee'),
                status='active',
            ).order_by('name'),
            'assignments': StudentBusAssignment.objects.filter(
                org=org, status='active',
            ).select_related('student', 'bus').order_by('bus__name', 'student__name'),
            'active_sessions': active_sessions,
        })

    def post(self, request):
        from handle.models import SchoolBus, StudentBusAssignment, member

        org = _get_org(request)
        action = request.POST.get('action')
        if action == 'create_bus':
            name = (request.POST.get('name') or '').strip()
            registration = (
                request.POST.get('registration_number') or ''
            ).strip().upper()
            if not name or not registration:
                messages.error(request, 'Bus name and registration number are required.')
                return redirect('schooladmin:transport_dashboard')
            driver = member.objects.filter(
                pk=request.POST.get('driver_id'),
                org=org,
                member_type='driver',
                status='active',
            ).first()
            branch = Branch.objects.filter(
                pk=request.POST.get('branch_id'), org=org,
            ).first()
            try:
                capacity = max(1, min(int(request.POST.get('capacity', 1)), 200))
            except (TypeError, ValueError):
                capacity = 1
            _, created = SchoolBus.objects.update_or_create(
                org=org,
                registration_number=registration,
                defaults={
                    'name': name[:120],
                    'route_name': (
                        request.POST.get('route_name') or ''
                    ).strip()[:180],
                    'capacity': capacity,
                    'driver': driver,
                    'branch': branch,
                    'is_active': True,
                },
            )
            messages.success(
                request,
                'Bus created.' if created else 'Bus details updated.',
            )
        elif action == 'assign_student':
            student = member.objects.filter(
                pk=request.POST.get('student_id'),
                org=org,
                member_type__in=('student', 'trainee'),
                status='active',
            ).first()
            bus = SchoolBus.objects.filter(
                pk=request.POST.get('bus_id'), org=org, is_active=True,
            ).first()
            if student is None or bus is None:
                messages.error(request, 'Choose a valid student and bus.')
                return redirect('schooladmin:transport_dashboard')
            try:
                stop_lat = (
                    float(request.POST['stop_latitude'])
                    if request.POST.get('stop_latitude') else None
                )
                stop_lon = (
                    float(request.POST['stop_longitude'])
                    if request.POST.get('stop_longitude') else None
                )
            except (TypeError, ValueError):
                messages.error(request, 'Enter valid stop coordinates.')
                return redirect('schooladmin:transport_dashboard')
            if (
                (stop_lat is None) != (stop_lon is None)
                or (
                    stop_lat is not None
                    and not (
                        -90 <= stop_lat <= 90
                        and -180 <= stop_lon <= 180
                    )
                )
            ):
                messages.error(
                    request,
                    'Enter both stop coordinates within the valid GPS range.',
                )
                return redirect('schooladmin:transport_dashboard')
            with transaction.atomic():
                StudentBusAssignment.objects.filter(
                    student=student, status='active',
                ).update(status='inactive')
                StudentBusAssignment.objects.create(
                    org=org,
                    student=student,
                    bus=bus,
                    stop_name=(
                        request.POST.get('stop_name') or ''
                    ).strip()[:180],
                    stop_latitude=stop_lat,
                    stop_longitude=stop_lon,
                )
            messages.success(request, f'{student.name} assigned to {bus.name}.')
        else:
            messages.error(request, 'Unsupported transport action.')
        return redirect('schooladmin:transport_dashboard')


class QRSessionStatusView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'qr_attendance'
    def get(self, request, session_id):
        org = _get_org(request)
        session = get_object_or_404(QRAttendanceSession, id=session_id, org=org)
        session.refresh_status()

        logs = session.scan_logs.select_related('member').order_by('-scanned_at')[:20]
        logs_data = []
        for log in logs:
            logs_data.append({
                'member_name': log.member.name if log.member else '—',
                'classification': log.member.classification.name if log.member and log.member.classification else '—',
                'status': log.status,
                'scanned_at': log.scanned_at.strftime('%H:%M:%S'),
                'failure_reason': log.failure_reason,
            })

        return JsonResponse({
            'session_id': session.id,
            'status': session.status,
            'is_active': session.is_valid(),
            'total_scans': session.total_scans,
            'successful_scans': session.successful_scans,
            'expires_at_ts': (
                int(session.expires_at.timestamp() * 1000)
                if session.expires_at else None
            ),
            'recent_scans': logs_data,
        })


class QRCloseSessionView(AdminRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'qr_attendance'
    def post(self, request, session_id):
        org = _get_org(request)
        session = get_object_or_404(QRAttendanceSession, id=session_id, org=org)
        if session.status == 'active':
            session.status = 'closed'
            session.closed_at = tz.now()
            session.save(update_fields=['status', 'closed_at'])
        return JsonResponse({'status': 'closed'})


# ─── Timesheet Admin Views ────────────────────────────────────────────────────

from handle.models import Timesheet as _Timesheet, TimesheetEntry as _TSEntry


class TimesheetAdminListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'timesheet'
    template_name = 'admin/timesheet/list.html'

    def get(self, request):
        org = _get_org(request)
        if not org or not org.feature_timesheet:
            messages.error(request, "Timesheet module is not enabled.")
            return redirect('schooladmin:dashboard')

        qs = _Timesheet.objects.filter(org=org).select_related('member', 'approved_by')

        # Filters
        status_f    = request.GET.get('status', '')
        member_f    = request.GET.get('member', '')
        date_from   = request.GET.get('date_from', '')
        date_to     = request.GET.get('date_to', '')

        if status_f:
            qs = qs.filter(status=status_f)
        if member_f:
            qs = qs.filter(member_id=member_f)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)

        from handle.models import member as _member
        members = _member.objects.filter(org=org, status='active').order_by('name')
        nepali_enabled = getattr(org, 'nepali_date', False)
        timesheets = list(qs)
        if nepali_enabled:
            for ts in timesheets:
                ts.date_np = to_bs_display(ts.date)

        return render(request, self.template_name, {
            'org': org,
            'timesheets': timesheets,
            'members': members,
            'nepali_enabled': nepali_enabled,
            'status_f': status_f,
            'member_f': member_f,
            'date_from': date_from,
            'date_to': date_to,
        })


class TimesheetAdminDetailView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'timesheet'
    template_name = 'admin/timesheet/detail.html'

    def get(self, request, pk):
        org = _get_org(request)
        ts = get_object_or_404(_Timesheet, pk=pk, org=org)
        nepali_enabled = getattr(org, 'nepali_date', False)
        if nepali_enabled:
            ts.date_np = to_bs_display(ts.date)
        entries = ts.entries.all()

        if request.headers.get('HX-Request'):
            return render(request, 'admin/timesheet/_detail_partial.html', {
                'org': org, 'ts': ts, 'entries': entries,
                'nepali_enabled': nepali_enabled,
            })
        return render(request, self.template_name, {
            'org': org, 'ts': ts, 'entries': entries,
            'nepali_enabled': nepali_enabled,
        })


class TimesheetAdminApproveView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'timesheet'
    def post(self, request, pk):
        from django.utils import timezone as _tz2
        org = _get_org(request)
        ts = get_object_or_404(_Timesheet, pk=pk, org=org)
        if ts.status == 'submitted':
            ts.status = 'approved'
            ts.approved_by = request.user
            ts.approved_at = _tz2.now()
            ts.admin_comment = ''
            ts.save()
        if request.headers.get('HX-Request'):
            return render(request, 'admin/timesheet/_status_badge.html', {'ts': ts})
        messages.success(request, f"Timesheet for {ts.member.name} approved.")
        return redirect('schooladmin:timesheet_detail', pk=pk)


class TimesheetAdminRejectView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'timesheet'
    def post(self, request, pk):
        org = _get_org(request)
        ts = get_object_or_404(_Timesheet, pk=pk, org=org)
        comment = request.POST.get('comment', '').strip()
        if ts.status == 'submitted':
            ts.status = 'rejected'
            ts.admin_comment = comment
            ts.approved_by = None
            ts.approved_at = None
            ts.save()
        if request.headers.get('HX-Request'):
            return render(request, 'admin/timesheet/_status_badge.html', {'ts': ts})
        messages.warning(request, f"Timesheet for {ts.member.name} rejected.")
        return redirect('schooladmin:timesheet_detail', pk=pk)


# ── ID Card Generation ──────────────────────────────────────────────────────

from handle.models import IDCardTemplate as _IDCardTemplate
from handle.models import CertificateTemplate as _CertificateTemplate
from handle.forms import (
    IDCardTemplateForm as _IDCardTemplateForm,
    CertificateTemplateForm as _CertificateTemplateForm,
    sanitize_certificate_html as _sanitize_certificate_html,
)


class IDCardSettingsView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """
    Manages the org's saved ID card designs — one IDCardTemplate row per
    design key (up to the 6 built-in designs). ?design=<key> selects which
    one the form is editing; defaults to the first design key.
    """
    required_feature = 'id_cards'
    template_name = 'admin/idcard/settings.html'

    @staticmethod
    def _safe_design_key(raw_key):
        choices = dict(_IDCardTemplate.DESIGN_CHOICES)
        return raw_key if raw_key in choices else _IDCardTemplate.DESIGN_CHOICES[0][0]

    def get(self, request):
        org = _get_org(request)
        design_key = self._safe_design_key(request.GET.get('design'))
        has_any_template = _IDCardTemplate.objects.filter(org=org).exists()
        template, created = _IDCardTemplate.objects.get_or_create(
            org=org,
            name=design_key,
            defaults={'is_default': not has_any_template},
        )
        preview_member = member.objects.filter(org=org, status='active').order_by('name').first()
        return render(request, self.template_name, {
            'org': org,
            'template': template,
            'form': _IDCardTemplateForm(instance=template),
            'design_choices': _IDCardTemplate.DESIGN_CHOICES,
            'selected_design': design_key,
            'saved_templates': {t.name: t for t in _IDCardTemplate.objects.filter(org=org)},
            'preview_member': preview_member,
            'preview_course_names': ', '.join(course.name for course in preview_member.courses.all()) if preview_member else '',
            'configured_design_count': _IDCardTemplate.objects.filter(org=org).count(),
            'active_member_count': member.objects.filter(org=org, status='active').count(),
        })

    def post(self, request):
        org = _get_org(request)
        design_key = self._safe_design_key(request.POST.get('design'))
        template, _ = _IDCardTemplate.objects.get_or_create(org=org, name=design_key)
        form = _IDCardTemplateForm(request.POST, request.FILES, instance=template)
        if form.is_valid():
            with transaction.atomic():
                saved = form.save(commit=False)
                saved.org = org
                saved.name = design_key
                saved.show_blood_group = False
                saved.show_signature = False
                saved.show_valid_until = False
                # Always keep one deterministic default. This prevents the
                # generation screen from changing designs unexpectedly.
                if not saved.is_default and not _IDCardTemplate.objects.filter(
                    org=org, is_default=True
                ).exclude(pk=saved.pk).exists():
                    saved.is_default = True
                saved.save()
                if saved.is_default:
                    _IDCardTemplate.objects.filter(org=org).exclude(pk=saved.pk).update(is_default=False)
            messages.success(request, f"'{saved.get_name_display()}' design saved.")
        else:
            messages.error(request, "Please correct the highlighted ID-card design fields.")
            return render(request, self.template_name, {
                'org': org,
                'template': template,
                'form': form,
                'design_choices': _IDCardTemplate.DESIGN_CHOICES,
                'selected_design': design_key,
                'saved_templates': {t.name: t for t in _IDCardTemplate.objects.filter(org=org)},
                'preview_member': member.objects.filter(org=org, status='active').order_by('name').first(),
                'configured_design_count': _IDCardTemplate.objects.filter(org=org).count(),
                'active_member_count': member.objects.filter(org=org, status='active').count(),
            }, status=400)
        return redirect(f"{reverse('schooladmin:idcard_settings')}?design={design_key}")


class IDCardGenerateView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'id_cards'
    template_name = 'admin/idcard/generate.html'

    def get(self, request):
        org = _get_org(request)
        existing_by_name = {
            item.name: item for item in _IDCardTemplate.objects.filter(org=org)
        }
        available_templates = []
        for design_key, _label in _IDCardTemplate.DESIGN_CHOICES:
            template_item = existing_by_name.get(design_key)
            if template_item is None:
                template_item = _IDCardTemplate.objects.create(
                    org=org,
                    name=design_key,
                    is_default=not existing_by_name and not available_templates,
                )
            available_templates.append(template_item)

        design_key = request.GET.get('design')
        if design_key and design_key in dict(_IDCardTemplate.DESIGN_CHOICES):
            template = next(item for item in available_templates if item.name == design_key)
        else:
            template = next((t for t in available_templates if t.is_default), available_templates[0])

        members_qs = member.objects.filter(org=org, status='active').select_related(
            'classification', 'branch', 'section'
        ).order_by('name')
        classification_id = request.GET.get('classification')
        branch_id = request.GET.get('branch')
        section_id = request.GET.get('section')
        member_type = request.GET.get('member_type', '').strip()
        search = request.GET.get('q', '').strip()

        if classification_id:
            members_qs = members_qs.filter(classification_id=classification_id)
        if branch_id:
            members_qs = members_qs.filter(branch_id=branch_id)
        if section_id:
            members_qs = members_qs.filter(section_id=section_id)
        if member_type in dict(member.MEMBER_TYPE_CHOICES):
            members_qs = members_qs.filter(member_type=member_type)
        else:
            member_type = ''
        if search:
            members_qs = members_qs.filter(
                Q(name__icontains=search)
                | Q(card__icontains=search)
                | Q(roll_number__icontains=search)
            )

        # Only actually render cards once the admin has picked a filter or
        # explicitly asked for everyone — avoids accidentally dumping the
        # whole org's cards on first page load.
        generate = request.GET.get('generate') == '1'
        result_count = members_qs.count()
        members = list(members_qs[:500]) if generate else []
        result_limited = generate and result_count > 500

        if generate and (template.show_qr_code or template.show_barcode):
            _attach_idcard_codes(members, template, org)

        width_mm, height_mm = template.card_dimensions_mm()
        photo_w_mm, photo_h_mm = template.photo_dimensions_mm()

        return render(request, self.template_name, {
            'org': org,
            'template': template,
            'design_path': f'admin/idcard/designs/{template.name}.html',
            'available_templates': available_templates,
            'selected_design': template.name,
            'classifications': Classification.objects.filter(org=org, status='active'),
            'branches': Branch.objects.filter(org=org, status='active'),
            'sections': _Section.objects.filter(org=org, status='active'),
            'selected_classification': classification_id,
            'selected_branch': branch_id,
            'selected_section': section_id,
            'selected_member_type': member_type,
            'member_type_choices': member.MEMBER_TYPE_CHOICES,
            'search': search,
            'members': members,
            'result_count': result_count,
            'result_limited': result_limited,
            'members_with_photo': sum(1 for item in members if item.photo),
            'generate': generate,
            'card_width_mm': width_mm,
            'card_height_mm': height_mm,
            'photo_width_mm': photo_w_mm,
            'photo_height_mm': photo_h_mm,
        })


def _attach_idcard_codes(members, template, org):
    """Generate QR/barcode PNGs (base64) for each member, attached as attrs the design templates read."""
    import base64
    import io
    if template.show_qr_code:
        import qrcode
        for m in members:
            buf = io.BytesIO()
            qrcode.make(f"MEMBER-{org.id}-{m.id}").save(buf, format='PNG')
            m.idc_qr_b64 = base64.b64encode(buf.getvalue()).decode()
    if template.show_barcode:
        try:
            import barcode
            from barcode.writer import ImageWriter
            for m in members:
                buf = io.BytesIO()
                barcode.get('code128', str(m.id), writer=ImageWriter()).write(
                    buf, options={'write_text': False, 'module_height': 8.0}
                )
                m.idc_barcode_b64 = base64.b64encode(buf.getvalue()).decode()
        except ImportError:
            pass


# ── Certificate Design & Bulk Generation ───────────────────────────────────

class CertificateSettingsView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'id_cards'
    template_name = 'admin/certificates/settings.html'

    def _get_selected(self, request, org):
        raw_id = request.GET.get('template')
        if raw_id:
            return get_object_or_404(_CertificateTemplate, pk=raw_id, org=org)
        if request.GET.get('new') == '1':
            return None
        return _CertificateTemplate.objects.filter(org=org).order_by('-is_default', 'name').first()

    def _context(self, org, selected, form, preview_member):
        from django.utils.html import escape

        raw_preview_body = (
            form.data.get('body_html') if form.is_bound
            else (selected.body_html if selected else _CertificateTemplate._meta.get_field('body_html').get_default())
        )
        editor_body_html = _sanitize_certificate_html(raw_preview_body)
        preview_body_html = editor_body_html
        if preview_member:
            preview_values = {
                '[[member_name]]': preview_member.name or '',
                '[[member_id]]': str(preview_member.pk),
                '[[roll_number]]': preview_member.roll_number or '',
                '[[classification]]': preview_member.classification.name if preview_member.classification else '',
                '[[section]]': preview_member.section.name if preview_member.section else '',
                '[[course]]': ', '.join(course.name for course in preview_member.courses.all()),
                '[[organization]]': org.name or '',
                '[[issue_date]]': timezone.localdate().strftime('%B %d, %Y'),
                '[[certificate_number]]': 'CERT-PREVIEW-0001',
            }
            for token, value in preview_values.items():
                preview_body_html = preview_body_html.replace(token, str(escape(value)))
        return {
            'org': org,
            'template': selected,
            'form': form,
            'templates': _CertificateTemplate.objects.filter(org=org),
            'preview_member': preview_member,
            'editor_body_html': editor_body_html,
            'preview_body_html': preview_body_html,
            'certificate_types': _CertificateTemplate.TYPE_CHOICES,
            'supported_tokens': (
                '[[member_name]]', '[[member_id]]', '[[roll_number]]',
                '[[classification]]', '[[section]]', '[[course]]',
                '[[organization]]', '[[issue_date]]', '[[certificate_number]]',
            ),
        }

    def get(self, request):
        org = _get_org(request)
        selected = self._get_selected(request, org)
        preview_member = member.objects.filter(org=org, status='active').select_related(
            'classification', 'section'
        ).prefetch_related('courses').order_by('name').first()
        initial = None
        if selected is None:
            initial = {
                'name': 'New Certificate Template',
                'certificate_type': 'completion',
                'title': 'Certificate of Completion',
                'subtitle': 'This certificate is proudly presented to',
            }
        form = _CertificateTemplateForm(instance=selected, initial=initial, org=org)
        return render(request, self.template_name, self._context(org, selected, form, preview_member))

    def post(self, request):
        org = _get_org(request)
        template_id = request.POST.get('template_id')
        selected = get_object_or_404(_CertificateTemplate, pk=template_id, org=org) if template_id else None
        form = _CertificateTemplateForm(
            request.POST, request.FILES, instance=selected, org=org,
        )
        preview_member = member.objects.filter(org=org, status='active').select_related(
            'classification', 'section'
        ).prefetch_related('courses').order_by('name').first()
        if not form.is_valid():
            messages.error(request, 'Please correct the highlighted certificate design fields.')
            return render(
                request, self.template_name,
                self._context(org, selected, form, preview_member), status=400,
            )

        with transaction.atomic():
            saved = form.save(commit=False)
            saved.org = org
            if not saved.pk:
                saved.created_by = request.user
            if not saved.is_default and not _CertificateTemplate.objects.filter(
                org=org, is_default=True
            ).exclude(pk=saved.pk).exists():
                saved.is_default = True
            saved.save()
            if saved.is_default:
                _CertificateTemplate.objects.filter(org=org).exclude(pk=saved.pk).update(is_default=False)
        messages.success(request, f"Certificate template '{saved.name}' saved.")
        return redirect(f"{reverse('schooladmin:certificate_settings')}?template={saved.pk}")


def _render_certificate_body(template, certificate_member, org, issue_date, certificate_number):
    """Replace a small explicit token allow-list inside already-sanitized rich text."""
    from django.utils.html import escape
    from django.utils.safestring import mark_safe

    course_names = ', '.join(c.name for c in certificate_member.courses.all())
    values = {
        '[[member_name]]': certificate_member.name or '',
        '[[member_id]]': str(certificate_member.pk),
        '[[roll_number]]': certificate_member.roll_number or '',
        '[[classification]]': certificate_member.classification.name if certificate_member.classification else '',
        '[[section]]': certificate_member.section.name if certificate_member.section else '',
        '[[course]]': course_names,
        '[[organization]]': org.name or '',
        '[[issue_date]]': issue_date.strftime('%B %d, %Y'),
        '[[certificate_number]]': certificate_number,
    }
    rendered = _sanitize_certificate_html(template.body_html)
    for token, value in values.items():
        rendered = rendered.replace(token, str(escape(value)))
    return mark_safe(rendered)


class CertificateGenerateView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'id_cards'
    template_name = 'admin/certificates/generate.html'

    def get(self, request):
        from django.utils.dateparse import parse_date

        org = _get_org(request)
        available_templates = _CertificateTemplate.objects.filter(org=org, is_active=True)
        template_id = request.GET.get('template')
        if template_id:
            template = get_object_or_404(available_templates, pk=template_id)
        else:
            template = available_templates.filter(is_default=True).first() or available_templates.first()

        members_qs = member.objects.filter(org=org, status='active').select_related(
            'classification', 'branch', 'section'
        ).prefetch_related('courses').order_by('name')
        classification_id = request.GET.get('classification')
        branch_id = request.GET.get('branch')
        section_id = request.GET.get('section')
        member_type = request.GET.get('member_type', '').strip()
        search = request.GET.get('q', '').strip()

        if classification_id:
            members_qs = members_qs.filter(classification_id=classification_id)
        if branch_id:
            members_qs = members_qs.filter(branch_id=branch_id)
        if section_id:
            members_qs = members_qs.filter(section_id=section_id)
        if member_type in dict(member.MEMBER_TYPE_CHOICES):
            members_qs = members_qs.filter(member_type=member_type)
        else:
            member_type = ''
        if search:
            members_qs = members_qs.filter(
                Q(name__icontains=search)
                | Q(card__icontains=search)
                | Q(roll_number__icontains=search)
            )

        issue_date = parse_date(request.GET.get('issue_date', '')) or timezone.localdate()
        generate = request.GET.get('generate') == '1' and template is not None
        result_count = members_qs.count()
        members = list(members_qs[:500]) if generate else []
        if template:
            for certificate_member in members:
                number = f"{template.serial_prefix or 'CERT'}-{issue_date:%Y%m%d}-{certificate_member.pk}"
                certificate_member.certificate_number = number
                certificate_member.certificate_body_html = _render_certificate_body(
                    template, certificate_member, org, issue_date, number,
                )

        return render(request, self.template_name, {
            'org': org,
            'template': template,
            'available_templates': available_templates,
            'classifications': Classification.objects.filter(org=org, status='active'),
            'branches': Branch.objects.filter(org=org, status='active'),
            'sections': _Section.objects.filter(org=org, status='active'),
            'selected_classification': classification_id,
            'selected_branch': branch_id,
            'selected_section': section_id,
            'selected_member_type': member_type,
            'member_type_choices': member.MEMBER_TYPE_CHOICES,
            'search': search,
            'issue_date': issue_date,
            'members': members,
            'result_count': result_count,
            'result_limited': generate and result_count > 500,
            'generate': generate,
        })


# ============================================================
# FIELD VISITS — Admin Views
# ============================================================

from handle.models import FieldVisit, FieldVisitReport, Client, ClientFollowUp


class FieldVisitListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'field_visits'
    template_name = 'admin/field_visits/list.html'

    def get(self, request):
        org = _task_org(request)
        visits = (
            FieldVisit.objects.filter(org=org)
            .select_related('member', 'report', 'client', 'created_by')
            .prefetch_related('follow_ups__created_by')
            .order_by('-visited_at')
        )

        member_id = request.GET.get('member', '')
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')
        if member_id:
            visits = visits.filter(member_id=member_id)
        if from_date:
            visits = visits.filter(visited_at__date__gte=from_date)
        if to_date:
            visits = visits.filter(visited_at__date__lte=to_date)

        members = member.objects.filter(org=org, status='active').order_by('name')
        clients = Client.objects.filter(org=org, is_active=True).order_by('client_org_name')
        nepali_enabled = org.nepali_date
        visits = list(visits[:200])
        if nepali_enabled:
            for v in visits:
                v.visited_at_np = to_bs_display(v.visited_at.date())
        ctx = dict(
            org=org, visits=visits, members=members, clients=clients,
            today_val=timezone.localdate().isoformat(),
            nepali_enabled=nepali_enabled,
            filters=dict(member=member_id, from_date=from_date, to_date=to_date),
        )
        return render(request, self.template_name, ctx)


class FieldVisitDetailView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'field_visits'
    template_name = 'admin/field_visits/detail.html'

    def get(self, request, pk):
        org = _task_org(request)
        visit = get_object_or_404(
            FieldVisit.objects.select_related('member', 'client', 'report'),
            pk=pk, org=org
        )
        linked_follow_up_ids = list(visit.follow_ups.values_list('id', flat=True))
        other_client_follow_ups = []
        if visit.client:
            other_client_follow_ups = list(
                visit.client.follow_ups.exclude(id__in=linked_follow_up_ids)
                .order_by('-follow_up_date', '-id')[:10]
            )
        follow_ups = list(visit.follow_ups.select_related(
            'created_by', 'visited_by',
        ).order_by('-follow_up_date', '-id'))

        nepali_enabled = org.nepali_date
        if nepali_enabled:
            visit.visited_at_np = to_bs_display(visit.visited_at.date())
            for fu in follow_ups + other_client_follow_ups:
                fu.date_np = to_bs_display(fu.follow_up_date)
                fu.next_date_np = to_bs_display(fu.next_follow_up_date)

        ctx = dict(
            org=org,
            visit=visit,
            nepali_enabled=nepali_enabled,
            follow_ups=follow_ups,
            other_client_follow_ups=other_client_follow_ups,
            existing_records=visit.attendance_records.order_by('scanned_time'),
            default_start=visit.member.effective_shift_start(visit.visited_at.date()),
            default_end=visit.member.effective_shift_end(visit.visited_at.date()),
        )
        return render(request, self.template_name, ctx)

    def post(self, request, pk):
        org = _task_org(request)
        visit = get_object_or_404(FieldVisit, pk=pk, org=org)
        action = request.POST.get('action')
        review_note = request.POST.get('review_note', '').strip()

        if action == 'reject':
            visit.status = 'rejected'
            visit.reviewed_by = request.user
            visit.reviewed_at = timezone.now()
            visit.review_note = review_note
            visit.save()
            messages.success(request, "Field visit rejected.")
            return redirect('schooladmin:field_visit_detail', pk=visit.pk)

        if action == 'approve':
            visit_date = visit.visited_at.date()
            shift_mode = request.POST.get('shift_mode', 'full')

            try:
                if shift_mode == 'custom':
                    checkin_str = request.POST.get('checkin_time', '').strip()
                    checkout_str = request.POST.get('checkout_time', '').strip()
                    if not checkin_str:
                        messages.error(request, "Check-in time is required for a custom entry.")
                        return redirect('schooladmin:field_visit_detail', pk=visit.pk)
                    checkin_dt = timezone.make_aware(
                        datetime.datetime.strptime(f"{visit_date} {checkin_str}", "%Y-%m-%d %H:%M")
                    )
                    checkout_dt = None
                    if checkout_str:
                        checkout_dt = timezone.make_aware(
                            datetime.datetime.strptime(f"{visit_date} {checkout_str}", "%Y-%m-%d %H:%M")
                        )
                else:
                    start_t = visit.member.effective_shift_start(visit_date)
                    end_t = visit.member.effective_shift_end(visit_date)
                    if not start_t or not end_t:
                        messages.error(request, "This member has no shift scheduled for that date — use a custom time instead.")
                        return redirect('schooladmin:field_visit_detail', pk=visit.pk)
                    checkin_dt = timezone.make_aware(datetime.datetime.combine(visit_date, start_t))
                    checkout_dt = timezone.make_aware(datetime.datetime.combine(visit_date, end_t))
                    if checkout_dt <= checkin_dt:
                        checkout_dt += datetime.timedelta(days=1)
            except ValueError:
                messages.error(request, "Invalid time format.")
                return redirect('schooladmin:field_visit_detail', pk=visit.pk)

            AttendanceRecord.objects.create(
                mem=visit.member, org=org, scanned_time=checkin_dt,
                attendance_method='field_visit', field_visit=visit,
            )
            if checkout_dt:
                AttendanceRecord.objects.create(
                    mem=visit.member, org=org, scanned_time=checkout_dt,
                    attendance_method='field_visit', field_visit=visit,
                )

            visit.status = 'approved'
            visit.reviewed_by = request.user
            visit.reviewed_at = timezone.now()
            visit.review_note = review_note
            visit.save()
            messages.success(request, f"Field visit approved — attendance recorded for {visit.member.name}.")
            return redirect('schooladmin:field_visit_detail', pk=visit.pk)

        messages.error(request, "Unknown action.")
        return redirect('schooladmin:field_visit_detail', pk=visit.pk)


class FieldVisitManualAddView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Lets an admin log a field visit on a staff member's behalf from
    scratch (no prior mobile submission) — entering check-in/check-out time
    directly, same as the "approve" flow's custom-time entry, but the visit
    is created already-approved with attendance recorded immediately.

    Supports logging a single day (leave End Date blank — original
    behaviour) or backfilling a whole date range in one submission: one
    FieldVisit + its AttendanceRecord(s) is created per calendar day in
    [start_date, end_date], all using the same check-in/check-out time,
    client, purpose, destination and note."""
    required_feature = 'field_visits'
    MAX_RANGE_DAYS = 62  # guards against a fat-fingered multi-year range

    def post(self, request):
        org = _task_org(request)

        member_id = request.POST.get('member_id')
        memb = get_object_or_404(member, pk=member_id, org=org)

        client_id = request.POST.get('client_id', '').strip()
        client_obj = get_object_or_404(Client, pk=client_id, org=org) if client_id else None

        start_str = request.POST.get('visit_date', '').strip()
        end_str = request.POST.get('end_date', '').strip()
        try:
            start_date = (
                datetime.datetime.strptime(start_str, "%Y-%m-%d").date()
                if start_str else timezone.localdate()
            )
            end_date = (
                datetime.datetime.strptime(end_str, "%Y-%m-%d").date()
                if end_str else start_date
            )
        except ValueError:
            messages.error(request, "Invalid visit date.")
            return redirect('schooladmin:field_visit_list')

        if end_date < start_date:
            messages.error(request, "End date can't be before the start date.")
            return redirect('schooladmin:field_visit_list')

        span_days = (end_date - start_date).days + 1
        if span_days > self.MAX_RANGE_DAYS:
            messages.error(request, f"That range is {span_days} days — please keep manual backfills under {self.MAX_RANGE_DAYS} days at a time.")
            return redirect('schooladmin:field_visit_list')

        checkin_str = request.POST.get('checkin_time', '').strip()
        checkout_str = request.POST.get('checkout_time', '').strip()
        if not checkin_str:
            messages.error(request, "Check-in time is required to log a manual visit.")
            return redirect('schooladmin:field_visit_list')

        purpose = request.POST.get('purpose', '').strip()
        destination = request.POST.get('destination', '').strip()
        note = request.POST.get('note', '').strip()

        visit = None
        created_days = 0
        try:
            with transaction.atomic():
                for offset in range(span_days):
                    day = start_date + datetime.timedelta(days=offset)
                    checkin_dt = timezone.make_aware(
                        datetime.datetime.strptime(f"{day} {checkin_str}", "%Y-%m-%d %H:%M")
                    )
                    checkout_dt = None
                    if checkout_str:
                        checkout_dt = timezone.make_aware(
                            datetime.datetime.strptime(f"{day} {checkout_str}", "%Y-%m-%d %H:%M")
                        )

                    visit = FieldVisit.objects.create(
                        org=org, member=memb, client=client_obj,
                        purpose=purpose, destination=destination,
                        latitude=None, longitude=None,
                        visit_state='completed', status='approved',
                        created_by=request.user, reviewed_by=request.user,
                        reviewed_at=timezone.now(),
                        review_note='Manually logged by an admin.',
                    )
                    AttendanceRecord.objects.create(
                        mem=memb, org=org, scanned_time=checkin_dt,
                        attendance_method='field_visit', field_visit=visit,
                    )
                    if checkout_dt:
                        AttendanceRecord.objects.create(
                            mem=memb, org=org, scanned_time=checkout_dt,
                            attendance_method='field_visit', field_visit=visit,
                        )
                    if note:
                        FieldVisitReport.objects.create(visit=visit, note=note)
                    created_days += 1
        except ValueError:
            messages.error(request, "Invalid time format.")
            return redirect('schooladmin:field_visit_list')

        if created_days > 1:
            messages.success(request, f"Field visits logged and attendance recorded for {memb.name} across {created_days} days ({start_date} to {end_date}).")
            return redirect('schooladmin:field_visit_list')

        messages.success(request, f"Field visit logged and attendance recorded for {memb.name}.")
        return redirect('schooladmin:field_visit_detail', pk=visit.pk)


# ============================================================
# CLIENT FOLLOW-UP — Admin Views
# ============================================================

def _clients_due_qs(org, priority=''):
    from django.db.models import Case, IntegerField, OuterRef, Subquery, Value, When
    today = datetime.date.today()
    # Only an OPEN follow-up's next-date counts as "due" — closing a case
    # (see ClientDetailView's close_followup action) removes it from this
    # queryset, which is the whole point of closing it.
    latest = ClientFollowUp.objects.filter(client=OuterRef('pk'), status='open').order_by('-follow_up_date', '-id')
    queryset = (
        Client.objects.filter(org=org, is_active=True)
        .annotate(latest_next=Subquery(latest.values('next_follow_up_date')[:1]))
        .filter(latest_next__isnull=False, latest_next__lte=today)
        .annotate(priority_order=Case(
            When(priority='high', then=Value(0)),
            When(priority='medium', then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        ))
        .order_by('priority_order', 'latest_next', 'client_org_name')
    )
    if priority in dict(Client.PRIORITY_CHOICES):
        queryset = queryset.filter(priority=priority)
    return queryset


def _create_follow_up(
    org, client, user, feedback, follow_up_date, next_follow_up_date,
    priority='medium', field_visit=None,
):
    memb = getattr(getattr(user, 'staff', None), 'member', None)
    if priority not in dict(ClientFollowUp.PRIORITY_CHOICES):
        priority = 'medium'
    return ClientFollowUp.objects.create(
        client=client, org=org, visited_by=memb, feedback=feedback,
        priority=priority,
        follow_up_date=follow_up_date, next_follow_up_date=next_follow_up_date,
        field_visit=field_visit,
        created_by=user,
    )


class ClientListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'clients'
    template_name = 'admin/clients/client_list.html'

    def get(self, request):
        org = _task_org(request)
        qs = Client.objects.filter(org=org, is_active=True)

        client_number = request.GET.get('client_number', '').strip()
        client_org = request.GET.get('client_org', '').strip()
        status = request.GET.get('status', '').strip()
        priority = request.GET.get('priority', '').strip()
        if client_number:
            qs = qs.filter(client_number__icontains=client_number)
        if client_org:
            qs = qs.filter(client_org_name__icontains=client_org)
        if status in dict(Client.STATUS_CHOICES):
            qs = qs.filter(status=status)
        if priority in dict(Client.PRIORITY_CHOICES):
            qs = qs.filter(priority=priority)

        qs = qs.select_related('branch').prefetch_related('follow_ups')
        clients = [{'client': c, 'count': c.follow_up_count(), 'latest': c.latest_follow_up()} for c in qs]
        due_count = _clients_due_qs(org).count()
        ctx = dict(
            org=org, clients=clients, due_count=due_count,
            filters=dict(client_number=client_number, client_org=client_org, status=status, priority=priority),
            status_choices=Client.STATUS_CHOICES, priority_choices=Client.PRIORITY_CHOICES,
        )
        return render(request, self.template_name, ctx)


class ClientCreateView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'clients'
    template_name = 'admin/clients/create_client.html'

    def get(self, request):
        org = _task_org(request)
        return render(request, self.template_name, {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'status_choices': Client.STATUS_CHOICES,
            'priority_choices': Client.PRIORITY_CHOICES,
        })

    def post(self, request):
        org = _task_org(request)
        client_number = request.POST.get('client_number', '').strip()
        client_org_name = request.POST.get('client_org_name', '').strip()

        if not client_org_name:
            messages.error(request, "Client organization name is required.")
            return redirect('schooladmin:create_client')
        if client_number and Client.objects.filter(org=org, client_number=client_number).exists():
            messages.error(request, f"Client number '{client_number}' already exists.")
            return redirect('schooladmin:create_client')

        status = request.POST.get('status', 'inquiry')
        priority = request.POST.get('priority', 'medium')
        if status not in dict(Client.STATUS_CHOICES):
            status = 'inquiry'
        if priority not in dict(Client.PRIORITY_CHOICES):
            priority = 'medium'
        org_branches = Branch.objects.filter(org=org, status='active')
        branch = org_branches.filter(pk=request.POST.get('branch')).first()
        if org_branches.exists() and not branch:
            messages.error(request, "Branch selection is required.")
            return redirect('schooladmin:create_client')

        client = Client.create_for_org(
            org=org, client_number=client_number, client_org_name=client_org_name,
            branch=branch, priority=priority,
            contact_person=request.POST.get('contact_person', '').strip(),
            phone=request.POST.get('phone') or None,
            email=request.POST.get('email') or None,
            address=request.POST.get('address', '').strip(),
            website=request.POST.get('website', '').strip(),
            industry=request.POST.get('industry', '').strip(),
            status=status,
            notes=request.POST.get('notes', '').strip(),
            billing_cycle=request.POST.get('billing_cycle') or None,
            billing_amount=request.POST.get('billing_amount') or None,
            contract_start=request.POST.get('contract_start') or None,
            contract_end=request.POST.get('contract_end') or None,
            next_billing_date=request.POST.get('next_billing_date') or None,
            monthly_target=request.POST.get('monthly_target') or None,
            yearly_target=request.POST.get('yearly_target') or None,
            created_by=request.user,
        )
        messages.success(request, f"Customer '{client_org_name}' created.")
        return redirect('schooladmin:client_detail', pk=client.pk)


class ClientDetailView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'clients'
    template_name = 'admin/clients/client_detail.html'

    def get(self, request, pk):
        org = _task_org(request)
        client = get_object_or_404(Client, pk=pk, org=org)
        follow_ups = client.follow_ups.select_related(
            'visited_by', 'created_by', 'field_visit',
        ).all()
        fu_status = request.GET.get('fu_status', 'open').strip()
        fu_priority = request.GET.get('fu_priority', '').strip()
        fu_from = request.GET.get('fu_from', '').strip()
        fu_to = request.GET.get('fu_to', '').strip()
        if fu_status in dict(ClientFollowUp.STATUS_CHOICES):
            follow_ups = follow_ups.filter(status=fu_status)
        if fu_priority in dict(Client.PRIORITY_CHOICES):
            follow_ups = follow_ups.filter(priority=fu_priority)
        if fu_from:
            follow_ups = follow_ups.filter(follow_up_date__gte=fu_from)
        if fu_to:
            follow_ups = follow_ups.filter(follow_up_date__lte=fu_to)
        fu_filters = dict(status=fu_status, priority=fu_priority, from_date=fu_from, to_date=fu_to)
        bills = client.customer_bills.select_related('branch').prefetch_related('payments').all()
        contracts = client.contracts.all()
        proposals = client.proposals.all()
        documents = client.documents.all()
        today = datetime.date.today()
        # billing reminder: next_billing_date within 7 days
        billing_alert = (
            client.next_billing_date and
            datetime.timedelta(0) <= (client.next_billing_date - today) <= datetime.timedelta(days=7)
        )
        bill_totals = bills.exclude(status='cancelled').aggregate(
            invoice_amount=Sum('amount'), paid=Sum('paid_amount'),
        )
        total_invoice_amount = bill_totals['invoice_amount'] or Decimal('0')
        total_paid = bill_totals['paid'] or Decimal('0')
        total_outstanding = max(total_invoice_amount - total_paid, Decimal('0'))
        total_overdue = sum(
            (bill.remaining_amount for bill in bills if bill.is_overdue),
            Decimal('0'),
        )
        payment_history = client.bill_payments.select_related(
            'bill', 'branch', 'income_transaction', 'created_by',
        ).order_by('-payment_date', '-id')

        nepali_enabled = org.nepali_date
        follow_ups = list(follow_ups)
        if nepali_enabled:
            for fu in follow_ups:
                fu.date_np = to_bs_display(fu.follow_up_date)
                fu.next_date_np = to_bs_display(fu.next_follow_up_date)

        ctx = dict(
            org=org, client=client, follow_ups=follow_ups,
            fu_filters=fu_filters, fu_status_choices=ClientFollowUp.STATUS_CHOICES,
            nepali_enabled=nepali_enabled,
            bills=bills, contracts=contracts, proposals=proposals, documents=documents,
            billing_alert=billing_alert, today=today,
            total_billed=total_paid, total_paid=total_paid,
            total_invoice_amount=total_invoice_amount,
            total_outstanding=total_outstanding, total_overdue=total_overdue,
            total_invoices=bills.exclude(status='cancelled').count(),
            payment_history=payment_history,
            branches=Branch.objects.filter(org=org, status='active'),
            status_choices=Client.STATUS_CHOICES,
            priority_choices=Client.PRIORITY_CHOICES,
        )

        # Sales tab — Sale lives behind the 'stock' feature, not 'clients',
        # so this is additionally gated (org may have Clients without Stock).
        if has_feature(org, 'stock'):
            ctx.update({
                'has_sales': True,
                'sales': client.sales.order_by('-sale_date'),
                'sale_payments': client.sale_payments.order_by('-payment_date'),
                'sales_returns': client.sales_returns.order_by('-return_date'),
                'total_sales': client.total_sales(),
                'sales_outstanding': client.outstanding_balance(),
                'recent_transactions': client.recent_transactions(),
            })
        return render(request, self.template_name, ctx)

    def post(self, request, pk):
        org = _task_org(request)
        client = get_object_or_404(Client, pk=pk, org=org)
        action = request.POST.get('action', 'follow_up')

        if action == 'update_info':
            org_branches = Branch.objects.filter(org=org, status='active')
            branch = org_branches.filter(pk=request.POST.get('branch')).first()
            if org_branches.exists() and not branch:
                messages.error(request, "Branch selection is required.")
                return redirect(f"{reverse('schooladmin:client_detail', kwargs={'pk': client.pk})}?tab=info")

            client.client_org_name = request.POST.get('client_org_name', client.client_org_name).strip()
            client.contact_person = request.POST.get('contact_person', client.contact_person).strip()
            client.phone = request.POST.get('phone') or None
            client.email = request.POST.get('email') or None
            client.address = request.POST.get('address', client.address).strip()
            client.website = request.POST.get('website', client.website).strip()
            client.industry = request.POST.get('industry', client.industry).strip()
            status = request.POST.get('status', client.status)
            priority = request.POST.get('priority', client.priority)
            if status in dict(Client.STATUS_CHOICES):
                client.status = status
            if priority in dict(Client.PRIORITY_CHOICES):
                client.priority = priority
            client.branch = branch
            client.notes = request.POST.get('notes', client.notes).strip()
            client.billing_cycle = request.POST.get('billing_cycle') or None
            client.billing_amount = request.POST.get('billing_amount') or None
            client.contract_start = request.POST.get('contract_start') or None
            client.contract_end = request.POST.get('contract_end') or None
            client.next_billing_date = request.POST.get('next_billing_date') or None
            client.monthly_target = request.POST.get('monthly_target') or None
            client.yearly_target = request.POST.get('yearly_target') or None
            client.save()
            messages.success(request, "Customer info updated.")

        elif action == 'add_bill':
            if not client.can_be_billed:
                messages.error(request, "Billing is unavailable for an Inquiry. Change the client status to Customer first.")
                return redirect(f"{reverse('schooladmin:client_detail', kwargs={'pk': client.pk})}?tab=bills")
            inv = request.POST.get('invoice_number', '').strip()
            amount = request.POST.get('amount')
            issue_date = request.POST.get('issue_date') or datetime.date.today()
            due_date = request.POST.get('due_date')
            try:
                amount_decimal = Decimal(str(amount))
            except (InvalidOperation, TypeError):
                amount_decimal = None
            if amount_decimal is None or amount_decimal <= 0 or not due_date:
                messages.error(request, "Amount and due date are required for a bill.")
            elif inv and CustomerBill.objects.filter(org=org, invoice_number=inv).exists():
                messages.error(request, f"Invoice number '{inv}' already exists in this organisation.")
            else:
                bill = CustomerBill.objects.create(
                    client=client, org=org, branch=client.branch,
                    invoice_number=inv or f"INV-{client.client_number or client.pk}-{timezone.now().strftime('%Y%m%d%H%M%S%f')}",
                    amount=amount_decimal,
                    description=request.POST.get('description', ''),
                    issue_date=issue_date, due_date=due_date,
                    status='draft' if request.POST.get('bill_status') == 'draft' else 'unpaid',
                    created_by=request.user,
                )
                if request.FILES.get('bill_image'):
                    bill.bill_image = request.FILES['bill_image']
                    bill.save(update_fields=['bill_image'])
                messages.success(request, "Bill added.")

        elif action == 'update_bill_status':
            bill_id = request.POST.get('bill_id')
            new_status = request.POST.get('bill_status')
            bill = get_object_or_404(CustomerBill, pk=bill_id, client=client, org=org)
            if new_status in ('paid', 'partial'):
                messages.error(request, "Paid status is calculated from recorded payments and cannot be set manually.")
            elif new_status not in ('draft', 'unpaid', 'overdue', 'cancelled'):
                messages.error(request, "Invalid bill status.")
            else:
                bill.status = new_status
                bill.save(update_fields=['status'])
                messages.success(request, "Bill status updated.")

        elif action == 'record_bill_payment':
            if not client.can_be_billed:
                messages.error(request, "Payments cannot be recorded for an Inquiry. Change the client status to Customer first.")
            else:
                try:
                    payment_amount = Decimal(request.POST.get('payment_amount', ''))
                except (InvalidOperation, TypeError):
                    payment_amount = Decimal('0')
                with transaction.atomic():
                    bill = get_object_or_404(
                        CustomerBill.objects.select_for_update(),
                        pk=request.POST.get('bill_id'), client=client, org=org,
                    )
                    if bill.status == 'cancelled':
                        messages.error(request, "A cancelled invoice cannot receive payments.")
                    elif payment_amount <= 0:
                        messages.error(request, "Payment amount must be greater than zero.")
                    elif payment_amount > bill.remaining_amount:
                        messages.error(request, "Payment cannot exceed the remaining invoice amount.")
                    else:
                        payment = CustomerBillPayment(
                            org=org, branch=bill.branch, bill=bill, client=client,
                            amount=payment_amount,
                            payment_date=request.POST.get('payment_date') or timezone.localdate(),
                            payment_method=request.POST.get('payment_method', 'cash'),
                            payment_reference=request.POST.get('payment_reference', '').strip(),
                            note=request.POST.get('payment_note', '').strip(),
                            created_by=request.user,
                        )
                        payment.full_clean()
                        payment.save()
                        bill.paid_amount += payment_amount
                        if bill.paid_amount == bill.amount:
                            bill.status = 'paid'
                            bill.paid_date = payment.payment_date
                        else:
                            bill.status = 'partial'
                            bill.paid_date = None
                        bill.full_clean()
                        bill.save(update_fields=['paid_amount', 'status', 'paid_date'])
                        if request.POST.get('add_to_income') == 'on':
                            category, _ = TransactionCategory.objects.get_or_create(
                                org=org, name='Customer Bill Payment', transaction_type='income',
                            )
                            income = FinancialTransaction.objects.create(
                                org=org, branch=bill.branch, category=category,
                                transaction_type='income',
                                title=f"Bill Payment — {bill.invoice_number} ({client.client_org_name})",
                                amount=payment_amount, transaction_date=payment.payment_date,
                                payment_method=payment.payment_method,
                                reference_number=payment.payment_reference or bill.invoice_number,
                                note=f"CRM customer payment for invoice {bill.invoice_number}.",
                                created_by=request.user,
                            )
                            payment.income_transaction = income
                            payment.save(update_fields=['income_transaction'])
                        messages.success(request, "Payment recorded safely.")

        elif action == 'add_contract':
            CustomerContract.objects.create(
                client=client, org=org,
                title=request.POST.get('contract_title', '').strip(),
                contract_value=request.POST.get('contract_value') or None,
                start_date=request.POST.get('start_date') or datetime.date.today(),
                end_date=request.POST.get('contract_end_date') or None,
                status=request.POST.get('contract_status', 'active'),
                document=request.FILES.get('contract_file'),
                notes=request.POST.get('contract_notes', ''),
                created_by=request.user,
            )
            messages.success(request, "Contract added.")

        elif action == 'add_proposal':
            CustomerProposal.objects.create(
                client=client, org=org,
                title=request.POST.get('proposal_title', '').strip(),
                amount=request.POST.get('proposal_amount') or None,
                sent_date=request.POST.get('sent_date') or None,
                valid_until=request.POST.get('valid_until') or None,
                status=request.POST.get('proposal_status', 'draft'),
                document=request.FILES.get('proposal_file'),
                notes=request.POST.get('proposal_notes', ''),
                created_by=request.user,
            )
            messages.success(request, "Proposal added.")

        elif action == 'add_document':
            f = request.FILES.get('doc_file')
            if not f:
                messages.error(request, "Please select a file.")
            else:
                CustomerDocument.objects.create(
                    client=client, org=org,
                    title=request.POST.get('doc_title', f.name),
                    doc_type=request.POST.get('doc_type', 'other'),
                    file=f,
                    notes=request.POST.get('doc_notes', ''),
                    uploaded_by=request.user,
                )
                messages.success(request, "Document uploaded.")

        elif action == 'follow_up':
            feedback = request.POST.get('feedback', '').strip()
            follow_up_date = request.POST.get('follow_up_date') or datetime.date.today()
            next_follow_up_date = request.POST.get('next_follow_up_date') or None
            if not feedback:
                messages.error(request, "Feedback is required.")
            else:
                _create_follow_up(
                    org, client, request.user, feedback,
                    follow_up_date, next_follow_up_date,
                    priority=request.POST.get('priority', 'medium'),
                )
                messages.success(request, "Follow-up logged successfully.")

        elif action == 'close_followup':
            fu = get_object_or_404(client.follow_ups, pk=request.POST.get('follow_up_id'))
            fu.status = 'closed'
            fu.closed_at = timezone.now()
            fu.closed_by = request.user
            fu.save(update_fields=['status', 'closed_at', 'closed_by'])
            messages.success(request, "Follow-up closed.")

        elif action == 'reopen_followup':
            fu = get_object_or_404(client.follow_ups, pk=request.POST.get('follow_up_id'))
            fu.status = 'open'
            fu.closed_at = None
            fu.closed_by = None
            fu.save(update_fields=['status', 'closed_at', 'closed_by'])
            messages.success(request, "Follow-up reopened.")

        elif action == 'convert_to_customer':
            client.status = 'customer'
            client.save(update_fields=['status'])
            messages.success(request, f"{client.client_org_name} converted to Customer. Info, Bills, Contracts, Proposals, Docs and Sales are now unlocked.")

        tab_map = {
            'update_info': 'info', 'add_bill': 'bills', 'update_bill_status': 'bills',
            'record_bill_payment': 'bills',
            'add_contract': 'contracts', 'add_proposal': 'proposals',
            'add_document': 'docs', 'follow_up': 'followup',
            'close_followup': 'followup', 'reopen_followup': 'followup',
            'convert_to_customer': 'followup',
        }
        tab = tab_map.get(action, 'followup')
        return redirect(f"{reverse('schooladmin:client_detail', kwargs={'pk': client.pk})}?tab={tab}")


class ClientFollowUpDueListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'clients'
    template_name = 'admin/clients/due_followups.html'

    def get(self, request):
        org = _task_org(request)
        priority = request.GET.get('priority', '').strip()
        clients_due = _clients_due_qs(org, priority=priority)
        ctx = dict(
            org=org,
            clients_due=clients_due,
            priority=priority,
            priority_choices=Client.PRIORITY_CHOICES,
        )
        return render(request, self.template_name, ctx)


class BillingReminderView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'clients'
    template_name = 'admin/clients/billing_reminders.html'

    def get(self, request):
        org = _task_org(request)
        today = datetime.date.today()
        in_7 = today + datetime.timedelta(days=7)
        in_30 = today + datetime.timedelta(days=30)
        # Clients with next_billing_date coming up
        billing_soon = Client.objects.filter(
            org=org, is_active=True, next_billing_date__isnull=False,
            next_billing_date__lte=in_30, next_billing_date__gte=today
        ).order_by('next_billing_date')
        overdue_bills = CustomerBill.objects.filter(
            org=org, status__in=('sent', 'draft'), due_date__lt=today
        ).select_related('client').order_by('due_date')
        expiring_contracts = CustomerContract.objects.filter(
            org=org, status='active', end_date__isnull=False,
            end_date__lte=in_30, end_date__gte=today
        ).select_related('client').order_by('end_date')
        ctx = dict(
            org=org, today=today, in_7=in_7, in_30=in_30,
            billing_soon=billing_soon, overdue_bills=overdue_bills,
            expiring_contracts=expiring_contracts,
        )
        return render(request, self.template_name, ctx)


# ═══════════════════════════════════════════════════════════════════════════════
# SHIFT MANAGEMENT (multi-window / split shifts)
# ═══════════════════════════════════════════════════════════════════════════════
from handle.models import AttendanceReminderPolicy, Shift, ShiftWindow


def _parse_shift_windows(request):
    """Read parallel start_time[]/end_time[] arrays from POST into a clean list."""
    starts = request.POST.getlist('start_time')
    ends = request.POST.getlist('end_time')
    windows = []
    for i, (s, e) in enumerate(zip(starts, ends)):
        s, e = (s or '').strip(), (e or '').strip()
        if s and e:
            windows.append((i + 1, s, e))
    return windows


class ShiftListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/shifts/list.html'

    def get(self, request):
        org = _task_org(request)
        shifts = Shift.objects.filter(org=org).prefetch_related('windows', 'members').order_by('name')
        reminder_policy = (
            AttendanceReminderPolicy.objects.filter(org=org).first()
            or AttendanceReminderPolicy(org=org)
        )
        return render(request, self.template_name, {
            'org': org,
            'shifts': shifts,
            'reminder_policy': reminder_policy,
        })

    def post(self, request):
        org = _task_org(request)
        policy, _ = AttendanceReminderPolicy.objects.get_or_create(org=org)
        policy.enabled = request.POST.get('enabled') == 'on'
        policy.checkin_enabled = request.POST.get('checkin_enabled') == 'on'
        policy.checkout_enabled = request.POST.get('checkout_enabled') == 'on'

        def parse_offsets(name):
            raw_values = (request.POST.get(name) or '').split(',')
            values = []
            for raw in raw_values:
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    value = int(raw)
                except ValueError as exc:
                    raise ValidationError(
                        'Reminder offsets must be whole minutes separated by commas.'
                    ) from exc
                if not 0 <= value <= 180:
                    raise ValidationError(
                        'Reminder offsets must be between 0 and 180 minutes.'
                    )
                values.append(value)
            return values

        try:
            policy.checkin_offsets = parse_offsets('checkin_offsets')
            policy.checkout_offsets = parse_offsets('checkout_offsets')
            policy.full_clean()
            policy.save()
        except ValidationError as exc:
            messages.error(request, '; '.join(exc.messages))
        else:
            messages.success(
                request,
                'Mobile attendance reminder timing updated.',
            )
        return redirect('schooladmin:shift_list')


class ShiftFormView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/shifts/form.html'

    def get(self, request, pk=None):
        org = _task_org(request)
        shift = get_object_or_404(Shift, pk=pk, org=org) if pk else None
        windows = shift.ordered_windows() if shift else []
        return render(request, self.template_name, {'org': org, 'shift': shift, 'windows': windows})

    def post(self, request, pk=None):
        org = _task_org(request)
        shift = get_object_or_404(Shift, pk=pk, org=org) if pk else None

        name = (request.POST.get('name') or '').strip()
        if not name:
            messages.error(request, "Shift name is required.")
            return redirect(request.path)

        parsed = _parse_shift_windows(request)
        if not parsed:
            messages.error(request, "Add at least one time window (start and end).")
            return redirect(request.path)

        is_active = request.POST.get('is_active') == 'on'
        if shift:
            shift.name = name
            shift.is_active = is_active
            shift.save()
            shift.windows.all().delete()
        else:
            shift = Shift.objects.create(org=org, name=name, is_active=is_active)

        for order, s, e in parsed:
            ShiftWindow.objects.create(shift=shift, order=order, start_time=s, end_time=e)

        messages.success(request, f"Shift '{shift.name}' saved.")
        return redirect('schooladmin:shift_list')


class ShiftDeleteView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'

    def post(self, request, pk):
        org = _task_org(request)
        shift = get_object_or_404(Shift, pk=pk, org=org)
        name = shift.name
        if shift.weekday_assignments.exists() or shift.members.exists() or shift.date_overrides.exists():
            messages.error(request, f"'{name}' is assigned to members. Reassign those schedules before deleting it.")
            return redirect('schooladmin:shift_list')
        shift.delete()
        messages.success(request, f"Shift '{name}' deleted.")
        return redirect('schooladmin:shift_list')


class ShiftAssignView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/shifts/assign.html'

    def get(self, request):
        org = _task_org(request)
        members = member.objects.filter(org=org, status='active').prefetch_related(
            'shifts', 'weekday_shifts__shift',
        ).order_by('name')
        shifts = Shift.objects.filter(org=org, is_active=True).order_by('name')
        return render(request, self.template_name, {'org': org, 'members': members, 'shifts': shifts})

    def post(self, request):
        org = _task_org(request)
        member_id = request.POST.get('member_id')
        from handle.models import MemberHistory, MemberWeekdayShift
        memb = get_object_or_404(member, pk=member_id, org=org)
        allowed = {str(s.pk): s for s in Shift.objects.filter(org=org, is_active=True)}
        with transaction.atomic():
            MemberWeekdayShift.objects.filter(org=org, member=memb).delete()
            rows = []
            seen = set()
            for weekday in range(7):
                for raw_id in request.POST.getlist(f'shift_{weekday}'):
                    shift = allowed.get(raw_id)
                    if shift and (weekday, shift.pk) not in seen:
                        seen.add((weekday, shift.pk))
                        rows.append(MemberWeekdayShift(
                            org=org, member=memb, weekday=weekday, shift=shift,
                        ))
            MemberWeekdayShift.objects.bulk_create(rows)
            # Keep the legacy M2M synchronized for older dashboards while all
            # date-aware reporting resolves the exact weekday assignment.
            memb.shifts.set({row.shift for row in rows})
            metadata = {}
            for row in rows:
                metadata.setdefault(str(row.weekday), []).append(row.shift.name)
            MemberHistory.objects.create(
                org=org, member=memb, action='weekday_shifts_changed',
                field_name='weekday_shifts', changed_by=request.user,
                description='Weekly shift schedule updated',
                metadata=metadata,
            )
        messages.success(request, f"Sunday-to-Saturday shifts updated for {memb.name}.")
        return redirect('schooladmin:shift_assign')


def _shift_week_context(request, org, memb):
    """Build the current Sun-Sat week's per-day shift + completion context for
    the 'This Week' modal, shared by the GET view and the override add/delete
    POST views (which re-render the same fragment)."""
    import datetime as _dt
    from schooladmin.payroll_service import compute_shift_breakdown
    from handle.models import MemberShiftOverride

    today = timezone.localdate()
    week_start = today - _dt.timedelta(days=member.weekday_number(today))

    grace_minutes = 0
    policy = getattr(org, 'payroll_policy', None)
    if policy:
        grace_minutes = policy.late_grace_minutes or 0

    overrides = {}
    override_qs = MemberShiftOverride.objects.filter(
        org=org, member=memb, date__range=(week_start, week_start + _dt.timedelta(days=6)),
    ).select_related('shift')
    for o in override_qs:
        overrides.setdefault(o.date, []).append(o)

    days = []
    for offset in range(7):
        d = week_start + _dt.timedelta(days=offset)
        # Recurring-pattern shifts only (not merged with overrides) so the
        # panel can show "weekly pattern" badges and "+ extra" chips distinctly
        # instead of double-displaying an overridden shift.
        weekday_rows = memb.weekday_shifts.filter(
            weekday=member.weekday_number(d), shift__is_active=True,
        ).select_related('shift')
        seen_ids = set()
        shifts_today = []
        for row in weekday_rows:
            if row.shift_id not in seen_ids:
                seen_ids.add(row.shift_id)
                shifts_today.append(row.shift)
        status = None
        if d <= today:
            b = compute_shift_breakdown(memb, d, grace_minutes)
            if b:
                windows = b['windows']
                complete_windows = sum(1 for w in windows if w.get('complete'))
                attended_windows = sum(1 for w in windows if w.get('attended'))
                if windows and complete_windows == len(windows):
                    status = 'complete'
                elif attended_windows:
                    status = 'incomplete'
                else:
                    status = 'absent'
        day_overrides = overrides.get(d, [])
        days.append({
            'date': d, 'weekday': member.weekday_number(d),
            'weekday_label': d.strftime('%A'),
            'is_today': d == today, 'is_future': d > today,
            'shifts': shifts_today,
            'has_shift': bool(shifts_today or day_overrides),
            'status': status,
            'overrides': day_overrides,
        })

    return {
        'org': org, 'member': memb, 'days': days,
        'week_start': week_start, 'week_end': week_start + _dt.timedelta(days=6),
        'all_shifts': Shift.objects.filter(org=org, is_active=True).order_by('name'),
    }


class ShiftMemberWeekView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Renders the 'This Week' modal fragment: real Sun-Sat dates for the
    current week, each day's resolved shift(s), and its completion status."""
    required_feature = 'hrms'
    template_name = 'admin/shifts/_week_panel.html'

    def get(self, request, member_id):
        org = _task_org(request)
        memb = get_object_or_404(member, pk=member_id, org=org)
        return render(request, self.template_name, _shift_week_context(request, org, memb))


class ShiftOverrideAddView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Adds a one-off extra shift for a member on a specific date, on top of
    (not replacing) their recurring weekday pattern."""
    required_feature = 'hrms'
    template_name = 'admin/shifts/_week_panel.html'

    def post(self, request, member_id):
        import datetime as _dt
        from handle.models import MemberHistory, MemberShiftOverride
        org = _task_org(request)
        memb = get_object_or_404(member, pk=member_id, org=org)
        shift_id = request.POST.get('shift_id', '')
        shift = Shift.objects.filter(pk=shift_id, org=org, is_active=True).first() if shift_id.isdigit() else None
        try:
            override_date = _dt.datetime.strptime(request.POST.get('date', ''), '%Y-%m-%d').date()
        except ValueError:
            override_date = None

        if not shift:
            messages.error(request, "Choose a valid shift for the extra day.")
        elif not override_date:
            messages.error(request, "Choose a valid date for the extra shift.")
        else:
            override, created = MemberShiftOverride.objects.get_or_create(
                org=org, member=memb, date=override_date, shift=shift,
            )
            if created:
                MemberHistory.objects.create(
                    org=org, member=memb, action='shift_override_added',
                    field_name='shift_overrides', changed_by=request.user,
                    description=f'Extra shift "{shift.name}" added for {override_date.isoformat()}',
                    metadata={'date': override_date.isoformat(), 'shift': shift.name},
                )
        return render(request, self.template_name, _shift_week_context(request, org, memb))


class ShiftOverrideDeleteView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/shifts/_week_panel.html'

    def post(self, request, pk):
        from handle.models import MemberHistory, MemberShiftOverride
        org = _task_org(request)
        override = get_object_or_404(MemberShiftOverride, pk=pk, org=org)
        memb = override.member
        MemberHistory.objects.create(
            org=org, member=memb, action='shift_override_removed',
            field_name='shift_overrides', changed_by=request.user,
            description=f'Extra shift "{override.shift.name}" removed for {override.date.isoformat()}',
            metadata={'date': override.date.isoformat(), 'shift': override.shift.name},
        )
        override.delete()
        return render(request, self.template_name, _shift_week_context(request, org, memb))


class ShiftReportView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Per-window attendance report for shift-assigned members: shows late-in,
    early-out and overtime for each shift window on each day."""
    required_feature = 'hrms'
    template_name = 'admin/shifts/report.html'

    def get(self, request):
        import datetime as _dt
        from schooladmin.payroll_service import compute_shift_breakdown
        org = _task_org(request)

        shift_members = member.objects.filter(
            Q(org=org, shifts__isnull=False) | Q(org=org, weekday_shifts__isnull=False)
            | Q(org=org, shift_overrides__isnull=False)
        ).exclude(status='dumped').prefetch_related('shifts', 'weekday_shifts__shift').distinct().order_by('name')

        member_id = request.GET.get('member', '').strip()
        from_str = request.GET.get('from_date', '').strip()
        to_str = request.GET.get('to_date', '').strip()
        preset = request.GET.get('preset', '').strip()
        nepali_enabled = bool(getattr(org, 'nepali_date', False))
        today = timezone.localdate()

        if preset:
            if preset == 'today':
                preset_from = preset_to = today
            elif preset == 'this_week':
                preset_from = today - _dt.timedelta(days=today.weekday())
                preset_to = today
            elif preset == 'last_month':
                previous_month_end = today.replace(day=1) - _dt.timedelta(days=1)
                preset_from = previous_month_end.replace(day=1)
                preset_to = previous_month_end
            else:
                preset = 'this_month'
                preset_from = today.replace(day=1)
                preset_to = today
            from_str = preset_from.isoformat()
            to_str = preset_to.isoformat()

        grace_minutes = 0
        policy = getattr(org, 'payroll_policy', None)
        if policy:
            grace_minutes = policy.late_grace_minutes or 0

        selected = None
        days = []
        totals = {'early_in': 0.0, 'late': 0.0, 'early': 0.0, 'late_out': 0.0, 'ot': 0.0, 'worked': 0.0}
        summary = {
            'scheduled_days': 0, 'present_days': 0, 'absent_days': 0,
            'complete_days': 0, 'incomplete_days': 0, 'attendance_rate': 0,
            'late_days': 0, 'early_out_days': 0, 'overtime_days': 0,
        }
        absence_dates = []
        range_error = ''

        if member_id:
            selected = shift_members.filter(id=member_id).first()
        elif shift_members.count() == 1:
            # Only one member has a shift — auto-select so the report just shows.
            selected = shift_members.first()
            member_id = str(selected.id)

        # If a member is chosen but no date range given, default to the current month
        # so the report shows data immediately instead of looking empty.
        if selected and not (from_str and to_str):
            from_str = from_str or today.replace(day=1).isoformat()
            to_str = to_str or today.isoformat()

        if selected and from_str and to_str:
            try:
                d_from = _dt.datetime.strptime(from_str, "%Y-%m-%d").date()
                d_to = _dt.datetime.strptime(to_str, "%Y-%m-%d").date()
            except ValueError:
                d_from = d_to = None

            if not d_from or not d_to:
                range_error = 'Enter a valid From and To date.'
            elif d_from > d_to:
                range_error = 'From date cannot be after To date.'
            elif (d_to - d_from).days > 366:
                range_error = 'For performance and readability, generate at most 367 days at a time.'
            else:
                cur = d_from
                while cur <= d_to:
                    b = compute_shift_breakdown(selected, cur, grace_minutes)
                    if b:
                        day_early_in = sum(w['early_in_min'] for w in b['windows'])
                        day_late = sum(w['late_min'] for w in b['windows'])
                        day_early = sum(w['early_min'] for w in b['windows'])
                        day_late_out = sum(w['late_out_min'] for w in b['windows'])
                        day_ot = sum(w['ot_min'] for w in b['windows'])
                        day_worked = float(b['worked_hours'])
                        windows = b['windows']
                        complete_windows = sum(1 for window in windows if window.get('complete'))
                        attended_windows = sum(1 for window in windows if window.get('attended'))
                        is_complete = bool(windows) and complete_windows == len(windows)
                        is_incomplete = attended_windows > 0 and complete_windows < len(windows)
                        day_status = 'complete' if is_complete else ('incomplete' if is_incomplete else 'absent')
                        shift_count = len({w['shift_id'] for w in windows if w.get('shift_id') is not None})
                        date_np = ''
                        if nepali_enabled:
                            try:
                                date_np = str(nepali_datetime.date.from_datetime_date(cur))
                            except (ValueError, TypeError):
                                date_np = ''
                        days.append({
                            'date': cur,
                            'date_np': date_np,
                            'windows': windows,
                            'attended_any': b['attended_any'],
                            'complete': is_complete,
                            'incomplete': is_incomplete,
                            'status': day_status,
                            'shift_count': shift_count,
                            'early_in': round(day_early_in, 1),
                            'late': round(day_late, 1),
                            'early': round(day_early, 1),
                            'late_out': round(day_late_out, 1),
                            'ot': round(day_ot, 1),
                            'worked': round(day_worked, 2),
                        })
                        totals['early_in'] += day_early_in
                        totals['late'] += day_late
                        totals['early'] += day_early
                        totals['late_out'] += day_late_out
                        totals['ot'] += day_ot
                        totals['worked'] += day_worked
                        summary['scheduled_days'] += 1
                        if b['attended_any']:
                            summary['present_days'] += 1
                        else:
                            summary['absent_days'] += 1
                            absence_dates.append({
                                'date': cur, 'date_np': date_np,
                                'weekday': cur.strftime('%A'),
                            })
                        if windows and complete_windows == len(windows):
                            summary['complete_days'] += 1
                        elif attended_windows:
                            summary['incomplete_days'] += 1
                        if day_late:
                            summary['late_days'] += 1
                        if day_early:
                            summary['early_out_days'] += 1
                        if day_ot:
                            summary['overtime_days'] += 1
                    cur += _dt.timedelta(days=1)
                totals = {k: round(v, 1) for k, v in totals.items()}
                if summary['scheduled_days']:
                    summary['attendance_rate'] = round(
                        summary['present_days'] / summary['scheduled_days'] * 100, 1
                    )

        period_label = ''
        if from_str and to_str:
            period_label = f'{from_str} to {to_str}'
            if nepali_enabled:
                try:
                    period_label = (
                        f"{nepali_datetime.date.from_datetime_date(_dt.date.fromisoformat(from_str))}"
                        f" to {nepali_datetime.date.from_datetime_date(_dt.date.fromisoformat(to_str))} BS"
                    )
                except (ValueError, TypeError):
                    pass

        if request.GET.get('export') == 'csv' and selected and days and not range_error:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = (
                f'attachment; filename="shift-report-{selected.id}-{from_str}-to-{to_str}.csv"'
            )
            writer = csv.writer(response)
            writer.writerow([
                'Member', 'Date (AD)', 'Date (BS)', 'Shift Name', 'Shift Window', 'Punch In', 'Punch Out',
                'Early In (min)', 'Late In (min)', 'Late After Grace (min)',
                'Early Out (min)', 'Late Out (min)', 'Overtime (min)', 'Status', 'Day Status',
            ])
            for day in days:
                for window in day['windows']:
                    status = 'Absent'
                    if window.get('complete'):
                        status = 'Complete'
                    elif window.get('attended'):
                        status = 'Incomplete'
                    writer.writerow([
                        selected.name, day['date'].isoformat(), day.get('date_np', ''),
                        window.get('shift_name', '') or '', window.get('label', ''), window.get('punch_in', ''),
                        window.get('punch_out', ''), window.get('early_in_min', 0),
                        window.get('late_min', 0), window.get('late_penalized_min', 0),
                        window.get('early_min', 0), window.get('late_out_min', 0),
                        window.get('ot_min', 0), status, day.get('status', '').capitalize(),
                    ])
            return response

        return render(request, self.template_name, {
            'org': org, 'shift_members': shift_members, 'selected': selected,
            'days': days, 'totals': totals, 'summary': summary,
            'absence_dates': absence_dates,
            'range_error': range_error, 'period_label': period_label,
            'nepali_enabled': nepali_enabled, 'active_preset': preset,
            'filters': {'member': member_id, 'from_date': from_str, 'to_date': to_str},
        })


# ═══════════════════════════════════════════════════════════════════════════════
# DUTY ROSTER — workforce scheduling table + temporary shift changes.
# Reuses the existing Shift/MemberWeekdayShift/TemporaryShiftAssignment
# engine (see handle.models.member.shift_windows_detailed) rather than a
# parallel scheduling system: this is a table view over the same data the
# Shift Management pages already manage, plus the ability to schedule a
# date-range change that temporarily replaces a member's regular pattern.
# ═══════════════════════════════════════════════════════════════════════════════

def _duty_roster_default_label(windows):
    """Compact label for a day with no Duty Roster override — the member's
    regular pattern for that day, or 'Off' if they have none."""
    if not windows:
        return 'Off'
    w = windows[0]
    if w.get('shift_name'):
        return w['shift_name']
    if w.get('start_time') and w.get('end_time'):
        return f"{w['start_time'].strftime('%H:%M')}–{w['end_time'].strftime('%H:%M')}"
    return 'Default'


class DutyRosterView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'
    template_name = 'admin/shifts/duty_roster.html'

    def get(self, request):
        from school.hierarchy import get_accessible_members, get_accessible_branches, get_accessible_classifications
        from handle.models import DutyType

        org = _task_org(request)
        nepali_enabled = bool(getattr(org, 'nepali_date', False))

        date_np_str = request.GET.get('date_np', '').strip()
        selected_date = None
        if nepali_enabled and date_np_str:
            try:
                y, m, d = map(int, date_np_str.replace('/', '-').strip().split('-'))
                selected_date = nepali_datetime.date(y, m, d).to_datetime_date()
            except (ValueError, TypeError):
                selected_date = None
        if selected_date is None:
            try:
                selected_date = datetime.datetime.strptime(request.GET.get('date', ''), '%Y-%m-%d').date()
            except ValueError:
                selected_date = timezone.localdate()
        selected_date_np = str(nepali_datetime.date.from_datetime_date(selected_date)) if nepali_enabled else ''

        branch_id = request.GET.get('branch', '')
        classification_id = request.GET.get('classification', '')
        active_view = request.GET.get('view', 'day')

        members_qs = get_accessible_members(request.user, org).exclude(status='dumped').select_related(
            'branch', 'classification', 'section',
        ).order_by('name')
        if branch_id:
            members_qs = members_qs.filter(branch_id=branch_id)
        if classification_id:
            members_qs = members_qs.filter(classification_id=classification_id)
        members_qs = list(members_qs)

        duty_types = DutyType.objects.filter(org=org, is_active=True).order_by('name')
        default_duty_type = duty_types.filter(is_default=True).first()

        rows = []
        for m in members_qs:
            windows = m.shift_windows_detailed(selected_date)
            rows.append({
                'member': m,
                'windows': windows,
                'is_off': not windows,
                'temp_assignment': m.temporary_shift_assignment_for(selected_date),
            })

        # --- Weekly grid: Sunday-first (Mero convention), matching every
        # other weekday-numbered feature (MemberWeekdayShift, calculate_
        # attendance_stats, etc). Saved/persisted as ordinary
        # TemporaryShiftAssignment rows (see WeeklyDutyRosterSaveView), so
        # "viewing a past week later" is just re-navigating to it - nothing
        # ephemeral or session-only.
        try:
            week_anchor = datetime.datetime.strptime(request.GET.get('week_start', ''), '%Y-%m-%d').date()
        except ValueError:
            week_anchor = timezone.localdate()
        week_start = week_anchor - datetime.timedelta(days=member.weekday_number(week_anchor))
        week_days_list = [week_start + datetime.timedelta(days=i) for i in range(7)]
        week_days = [
            {
                'date': d,
                'date_np': str(nepali_datetime.date.from_datetime_date(d)) if nepali_enabled else '',
            }
            for d in week_days_list
        ]

        week_rows = []
        for m in members_qs:
            cells = []
            for d in week_days_list:
                ta = m.temporary_shift_assignment_for(d)
                if ta:
                    cells.append({
                        'date': d,
                        'value': 'off' if ta.shift_id is None else str(ta.shift_id),
                        'label': 'Off' if ta.shift_id is None else ta.shift.name,
                        'is_override': True,
                        'duty_value': str(ta.duty_type_id) if ta.duty_type_id else '',
                        'duty_label': ta.duty_type.name if ta.duty_type_id else '',
                    })
                else:
                    windows = m.shift_windows_detailed(d)
                    cells.append({
                        'date': d, 'value': '',
                        'label': _duty_roster_default_label(windows),
                        'is_override': False,
                        'duty_value': '', 'duty_label': '',
                    })
            week_rows.append({'member': m, 'cells': cells})

        week_end = week_days_list[-1]
        week_period_label_np = ''
        if nepali_enabled:
            start_np = nepali_datetime.date.from_datetime_date(week_start)
            end_np = nepali_datetime.date.from_datetime_date(week_end)
            week_period_label_np = f"{start_np} – {end_np} (BS)"

        return render(request, self.template_name, {
            'org': org, 'rows': rows, 'selected_date': selected_date,
            'nepali_enabled': nepali_enabled,
            'selected_date_np': selected_date_np,
            'branches': get_accessible_branches(request.user, org),
            'classifications': get_accessible_classifications(request.user, org),
            'selected_branch': branch_id,
            'selected_classification': classification_id,
            'active_view': active_view,
            'shifts': Shift.objects.filter(org=org, is_active=True).order_by('name'),
            'duty_types': duty_types,
            'default_duty_type': default_duty_type,
            'week_start': week_start,
            'week_end': week_end,
            'week_days': week_days,
            'week_rows': week_rows,
            'week_period_label_np': week_period_label_np,
            'prev_week': week_start - datetime.timedelta(days=7),
            'next_week': week_start + datetime.timedelta(days=7),
            'this_week': timezone.localdate() - datetime.timedelta(
                days=member.weekday_number(timezone.localdate()),
            ),
        })

    def post(self, request):
        from school.hierarchy import get_accessible_members
        from handle.models import MemberHistory, TemporaryShiftAssignment, DutyType

        org = _task_org(request)
        memb = get_object_or_404(get_accessible_members(request.user, org), pk=request.POST.get('member_id'))

        try:
            start_date = datetime.datetime.strptime(request.POST.get('start_date', ''), '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Choose a valid start date.")
            return redirect('schooladmin:duty_roster')

        end_date = None
        end_str = request.POST.get('end_date', '').strip()
        if end_str:
            try:
                end_date = datetime.datetime.strptime(end_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Choose a valid end date.")
                return redirect('schooladmin:duty_roster')
            if end_date < start_date:
                messages.error(request, "End date cannot be before start date.")
                return redirect('schooladmin:duty_roster')

        shift = None
        shift_id = request.POST.get('shift_id', '').strip()
        if shift_id:
            shift = Shift.objects.filter(org=org, pk=shift_id, is_active=True).first()
            if not shift:
                messages.error(request, "Choose a valid shift, or leave it blank to mark the member off-duty.")
                return redirect('schooladmin:duty_roster')

        duty_type = None
        duty_type_id = request.POST.get('duty_type_id', '').strip()
        if duty_type_id:
            duty_type = DutyType.objects.filter(org=org, pk=duty_type_id, is_active=True).first()

        notes = request.POST.get('notes', '').strip()
        TemporaryShiftAssignment.objects.create(
            org=org, member=memb, start_date=start_date, end_date=end_date,
            shift=shift, duty_type=duty_type, notes=notes, created_by=request.user,
        )
        span = f"{start_date}" + (f" → {end_date}" if end_date else " onward")
        label = shift.name if shift else 'Off duty'
        if duty_type:
            label += f" / {duty_type.name}"
        MemberHistory.objects.create(
            org=org, member=memb, action='duty_roster_change', changed_by=request.user,
            description=f"Duty roster: {label} ({span})" + (f" — {notes}" if notes else ''),
        )
        messages.success(request, f"Scheduled {'a shift change' if shift else 'an off-duty period'} for {memb.name}.")
        return redirect('schooladmin:duty_roster')


class DutyRosterCancelView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'hrms'

    def post(self, request, pk):
        from handle.models import MemberHistory, TemporaryShiftAssignment

        org = _task_org(request)
        assignment = get_object_or_404(TemporaryShiftAssignment, pk=pk, org=org)
        assignment.is_active = False
        assignment.save(update_fields=['is_active'])
        span = f"{assignment.start_date} → {assignment.end_date or 'ongoing'}"
        MemberHistory.objects.create(
            org=org, member=assignment.member, action='duty_roster_change', changed_by=request.user,
            description=f"Duty roster change cancelled: {assignment.shift.name if assignment.shift else 'Off duty'} ({span})",
        )
        messages.success(request, "Duty roster change cancelled.")
        return redirect('schooladmin:duty_roster')


class DutyTypeManageView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Add / set-default / enable-disable org-wide Duty Types (e.g.
    "Regular", "On-Call", "Field Duty") from the Duty Roster page's Manage
    Duty Types modal. A DutyType is independent of Shift — it labels WHAT
    KIND of duty a roster entry is, not WHEN. At most one DutyType per org
    is flagged `is_default` (pre-selected in the roster-making dropdowns);
    enforced here by clearing the flag off every other row, not at the DB
    level, matching how other single-flag-per-org settings in this app work."""
    required_feature = 'hrms'

    def post(self, request):
        from handle.models import DutyType

        org = _task_org(request)
        action = request.POST.get('action', '').strip()

        if action == 'create':
            name = request.POST.get('name', '').strip()
            if not name:
                messages.error(request, "Duty type name is required.")
                return redirect('schooladmin:duty_roster')
            if DutyType.objects.filter(org=org, name__iexact=name).exists():
                messages.error(request, f'"{name}" already exists.')
                return redirect('schooladmin:duty_roster')
            is_default = request.POST.get('is_default') == 'on'
            duty_type = DutyType.objects.create(org=org, name=name, is_default=is_default)
            if is_default:
                DutyType.objects.filter(org=org).exclude(pk=duty_type.pk).update(is_default=False)
            messages.success(request, f'Added duty type "{name}".')
            return redirect('schooladmin:duty_roster')

        duty_type = get_object_or_404(DutyType, pk=request.POST.get('duty_type_id'), org=org)

        if action == 'set_default':
            DutyType.objects.filter(org=org).exclude(pk=duty_type.pk).update(is_default=False)
            duty_type.is_default = True
            duty_type.save(update_fields=['is_default'])
            messages.success(request, f'"{duty_type.name}" is now the default duty type.')
            return redirect('schooladmin:duty_roster')

        if action == 'toggle_active':
            duty_type.is_active = not duty_type.is_active
            update_fields = ['is_active']
            if not duty_type.is_active and duty_type.is_default:
                duty_type.is_default = False
                update_fields.append('is_default')
            duty_type.save(update_fields=update_fields)
            messages.success(request, f'"{duty_type.name}" {"enabled" if duty_type.is_active else "disabled"}.')
            return redirect('schooladmin:duty_roster')

        messages.error(request, "Unknown action.")
        return redirect('schooladmin:duty_roster')


class WeeklyDutyRosterSaveView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Bulk-saves one week's worth of duty-roster cells from
    templates/admin/shifts/duty_roster.html's Week View grid — one
    <select> per member per day, named `shift_<member_id>_<iso-date>`,
    all inside one form with one Save button (unlike the Day View's
    one-modal-per-member flow, which doesn't scale to a 7-day sheet).

    Each cell becomes an ordinary single-day TemporaryShiftAssignment
    (start_date == end_date), so "the week is saved and can be viewed
    later" falls out of the existing model for free - re-opening this same
    week just re-reads these rows. A cell submitted unchanged from what's
    already in effect is a no-op (idempotent re-save); a cell that differs
    deactivates whatever TemporaryShiftAssignment currently covers that
    exact day for that member (even a multi-day one from the older
    single-day modal flow) and, unless the new value is "Default", creates
    a fresh single-day row - the weekly grid always wins for the specific
    day being edited rather than attempting to split date ranges.
    """
    required_feature = 'hrms'

    def post(self, request):
        from school.hierarchy import get_accessible_members
        from handle.models import MemberHistory, TemporaryShiftAssignment, DutyType

        org = _task_org(request)
        try:
            week_start = datetime.datetime.strptime(request.POST.get('week_start', ''), '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid week.")
            return redirect('schooladmin:duty_roster')

        week_days = [week_start + datetime.timedelta(days=i) for i in range(7)]
        members_qs = get_accessible_members(request.user, org).exclude(status='dumped')
        shift_ids = set(str(pk) for pk in Shift.objects.filter(org=org, is_active=True).values_list('id', flat=True))
        duty_type_ids = set(str(pk) for pk in DutyType.objects.filter(org=org, is_active=True).values_list('id', flat=True))

        changed_members = 0
        with transaction.atomic():
            for m in members_qs:
                change_notes = []
                for d in week_days:
                    field = f"shift_{m.id}_{d.isoformat()}"
                    if field not in request.POST:
                        continue  # cell wasn't part of the submitted grid (e.g. filtered out)
                    v = request.POST.get(field, '').strip()
                    if v not in ('', 'off') and v not in shift_ids:
                        continue  # bogus/cross-org shift id - ignore rather than 400 the whole save

                    duty_field = f"dutytype_{m.id}_{d.isoformat()}"
                    v_duty = request.POST.get(duty_field, '').strip()
                    if v_duty and v_duty not in duty_type_ids:
                        v_duty = ''  # bogus/cross-org duty type id - ignore, not a hard error

                    existing = m.temporary_shift_assignment_for(d)
                    existing_value = '' if existing is None else ('off' if existing.shift_id is None else str(existing.shift_id))
                    existing_duty_value = '' if (existing is None or existing.duty_type_id is None) else str(existing.duty_type_id)
                    if v == existing_value and v_duty == existing_duty_value:
                        continue

                    if existing is not None:
                        existing.is_active = False
                        existing.save(update_fields=['is_active'])

                    duty_obj = DutyType.objects.filter(org=org, pk=v_duty, is_active=True).first() if v_duty else None
                    duty_suffix = f" / {duty_obj.name}" if duty_obj else ""

                    if v == 'off':
                        TemporaryShiftAssignment.objects.create(
                            org=org, member=m, start_date=d, end_date=d, shift=None, duty_type=duty_obj, created_by=request.user,
                        )
                        change_notes.append(f"{d.strftime('%a')}: Off{duty_suffix}")
                    elif v:
                        shift = Shift.objects.filter(org=org, pk=v, is_active=True).first()
                        if shift:
                            TemporaryShiftAssignment.objects.create(
                                org=org, member=m, start_date=d, end_date=d, shift=shift, duty_type=duty_obj, created_by=request.user,
                            )
                            change_notes.append(f"{d.strftime('%a')}: {shift.name}{duty_suffix}")
                    else:
                        # v == '' (Default). A duty type alone, with no shift/off decision
                        # for the day, has nothing to attach to - shift=None already means
                        # "off duty" on this model, so it can't also mean "keep the normal
                        # pattern" - deliberately not saved rather than risk marking someone
                        # off duty just because a duty type was picked.
                        change_notes.append(f"{d.strftime('%a')}: Default")

                if change_notes:
                    changed_members += 1
                    MemberHistory.objects.create(
                        org=org, member=m, action='duty_roster_change', changed_by=request.user,
                        description=f"Weekly duty roster ({week_start} – {week_days[-1]}): " + ', '.join(change_notes),
                    )

        if changed_members:
            messages.success(request, f"Saved the weekly duty roster — updated {changed_members} member(s).")
        else:
            messages.info(request, "No changes to save.")
        return redirect(f"{reverse('schooladmin:duty_roster')}?view=week&week_start={week_start.isoformat()}")


# ═══════════════════════════════════════════════════════════════════════════════
# FACIAL RECOGNITION ATTENDANCE (client-side face-api.js)
# ═══════════════════════════════════════════════════════════════════════════════
import json as _json
from django.http import JsonResponse
from handle.models import MemberFace


class FaceEnrollListView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'face_attendance'
    template_name = 'admin/face/enroll_list.html'

    def get(self, request):
        org = _task_org(request)
        members = member.objects.filter(org=org, status='active').select_related('face_profile').order_by('name')
        rows = [{'m': m, 'face': getattr(m, 'face_profile', None)} for m in members]
        return render(request, self.template_name, {'org': org, 'rows': rows})


class FaceEnrollView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'face_attendance'
    template_name = 'admin/face/enroll.html'

    def get(self, request, member_id):
        org = _task_org(request)
        memb = get_object_or_404(member, pk=member_id, org=org)
        face = getattr(memb, 'face_profile', None)
        return render(request, self.template_name, {'org': org, 'memb': memb, 'face': face})

    def post(self, request, member_id):
        org = _task_org(request)
        memb = get_object_or_404(member, pk=member_id, org=org)
        try:
            descriptors = _json.loads(request.POST.get('descriptors', '[]'))
        except ValueError:
            return JsonResponse({'ok': False, 'error': 'Invalid descriptor data.'}, status=400)

        # Validate: list of 128-float arrays
        clean = []
        for d in descriptors:
            if isinstance(d, list) and len(d) == 128:
                clean.append([float(x) for x in d])
        if not clean:
            return JsonResponse({'ok': False, 'error': 'No valid face captured. Try again in good lighting.'}, status=400)

        face, _ = MemberFace.objects.get_or_create(member=memb, defaults={'org': org})
        face.org = org
        face.descriptors = clean
        img = request.FILES.get('sample_image')
        if img:
            face.sample_image = img
        face.save()
        return JsonResponse({'ok': True, 'samples': len(clean)})


class FaceEnrolledDataView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """JSON of all enrolled face descriptors for the org (used by the attendance page)."""
    required_feature = 'face_attendance'

    def get(self, request):
        org = _task_org(request)
        faces = MemberFace.objects.filter(org=org).select_related('member')
        data = [
            {'member_id': f.member_id, 'name': f.member.name, 'descriptors': f.descriptors or []}
            for f in faces if f.descriptors
        ]
        return JsonResponse({'members': data})


class FaceAttendanceView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'face_attendance'
    template_name = 'admin/face/attendance.html'

    def get(self, request):
        org = _task_org(request)
        return render(request, self.template_name, {'org': org})

    def post(self, request):
        org = _task_org(request)
        member_id = request.POST.get('member_id')
        memb = member.objects.filter(pk=member_id, org=org).exclude(status='dumped').first()
        if not memb:
            return JsonResponse({'ok': False, 'error': 'Unknown member.'}, status=404)
        if not MemberFace.objects.filter(member=memb, org=org).exists():
            return JsonResponse({'ok': False, 'error': 'Member is not enrolled.'}, status=400)

        now = timezone.now()
        # Prevent duplicate facial punches within 60s.
        recent = AttendanceRecord.objects.filter(
            mem=memb, org=org, attendance_method='facial',
            scanned_time__gte=now - datetime.timedelta(seconds=60)
        ).exists()
        if recent:
            return JsonResponse({'ok': True, 'duplicate': True, 'name': memb.name})

        AttendanceRecord.objects.create(
            mem=memb, org=org, scanned_time=now, attendance_method='facial',
        )
        return JsonResponse({'ok': True, 'name': memb.name, 'time': timezone.localtime(now).strftime('%I:%M:%S %p')})


# ═══════════════════════════════════════════════════════════════════════════════
# LIVE LOCATION TRACKING (marketer / field-staff trail)
# ═══════════════════════════════════════════════════════════════════════════════
from handle.models import LocationPing


class LiveTrackingView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    required_feature = 'field_visits'
    template_name = 'admin/tracking/live.html'

    def get(self, request):
        org = _task_org(request)
        members = member.objects.filter(org=org, status='active').order_by('name')
        rows = []
        for m in members:
            last = LocationPing.objects.filter(member=m).order_by('-tracked_at').first()
            rows.append({'m': m, 'last': last})
        selected_id = request.GET.get('member', '')
        selected = members.filter(id=selected_id).first() if selected_id else None
        return render(request, self.template_name, {
            'org': org, 'rows': rows, 'selected': selected,
            'date': request.GET.get('date', ''),
        })

    def post(self, request):
        """Toggle live tracking for a member."""
        org = _task_org(request)
        memb = get_object_or_404(member, pk=request.POST.get('member_id'), org=org)
        memb.live_tracking_enabled = not memb.live_tracking_enabled
        memb.save(update_fields=['live_tracking_enabled'])
        messages.success(
            request,
            f"Live tracking {'enabled' if memb.live_tracking_enabled else 'disabled'} for {memb.name}."
        )
        return redirect('schooladmin:live_tracking')


class LiveTrackingDataView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """JSON trail (ordered pings) for a member on a given date."""
    required_feature = 'field_visits'

    def get(self, request):
        org = _task_org(request)
        memb = member.objects.filter(pk=request.GET.get('member').exclude(status='dumped'), org=org).first()
        if not memb:
            return JsonResponse({'points': []})
        pings = LocationPing.objects.filter(member=memb, org=org)
        date_str = request.GET.get('date', '')
        if date_str:
            pings = pings.filter(tracked_at__date=date_str)
        else:
            latest = pings.order_by('-tracked_at').first()
            if latest:
                pings = pings.filter(tracked_at__date=timezone.localtime(latest.tracked_at).date())
        points = [
            {
                'lat': p.latitude, 'lng': p.longitude,
                'acc': p.accuracy_meters,
                'time': timezone.localtime(p.tracked_at).strftime('%H:%M:%S'),
            }
            for p in pings.order_by('tracked_at')
        ]

        # Total walked distance (metres) via haversine over consecutive points.
        import math
        total_m = 0.0
        for a, b in zip(points, points[1:]):
            R = 6371000.0
            dlat = math.radians(b['lat'] - a['lat'])
            dlon = math.radians(b['lng'] - a['lng'])
            h = (math.sin(dlat / 2) ** 2
                 + math.cos(math.radians(a['lat'])) * math.cos(math.radians(b['lat'])) * math.sin(dlon / 2) ** 2)
            total_m += 2 * R * math.asin(min(1.0, math.sqrt(h)))

        return JsonResponse({
            'points': points, 'name': memb.name,
            'total_meters': round(total_m, 1),
            'total_km': round(total_m / 1000.0, 2),
        })


class LiveTrackingDetailView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Full movement report for one member on a chosen date: user details, totals,
    area covered and the mapped trail."""
    required_feature = 'field_visits'
    template_name = 'admin/tracking/detail.html'

    def get(self, request, member_id):
        import math, json as _json, datetime as _dt
        org = _task_org(request)
        memb = get_object_or_404(member, pk=member_id, org=org)

        pings_all = LocationPing.objects.filter(member=memb, org=org)
        date_str = request.GET.get('date', '')
        if not date_str:
            latest = pings_all.order_by('-tracked_at').first()
            if latest:
                date_str = timezone.localtime(latest.tracked_at).date().isoformat()

        pings = pings_all
        if date_str:
            pings = pings.filter(tracked_at__date=date_str)
        pings = list(pings.order_by('tracked_at'))

        points = [{
            'lat': p.latitude, 'lng': p.longitude,
            'acc': p.accuracy_meters,
            'time': timezone.localtime(p.tracked_at).strftime('%H:%M:%S'),
        } for p in pings]

        # Distance (haversine)
        total_m = 0.0
        for a, b in zip(points, points[1:]):
            R = 6371000.0
            dlat = math.radians(b['lat'] - a['lat'])
            dlon = math.radians(b['lng'] - a['lng'])
            h = (math.sin(dlat / 2) ** 2
                 + math.cos(math.radians(a['lat'])) * math.cos(math.radians(b['lat'])) * math.sin(dlon / 2) ** 2)
            total_m += 2 * R * math.asin(min(1.0, math.sqrt(h)))

        # Area covered = bounding-box rectangle (km²)
        area_km2 = 0.0
        if len(points) >= 2:
            lats = [p['lat'] for p in points]
            lngs = [p['lng'] for p in points]
            mean_lat = sum(lats) / len(lats)
            lat_m = (max(lats) - min(lats)) * 111320.0
            lng_m = (max(lngs) - min(lngs)) * 111320.0 * math.cos(math.radians(mean_lat))
            area_km2 = round(abs(lat_m * lng_m) / 1_000_000.0, 4)

        # Tracking duration (first → last ping)
        duration_str = "—"
        if len(pings) >= 2:
            secs = (pings[-1].tracked_at - pings[0].tracked_at).total_seconds()
            h, rem = divmod(int(secs), 3600)
            m, _ = divmod(rem, 60)
            duration_str = (f"{h}h " if h else "") + f"{m}m"

        # Work hours from attendance punches on that date (first → last punch)
        work_str = "—"
        if date_str:
            try:
                d = _dt.datetime.strptime(date_str, "%Y-%m-%d").date()
                punches = memb.day_punch_times(d)
                if len(punches) >= 2:
                    s = _dt.datetime.combine(d, punches[0])
                    e = _dt.datetime.combine(d, punches[-1])
                    wh = (e - s).total_seconds() / 3600.0
                    work_str = f"{wh:.2f} h"
            except Exception:
                pass

        ctx = {
            'org': org, 'memb': memb, 'date': date_str,
            'points': points,
            'point_count': len(points),
            'total_m': round(total_m, 1),
            'total_km': round(total_m / 1000.0, 2),
            'area_km2': area_km2,
            'duration_str': duration_str,
            'work_str': work_str,
            'first_time': points[0]['time'] if points else '—',
            'last_time': points[-1]['time'] if points else '—',
            'attendance_covered': bool(date_str and AttendanceRecord.objects.filter(
                org=org, mem=memb, scanned_time__date=date_str,
            ).exists()),
        }
        return render(request, self.template_name, ctx)


class LiveTrackingAttendanceView(AdminRequiredMixin, LoginRequiredMixin, FeatureRequiredMixin, View):
    """Convert a verified GPS work trail into attendance without fabricated times."""
    required_feature = 'field_visits'

    def post(self, request, member_id):
        import datetime as _dt
        from handle.models import LocationPing, MemberHistory
        org = _task_org(request)
        memb = get_object_or_404(member, pk=member_id, org=org)
        try:
            work_date = _dt.date.fromisoformat(request.POST.get('date', ''))
        except ValueError:
            messages.error(request, 'Choose a valid tracking date.')
            return redirect('schooladmin:live_tracking_detail', member_id=member_id)

        pings = list(LocationPing.objects.filter(
            org=org, member=memb, tracked_at__date=work_date,
        ).order_by('tracked_at'))
        if not pings:
            messages.error(request, 'Attendance cannot be marked without location evidence for that date.')
        elif AttendanceRecord.objects.filter(
            org=org, mem=memb, scanned_time__date=work_date,
        ).exists():
            messages.info(request, 'Attendance is already covered for this member and date.')
        else:
            with transaction.atomic():
                AttendanceRecord.objects.create(
                    org=org, mem=memb, scanned_time=pings[0].tracked_at,
                    attendance_method='field_visit',
                )
                if len(pings) > 1 and pings[-1].tracked_at != pings[0].tracked_at:
                    AttendanceRecord.objects.create(
                        org=org, mem=memb, scanned_time=pings[-1].tracked_at,
                        attendance_method='field_visit',
                    )
                MemberHistory.objects.create(
                    org=org, member=memb, action='tracking_attendance_approved',
                    field_name='attendance',
                    description=f'Outside-work attendance approved from live tracking for {work_date}',
                    metadata={
                        'date': work_date.isoformat(), 'ping_count': len(pings),
                        'first_ping': pings[0].tracked_at.isoformat(),
                        'last_ping': pings[-1].tracked_at.isoformat(),
                    },
                    changed_by=request.user,
                )
            messages.success(request, 'Outside-work attendance marked from the verified tracking trail.')
        return redirect(f"{reverse('schooladmin:live_tracking_detail', args=[member_id])}?date={work_date}")


# ═══════════════════════════════════════════════════════════════════════════════
# MONTHLY ATTENDANCE REPORT (matrix + summary, all employees at once)
# ═══════════════════════════════════════════════════════════════════════════════
class MonthlyAttendanceReportView(AdminRequiredMixin, LoginRequiredMixin, View):
    template_name = 'admin/monthly_report.html'

    def _range(self, request):
        import datetime as _dt
        today = timezone.localdate()

        # Get org and nepali_enabled from request
        try:
            org = request.user.schooladmin.org if hasattr(request.user, 'schooladmin') else None
            nepali_enabled = getattr(org, 'nepali_date', False) if org else False
        except:
            nepali_enabled = False

        preset = request.GET.get('preset', '')
        from_str = request.GET.get('from_date', '')
        to_str = request.GET.get('to_date', '')
        from_np = request.GET.get('from_date_np', '')
        to_np = request.GET.get('to_date_np', '')

        if nepali_enabled and preset in ('last_month', 'this_month'):
            today_np = nepali_datetime.date.from_datetime_date(today)
            if preset == 'this_month':
                start = nepali_datetime.date(
                    today_np.year, today_np.month, 1,
                ).to_datetime_date()
                return start, today, 'this_month'

            if today_np.month == 1:
                previous_year, previous_month = today_np.year - 1, 12
            else:
                previous_year, previous_month = today_np.year, today_np.month - 1
            start = nepali_datetime.date(
                previous_year, previous_month, 1,
            ).to_datetime_date()
            # The day before the current BS month starts is the final day of
            # the previous BS month, regardless of its 29/30/31/32-day length.
            current_month_start = nepali_datetime.date(
                today_np.year, today_np.month, 1,
            ).to_datetime_date()
            end = current_month_start - _dt.timedelta(days=1)
            return start, end, 'last_month'

        if preset == 'last_month':
            first_this = today.replace(day=1)
            end = first_this - _dt.timedelta(days=1)
            start = end.replace(day=1)
            return start, end, 'last_month'
        if preset == 'this_month':
            return today.replace(day=1), today, 'this_month'

        # Try Nepali dates first if enabled
        if nepali_enabled and from_np and to_np:
            try:
                y, m, d = map(int, from_np.replace('/', '-').strip().split('-'))
                start = nepali_datetime.date(y, m, d).to_datetime_date()
                y, m, d = map(int, to_np.replace('/', '-').strip().split('-'))
                end = nepali_datetime.date(y, m, d).to_datetime_date()
                if start <= end:
                    return start, end, ''
            except (ValueError, TypeError):
                pass

        # Fall back to AD dates
        if from_str and to_str:
            try:
                start = _dt.datetime.strptime(from_str, "%Y-%m-%d").date()
                end = _dt.datetime.strptime(to_str, "%Y-%m-%d").date()
                if start <= end:
                    return start, end, ''
            except ValueError:
                pass
        # Default to the organisation's configured calendar month.
        if nepali_enabled:
            today_np = nepali_datetime.date.from_datetime_date(today)
            return (
                nepali_datetime.date(today_np.year, today_np.month, 1).to_datetime_date(),
                today,
                'this_month',
            )
        return today.replace(day=1), today, 'this_month'

    def get(self, request):
        import datetime as _dt
        from schooladmin.payroll_service import calculate_attendance_stats
        org = _task_org(request)
        start, end, active_preset = self._range(request)

        nepali_enabled = bool(getattr(org, 'nepali_date', False))

        # Day columns. Queries remain Gregorian, but presentation follows the
        # organisation calendar so a BS month is not split across AD months.
        days = []
        d = start
        while d <= end:
            if nepali_enabled:
                np_date = nepali_datetime.date.from_datetime_date(d)
                days.append({
                    'date': d,
                    'date_np': str(np_date),
                    'dow': np_date.strftime('%a'),
                    'dom': f'{np_date.day:02d}',
                })
            else:
                days.append({
                    'date': d, 'date_np': '',
                    'dow': d.strftime('%a'), 'dom': d.strftime('%d'),
                })
            d += _dt.timedelta(days=1)

        from school.hierarchy import get_accessible_members
        members = get_accessible_members(request.user, org).filter(status='active').select_related('section').order_by('name')

        # Branch / classification filters
        branch_id = request.GET.get('branch', '')
        classification_id = request.GET.get('classification', '')
        if branch_id:
            members = members.filter(branch_id=branch_id)
        if classification_id:
            members = members.filter(classification_id=classification_id)

        # "F" (field visit) column — only for orgs/users with field-visit access.
        show_field_visits = has_feature(org, 'field_visits') and has_perm(request.user, 'can_view_field_visits')

        rows = []
        for m in members:
            stats, logs = calculate_attendance_stats(
                m, start, end, org, nepali_enabled=nepali_enabled, show_field_visits=show_field_visits,
            )
            cells = [{'code': l['code'], 'cls': l['code_class']} for l in logs]
            counts = {'P': 0, 'A': 0, 'L': 0, 'H': 0, 'W': 0, 'F': 0}
            late_days = early_days = 0
            for l in logs:
                counts[l['code']] = counts.get(l['code'], 0) + 1
                if float(l.get('late_hours_dec') or 0) > 0:
                    late_days += 1
                if float(l.get('early_hours_dec') or 0) > 0:
                    early_days += 1
            # A field-visit day ('F') is a present/worked day that's just
            # displayed with a different code — count it alongside 'P' here
            # so reclassifying a day from P to F never changes the
            # attendance percentage, only which badge shows in the grid.
            attended_days = counts['P'] + counts['F']
            expected_days = attended_days + counts['A']
            attendance_pct = round(attended_days / expected_days * 100, 1) if expected_days else None
            rows.append({
                'id': m.id, 'name': m.name, 'cells': cells,
                'photo_url': m.photo.url if m.photo else '',
                'section': m.section.name if m.section_id else '',
                'total': len(logs),
                'present': counts['P'], 'absent': counts['A'], 'leave': counts['L'],
                'paid_leave': stats['days_paid_leave'], 'unpaid_leave': stats['days_unpaid_leave'],
                'holiday': counts['H'], 'weekend': counts['W'], 'field_visit': counts['F'],
                'worked': round(float(stats['total_hours_worked']), 2),
                'ot': round(float(stats['total_overtime_hours']), 2),
                'late_days': late_days, 'early_days': early_days,
                'attendance_pct': attendance_pct,
            })

        # CSV export of the summary
        if request.GET.get('export') == 'csv':
            resp = HttpResponse(content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="attendance_summary_{start}_{end}.csv"'
            w = csv.writer(resp)
            headers = ['Code', 'Name', 'Total', 'Present', 'Absent', 'Leave', 'Holiday',
                       'Weekend', 'Worked (H)', 'OT (H)', 'Late-in days', 'Early-out days']
            if show_field_visits:
                headers.insert(6, 'Field Visit')
            headers += ['Section', 'Paid Leave', 'Unpaid Leave', 'Attendance %']
            w.writerow(headers)
            for r in rows:
                row = [r['id'], r['name'], r['total'], r['present'], r['absent'], r['leave'],
                       r['holiday'], r['weekend'], r['worked'], r['ot'], r['late_days'], r['early_days']]
                if show_field_visits:
                    row.insert(6, r['field_visit'])
                row += [r['section'], r['paid_leave'], r['unpaid_leave'],
                        r['attendance_pct'] if r['attendance_pct'] is not None else '']
                w.writerow(row)
            return resp

        start_np = end_np = ''
        period_label = f"{start.strftime('%b %d, %Y')} → {end.strftime('%b %d, %Y')} (AD)"
        if nepali_enabled:
            start_np_date = nepali_datetime.date.from_datetime_date(start)
            end_np_date = nepali_datetime.date.from_datetime_date(end)
            start_np, end_np = str(start_np_date), str(end_np_date)
            period_label = (
                f"{start_np_date.strftime('%B %d, %Y')} → "
                f"{end_np_date.strftime('%B %d, %Y')} (BS)"
            )

        from school.print_settings import get_print_preference
        return render(request, self.template_name, {
            'org': org, 'days': days, 'rows': rows,
            'start': start, 'end': end, 'active_preset': active_preset,
            'period_label': period_label,
            'filters': {
                'from_date': start.isoformat(), 'to_date': end.isoformat(),
                'from_date_np': start_np, 'to_date_np': end_np,
            },
            'branches': Branch.objects.filter(org=org, status='active').order_by('name'),
            'classifications': Classification.objects.filter(org=org).order_by('name'),
            'selected_branch': branch_id,
            'selected_classification': classification_id,
            'nepali_enabled': nepali_enabled,
            'show_field_visits': show_field_visits,
            'print_preference': get_print_preference(request.user, 'monthly_report', org=org),
            'calendar_members': [{'id': m.id, 'name': m.name} for m in members],
            'leave_types': list(LeaveType.objects.filter(org=org).values('id', 'name', 'is_paid')),
        })


class MemberCalendarDataView(AdminRequiredMixin, LoginRequiredMixin, View):
    """AJAX: full calendar-wise data for one member over a date range, for
    the Monthly Report's Calendar View tab (and reused by the Gap Report
    calendar). Kept as its own small endpoint — rather than embedding every
    member's full daily log in the Monthly Report page itself — so opening
    the page for an org with 100+ members doesn't multiply the payload by
    every member's whole month of per-day detail just to show one member's
    calendar at a time."""

    def get(self, request, member_id):
        org = _task_org(request)
        mem = get_object_or_404(member, id=member_id, org=org)
        nepali_enabled = bool(getattr(org, 'nepali_date', False))

        from_str = request.GET.get('from_date', '')
        to_str = request.GET.get('to_date', '')
        try:
            start = datetime.datetime.strptime(from_str, "%Y-%m-%d").date()
            end = datetime.datetime.strptime(to_str, "%Y-%m-%d").date()
        except ValueError:
            today = timezone.localdate()
            start, end = today.replace(day=1), today

        from schooladmin.calendar_service import build_member_calendar, serialize_day
        days = build_member_calendar(mem, start, end, org, nepali_enabled=nepali_enabled)

        return JsonResponse({
            'member': {
                'id': mem.id, 'name': mem.name,
                'photo_url': mem.photo.url if mem.photo else '',
                'classification': mem.classification.name if mem.classification_id else '',
            },
            'start': start.isoformat(), 'end': end.isoformat(),
            'days': [serialize_day(d) for d in days],
        })


class DayQuickActionView(AdminRequiredMixin, LoginRequiredMixin, View):
    """AJAX: click a calendar day -> mark it as Leave or attach a Note,
    without leaving the report page. Shared by Monthly Report's Calendar
    View tab and the Gap Report calendar."""

    def post(self, request):
        org = _task_org(request)
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except (ValueError, UnicodeDecodeError):
            return JsonResponse({'ok': False, 'message': 'Invalid request.'}, status=400)

        member_id = payload.get('member_id')
        date_str = payload.get('date')
        action = payload.get('action')
        mem = get_object_or_404(member, id=member_id, org=org)
        try:
            day = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return JsonResponse({'ok': False, 'message': 'Invalid date.'}, status=400)

        if action == 'leave':
            leave_type_id = payload.get('leave_type_id')
            reason = (payload.get('reason') or 'Admin assigned leave.').strip()
            l_type = get_object_or_404(LeaveType, id=leave_type_id, org=org)
            ok, msg, _leave = create_admin_leave(org, mem, l_type, day, day, reason)
            status_code = 200 if ok else 422
            return JsonResponse({'ok': ok, 'message': msg}, status=status_code)

        if action == 'note':
            text = (payload.get('text') or '').strip()
            if not text:
                return JsonResponse({'ok': False, 'message': 'Note text is required.'}, status=422)
            from handle.models import DailyNote
            DailyNote.objects.update_or_create(
                member=mem, date=day,
                defaults={'org': org, 'text': text[:500], 'created_by': request.user},
            )
            return JsonResponse({'ok': True, 'message': 'Note saved.'})

        return JsonResponse({'ok': False, 'message': 'Unknown action.'}, status=400)


class DynamicFeatureView(AdminRequiredMixin, View):
    """
    Generic landing page for any superadmin-defined DynamicFeature — the
    zero-code fallback so a brand new feature is reachable from the sidebar
    immediately, before (if ever) a bespoke module is built for it.
    """
    template_name = 'admin/dynamic_feature.html'

    def get(self, request, feature_key):
        from handle.models import DynamicFeature
        from school.features import has_feature
        org = _get_org(request)
        feature = get_object_or_404(DynamicFeature, key=feature_key, is_active=True)
        if not has_feature(org, feature_key):
            return render(request, '403.html', {
                'reason': f"The '{feature.label}' module is not enabled for your organization.",
            }, status=403)
        return render(request, self.template_name, {
            'org': org, 'feature': feature, 'permissions': feature.permissions.all(),
        })


# ═══════════════════════════════════════════════════════════════════════════
# NOTICE BOARD
# ═══════════════════════════════════════════════════════════════════════════

_NOTICE_AUDIENCE_FK = {
    'branch': 'branch', 'department': 'classification', 'section': 'section',
    'course': 'course', 'shift': 'shift', 'member': 'target_member',
}


def _notice_form_context(org, notice=None):
    """Choice lists for the create/edit form, all scoped to this org.

    `selected_target_id` is resolved here (once, server-side) rather than in
    the template: a chained `{{ a|default:b|default:c }}` filter only
    protects the *first* variable from a None/missing lookup — Django
    resolves every subsequent `default:` argument via Variable.resolve()
    directly, which does not swallow VariableDoesNotExist. With `notice=None`
    on the create page, that raised an uncaught exception. Computing the
    single winning id in Python sidesteps that template gotcha entirely and
    works identically for create (notice=None) and edit.
    """
    from handle.models import Notice, Shift

    selected_target_id = ''
    if notice is not None:
        for attr in ('branch_id', 'classification_id', 'section_id', 'course_id', 'shift_id', 'target_member_id'):
            value = getattr(notice, attr, None)
            if value:
                selected_target_id = value
                break

    return {
        'org': org,
        'priority_choices': Notice.PRIORITY_CHOICES,
        'audience_choices': Notice.AUDIENCE_CHOICES,
        'branches': Branch.objects.filter(org=org, status='active').order_by('name'),
        'classifications': Classification.objects.filter(org=org, status='active').order_by('name'),
        'sections': Section.objects.filter(org=org, status='active').order_by('classification__name', 'name'),
        'courses': Course.objects.filter(org=org, status='active').order_by('name'),
        'shifts': Shift.objects.filter(org=org, is_active=True).order_by('name'),
        'members': member.objects.filter(org=org, status='active').order_by('name'),
        'selected_target_id': selected_target_id,
    }


def _parse_dt(raw):
    """Accept the datetime-local format the form posts, else None."""
    raw = (raw or '').strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.datetime.strptime(raw, fmt)
            return timezone.make_aware(parsed) if timezone.is_naive(parsed) else parsed
        except ValueError:
            continue
    return None


def _apply_notice_post(notice, request, org):
    """Copy POSTed fields onto `notice`. Only the FK matching the chosen
    audience is kept — the others are cleared so a notice can never carry a
    stale target from a previous edit."""
    from handle.models import Notice, Shift

    notice.title = (request.POST.get('title') or '').strip()
    notice.body = (request.POST.get('body') or '').strip()

    priority = request.POST.get('priority', 'normal')
    notice.priority = priority if priority in dict(Notice.PRIORITY_CHOICES) else 'normal'

    audience = request.POST.get('audience', 'org')
    notice.audience = audience if audience in dict(Notice.AUDIENCE_CHOICES) else 'org'

    # Clear every target, then set only the relevant one.
    notice.branch = notice.classification = notice.section = None
    notice.course = notice.shift = notice.target_member = None

    target_id = request.POST.get('target_id') or None
    if target_id and notice.audience in _NOTICE_AUDIENCE_FK:
        # Every lookup is org-scoped, so a tampered id can't target another org.
        model_for = {
            'branch': (Branch, 'branch'),
            'department': (Classification, 'classification'),
            'section': (Section, 'section'),
            'course': (Course, 'course'),
            'shift': (Shift, 'shift'),
            'member': (member, 'target_member'),
        }[notice.audience]
        model_cls, attr = model_for
        obj = model_cls.objects.filter(org=org, id=target_id).first()
        if obj is None:
            return f"The selected {notice.get_audience_display().lower()} was not found in your organization."
        setattr(notice, attr, obj)
    elif notice.audience in _NOTICE_AUDIENCE_FK:
        return f"Please choose a target for “{notice.get_audience_display()}”."

    publish_at = _parse_dt(request.POST.get('publish_at'))
    notice.publish_at = publish_at or timezone.now()
    notice.expires_at = _parse_dt(request.POST.get('expires_at'))

    if notice.expires_at and notice.expires_at <= notice.publish_at:
        return "Expiry must be after the publish date."

    notice.send_email = request.POST.get('send_email') == 'on'

    if request.FILES.get('attachment'):
        notice.attachment = request.FILES['attachment']

    if not notice.title or not notice.body:
        return "Title and message are both required."
    return None


def _dispatch_notice_email(notice, force=False):
    """Email a notice to its recipients if it's live and email was requested.
    Scheduled notices are skipped — they get sent when they go live."""
    if not notice.send_email or not notice.is_live():
        return 0
    emails = [
        e for e in notice.recipient_members().values_list('email', flat=True) if e
    ]
    if not emails:
        return 0
    from school.email_utils import send_notice_email
    send_notice_email(
        recipient_emails=emails,
        title=notice.title,
        body=notice.body,
        priority=notice.priority,
        org_name=notice.org.name,
        published_on=notice.publish_at.strftime('%B %d, %Y') if notice.publish_at else '',
        org=notice.org,
        related_object_id=notice.id,
        force=force,
    )
    notice.email_sent_at = timezone.now()
    notice.save(update_fields=['email_sent_at'])
    return len(emails)


class NoticeListView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'notices'
    required_perm = 'can_view_notices'
    template_name = 'admin/notices/list.html'

    def get(self, request):
        from handle.models import Notice
        org = _get_org(request)
        qs = Notice.objects.filter(org=org).select_related(
            'branch', 'classification', 'section', 'course', 'shift',
            'target_member', 'created_by',
        )

        state = request.GET.get('state', '')
        priority = request.GET.get('priority', '')
        audience = request.GET.get('audience', '')
        if priority:
            qs = qs.filter(priority=priority)
        if audience:
            qs = qs.filter(audience=audience)

        notices = list(qs)
        if state in ('live', 'scheduled', 'expired'):
            notices = [n for n in notices if n.state() == state]

        counts = {'live': 0, 'scheduled': 0, 'expired': 0}
        print("DEBUG: Counting notice states...")
        print(f"DEBUG: Total notices: {len(notices)}")
        print(f"DEBUG: Filtered by state='{state}', priority='{priority}', audience='{audience}'")
        print(notices)
        for n in qs:
            counts[n.state()] += 1

        from django.core.paginator import Paginator
        page_obj = Paginator(notices, 20).get_page(request.GET.get('page'))

        from handle.models import Notice as N
        return render(request, self.template_name, {
            'org': org,
            'notices': page_obj,
            'page_obj': page_obj,
            'total_count': len(notices),
            'counts': counts,
            'priority_choices': N.PRIORITY_CHOICES,
            'audience_choices': N.AUDIENCE_CHOICES,
            'selected_state': state,
            'selected_priority': priority,
            'selected_audience': audience,
        })


class NoticeCreateView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'notices'
    required_perm = 'can_manage_notices'
    template_name = 'admin/notices/form.html'

    def get(self, request):
        org = _get_org(request)
        ctx = _notice_form_context(org)
        ctx.update({'notice': None, 'is_edit': False})
        return render(request, self.template_name, ctx)

    def post(self, request):
        from handle.models import Notice
        org = _get_org(request)
        notice = Notice(org=org, created_by=request.user)
        error = _apply_notice_post(notice, request, org)
        if error:
            messages.error(request, error)
            ctx = _notice_form_context(org, notice=notice)
            ctx.update({'notice': notice, 'is_edit': False})
            return render(request, self.template_name, ctx)

        notice.save()
        sent = _dispatch_notice_email(notice)
        msg = f"Notice “{notice.title}” created."
        if notice.is_scheduled():
            msg += f" It will publish on {notice.publish_at.strftime('%b %d, %Y %I:%M %p')}."
        if sent:
            msg += f" Emailed to {sent} recipient(s)."
        messages.success(request, msg)
        return redirect('schooladmin:notice_list')


class NoticeEditView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'notices'
    required_perm = 'can_manage_notices'
    template_name = 'admin/notices/form.html'

    def get(self, request, pk):
        from handle.models import Notice
        org = _get_org(request)
        notice = get_object_or_404(Notice, pk=pk, org=org)
        ctx = _notice_form_context(org, notice=notice)
        ctx.update({'notice': notice, 'is_edit': True})
        return render(request, self.template_name, ctx)

    def post(self, request, pk):
        from handle.models import Notice
        org = _get_org(request)
        notice = get_object_or_404(Notice, pk=pk, org=org)
        error = _apply_notice_post(notice, request, org)
        if error:
            messages.error(request, error)
            ctx = _notice_form_context(org, notice=notice)
            ctx.update({'notice': notice, 'is_edit': True})
            return render(request, self.template_name, ctx)

        notice.save()
        # Only send if it hasn't already gone out; dedup in _send_email also
        # guards this, but skipping avoids pointless work on every edit.
        sent = 0 if notice.email_sent_at else _dispatch_notice_email(notice)
        msg = f"Notice “{notice.title}” updated."
        if sent:
            msg += f" Emailed to {sent} recipient(s)."
        messages.success(request, msg)
        return redirect('schooladmin:notice_list')


class NoticeDetailView(FeatureRequiredMixin, PermRequiredMixin, View):
    required_feature = 'notices'
    required_perm = 'can_view_notices'
    template_name = 'admin/notices/detail.html'

    def get(self, request, pk):
        from handle.models import Notice
        org = _get_org(request)
        notice = get_object_or_404(Notice, pk=pk, org=org)
        recipients = notice.recipient_members().order_by('name')
        read_ids = set(notice.reads.values_list('member_id', flat=True))
        rows = [{'member': m, 'has_read': m.id in read_ids} for m in recipients]
        return render(request, self.template_name, {
            'org': org,
            'notice': notice,
            'rows': rows,
            'recipient_count': len(rows),
            'read_count': sum(1 for r in rows if r['has_read']),
        })

    def post(self, request, pk):
        """Resend the notice email on demand."""
        from handle.models import Notice
        org = _get_org(request)
        notice = get_object_or_404(Notice, pk=pk, org=org)
        if request.POST.get('action') == 'resend_email':
            if not notice.is_live():
                messages.error(request, "Only a live notice can be emailed.")
            else:
                notice.send_email = True
                sent = _dispatch_notice_email(notice, force=True)
                if sent:
                    messages.success(request, f"Notice emailed to {sent} recipient(s).")
                else:
                    messages.error(request, "No recipients with an email address on file.")
        return redirect('schooladmin:notice_detail', pk=pk)


@feature_required('notices')
@perm_required('can_manage_notices')
def delete_notice(request, pk):
    from handle.models import Notice
    org = _get_org(request)
    notice = get_object_or_404(Notice, pk=pk, org=org)
    title = notice.title
    notice.delete()
    messages.success(request, f"Notice “{title}” deleted.")
    return redirect('schooladmin:notice_list')
