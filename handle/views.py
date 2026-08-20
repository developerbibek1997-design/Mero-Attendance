
import csv
import re
import uuid
import datetime as dt
from datetime import timedelta

from management.models import CustomUser, Organization
from .models import (
    AttendanceRecord,
    Branch,
    Classification,
    Course,
    Device,
    FinancialTransaction,
    MemberHistory,
    Section,
    Staff,
    StaffDocument,
    StockCategory,
    StockItem,
    StockMovement,
    TransactionCategory,
    member,
    resolve_default_shift,
)
from .member_archive import archive_member, restore_member
from .services import provision_staff_account
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import redirect, render, get_object_or_404
from django.views import View
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse
from django.core.mail import send_mail
from django.conf import settings
from django.db import transaction, IntegrityError
from django.utils import timezone
from .forms import (
    BranchForm,
    ClassificationForm,
    CourseForm,
    DeviceForm,
    FinancialTransactionForm,
    FormChangePassword,
    MemberForm,
    SectionForm,
    StockCategoryForm,
    StockItemForm,
    StockMovementForm,
    TransactionCategoryForm,
)
# from zk import ZK
from django.db.models import Q, Sum
from school.decorators import perm_required, PermRequiredMixin

def current_org_for_user(user):
    if user.user_type == "2" and hasattr(user, 'schooladmin'):
        return user.schooladmin.org
    if user.user_type == "3" and hasattr(user, 'staff'):
        return user.staff.org
    return None


class PremiumNotificationListView(LoginRequiredMixin, View):
    """Shared notification centre for school admin, staff, teacher and student."""
    template_name = 'notifications/list.html'

    def get(self, request):
        from django.core.paginator import Paginator
        from handle.notifications import notifications_for_user

        org = current_org_for_user(request.user)
        if org is None:
            messages.error(request, 'Notifications are not available for this account.')
            return redirect('management:homepage')

        notifications = notifications_for_user(request.user, org)
        selected_filter = request.GET.get('filter', 'all')
        if selected_filter == 'unread':
            notifications = notifications.filter(is_read=False)
        elif selected_filter == 'tasks':
            notifications = notifications.filter(event_type__startswith='task_')
        elif selected_filter == 'academic':
            notifications = notifications.exclude(event_type__startswith='task_')
        else:
            selected_filter = 'all'

        paginator = Paginator(notifications, 20)
        return render(request, self.template_name, {
            'org': org,
            'notifications': paginator.get_page(request.GET.get('page')),
            'selected_filter': selected_filter,
            'unread_total': notifications_for_user(
                request.user, org,
            ).filter(is_read=False).count(),
        })


@login_required(login_url='/login/')
def open_notification(request, pk):
    """Mark an owned notification read and follow only a safe local action URL."""
    from django.utils.http import url_has_allowed_host_and_scheme
    from handle.notifications import mark_read, notifications_for_user

    org = current_org_for_user(request.user)
    notification = get_object_or_404(
        notifications_for_user(request.user, org), pk=pk,
    )
    mark_read(notification)
    target = (notification.link_url or '').strip()
    if (
        target.startswith('/')
        and not target.startswith('//')
        and url_has_allowed_host_and_scheme(
            target,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        )
    ):
        return redirect(target)
    return redirect('handle:notifications')


@login_required(login_url='/login/')
def mark_all_notifications_read(request):
    from handle.notifications import notifications_for_user

    if request.method != 'POST':
        return JsonResponse({'detail': 'Method not allowed.'}, status=405)
    org = current_org_for_user(request.user)
    updated = notifications_for_user(
        request.user, org,
    ).filter(is_read=False).update(is_read=True, read_at=timezone.now())
    messages.success(request, f'{updated} notification(s) marked as read.')
    return redirect('handle:notifications')


@login_required(login_url='/login/')
def save_print_preference(request, report_key):
    """AJAX: persist one user's print configuration for a given report."""
    import json
    from school.print_settings import save_print_preference as _save

    if request.method != 'POST':
        return JsonResponse({'detail': 'Method not allowed.'}, status=405)
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'detail': 'Invalid JSON.'}, status=400)
    saved = _save(request.user, report_key, payload.get('settings'))
    return JsonResponse({'settings': saved})


@login_required(login_url='/login/')
def save_org_print_default(request, report_key):
    """AJAX: an admin sets the org-wide default print configuration for a
    report (Organization Profile's Print Defaults section). Admin-only —
    unlike save_print_preference, this changes what every user in the org
    sees on their first visit, not just the caller's own preference."""
    import json
    from school.print_settings import save_org_print_default as _save

    if request.method != 'POST':
        return JsonResponse({'detail': 'Method not allowed.'}, status=405)
    org = current_org_for_user(request.user)
    if request.user.user_type != '2' or not org:
        return JsonResponse({'detail': 'Admins only.'}, status=403)
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (ValueError, UnicodeDecodeError):
        return JsonResponse({'detail': 'Invalid JSON.'}, status=400)
    saved = _save(org, report_key, payload.get('settings'))
    return JsonResponse({'settings': saved})


def _authoritative_member_dates(post_data, org):
    """`date_of_birth`/`admission_date` go through the BS date picker
    (static/assets/js/global-date-picker.js), which converts BS -> AD
    client-side for responsiveness. That conversion is only exact within a
    verified table of recent years; for older dates (routine for a date of
    birth) it falls back to an approximation. Re-derive the AD value here
    from the raw BS text using the authoritative `nepali_datetime` library —
    same pattern as staff/views.py's `_ts_to_ad` for timesheets — so what
    gets saved is always correct regardless of the browser-side estimate."""
    if not org.nepali_date:
        return post_data
    from school.nepali_utils import from_bs_display
    corrected = post_data.copy()
    for field in ('date_of_birth', 'admission_date'):
        raw_bs = post_data.get(f'{field}_np_display')
        if raw_bs:
            ad_date = from_bs_display(raw_bs)
            if ad_date:
                corrected[field] = ad_date.isoformat()
    return corrected


class AddMember(PermRequiredMixin, View):
    required_perm = 'can_add_members'
    form_class = MemberForm
    template_name = 'handle/addMember.html'
    SESSION_TOKEN_KEY = 'add_member_token'

    def get(self, request, *args, **kwargs):
        from school.hierarchy import get_accessible_branches
        auser = request.user
        org = current_org_for_user(auser)

        total_member = member.objects.filter(org=org).exclude(status='dumped').count()
        form = self.form_class(org=org, user=auser)

        # One-time token: a stale/back-cached copy of this page can never be
        # resubmitted successfully once the matching session value has moved on.
        form_token = uuid.uuid4().hex
        request.session[self.SESSION_TOKEN_KEY] = form_token

        dist = {
            'form': form,
            'org': org,
            'classi': Classification.objects.filter(org=org, status='active'),
            'branches': get_accessible_branches(auser, org),
            'sections': Section.objects.filter(org=org, status='active'),
            'courses': Course.objects.filter(org=org, status='active'),
            'total_member': total_member,
            'form_token': form_token,
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        auser = request.user
        org = current_org_for_user(auser)

        # Consume the token immediately so a double-click or a resubmit from
        # browser back/forward cache can't create a second member.
        expected_token = request.session.pop(self.SESSION_TOKEN_KEY, None)
        submitted_token = request.POST.get('form_token')
        if not submitted_token or submitted_token != expected_token:
            messages.warning(request, "This form was already submitted. Check the member list before adding again.")
            return HttpResponseRedirect(reverse('handle:addMember'))

        form = self.form_class(_authoritative_member_dates(request.POST, org), request.FILES, org=org, user=auser)

        if not form.is_valid():
            messages.error(request, "Failed adding Member: " + form.errors.as_text())
            return HttpResponseRedirect(reverse('handle:addMember'))

        current_count = member.objects.filter(org=org).exclude(status='dumped').count()
        if current_count >= org.member_limit:
            messages.error(request, f"Member limit reached ({org.member_limit}). Upgrade your plan or free up a slot to add more.")
            return HttpResponseRedirect(reverse('handle:addMember'))

        staff_result = None
        try:
            with transaction.atomic():
                new_form = form.save(commit=False)

                # Generate Unique Device ID
                device_id_candidate = int(member.objects.filter(org=org).exclude(status='dumped').count()) + 1
                while member.objects.filter(org=org, device_id=device_id_candidate).exclude(status='dumped').exists():
                    device_id_candidate += 1

                new_form.device_id = device_id_candidate
                new_form.org = org

                # Capture tax percentage (assuming it's in your form/model)
                new_form.tax_percentage = request.POST.get('tax_percentage', 0)

                new_form.save()
                form.save_m2m()

                # Save uploaded documents (up to 3)
                for i in ('1', '2', '3'):
                    doc_file = request.FILES.get(f'doc_file_{i}')
                    if doc_file:
                        doc_type = request.POST.get(f'doc_type_{i}', 'other') or 'other'
                        doc_title = request.POST.get(f'doc_title_{i}', doc_file.name) or doc_file.name
                        StaffDocument.objects.create(
                            org=org, member=new_form,
                            document_type=doc_type, title=doc_title, file=doc_file,
                        )

                if new_form.make_staff:
                    staff_result = provision_staff_account(new_form, org)
        except IntegrityError:
            messages.error(request, "Couldn't add this member — a record with the same email or card already exists. Please check and try again.")
            return HttpResponseRedirect(reverse('handle:addMember'))

        # Only email credentials once the transaction above has actually committed.
        if staff_result is None:
            messages.success(request, f"Successfully added Member: {new_form.name}.")
        elif staff_result.status == 'created':
            from school.email_utils import send_welcome_email
            send_welcome_email(new_form.email, new_form.name, staff_result.password, org.name, org=org)
            messages.success(request, f"Successfully added {new_form.name}, granted dashboard access, and sent login credentials via email.")
        else:
            messages.warning(request, f"Member added, but dashboard access wasn't granted: {staff_result.message}")

        return HttpResponseRedirect(reverse('handle:addMember'))



@perm_required('can_edit_members')
def memberEdit(request, id):
    from school.hierarchy import get_accessible_members, get_accessible_branches
    auser = request.user
    org = current_org_for_user(auser)
    mem = get_object_or_404(get_accessible_members(auser, org), id=id)

    if request.method == 'POST':
        previous_email = mem.email
        form = MemberForm(_authoritative_member_dates(request.POST, org), request.FILES, instance=mem, org=org, user=auser)
        if not form.is_valid():
            messages.error(request, "Failed updating Member: " + form.errors.as_text())
            return HttpResponseRedirect(reverse('handle:memberEdit', args=[mem.id]))

        staff_result = None
        email_sync_message = None
        try:
            with transaction.atomic():
                mem = form.save(commit=False)
                mem.org = org
                mem.save()
                form.save_m2m()

                # Save any newly-uploaded documents (up to 3 per submit, same as AddMember)
                for i in ('1', '2', '3'):
                    doc_file = request.FILES.get(f'doc_file_{i}')
                    if doc_file:
                        doc_type = request.POST.get(f'doc_type_{i}', 'other') or 'other'
                        doc_title = request.POST.get(f'doc_title_{i}', doc_file.name) or doc_file.name
                        StaffDocument.objects.create(
                            org=org, member=mem,
                            document_type=doc_type, title=doc_title, file=doc_file,
                        )

                # If this member already has a linked dashboard account, keep
                # its login email in sync — otherwise they'd be locked out the
                # moment their email changes here, since login authenticates
                # against the User's username/email, not the Member's.
                if mem.email and mem.email != previous_email:
                    linked_staff = Staff.objects.filter(member=mem).select_related('admin').first()
                    if linked_staff:
                        linked_user = linked_staff.admin
                        if CustomUser.objects.exclude(pk=linked_user.pk).filter(email=mem.email).exists():
                            email_sync_message = (
                                'warning',
                                f"Member email updated, but the dashboard login still uses '{previous_email}' "
                                f"— '{mem.email}' is already used by another account.",
                            )
                        else:
                            linked_user.email = mem.email
                            linked_user.username = mem.email
                            linked_user.save(update_fields=['email', 'username'])
                            email_sync_message = ('info', "Dashboard login email updated to match.")

                # Handle granting Staff later in edit phase
                if mem.make_staff:
                    staff_result = provision_staff_account(mem, org)
        except IntegrityError:
            messages.error(request, "Couldn't save changes — the email or card is already used by another member.")
            return HttpResponseRedirect(reverse('handle:memberEdit', args=[mem.id]))

        if email_sync_message:
            level, text = email_sync_message
            getattr(messages, level)(request, text)

        # Only email credentials once the transaction above has actually committed.
        if staff_result is not None:
            if staff_result.status == 'created':
                from school.email_utils import send_welcome_email
                send_welcome_email(mem.email, mem.name, staff_result.password, org.name, org=org)
                messages.success(request, f"Granted Staff access to {mem.name} and sent credentials to their email.")
            elif staff_result.status == 'already_exists':
                messages.info(request, staff_result.message)
            else:
                messages.warning(request, staff_result.message)

        messages.success(request, f"Successfully Updated {mem.name}'s details.")
        return HttpResponseRedirect(reverse('handle:memberEdit', args=[mem.id]))

    form = MemberForm(instance=mem, org=org, user=auser)
    dist ={
        'form': form,
        'mem': mem,
        'classi': Classification.objects.filter(org=org, status='active'),
        'sections': Section.objects.filter(org=org, status='active'),
        'branches': get_accessible_branches(auser, org),
        'courses': Course.objects.filter(org=org, status='active'),
        'documents': mem.documents.all().order_by('-uploaded_at'),
        'org': org
    }

    return render(request, "handle/editMember.html", dist)


@perm_required('can_edit_members')
def deleteMemberDocument(request, id):
    auser = request.user
    org = current_org_for_user(auser)
    doc = get_object_or_404(StaffDocument, id=id, org=org)
    member_id = doc.member_id
    doc.file.delete(save=False)
    doc.delete()
    messages.success(request, "Document deleted.")
    return HttpResponseRedirect(reverse('handle:memberEdit', args=[member_id]))

class AddClassification(LoginRequiredMixin, PermRequiredMixin, View):
    required_perm = ('can_manage_branches', 'can_edit_members')
    template_name = "handle/addClassification.html"
    form_class = ClassificationForm

    def get(self, request, *args, **kwargs):
        org = current_org_for_user(request.user)
        form = self.form_class(org=org)
        clas = Classification.objects.filter(org=org).prefetch_related('branches')
        context = {
            'form': form,
            'branch_form': BranchForm(org=org),
            'section_form': SectionForm(org=org),
            'course_form': CourseForm(org=org),
            'branches': Branch.objects.filter(org=org),
            'sections': Section.objects.filter(org=org).select_related('branch', 'classification'),
            'courses': Course.objects.filter(org=org).select_related('branch', 'teacher').prefetch_related('classifications', 'sections'),
            'clas': clas,
            'org': org,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        org = current_org_for_user(request.user)

        if 'add_branch' in request.POST:
            form = BranchForm(request.POST, org=org)
            if form.is_valid():
                branch = form.save(commit=False)
                branch.org = org
                branch.save()
                messages.success(request, "Successfully added Branch")
            else:
                messages.error(request, "Failed adding Branch: " + form.errors.as_text())
            return HttpResponseRedirect(reverse('handle:addClassification'))

        if 'add_section' in request.POST:
            form = SectionForm(request.POST, org=org)
            if form.is_valid():
                section = form.save(commit=False)
                section.org = org
                if not section.branch and section.classification:
                    # Only auto-inherit when the classification maps to exactly
                    # one branch — an org-wide or multi-branch classification
                    # gives no single unambiguous branch to default to.
                    classification_branches = list(section.classification.branches.all())
                    if len(classification_branches) == 1:
                        section.branch = classification_branches[0]
                section.save()
                messages.success(request, "Successfully added Section")
            else:
                messages.error(request, "Failed adding Section: " + form.errors.as_text())
            return HttpResponseRedirect(reverse('handle:addClassification'))

        if 'add_course' in request.POST:
            from school.features import has_feature
            if not has_feature(org, 'courses'):
                messages.error(request, "Course Management isn't enabled for this organization.")
                return HttpResponseRedirect(reverse('handle:addClassification'))
            form = CourseForm(request.POST, org=org)
            if form.is_valid():
                course = form.save(commit=False)
                course.org = org
                course.save()
                form.save_m2m()
                messages.success(request, "Successfully added Course")
            else:
                messages.error(request, "Failed adding Course: " + form.errors.as_text())
            return HttpResponseRedirect(reverse('handle:addClassification'))

        form = self.form_class(request.POST, org=org)
        if form.is_valid():
            classi = form.save(commit=False)
            classi.org = org
            classi.save()
            form.save_m2m()
            messages.success(request, "Successfully added Classification")
            return HttpResponseRedirect(reverse('handle:addClassification'))
        else:
            messages.error(request, "Failed adding Classification: " + form.errors.as_text())
            return HttpResponseRedirect(reverse('handle:addClassification'))

class AddDevice(LoginRequiredMixin, View):
    form_class = DeviceForm
    template_name = 'handle/addDevice.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        org = current_org_for_user(request.user)
        dist = {
            'form':form,
            'org':org
        }
        return render(request, self.template_name, dist)

    def post(self, request, *args, **kwargs):
        auser = request.user
        org = current_org_for_user(auser)
        form = self.form_class(request.POST)
        if form.is_valid():
            de_f = form.save(commit=False)
            de_f.org = org
            de_f.save()
            messages.success(request,"Successfully Added Devices")
            return HttpResponseRedirect(reverse('handle:addDevice'))
        else:
            messages.error(request,"Failed Adding Devices")
            return HttpResponseRedirect(reverse('handle:addDevice'))


@login_required
def editDevice(request, id):
    template_name = 'handle/editDevice.html'
    org = current_org_for_user(request.user)
    devi = get_object_or_404(Device, id=id, org=org)

    if request.method == "POST":
        form = DeviceForm(request.POST, instance=devi)
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Edited Device")
            return HttpResponseRedirect(reverse('handle:editDevice', args=[devi.id]))
        messages.error(request, "Please correct the highlighted device fields.")
    else:
        form = DeviceForm(instance=devi)
    dist = {
        'form':form,
        'devi':devi,
        'org':org
    }

    return render(request, template_name, dist)

@login_required
def deleteDevice(request, id):
    if request.method != 'POST':
        return HttpResponseRedirect(reverse('schooladmin:dashboard'))
    org = current_org_for_user(request.user)
    devi = get_object_or_404(Device, id=id, org=org)
    devi.delete()
    messages.success(request, "Successfully Deleted Device")
    return HttpResponseRedirect(reverse('schooladmin:dashboard'))

@perm_required('can_edit_members')
def deleteMember(request, id):
    from school.hierarchy import get_accessible_members
    org = current_org_for_user(request.user)
    devi = get_object_or_404(get_accessible_members(request.user, org), id=id)
    name = devi.name
    action = request.POST.get('action') or request.GET.get('action') or 'dump'

    if action == 'dump':
        archive_member(devi, performed_by=request.user, reason=request.POST.get('reason', ''))
        messages.success(request, "Moved " + name + " to Dumped Members. All their records were archived and can be restored on rejoin.")
    else:
        try:
            devi.delete()
            messages.success(request, "Successfully Deleted Member " +name)
        except:
            messages.success(request, "There are attendance records of the user, the system will delete your user after deleting records which can take a while!")

    return HttpResponseRedirect(reverse('handle:memberReport'))


@perm_required('can_edit_members')
def rejoinMember(request, id):
    from school.hierarchy import get_accessible_members
    org = current_org_for_user(request.user)
    devi = get_object_or_404(get_accessible_members(request.user, org).filter(status='dumped'), id=id)

    restore_member(devi, performed_by=request.user)
    messages.success(request, devi.name + " has been rejoined — all archived records were restored.")

    return HttpResponseRedirect(reverse('handle:memberReport'))

@login_required(login_url='/login/')
@perm_required(('can_manage_branches', 'can_edit_members'))
def editClassification(request, id):
    org = current_org_for_user(request.user)
    clas = get_object_or_404(Classification, id=id, org=org)

    # Handle Bulk Transfer
    if request.method == 'POST' and 'transfer_members' in request.POST:
        target_class_id = request.POST.get('target_classification')
        selected_member_ids = request.POST.getlist('member_select')

        if target_class_id and selected_member_ids:
            target_class = get_object_or_404(Classification, id=target_class_id, org=org)
            member.objects.filter(id__in=selected_member_ids, org=org).exclude(status='dumped').update(classification=target_class)
            messages.success(request, f"Successfully transferred {len(selected_member_ids)} members to {target_class.name}")
        return HttpResponseRedirect(reverse('handle:editClassification', args=[clas.id]))

    # Handle Name Update
    elif request.method == 'POST' and 'update_name' in request.POST:
        form = ClassificationForm(request.POST, instance=clas, org=org)
        if form.is_valid():
            form.save()
            messages.success(request, "Successfully Updated Classification")
        else:
            messages.error(request, "Failed updating Classification: " + form.errors.as_text())
        return HttpResponseRedirect(reverse('handle:editClassification', args=[clas.id]))

    # GET Request
    form = ClassificationForm(instance=clas, org=org)
    dist = {
        'form': form,
        'org': org,
        'clas': clas,
        'all_classifications': Classification.objects.filter(org=org).exclude(id=clas.id)
    }
    return render(request, "handle/editClassification.html", dist)


@login_required(login_url='/login/')
@perm_required(('can_manage_branches', 'can_edit_members'))
def deleteClassification(request, id):
    if request.method != 'POST':
        return HttpResponse(status=405)
    org = current_org_for_user(request.user)
    clas = get_object_or_404(Classification, id=id, org=org)
    clas.delete()
    messages.success(request, "Successfully deleted classification")
    return HttpResponseRedirect(reverse('handle:addClassification'))


@login_required(login_url='/login/')
def default_shift_for_selection(request):
    """Suggested check-in/out time for a new member given the branch/classification
    picked so far in the Add Member wizard. See `resolve_default_shift` for priority."""
    org = current_org_for_user(request.user)
    branch = Branch.objects.filter(org=org, pk=request.GET.get('branch_id')).first() if request.GET.get('branch_id') else None
    classification = Classification.objects.filter(org=org, pk=request.GET.get('classification_id')).first() if request.GET.get('classification_id') else None
    start, end = resolve_default_shift(org, branch=branch, classification=classification)
    return JsonResponse({'shift_start_time': start.strftime('%H:%M'), 'shift_end_time': end.strftime('%H:%M')})


@login_required(login_url='/login/')
def classifications_by_branch(request):
    """Branch -> Classification cascade: classifications available to the
    given branch (or every org-wide + branch-scoped one when no branch is
    passed), scoped to what the requesting user is allowed to see."""
    from school.hierarchy import get_accessible_classifications

    org = current_org_for_user(request.user)
    branch_id = request.GET.get('branch_id')
    branch = Branch.objects.filter(org=org, pk=branch_id).first() if branch_id else None
    classifications = get_accessible_classifications(request.user, org, branch=branch)

    data = [{'id': item.id, 'name': item.name} for item in classifications.order_by('name')]
    return JsonResponse({'classifications': data})


@login_required(login_url='/login/')
def sections_by_classification(request):
    from school.hierarchy import get_accessible_sections

    org = current_org_for_user(request.user)
    classification_id = request.GET.get('classification_id')
    branch_id = request.GET.get('branch_id')
    classification = Classification.objects.filter(org=org, pk=classification_id).first() if classification_id else None
    branch = Branch.objects.filter(org=org, pk=branch_id).first() if branch_id else None
    sections = get_accessible_sections(request.user, org, classification=classification, branch=branch)

    data = [{'id': item.id, 'name': item.name} for item in sections.order_by('name')]
    return JsonResponse({'sections': data})


class BranchHierarchyTree(LoginRequiredMixin, PermRequiredMixin, View):
    """Organization -> Branch -> Classification -> Section -> Member, as an
    expand-on-demand tree. Each level is fetched lazily via `hierarchy_tree_children`
    so opening the page never loads the whole member list at once."""
    required_perm = ('can_manage_branches', 'can_view_branches', 'can_view_members')
    template_name = 'handle/branch_tree.html'

    def get(self, request, *args, **kwargs):
        org = current_org_for_user(request.user)
        return render(request, self.template_name, {'org': org})


@login_required(login_url='/login/')
def hierarchy_tree_children(request):
    """Single reusable lazy-load endpoint for BranchHierarchyTree — avoids a
    bespoke AJAX view per level. `level` is the kind of node being expanded
    (its *children* are returned); `parent_id` is that node's id; `branch_id`
    carries the ancestor branch context down through classification/section
    (a classification/section can be shared by many branches, so the branch
    that was actually expanded to get here has to travel with the request)."""
    from school.hierarchy import get_accessible_branches, get_accessible_classifications, get_accessible_sections, get_accessible_members

    org = current_org_for_user(request.user)
    level = request.GET.get('level')
    parent_id = request.GET.get('parent_id')
    branch_id = request.GET.get('branch_id')
    branch = Branch.objects.filter(org=org, pk=branch_id).first() if branch_id else None

    if level == 'branch':
        nodes = []
        for b in get_accessible_branches(request.user, org).order_by('name'):
            nodes.append({
                'id': b.id, 'type': 'classification', 'name': b.name, 'subtitle': b.code,
                'status': b.status,
                'counts': {
                    'classifications': get_accessible_classifications(request.user, org, branch=b).count(),
                    'sections': Section.objects.filter(org=org, branch=b, status='active').count(),
                    'members': member.objects.filter(org=org, branch=b).exclude(status='dumped').count(),
                },
                'branch_id': b.id,
            })
        return JsonResponse({'nodes': nodes})

    if level == 'classification' and branch:
        nodes = []
        for c in get_accessible_classifications(request.user, org, branch=branch).order_by('name'):
            nodes.append({
                'id': c.id, 'type': 'section', 'name': c.name,
                'subtitle': 'All branches' if not c.branches.exists() else None,
                'status': c.status,
                'counts': {
                    'sections': Section.objects.filter(org=org, classification=c, status='active').filter(
                        Q(branch=branch) | Q(branch__isnull=True)
                    ).count(),
                    'members': member.objects.filter(org=org, classification=c, branch=branch).exclude(status='dumped').count(),
                },
                'branch_id': branch.id,
            })
        return JsonResponse({'nodes': nodes})

    if level == 'section' and parent_id and branch:
        classification = Classification.objects.filter(org=org, pk=parent_id).first()
        if not classification:
            return JsonResponse({'nodes': []})
        sections = get_accessible_sections(request.user, org, classification=classification, branch=branch)
        nodes = []
        for s in sections.order_by('name'):
            nodes.append({
                'id': s.id, 'type': 'member', 'name': s.name, 'subtitle': s.code,
                'status': s.status,
                'counts': {
                    'members': member.objects.filter(org=org, section=s, branch=branch).exclude(status='dumped').count(),
                },
                'branch_id': branch.id,
            })
        return JsonResponse({'nodes': nodes})

    if level == 'member' and parent_id and branch:
        members = get_accessible_members(request.user, org, branch=branch).filter(
            section_id=parent_id,
        ).exclude(status='dumped').order_by('name')
        total = members.count()
        page = members[:50]
        nodes = [{
            'id': m.id, 'type': None, 'name': m.name,
            'subtitle': m.designation or m.get_member_type_display(),
            'status': m.status,
            'photo_url': m.photo.url if m.photo else None,
            'detail_url': reverse('handle:memberEdit', args=[m.id]),
        } for m in page]
        return JsonResponse({'nodes': nodes, 'total': total, 'truncated': total > 50})

    return JsonResponse({'nodes': []})


class OperationsCenter(LoginRequiredMixin, View):
    template_name = "handle/operations.html"

    def get(self, request, *args, **kwargs):
        org = current_org_for_user(request.user)
        if request.GET.get('export') == 'stock':
            return self.export_stock(org)
        if request.GET.get('export') == 'transactions':
            return self.export_transactions(org)
        if request.GET.get('export') == 'birthdays':
            return self.export_birthdays(org)
        return render(request, self.template_name, self.context_data(request, org))

    def post(self, request, *args, **kwargs):
        org = current_org_for_user(request.user)
        if not org:
            messages.error(request, "No organization found for this user.")
            return HttpResponseRedirect(reverse('handle:operations'))

        action = request.POST.get('action')

        if action == 'send_message':
            subject = request.POST.get('msg_subject', '').strip()
            body = request.POST.get('msg_body', '').strip()
            recipient_filter = request.POST.get('msg_recipient', 'all')
            send_email = request.POST.get('send_email') == 'on'
            if not subject or not body:
                messages.error(request, "Subject and message body are required.")
            elif recipient_filter == 'specific' and not request.POST.get('msg_member_id'):
                messages.error(request, "Choose a member to send this message to.")
            else:
                qs = member.objects.filter(org=org, email__isnull=False).exclude(status='dumped').exclude(email='')
                if recipient_filter == 'staff':
                    qs = qs.filter(make_staff=True)
                elif recipient_filter == 'student':
                    qs = qs.filter(member_type='student')
                elif recipient_filter == 'employee':
                    qs = qs.filter(member_type='employee')
                elif recipient_filter == 'specific':
                    qs = qs.filter(pk=request.POST.get('msg_member_id'))
                if send_email:
                    email_list = list(qs.values_list('email', flat=True))
                    if email_list:
                        from school.email_utils import send_broadcast_message_email
                        send_broadcast_message_email(email_list, subject, body, org.name, org=org)
                        messages.success(
                            request,
                            f"Message queued for {len(email_list)} member(s) — check Super Admin's Email Logs "
                            "for delivery status, since sending happens in the background.",
                        )
                    else:
                        messages.warning(request, "No recipients in this group have an email address on file.")
                else:
                    messages.success(request, f"Message noted for {qs.count()} member(s) (email delivery was not selected).")
            return HttpResponseRedirect(reverse('handle:operations') + '?tab=messages')

        form_map = {
            'stock_category': StockCategoryForm,
            'stock_item': StockItemForm,
            'stock_movement': StockMovementForm,
            'transaction_category': TransactionCategoryForm,
            'financial_transaction': FinancialTransactionForm,
        }
        form_class = form_map.get(action)
        if not form_class:
            messages.error(request, "Invalid operations action.")
            return HttpResponseRedirect(reverse('handle:operations'))

        from school.features import has_feature
        required_feature = 'stock' if action.startswith('stock_') else 'finance'
        if not has_feature(org, required_feature):
            messages.error(request, f"The '{required_feature}' module isn't enabled for this organization.")
            return HttpResponseRedirect(reverse('handle:operations'))

        if action == 'financial_transaction':
            form = form_class(request.POST, request.FILES, org=org)
        elif action in ('stock_item', 'stock_movement'):
            form = form_class(request.POST, org=org)
        else:
            form = form_class(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.org = org
            if hasattr(obj, 'created_by_id'):
                obj.created_by = request.user
            if hasattr(obj, 'branch') and not obj.branch and action == 'stock_movement':
                obj.branch = obj.item.branch
            obj.save()
            messages.success(request, "Operations record saved successfully.")
        else:
            messages.error(request, "Could not save record: " + form.errors.as_text())
        return HttpResponseRedirect(reverse('handle:operations'))

    def context_data(self, request, org):
        branch_id = request.GET.get('branch') or ''
        category_id = request.GET.get('category') or ''
        tx_type = request.GET.get('transaction_type') or ''
        start_date = request.GET.get('start_date') or ''
        end_date = request.GET.get('end_date') or ''

        stock_items = StockItem.objects.filter(org=org).select_related('branch', 'category')
        movements = StockMovement.objects.filter(org=org).select_related('branch', 'item', 'created_by')
        transactions = FinancialTransaction.objects.filter(org=org).select_related('branch', 'category', 'created_by')
        birthdays = member.objects.filter(org=org, date_of_birth__isnull=False).exclude(status='dumped').select_related('branch', 'classification', 'section')

        if branch_id:
            stock_items = stock_items.filter(branch_id=branch_id)
            movements = movements.filter(branch_id=branch_id)
            transactions = transactions.filter(branch_id=branch_id)
            birthdays = birthdays.filter(branch_id=branch_id)
        if category_id:
            stock_items = stock_items.filter(category_id=category_id)
        if tx_type:
            transactions = transactions.filter(transaction_type=tx_type)
        if start_date:
            movements = movements.filter(movement_date__gte=start_date)
            transactions = transactions.filter(transaction_date__gte=start_date)
        if end_date:
            movements = movements.filter(movement_date__lte=end_date)
            transactions = transactions.filter(transaction_date__lte=end_date)

        today = timezone.localdate()
        next_30 = today + timedelta(days=30)
        birthday_pool = list(birthdays)
        todays_birthdays = [m for m in birthday_pool if m.date_of_birth.month == today.month and m.date_of_birth.day == today.day]
        upcoming_birthdays = [
            m for m in birthday_pool
            if today < self.next_birthday_date(m.date_of_birth, today) <= next_30
        ]
        upcoming_birthdays.sort(key=lambda m: self.next_birthday_date(m.date_of_birth, today))

        income_total = transactions.filter(transaction_type='income').aggregate(total=Sum('amount'))['total'] or 0
        expense_total = transactions.filter(transaction_type='expense').aggregate(total=Sum('amount'))['total'] or 0

        return {
            'org': org,
            'branches': Branch.objects.filter(org=org, status='active'),
            'stock_categories': StockCategory.objects.filter(org=org),
            'transaction_categories': TransactionCategory.objects.filter(org=org),
            'stock_category_form': StockCategoryForm(),
            'stock_item_form': StockItemForm(org=org),
            'stock_movement_form': StockMovementForm(org=org),
            'transaction_category_form': TransactionCategoryForm(),
            'financial_transaction_form': FinancialTransactionForm(org=org),
            'stock_items': stock_items,
            'stock_movements': movements[:25],
            'transactions': transactions[:25],
            'low_stock_items': [item for item in stock_items if item.is_low_stock],
            'income_total': income_total,
            'expense_total': expense_total,
            'net_total': income_total - expense_total,
            'todays_birthdays': todays_birthdays,
            'upcoming_birthdays': upcoming_birthdays[:20],
            'total_members_count': member.objects.filter(org=org).exclude(status='dumped').count(),
            'members_with_email': member.objects.filter(org=org, email__isnull=False).exclude(status='dumped').exclude(email='').count(),
            'staff_user_count': member.objects.filter(org=org, make_staff=True).exclude(status='dumped').count(),
            'messageable_members': member.objects.filter(org=org, email__isnull=False).exclude(status='dumped').exclude(email='').order_by('name'),
            'filters': {
                'branch': branch_id,
                'category': category_id,
                'transaction_type': tx_type,
                'start_date': start_date,
                'end_date': end_date,
            }
        }

    def next_birthday_date(self, birthday, today):
        try:
            candidate = birthday.replace(year=today.year)
        except ValueError:
            candidate = birthday.replace(year=today.year, day=28)
        if candidate < today:
            try:
                candidate = birthday.replace(year=today.year + 1)
            except ValueError:
                candidate = birthday.replace(year=today.year + 1, day=28)
        return candidate

    def export_stock(self, org):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="stock-report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Item', 'Branch', 'Category', 'SKU', 'Quantity', 'Unit', 'Low Stock Threshold', 'Supplier', 'Purchase Cost', 'Status'])
        for item in StockItem.objects.filter(org=org).select_related('branch', 'category'):
            writer.writerow([
                item.name, item.branch or '', item.category or '', item.sku or '',
                item.quantity, item.unit, item.low_stock_threshold,
                item.supplier or '', item.purchase_cost, item.status
            ])
        return response

    def export_transactions(self, org):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transactions-report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Date', 'Type', 'Title', 'Category', 'Branch', 'Amount', 'Payment Method', 'Reference'])
        for item in FinancialTransaction.objects.filter(org=org).select_related('branch', 'category'):
            writer.writerow([
                item.transaction_date, item.transaction_type, item.title,
                item.category or '', item.branch or '', item.amount,
                item.payment_method, item.reference_number or ''
            ])
        return response

    def export_birthdays(self, org):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="birthday-report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Birthday', 'Branch', 'Classification', 'Section', 'Phone', 'Email'])
        for item in member.objects.filter(org=org, date_of_birth__isnull=False).exclude(status='dumped').select_related('branch', 'classification', 'section'):
            writer.writerow([
                item.name, item.date_of_birth, item.branch or '',
                item.classification or '', item.section or '', item.phone or '', item.email or ''
            ])
        return response



class Search(LoginRequiredMixin, View):
    template_name = 'handle/search_result.html'
    model = member

    def get(self, request, *args, **kwargs):
        org = current_org_for_user(request.user)
        name = request.GET.get('member', '')
        memb = member.objects.filter(Q(name__icontains=name) | Q(card=name)).filter(org=org).exclude(status='dumped')
        dist = {
            'object_list':memb,
            'org':org
        }
        return render(request, self.template_name, dist)


@login_required
def member_search_suggestions(request):
    """Fast organization-scoped member results for the global navbar search."""
    from school.features import has_perm
    org = current_org_for_user(request.user)
    query = request.GET.get('q', '').strip()
    if not org or len(query) < 2:
        return JsonResponse({'members': []})
    qs = member.objects.filter(org=org).exclude(status='dumped').filter(
        Q(name__icontains=query) | Q(card__icontains=query)
        | Q(roll_number__icontains=query) | Q(email__icontains=query)
        | Q(phone__icontains=query)
    ).select_related('classification', 'section', 'branch').order_by('name')[:8]
    can_edit = request.user.user_type == '2' or has_perm(request.user, 'can_edit_members')
    return JsonResponse({'members': [{
        'id': item.id,
        'name': item.name,
        'meta': ' · '.join(filter(None, [
            item.get_member_type_display(), item.card,
            item.classification.name if item.classification else None,
            item.section.name if item.section else None,
        ])),
        'profile_url': reverse('schooladmin:member_profile', args=[item.id]),
        'edit_url': reverse('handle:memberEdit', args=[item.id]) if can_edit else '',
    } for item in qs]})

class MemberReport(PermRequiredMixin, View):
    required_perm = 'can_view_members'
    template_name = 'handle/member_Report.html'

    def get(self, request, *args, **kwargs):
        from django.db.models import Prefetch
        from school.hierarchy import get_accessible_members, get_accessible_branches
        auser = request.user
        org = current_org_for_user(auser)

        memb = get_accessible_members(auser, org).exclude(status='dumped').select_related(
            'branch', 'classification', 'section',
        ).order_by('-id')
        dumped_members = get_accessible_members(auser, org).filter(status='dumped').order_by('-updated_date').prefetch_related(
            Prefetch('change_history', queryset=MemberHistory.objects.select_related('changed_by')),
        )

        dist = {
            'mem':memb,
            'clas':Classification.objects.filter(org=org, status='active'),
            'branches': get_accessible_branches(auser, org),
            'sections': Section.objects.filter(org=org, status='active'),
            'thisone':'All',
            'selected_branch': 'All',
            'selected_section': 'All',
            'dumped_members': dumped_members,
            'org':org
        }
        return render(request, self.template_name, dist)

    def post(self, request, *agrs, **kwargs):
        from school.hierarchy import get_accessible_members, get_accessible_branches
        clas = request.POST.get('classification', 'All')
        branch = request.POST.get('branch', 'All')
        section = request.POST.get('section', 'All')
        auser = request.user
        org = current_org_for_user(auser)

        from django.db.models import Prefetch
        cl=Classification.objects.filter(org=org, status='active')
        mem = get_accessible_members(auser, org).exclude(status='dumped').select_related(
            'branch', 'classification', 'section',
        )
        if branch != 'All':
            mem = mem.filter(branch_id=branch)
        if section != 'All':
            mem = mem.filter(section_id=section)
        if clas == 'All':
            th = 'All'
        else:
            mem = mem.filter(classification=clas)
            th = Classification.objects.filter(id=clas, org=org).values_list('name', flat=True).first() or 'Unknown'
        dist = {
            'mem':mem,
            'clas':cl,
            'branches': get_accessible_branches(auser, org),
            'sections': Section.objects.filter(org=org, status='active'),
            'thisone': th,
            'selected_branch': branch,
            'selected_section': section,
            'dumped_members': get_accessible_members(auser, org).filter(status='dumped').order_by('-updated_date').prefetch_related(
                Prefetch('change_history', queryset=MemberHistory.objects.select_related('changed_by')),
            ),
            'org':org
        }
        return render(request, self.template_name, dist)
    
from django.contrib.auth import update_session_auth_hash

@login_required(login_url='/')
def changePassword(request):
    if request.method == 'POST':
        form = FormChangePassword(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('handle:changePassword')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = FormChangePassword(request.user)
    return render(request, 'handle/changePassword.html', {
        'form': form,
    })


# ─── Bulk Member Import (Excel / CSV) ────────────────────────────────────────────

MEMBER_IMPORT_HEADERS = [
    'name', 'member_type', 'gender', 'phone', 'email', 'address',
    'card', 'roll_number', 'designation', 'admission_date', 'salary_type', 'salary_amount',
    'branch', 'classification', 'section', 'dashboard_access',
]


MEMBER_IMPORT_DISPLAY_HEADERS = [
    'Name', 'Member Type', 'Gender', 'Phone', 'Email', 'Address',
    'Card / RFID', 'Roll Number', 'Designation', 'Joining Date', 'Salary Type', 'Salary Amount',
    'Branch', 'Classification', 'Section', 'Dashboard Access',
]

MEMBER_TYPE_OPTIONS = 'student / employee / staff / intern / trainee / teacher / worker / member'
SALARY_TYPE_OPTIONS = 'monthly / daily / hourly / weekly'

# Column headers are matched after stripping everything but letters/digits
# (so "Card / RFID", "card-rfid", "Card_RFID" etc. all normalize the same
# way) and then resolved through this alias table onto the canonical field
# name the importer actually understands. This is what makes the downloaded
# sample template (which shows the friendly display headers, not the raw
# field names) — and hand-edited variants of it — parse correctly instead of
# silently dropping columns like "Card / RFID" into a key nothing matches.
MEMBER_IMPORT_HEADER_ALIASES = {
    'card_rfid': 'card', 'rfid': 'card', 'card_number': 'card', 'card_no': 'card', 'card_id': 'card',
    'employee_id': 'roll_number', 'emp_id': 'roll_number', 'staff_id': 'roll_number',
    'joining_date': 'admission_date', 'join_date': 'admission_date', 'date_of_joining': 'admission_date',
    'member_type': 'member_type', 'type': 'member_type',
    'dashboard_access': 'dashboard_access', 'portal_access': 'dashboard_access', 'login_access': 'dashboard_access',
}


def _normalize_import_header(raw):
    key = re.sub(r'[^a-z0-9]+', '_', str(raw or '').strip().lower()).strip('_')
    return MEMBER_IMPORT_HEADER_ALIASES.get(key, key)


def _parse_import_date(val):
    if val in (None, ''):
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    text = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_import_bool(val):
    return str(val or '').strip().lower() in ('yes', 'y', 'true', '1')


class MemberImportDuplicateRow(Exception):
    """Raised for a row that collides with an existing member (card/email) —
    reported to the user as a distinct 'duplicate' bucket, not a generic error."""


def member_import_sample(request):
    """Download a styled Excel template with column titles, hints row, and one example row."""
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    # Row 1: column titles (display names, Title Case) — these are what the parser reads
    header_row = ws.append(MEMBER_IMPORT_HEADERS)  # parser needs lowercase/underscore names
    header_fill = PatternFill('solid', fgColor='1E40AF')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col_idx in range(1, len(MEMBER_IMPORT_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        # show friendly display name to user but keep raw header value for parser
        cell.value = MEMBER_IMPORT_DISPLAY_HEADERS[col_idx - 1]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 2: hint / allowed-values row
    hints = [
        'Required', MEMBER_TYPE_OPTIONS, 'Male / Female',
        'Numbers only', 'Email address', 'City or full address',
        'Card / RFID ID (unique in org)', 'Roll / registration no.', 'Job title / role', 'YYYY-MM-DD (optional)',
        SALARY_TYPE_OPTIONS, 'Number (e.g. 30000)',
        'Branch name (optional)', 'Class / Dept name', 'Section (optional)',
        'Yes / No — creates login + emails credentials (needs Email + Phone)',
    ]
    ws.append(hints)
    hint_fill = PatternFill('solid', fgColor='DBEAFE')
    hint_font = Font(italic=True, color='1E3A8A', size=9)
    for col_idx in range(1, len(hints) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = hint_fill
        cell.font = hint_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Row 3: example data
    ws.append([
        'John Doe', 'staff', 'Male', '9800000000', 'john@example.com', 'Kathmandu',
        'CARD-001', 'R-001', 'Cashier', '2024-01-15', 'monthly', '30000', '', '', '', 'No',
    ])
    example_font = Font(color='374151', size=10)
    for col_idx in range(1, len(MEMBER_IMPORT_HEADERS) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = example_font

    # Column widths
    col_widths = [18, 24, 12, 16, 24, 20, 16, 14, 16, 16, 16, 16, 16, 20, 14, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 40

    ws.freeze_panes = 'A3'

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="member_import_sample.xlsx"'
    wb.save(resp)
    return resp


class MemberImport(View):
    template_name = 'handle/member_import.html'

    def _org(self, request):
        if request.user.user_type == '2' and hasattr(request.user, 'schooladmin'):
            return request.user.schooladmin.org
        return None

    def get(self, request, *args, **kwargs):
        org = self._org(request)
        total_member = member.objects.filter(org=org).exclude(status='dumped').count() if org else 0
        return render(request, self.template_name, {
            'org': org, 'headers': MEMBER_IMPORT_HEADERS, 'total_member': total_member,
        })

    def post(self, request, *args, **kwargs):
        org = self._org(request)
        if org is None:
            messages.error(request, "Only organization admins can import members.")
            return redirect('handle:addMember')

        f = request.FILES.get('file')
        if not f:
            messages.error(request, "Please choose a file to upload.")
            return redirect('handle:member_import')

        # Parse rows into list of dicts keyed by canonical header (see
        # _normalize_import_header — resolves display-header variants like
        # "Card / RFID" back onto the field name the importer understands).
        rows = []
        fname = f.name.lower()
        try:
            if fname.endswith('.csv'):
                import io
                text = io.TextIOWrapper(f.file, encoding='utf-8', errors='ignore')
                reader = csv.DictReader(text)
                for r in reader:
                    rows.append({_normalize_import_header(k): v for k, v in r.items()})
            else:
                import openpyxl
                wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
                ws = wb.active
                it = ws.iter_rows(values_only=True)
                headers = [_normalize_import_header(h) if h is not None else '' for h in next(it)]
                # auto-skip the hints row (first cell == "Required") from the sample template
                first_data = next(it, None)
                if first_data and str(first_data[0] or '').strip().lower() == 'required':
                    first_data = next(it, None)  # consume hint row, advance to real data
                if first_data is not None:
                    rows.append({headers[i]: first_data[i] for i in range(min(len(headers), len(first_data)))})
                for raw in it:
                    if raw is None:
                        continue
                    row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
                    rows.append(row)
        except Exception as e:
            messages.error(request, f"Could not read the file: {e}")
            return redirect('handle:member_import')

        valid_types = {c[0] for c in member.MEMBER_TYPE_CHOICES}
        valid_salary = {c[0] for c in member.SALARY_CHOICES}
        created, duplicates, invalid, skipped_for_limit, accounts_created = 0, [], [], 0, 0
        current_count = member.objects.filter(org=org).exclude(status='dumped').count()

        def lookup(model, name_val):
            if not name_val:
                return None
            return model.objects.filter(org=org, name__iexact=str(name_val).strip()).first()

        for i, row in enumerate(rows, start=2):
            name = (str(row.get('name') or '')).strip()
            if not name:
                continue  # skip blank lines
            if current_count + created >= org.member_limit:
                skipped_for_limit += 1
                continue
            try:
                mtype = (str(row.get('member_type') or 'member')).strip().lower()
                if mtype not in valid_types:
                    mtype = 'member'
                gender = (str(row.get('gender') or 'Male')).strip().title()
                if gender not in ('Male', 'Female'):
                    gender = 'Male'
                stype = (str(row.get('salary_type') or 'monthly')).strip().lower()
                if stype not in valid_salary:
                    stype = 'monthly'

                salary_amount = row.get('salary_amount')
                try:
                    salary_amount = float(salary_amount) if salary_amount not in (None, '') else 0
                except (TypeError, ValueError):
                    salary_amount = 0

                raw_phone = row.get('phone')
                phone = None
                if raw_phone not in (None, ''):
                    digits = ''.join(ch for ch in str(raw_phone) if ch.isdigit())
                    phone = int(digits) if digits else None

                device_id = int(member.objects.filter(org=org).exclude(status='dumped').count()) + 1 + created
                while member.objects.filter(org=org, device_id=device_id).exclude(status='dumped').exists():
                    device_id += 1

                card = str(row.get('card') or '').strip() or None
                if card and member.objects.filter(org=org, card=card).exclude(status='dumped').exists():
                    raise MemberImportDuplicateRow(f"Card / RFID '{card}' is already assigned in this organization")

                email_val = str(row.get('email')).strip() if row.get('email') else None
                if email_val and member.objects.filter(email=email_val).exists():
                    raise MemberImportDuplicateRow(f"Email '{email_val}' is already in use")

                resolved_branch = lookup(Branch, row.get('branch'))
                resolved_classification = lookup(Classification, row.get('classification'))
                shift_start, shift_end = resolve_default_shift(org, branch=resolved_branch, classification=resolved_classification)

                with transaction.atomic():
                    new_member = member.objects.create(
                        org=org, name=name, member_type=mtype, gender=gender,
                        status='active', device_id=device_id,
                        card=card,
                        phone=phone,
                        email=email_val,
                        address=(str(row.get('address')).strip() if row.get('address') else ''),
                        roll_number=(str(row.get('roll_number')).strip() if row.get('roll_number') else None),
                        designation=(str(row.get('designation')).strip() if row.get('designation') else None),
                        admission_date=_parse_import_date(row.get('admission_date')),
                        salary_type=stype, salary_amount=salary_amount,
                        branch=resolved_branch,
                        classification=resolved_classification,
                        section=lookup(Section, row.get('section')),
                        shift_start_time=shift_start,
                        shift_end_time=shift_end,
                    )
                    created += 1

                    staff_result = None
                    if _parse_import_bool(row.get('dashboard_access')):
                        staff_result = provision_staff_account(new_member, org)

                if staff_result is not None:
                    if staff_result.status == 'created':
                        from school.email_utils import send_welcome_email
                        send_welcome_email(new_member.email, new_member.name, staff_result.password, org.name, org=org)
                        accounts_created += 1
                    else:
                        invalid.append(f"Row {i} ({name}): imported, but dashboard access wasn't granted — {staff_result.message}")
            except MemberImportDuplicateRow as e:
                duplicates.append(f"Row {i} ({name}): {e}")
            except IntegrityError as e:
                duplicates.append(f"Row {i} ({name}): {e}")
            except Exception as e:
                invalid.append(f"Row {i} ({name}): {e}")

        if created:
            msg = f"Imported {created} member{'s' if created != 1 else ''} successfully."
            if accounts_created:
                msg += f" {accounts_created} dashboard account{'s' if accounts_created != 1 else ''} created and emailed."
            messages.success(request, msg)
        if duplicates:
            messages.warning(request, f"{len(duplicates)} row(s) skipped as duplicates: " + " | ".join(duplicates[:8]))
        if invalid:
            messages.error(request, f"{len(invalid)} row(s) invalid: " + " | ".join(invalid[:8]))
        if skipped_for_limit:
            messages.error(request, f"{skipped_for_limit} row(s) were not imported — organization member limit ({org.member_limit}) reached.")
        if not created and not duplicates and not invalid and not skipped_for_limit:
            messages.error(request, "No valid rows found. Make sure the 'name' column is filled.")
        return redirect('handle:member_import')
